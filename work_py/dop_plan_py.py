import json
import copy
from openpyxl.styles import Font, Alignment
import data_list

from PyQt5.Qt import *
from PyQt5.QtGui import QIntValidator, QDoubleValidator
from PyQt5.QtWidgets import QMessageBox, QWidget, QLabel, QComboBox, QLineEdit, QGridLayout, \
    QPushButton, QTextEdit, QDateEdit
from PyQt5.QtCore import Qt, pyqtSignal
from datetime import datetime

from data_base.config_base import connection_to_database, WorkDatabaseWell
from decrypt import decrypt
from find import FindIndexPZ
from log_files.log import logger
from server_response import ApiClient

from work_py.advanted_file import merge_overlapping_intervals
from work_py.parent_work import TabPageUnion, TabWidgetUnion, WindowUnion


class TabPageDp(TabPageUnion):
    def __init__(self, data_well, tableWidget, old_index, parent=None):
        super().__init__(data_well)
        try:
            self.tableWidget = tableWidget
            self.old_index = old_index

            self.validator_int = QIntValidator(0, 8000)
            self.validator_float = QDoubleValidator(0, 8000, 1)
            self.work_plan = data_well.work_plan

            self.well_number_label = QLabel('номер скважины')
            self.well_number_edit = QLineEdit(self)
            # self.well_number_edit.setValidator(self.validator_int)

            self.well_area_label = QLabel('площадь скважины')
            self.well_area_edit = QLineEdit(self)
            self.well_area_edit.textChanged[str].connect(self.update_well_area)

            self.number_DP_label = QLabel('номер \nдополнительного плана')
            self.number_DP_Combo = QComboBox(self)

            self.number_DP_Combo.addItems(["", '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'])
            if self.data_well.number_dp != 0:
                self.number_DP_Combo.setCurrentIndex(int(self.data_well.number_dp) - 1)

            self.current_bottom_label = QLabel('Забой текущий')
            self.current_bottom_edit = QLineEdit(self)
            self.current_bottom_edit.setValidator(self.validator_float)
            # self.current_bottom_edit.setText(f'{self.data_well.current_bottom}')

            self.current_bottom_date_label = QLabel('дата определения забоя')
            self.current_bottom_date_edit = QDateEdit(self)
            self.current_bottom_date_edit.setDisplayFormat("dd.MM.yyyy")
            self.current_bottom_date_edit.setDate(datetime.now())

            self.method_bottom_label = QLabel('Метод определения \nзабоя')
            self.method_bottom_combo = QComboBox(self)
            self.method_bottom_combo.addItems(['', 'ГИС', 'НКТ'])

            self.template_depth_label = QLabel('Глубина \nшаблонирования ЭК')
            self.template_depth_edit = QLineEdit(self)
            self.template_depth_edit.setValidator(self.validator_float)
            # self.template_depth_edit.setText(str(self.data_well.template_depth))

            self.template_length_label = QLabel('Длина шаблона')
            self.template_length_edit = QLineEdit(self)
            self.template_length_edit.setValidator(self.validator_float)

            self.change_pvr_combo_label = QLabel('Были ли изменения \nв интервале перфорации')
            self.change_pvr_combo = QComboBox(self)
            self.change_pvr_combo.addItems(['', 'Нет', 'Да'])

            self.skm_interval_label = QLabel('интервалы \nскреперования')
            self.skm_interval_edit = QLineEdit(self)

            self.raiding_interval_label = QLabel('интервалы \n райбирования')
            self.raiding_interval_edit = QLineEdit(self)

            self.table_name = ''

            self.fluid_label = QLabel("уд.вес жидкости глушения", self)
            self.fluid_edit = QTextEdit(self)
            self.fluid_edit.setAlignment(Qt.AlignLeft)

            self.plast_label = QLabel("пласта")
            self.plast_line = QLineEdit(self)
            self.vertical_label = QLabel("по вертикале")
            self.vertical_line = QLineEdit(self)
            self.roof_label = QLabel("Кровля")
            self.roof_edit = QLineEdit(self)
            self.sole_label = QLabel("Подошва")
            self.sole_edit = QLineEdit(self)
            self.date_pvr_label = QLabel("Дата")
            self.date_pvr_edit = QDateEdit(self)
            self.date_pvr_edit.setDisplayFormat("dd.MM.yyyy")
            self.count_pvr_label = QLabel("Кол-во отв")
            self.count_pvr_edit = QLineEdit(self)
            self.type_pvr_label = QLabel("Тип перфоратора")
            self.type_pvr_edit = QLineEdit(self)
            self.pressure_pvr_label = QLabel("Давление")
            self.pressure_pvr_edit = QLineEdit(self)
            self.date_pressure_label = QLabel("Дата замера")
            self.date_pressure_edit = QDateEdit(self)
            self.date_pressure_edit.setDisplayFormat("dd.MM.yyyy")

            # if self.data_well.fluid_work == '':
            #     self.fluid_edit.setText(f'{TabPageGno.calc_fluid(self.work_plan, self.data_well.current_bottom)}')
            # else:
            #     self.fluid_edit.setText(f'{self.data_well.fluid_work}')

            self.work_label = QLabel("Ранее проведенные работы:", self)
            self.work_edit = QTextEdit(self)

            # self.work_edit.setFixedWidth(300)
            self.work_edit.setAlignment(Qt.AlignLeft)

            # self.grid = QGridLayout(self)

            self.grid.addWidget(self.well_number_label, 2, 1)
            self.grid.addWidget(self.well_number_edit, 3, 1)

            self.grid.addWidget(self.well_area_label, 2, 2)
            self.grid.addWidget(self.well_area_edit, 3, 2)

            self.grid.addWidget(self.well_area_label, 2, 3)
            self.grid.addWidget(self.well_area_edit, 3, 3)

            self.grid.addWidget(self.number_DP_label, 2, 4)
            self.grid.addWidget(self.number_DP_Combo, 3, 4)

            self.grid.addWidget(self.current_bottom_label, 4, 1)
            self.grid.addWidget(self.current_bottom_edit, 5, 1)
            self.grid.addWidget(self.current_bottom_date_label, 4, 2)
            self.grid.addWidget(self.current_bottom_date_edit, 5, 2)
            self.grid.addWidget(self.method_bottom_label, 4, 3)
            self.grid.addWidget(self.method_bottom_combo, 5, 3)

            self.grid.addWidget(self.fluid_label, 4, 4, 1, 6)
            self.grid.addWidget(self.fluid_edit, 5, 4, 1, 6)
            self.grid.addWidget(self.template_depth_label, 6, 1)
            self.grid.addWidget(self.template_depth_edit, 7, 1)
            self.grid.addWidget(self.template_length_label, 6, 2)
            self.grid.addWidget(self.template_length_edit, 7, 2)
            self.grid.addWidget(self.skm_interval_label, 8, 1, 1, 2)
            self.grid.addWidget(self.skm_interval_edit, 9, 1, 1, 2)
            self.grid.addWidget(self.raiding_interval_label, 8, 3, 1, 3)
            self.grid.addWidget(self.raiding_interval_edit, 9, 3, 1, 3)
            self.grid.addWidget(self.change_pvr_combo_label, 10, 1)
            self.grid.addWidget(self.change_pvr_combo, 11, 1)

            self.grid.addWidget(self.work_label, 25, 1)
            self.grid.addWidget(self.work_edit, 26, 1, 2, 4)

            self.well_number_edit.editingFinished.connect(self.update_well)
            try:
                self.well_area_edit.setText(f'{self.data_well.well_area.get_value}')
                self.well_number_edit.setText(f'{self.data_well.well_number.get_value}')
                self.well_area_edit.setEnabled(False)
                # self.well_number_edit.setEnabled(False)
            except Exception:
                pass
            if self.work_plan in 'dop_plan':
                if self.data_well.column_additional:
                    self.template_depth_addition_label = QLabel('Глубина спуска шаблона в доп колонне')
                    self.template_depth_addition_edit = QLineEdit(self)
                    self.template_depth_addition_edit.setValidator(self.validator_float)
                    self.template_depth_addition_edit.setText(str(self.data_well.template_depth_addition))

                    self.template_length_addition_label = QLabel('Длина шаблона в доп колонне')
                    self.template_length_addition_edit = QLineEdit(self)
                    self.template_length_addition_edit.setValidator(self.validator_float)
                    self.template_length_addition_edit.setText(str(self.data_well.template_length_addition))
                    self.grid.addWidget(self.template_depth_addition_label, 6, 4)
                    self.grid.addWidget(self.template_depth_addition_edit, 7, 4)
                    self.grid.addWidget(self.template_length_addition_label, 6, 5)
                    self.grid.addWidget(self.template_length_addition_edit, 7, 5)

            # self.well_area_edit.textChanged[str].connect(self.update_well)

            self.change_pvr_combo.currentTextChanged.connect(self.update_change_pvr)
            self.change_pvr_combo.setCurrentIndex(1)
            self.change_pvr_combo.setCurrentIndex(0)

            # Виджет выбора файла из базы должен существовать всегда: его используют в add_work/add_perforation_project.
            self.well_data_label = QLabel('файл excel в базе')
            self.well_data_in_base_combo = QComboBox()
            self.well_data_in_base_combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)
            self.well_data_in_base_combo.currentTextChanged.connect(self.update_well_data_in_base_combo)

            if data_list.data_in_base:
                # self.table_in_base_label = QLabel('данные по скважине')
                # self.table_in_base_combo = QComboBox()
                # self.table_in_base_combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)

                self.index_change_label = QLabel('пункт после которого происходят изменения')
                self.index_change_line = QLineEdit(self)
                self.index_change_line.setValidator(self.validator_int)

                self.grid.addWidget(self.well_data_label, 2, 6)
                self.grid.addWidget(self.well_data_in_base_combo, 3, 6)
                self.grid.addWidget(self.index_change_label, 2, 7)
                self.grid.addWidget(self.index_change_line, 3, 7)

                self.index_change_line.editingFinished.connect(self.update_table_in_base_combo)
        except Exception as e:
            # Главное: не падать всей вкладкой при проблемах с данными/виджетами.
            logger.exception("TabPageDp initialization failed")
            try:
                QMessageBox.critical(
                    self,
                    'Ошибка',
                    f'Ошибка инициализации вкладки доп. плана работ: {e}'
                )
            except Exception:
                pass
            if hasattr(self, "tableWidget"):
                try:
                    self.tableWidget.hide()
                except Exception:
                    pass

    def update_well_data_in_base_combo(self, index):
        try:
            if index:
                if index.split(' ')[3] not in ['ПР', "ПРС", "ПРизм"]:
                    number_dp_in_base = [num for num in index.split(' ')[3] if num.isdigit()]
                    if number_dp_in_base:
                        self.number_DP_Combo.setCurrentIndex(int(number_dp_in_base[0]))
                    else:
                        QMessageBox.warning(
                            self,
                            'Внимание',
                            'В данных из базы не найден номер дополнительного плана. Нужно выбрать номер дополнительного плана вручную.'
                        )
                        self.number_DP_Combo.setCurrentIndex(0)
        except Exception as e:
            logger.exception("update_well_data_in_base_combo failed")
            try:
                QMessageBox.warning(
                    self,
                    'Ошибка',
                    f'Не удалось обновить номер доп. плана из базы: {e}'
                )
            except Exception:
                pass

    def update_change_pvr(self, index):
        try:
            if self.old_index == 0:
                self.tableWidget.setHorizontalHeaderLabels(
                    ["Пласт/\nГоризонт", "по вертикали", "кровля,\n м", "подошва, \nм",
                     "Дата\n вскрытия", "Дата \nотключения", "кол-во \nотвер", "Тип \nперф", 'Удлинение,\nм',
                     "Рпл, \nатм", "Дата \nзамера"])
                for i in range(11):
                    self.tableWidget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)
            else:
                self.tableWidget.setHorizontalHeaderLabels(
                    ["Пласт/\nГоризонт", "по вертикали", "кровля,\n м", "подошва, \nм",
                     "Дата\n вскрытия\отключения", "кол-во \nотвер", "Тип \nперф", 'Удлинение,\nм',
                     "Рпл, \nатм", "Дата \nзамера"])
                for i in range(10):
                    self.tableWidget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)

            if index == 'Да':
                self.grid.addWidget(self.plast_label, 12, 1)
                self.grid.addWidget(self.plast_line, 13, 1)
                self.grid.addWidget(self.vertical_label, 12, 2)
                self.grid.addWidget(self.vertical_line, 13, 2)
                self.grid.addWidget(self.roof_label, 12, 3)
                self.grid.addWidget(self.roof_edit, 13, 3)
                self.grid.addWidget(self.sole_label, 12, 4)
                self.grid.addWidget(self.sole_edit, 13, 4)
                self.grid.addWidget(self.date_pvr_label, 12, 5)
                self.grid.addWidget(self.date_pvr_edit, 13, 5)
                self.grid.addWidget(self.count_pvr_label, 12, 6)
                self.grid.addWidget(self.count_pvr_edit, 13, 6)
                self.grid.addWidget(self.type_pvr_label, 12, 7)
                self.grid.addWidget(self.type_pvr_edit, 13, 7)
                self.grid.addWidget(self.pressure_pvr_label, 12, 8)
                self.grid.addWidget(self.pressure_pvr_edit, 13, 8)
                self.grid.addWidget(self.date_pressure_label, 12, 9)
                self.grid.addWidget(self.date_pressure_edit, 13, 9)
                self.tableWidget.show()
            else:
                self.pressure_pvr_label.setParent(None)
                self.date_pressure_label.setParent(None)
                self.pressure_pvr_edit.setParent(None)
                self.date_pressure_edit.setParent(None)
                self.vertical_line.setParent(None)
                self.plast_label.setParent(None)
                self.plast_line.setParent(None)
                self.vertical_label.setParent(None)

                self.roof_label.setParent(None)
                self.roof_edit.setParent(None)
                self.sole_label.setParent(None)
                self.sole_edit.setParent(None)
                self.date_pvr_label.setParent(None)
                self.date_pvr_edit.setParent(None)
                self.count_pvr_label.setParent(None)
                self.count_pvr_edit.setParent(None)
                self.type_pvr_label.setParent(None)
                self.type_pvr_edit.setParent(None)
                self.tableWidget.hide()

                # self.table_in_base_combo.currentTextChanged.connect(self.update_table_in_base_combo)
        except Exception as e:
            logger.exception("update_change_pvr failed")
            try:
                QMessageBox.warning(
                    self,
                    'Ошибка',
                    f'Не удалось обновить блок ПВР (перфорация): {e}'
                )
            except Exception:
                pass

    def update_table_name(self):
        try:
            self.index_change_line.setText('0')
        except Exception as e:
            logger.exception("update_table_name failed")
            try:
                QMessageBox.warning(
                    self,
                    'Ошибка',
                    f'Не удалось сбросить поле индекса изменения: {e}'
                )
            except Exception:
                pass


class TabWidget(TabWidgetUnion):
    def __init__(self, work_plan, tableWidget=0, old_index=0):
        super().__init__()
        self.addTab(TabPageDp(work_plan, tableWidget, old_index), 'Дополнительный план работ')


class DopPlanWindow(WindowUnion):
    def __init__(self, data_well, table_widget, parent=None):
        super(DopPlanWindow, self).__init__(data_well)

        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)

        self.table_widget = table_widget
        self.work_plan = self.data_well.work_plan
        if self.work_plan == 'dop_plan_in_base':
            self.data_well.number_dp = 0
            self.insert_index = 0

        else:
            self.insert_index = self.data_well.insert_index

        self.data, self.row_heights, self.col_width, self.boundaries_dict = None, None, None, None
        self.target_row_index = None
        self.target_row_index_cancel = None
        self.old_index = 0
        self.tableWidget = QTableWidget(0, 12)

        self.tab_widget = TabWidget(self.data_well, self.tableWidget, self.old_index)
        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.setAlternatingRowColors(True)

        self.buttonAdd = QPushButton('Добавить интервалы перфорации в таблицу')
        self.buttonAdd.clicked[bool].connect(lambda _checked: self.add_row_table())
        self.buttonDel = QPushButton('Удалить интервалы перфорации в таблице')
        self.buttonDel.clicked[bool].connect(lambda _checked: self.del_row_table())
        self.buttonadd_work = QPushButton('Создать доп план работ')
        self.buttonadd_work.clicked[bool].connect(lambda _checked: self.add_work(), Qt.QueuedConnection)
        self.buttonAddProject = QPushButton('Добавить проектные интервалы перфорации')
        self.buttonAddProject.clicked[bool].connect(lambda _checked: self.add_perforation_project())

        vbox = QGridLayout(self.centralWidget)
        vbox.addWidget(self.tab_widget, 0, 0, 1, 2)
        vbox.addWidget(self.tableWidget, 1, 0, 1, 2)
        vbox.addWidget(self.buttonAdd, 2, 0)
        vbox.addWidget(self.buttonDel, 2, 1)
        vbox.addWidget(self.buttonadd_work, 3, 0)
        vbox.addWidget(self.buttonAddProject, 3, 1)

        self.perforation_list = []

    def add_row_table(self):
        try:
            current_widget = self.tab_widget.currentWidget()

            self.plast_line = current_widget.plast_line.text()
            self.roof_edit = current_widget.roof_edit.text().replace(',', '.')
            self.sole_edit = current_widget.sole_edit.text().replace(',', '.')
            self.date_pvr_edit = current_widget.date_pvr_edit.text()
            self.count_pvr_edit = current_widget.count_pvr_edit.text()
            self.type_pvr_edit = current_widget.type_pvr_edit.text()
            self.pressure_pvr_edit = current_widget.pressure_pvr_edit.text().replace(',', '.')
            self.date_pressure_edit = current_widget.date_pressure_edit.text()
            vertical_line = current_widget.vertical_line.text().replace(',', '.')

            if '' in [self.plast_line, self.roof_edit, self.sole_edit, self.count_pvr_edit, self.type_pvr_edit, vertical_line]:
                QMessageBox.warning(self, 'Ошибка', 'Не введены все данные')
                return
            udlin = round(float(self.roof_edit) - float(vertical_line), 1)
            if len(self.perforation_list) == 0:
                QMessageBox.warning(self, 'Ошибка', 'Нужно сначало добавить проектные интервалы перфорации')
                return

            if [self.plast_line, vertical_line, self.roof_edit, self.sole_edit, self.date_pvr_edit, self.count_pvr_edit,
                self.type_pvr_edit, self.pressure_pvr_edit, self.date_pressure_edit] not in self.perforation_list:
                self.perforation_list.append(
                    [self.plast_line, vertical_line, self.roof_edit, self.sole_edit, self.date_pvr_edit,
                     self.count_pvr_edit,
                     self.type_pvr_edit, self.pressure_pvr_edit, self.date_pressure_edit])
            rows = 0
            self.tableWidget.insertRow(rows)

            self.tableWidget.setItem(rows, 0, QTableWidgetItem(str(self.plast_line)))
            self.tableWidget.setItem(rows, 1, QTableWidgetItem(str(vertical_line)))
            self.tableWidget.setItem(rows, 2, QTableWidgetItem(str(self.roof_edit)))
            self.tableWidget.setItem(rows, 3, QTableWidgetItem(str(self.sole_edit)))
            self.tableWidget.setItem(rows, 4, QTableWidgetItem(str(self.date_pvr_edit)))

            if self.old_index == 0:
                self.tableWidget.setItem(rows, 6, QTableWidgetItem(str(self.count_pvr_edit)))
                self.tableWidget.setItem(rows, 7, QTableWidgetItem(str(self.type_pvr_edit)))
                self.tableWidget.setItem(rows, 8, QTableWidgetItem(str(udlin)))
                self.tableWidget.setItem(rows, 9, QTableWidgetItem(str(self.pressure_pvr_edit)))
                self.tableWidget.setItem(rows, 10, QTableWidgetItem(str(self.date_pressure_edit)))
            else:
                self.tableWidget.setItem(rows, 5, QTableWidgetItem(str(self.count_pvr_edit)))
                self.tableWidget.setItem(rows, 6, QTableWidgetItem(str(self.type_pvr_edit)))
                self.tableWidget.setItem(rows, 7, QTableWidgetItem(str(udlin)))
                self.tableWidget.setItem(rows, 8, QTableWidgetItem(str(self.pressure_pvr_edit)))
                self.tableWidget.setItem(rows, 9, QTableWidgetItem(str(self.date_pressure_edit)))
        except Exception as e:
            logger.critical(e)

    def add_perforation_project(self):
        current_widget = self.tab_widget.currentWidget()
        well_data_combo = getattr(current_widget, "well_data_in_base_combo", None)
        if well_data_combo is None:
            QMessageBox.critical(
                self,
                'ошибка',
                'Не удалось определить данные скважины из базы. Обновите вкладку и повторите.'
            )
            logger.error("add_perforation_project: well_data_in_base_combo is None")
            return

        table_in_base_combo = str(well_data_combo.currentText())

        if ' от' in table_in_base_combo:
            table_in_base = table_in_base_combo.split(' ')[3].replace('krs', 'ПР').replace('dop_plan', 'ДП').replace(
                'dop_plan_in_base', 'ДП')
            type_kr = table_in_base_combo.split(' ')[2]
        else:
            return
        well_number = current_widget.well_number_edit.text()
        well_area = current_widget.well_area_edit.text()
        if well_number == '' or well_area == '':
            QMessageBox.critical(self, 'ошибка', 'Введите номер площадь скважины')
            return
        self.work_with_excel(well_number, well_area, table_in_base, type_kr)

    def work_with_excel(self, well_number, well_area, work_plan, type_kr):

        self.data_well.gips_in_well = False

        if data_list.connect_in_base:
            data_well = FindIndexPZ.excel_json
        else:
            db = connection_to_database(decrypt("DB_WELL_DATA"))
            data_well_base = WorkDatabaseWell(db)

            data_well = data_well_base.read_excel_in_base(well_number, well_area, work_plan, type_kr)
        self.data, self.row_heights, self.col_width, self.boundaries_dict = \
            self.read_excel_in_base(data_well)

        self.target_row_index = 5000
        self.target_row_index_cancel = 5000
        self.bottom_row_index = 5000
        self.perforation_list = []
        self.insert_index2 = None

        plan_work_header_found = False
        for i, row in self.data.items():
            if plan_work_header_found:
                break
            if i != 'image':
                list_row = []
                for col in range(len(row)):
                    value = row[col]['value']
                    if 'оризонт' in str(row[1]['value']) or 'пласт/' in str(row[col]['value']).lower():
                        self.target_row_index = int(i) + 1
                    elif 'вскрытия/отключения' in str(row[col]['value']):
                        self.old_index = 1
                    elif 'II. История эксплуатации скважины' in str(row[1]['value']) and \
                            self.data_well.work_plan not in ['plan_change']:
                        self.data_well.data_pvr_max = data_list.ProtectedIsDigit(int(i) - 1)
                        self.target_row_index_cancel = int(i) - 1
                        break
                    elif 'внутренний диаметр ( d шарошечного долота) не обсаженной части ствола' in str(
                            row[col]['value']) and \
                            self.data_well.work_plan not in ['plan_change']:
                        self.target_row_index_cancel = int(i)
                        break
                    elif 'II. История эксплуатации скважины' in str(
                            row[1]['value']):
                        self.data_well.data_pvr_max = data_list.ProtectedIsDigit(int(i) - 1)
                        break
                    elif 'Оборудование скважины' in str(row[1]['value']):
                        self.data_well.data_fond_min = data_list.ProtectedIsDigit(int(i) - 1)
                        break
                    elif ('Ранее проведенные' in str(row[1]['value']) or 'Ранее проведенные' in str(row[2]['value'])) \
                            and self.data_well.work_plan != "plan_change":
                        row[1]['value'] = None
                        row[2]['value'] = None
                    elif 'Наименование работ' in str(row[2]['value']):
                        self.data_well.data_x_max = data_list.ProtectedIsDigit(int(i) - 1)
                        self.data_well.insert_index2 = int(i) + 1
                        self.insert_index2 = int(1) + 1
                        plan_work_header_found = True
                        break

                    elif 'ИТОГО:' in str(row[col]['value']) and self.data_well.work_plan in ['plan_change']:
                        self.target_row_index_cancel = int(i) + 1
                        break
                    elif 'Текущий забой ' == str(row[col]['value']):
                        self.bottom_row_index = int(i)

                    if int(i) > self.target_row_index:
                        list_row.append(row[col]['value'])

                    if int(i) > self.target_row_index_cancel and plan_work_header_found:
                        break
            else:

                self.data_well.image_list = row
            self.count_diam = 0

            if len(list_row) > 4:

                if list_row:
                    if 'внутренний диаметр' not in str(list_row[1]):
                        if all([col is None or col == '' for col in list_row]) is False:
                            if list_row not in self.perforation_list:
                                self.perforation_list.append(list_row)
                    else:
                        self.count_diam = 1
        # self.data_well.insert_index2 = self.data_well.data_x_max.get_value -1
        self.data_well.count_template = 1

        if self.data_well.work_plan != 'plan_change':
            self.tableWidget.setSortingEnabled(False)
            rows = self.tableWidget.rowCount()
            for row_pvr in self.perforation_list[::-1]:
                self.tableWidget.insertRow(rows)
                for index_col, col_pvr in enumerate(row_pvr):
                    if col_pvr is not None:
                        self.tableWidget.setItem(rows, index_col - 1, QTableWidgetItem(str(col_pvr)))

        self.definition_open_trunk_well()

    def change_pvr_in_bottom(self, data, row_heights, col_width, boundaries_dict, current_bottom=0,
                             current_bottom_date_edit=0, method_bottom_combo=0):

        for i, row in data.items():
            if i != 'image':
                for col in range(len(row)):
                    if 'Текущий забой ' == str(row[col]['value']):
                        self.bottom_row_index = int(i)
                        break
        if 0 not in [current_bottom, current_bottom_date_edit, method_bottom_combo]:
            data[str(self.bottom_row_index)][3]['value'] = current_bottom
            data[str(self.bottom_row_index)][5]['value'] = current_bottom_date_edit
            data[str(self.bottom_row_index)][-1]['value'] = method_bottom_combo

        return data, row_heights, col_width, boundaries_dict

    def insert_row_in_pvr(self, data, row_heights, col_width, boundaries_dict, plast_list, current_bottom,
                          current_bottom_date_edit, method_bottom_combo):
        count_row_insert = len(plast_list)
        count_row_in_plan = self.target_row_index_cancel - self.target_row_index - 1
        boundaries_dict_new = {}
        if self.target_row_index != 5000:
            n = len(data) - 1
            if count_row_in_plan < count_row_insert:
                count_row_insert = count_row_insert - count_row_in_plan
                while n + 1 != int(self.target_row_index):
                    data.update({str(n + count_row_insert): data[str(n)]})
                    n -= 1
                row_heights.insert(int(self.target_row_index) + count_row_insert, None)

                for key, value in boundaries_dict.items():
                    if value[1] >= int(self.target_row_index) + 1:
                        value = [value[0], value[1] + count_row_insert, value[2], value[3] + count_row_insert]
                    boundaries_dict_new[key] = value

            else:

                for key, value in boundaries_dict.items():

                    if int(self.target_row_index) + 1 <= value[1] <= int(self.target_row_index_cancel - 4):
                        value = [value[0], value[1], value[0], value[3]]
                    boundaries_dict_new[key] = value
            len_rows = len(plast_list) - (self.target_row_index_cancel - self.target_row_index) + self.count_diam

            index_key_change = self.target_row_index + len_rows

            while index_key_change != self.target_row_index:
                data[str(index_key_change)] = data[str(index_key_change - len_rows)]
                index_key_change -= 1
            for index_row, plast in enumerate(plast_list):
                row_index = str(int(self.target_row_index + index_row + 1))
                a = plast[0]
                data[row_index] = [
                    {'value': '', 'font': {'name': 'Times New Roman Cyr', 'size': 10.0, 'bold': False, 'italic': False},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': None, 'right': None, 'top': None, 'bottom': None},
                     'alignment': {'horizontal': None, 'vertical': None, 'wrap_text': True}},
                    {'value': plast[0], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[1], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': None, 'vertical': None, 'wrap_text': None}},
                    {'value': plast[2], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[3], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[4], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': None, 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': None, 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[5], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': None, 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': None, 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[6], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[7], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[8], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[9], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': None}},
                    {'value': plast[10], 'font': {'name': 'Arial Cyr', 'size': 12.0, 'bold': False, 'italic': True},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'},
                     'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': None}},
                    {'value': None, 'font': {'name': 'Calibri', 'size': 11.0, 'bold': False, 'italic': False},
                     'fill': {'color': 'RGB(00000000)'},
                     'borders': {'left': None, 'right': None, 'top': None, 'bottom': None},
                     'alignment': {'horizontal': None, 'vertical': None, 'wrap_text': None}}]

        data[str(self.bottom_row_index)][3]['value'] = current_bottom
        data[str(self.bottom_row_index)][5]['value'] = current_bottom_date_edit
        data[str(self.bottom_row_index)][10 - self.old_index]['value'] = method_bottom_combo
        #     print(i)

        return data, row_heights, col_width, boundaries_dict_new

    def del_row_table(self):
        row = self.tableWidget.currentRow()
        if row == -1:
            QMessageBox.information(self, 'Внимание', 'Выберите строку для удаления')
            return
        self.tableWidget.removeRow(row)

    def add_work(self):
        from data_base.work_with_base import round_cell
        from data_list import ProtectedIsNonNone
        from work_py.advanted_file import definition_plast_work
        try:
            self.data_well.data_list = []
            current_widget = self.tab_widget.currentWidget()
            if hasattr(current_widget, "well_data_in_base_combo"):
                well_data_combo = getattr(current_widget, "well_data_in_base_combo", None)
                if well_data_combo is None:
                    QMessageBox.critical(
                        self,
                        'ошибка',
                        'Не удалось определить данные скважины из базы. Обновите вкладку и повторите.'
                    )
                    logger.error("add_work: well_data_in_base_combo is None")
                    return

                well_data_in_base_combo = well_data_combo.currentText()
                if well_data_in_base_combo == '' and self.data_well.work_plan != 'dop_plan':
                    QMessageBox.critical(self, 'База данных', 'Необходимо выбрать план работ')
                    return

                if ' от' in well_data_in_base_combo:
                    work_plan_in_base = well_data_in_base_combo.split(' ')[3]
                    type_kr = well_data_in_base_combo.split(' ')[2]

            well_number = current_widget.well_number_edit.text().strip()
            well_area = current_widget.well_area_edit.text().strip()
            if data_list.data_in_base and (not well_number or not well_area):
                QMessageBox.critical(self, 'База данных', 'Введите номер и площадь скважины')
                return

            if well_number != '' and well_area != '':
                self.data_well.well_number, self.data_well.well_area = \
                    ProtectedIsNonNone(well_number), ProtectedIsNonNone(well_area)

            method_bottom_combo = current_widget.method_bottom_combo.currentText()
            if method_bottom_combo == '':
                QMessageBox.critical(self, 'Забой', 'Выберете метод определения забоя')
                return

            change_pvr_combo = current_widget.change_pvr_combo.currentText()
            if change_pvr_combo == '':
                QMessageBox.warning(self, 'Ошибка', 'Нужно выбрать пункт изменен ли интервалы ПВР')
                return
            if data_list.data_in_base:
                fluid = current_widget.fluid_edit.toPlainText().replace(',', '.')
                current_bottom = current_widget.current_bottom_edit.text()

                if current_bottom != '':
                    current_bottom = round_cell(current_bottom.replace(',', '.'))

                work_earlier = current_widget.work_edit.toPlainText()
                number_dp = current_widget.number_DP_Combo.currentText()

                current_bottom_date_edit = current_widget.current_bottom_date_edit.text()

                template_depth_edit = int(float(current_widget.template_depth_edit.text().replace(',', '.')))
                self.data_well.need_depth = current_bottom
                if str(template_depth_edit).isdigit() is False:
                    QMessageBox.critical(self, 'Ошибка', 'ошибка в глубине диаметра')
                    return
                template_length_edit = current_widget.template_length_edit.text()
                if self.data_well.column_additional:
                    template_depth_addition_edit = current_widget.template_depth_addition_edit.text()
                    template_length_addition_edit = current_widget.template_length_addition_edit.text()
                skm_interval_edit = current_widget.skm_interval_edit.text()
                raiding_interval_edit = current_widget.raiding_interval_edit.text()

                if current_bottom == '' or fluid == '' or work_earlier == '' or \
                        template_depth_edit == '' or template_length_edit == '':
                    # print(current_bottom, fluid, work_earlier)
                    QMessageBox.critical(self, 'Забой', 'не все значения введены')
                    return
                if template_length_edit == '0' or template_length_edit == '':
                    mes = QMessageBox.question(self, 'Длина шаблона', 'в скважину во время ремонта не был спущен шаблон, '
                                                                      'так ли это?')
                    if mes == QMessageBox.StandardButton.No:
                        return
                if float(template_depth_edit) > float(self.data_well.bottom_hole_drill.get_value):
                    QMessageBox.critical(self, 'Забой', 'Шаблонирование не может быть ниже искусственного забоя забоя')
                    return

                if number_dp != '':
                    self.data_well.number_dp = int(float(number_dp))
                else:
                    QMessageBox.warning(self, 'Внимание', 'Нужно выбрать номер дополнительного плана вручную.')
                    return

                if (0.87 <= float(fluid[:3].replace(',', '.')) <= 1.64) == False:
                    QMessageBox.critical(self, 'рабочая жидкость',
                                         'уд. вес рабочей жидкости не может быть меньше 0,87 и больше 1,64')
                    return

                if data_list.data_in_base:
                    if 'г/см3' not in fluid:
                        QMessageBox.critical(self, 'уд.вес', 'нужно добавить значение "г/см3" в уд.вес')
                        return
                    self.data_well.fluid_work = fluid
                    self.data_well.well_fluid_in_pz = [fluid]
                    self.data_well.fluid_work_short = fluid[:7]

                    self.data_well.fluid = float(fluid[:4].replace('г', ''))
                else:
                    self.data_well.fluid = float(fluid)

                    if float(current_bottom) > self.data_well.bottom_hole_drill.get_value:
                        QMessageBox.critical(self, 'Забой', 'Текущий забой больше пробуренного забоя')
                        return
                    self.data_well.fluid_work, self.data_well.fluid_work_short = self.calc_work_fluid(fluid)

                self.data_well.template_depth = float(template_depth_edit)
                self.data_well.template_length = float(template_length_edit)
                if self.data_well.column_additional:
                    self.data_well.template_depth = float(template_depth_addition_edit)
                    self.data_well.template_length = float(template_length_addition_edit)

                try:
                    skm_interval = []
                    if ',' in skm_interval_edit:
                        for skm in skm_interval_edit.split(','):
                            if '-' in skm:
                                skm_interval.append(list(map(int, skm.split('-'))))
                    else:
                        if '-' in skm_interval_edit:
                            skm_interval.append(list(map(int, skm_interval_edit.split('-'))))

                    skm_interval_new = merge_overlapping_intervals(skm_interval)

                    self.data_well.skm_interval = skm_interval_new

                except Exception as e:
                    QMessageBox.warning(self, 'Ошибка',
                                        f'в интервале скреперования отсутствует '
                                        f'корректные интервалы скреперования \n {e}')
                    return

                try:
                    raiding_interval = []
                    if ',' in raiding_interval_edit:
                        for skm in raiding_interval_edit.split(','):
                            if '-' in skm:
                                raiding_interval.append(list(map(int, skm.split('-'))))
                    else:
                        if '-' in raiding_interval_edit:
                            raid = raiding_interval_edit.split('-')
                            raiding_interval.append(list(map(int, raid)))

                    raiding_interval_new = merge_overlapping_intervals(raiding_interval)

                    self.data_well.ribbing_interval = raiding_interval_new

                except Exception as e:
                    QMessageBox.warning(self, 'Ошибка',
                                        f'в интервале райбирования отсутствует корректные интервалы '
                                        f' {e}')
                    return

                if len(self.data_well.skm_interval) == 0:
                    mes = QMessageBox.question(self, 'Ошибка',
                                               'Интервалы скреперования отсутствуют, так ли это?')
                    if mes == QMessageBox.StandardButton.No:
                        return

                if len(self.data_well.ribbing_interval) == 0:
                    mes = QMessageBox.question(self, 'Ошибка',
                                               'Интервалы Райбирования отсутствуют, так ли это?')
                    if mes == QMessageBox.StandardButton.No:
                        return

                self.data_well.count_template = 1

                list_dop_plan = [current_widget.well_data_in_base_combo.itemText(i) for i in range(current_widget.well_data_in_base_combo.count())]
                if any([f'ДП№{number_dp}' in dop_plan or f'ДП№ {number_dp}' in dop_plan for dop_plan in list_dop_plan]):
                    question = QMessageBox.question(self, 'Ошибка', f'дополнительный план работ № {number_dp} '
                                                                    f'есть в базе, обновить доп план?')
                    if question == QMessageBox.StandardButton.No:
                        return

                index_change_line = current_widget.index_change_line.text()

                if index_change_line != '':
                    index_change_line = int(float(index_change_line))
                else:
                    QMessageBox.critical(self, 'пункт', 'Необходимо выбрать пункт плана работ')
                    return

                self.data_well.current_bottom = current_bottom

                if skm_interval_edit not in ['', 0, '0-0'] and '-' in skm_interval_edit:
                    if ',' in skm_interval_edit:
                        for skm_int in skm_interval_edit.split(','):
                            self.data_well.skm_interval.append(list(map(int, skm_int.split('-'))))
                    else:
                        self.data_well.skm_interval.append(list(map(int, skm_interval_edit.split('-'))))

                rows = self.tableWidget.rowCount()
                table_pvr_rows = []

                if hasattr(current_widget, "well_data_in_base_combo"):
                    self.work_with_excel(well_number, well_area, work_plan_in_base, type_kr)

                if change_pvr_combo == 'Да':
                    if rows == 0:
                        QMessageBox.warning(self, 'Ошибка', 'Нужно загрузить интервалы перфорации')
                        return

                    plast_row = []
                    for row in range(rows):
                        list_row = []
                        for col in range(13):
                            item = self.tableWidget.item(row, col)
                            if item is not None:
                                list_row.append(item.text())
                            else:
                                list_row.append(None)
                        plast_row.append(list_row)
                    table_pvr_rows = plast_row
                    data_list.data, data_list.row_heights, data_list.col_width, data_list.boundaries_dict = \
                        self.insert_row_in_pvr(self.data, self.row_heights, self.col_width, self.boundaries_dict, plast_row,
                                               current_bottom, current_bottom_date_edit, method_bottom_combo)
                else:

                    data_list.data, data_list.row_heights, data_list.col_width, data_list.boundaries_dict = \
                        self.change_pvr_in_bottom(self.data, self.row_heights, self.col_width, self.boundaries_dict,
                                                  current_bottom, current_bottom_date_edit, method_bottom_combo)
                work_list = self.work_list(work_earlier)
                data_list.dop_work_list = work_list

                # Для `dop_plan_in_base` итоговый Excel собирается из `data_list.data`
                # (см. `main.py` → `insert_data_new_excel_file(...)`), поэтому нужно вставлять строки
                # и в Excel-JSON, а не только в UI-таблицу.
                if data_list.data_in_base:
                    insert_index2 = getattr(self.data_well, "insert_index2", None)
                    if (
                        self.data_well is None or insert_index2 in (None, "")
                    ):
                        QMessageBox.warning(
                            self,
                            "Ошибка",
                            "Не выбрана скважина для доп. плана из базы.",
                        )
                        return
                    data_list.data, data_list.row_heights, data_list.boundaries_dict = (
                        self._insert_rows_into_excel_json(
                            data_list.data,
                            data_list.row_heights,
                            data_list.boundaries_dict,
                            insert_row=insert_index2,
                            rows_values=work_list,
                        )
                    )
                else:
                    self.populate_work_rows_with_remote_fallback(
                        "dop_plan_py",
                        {},
                        self.table_widget,
                        work_list,
                        work_plan=self.work_plan,
                        insert_index=self.insert_index + 3,
                    )

                # Строки ПВР из таблицы: пласт, вертикаль, кровля, подошва, … (см. колонки TabPageDp).
                # Раньше по ошибке шёл обход dict_perforation как списка кортежей — при итерации по dict
                # приходили ключи-символы строки (например 5 букв → «expected 9, got 5»).
                if table_pvr_rows:
                    for row_cells in table_pvr_rows:
                        if not row_cells or len(row_cells) < 4:
                            continue
                        plast = row_cells[0]
                        if plast is None or str(plast).strip() == '':
                            continue
                        try:
                            roof_num = float(str(row_cells[2]).replace(",", "."))
                            sole_num = float(str(row_cells[3]).replace(",", "."))
                        except (TypeError, ValueError, AttributeError):
                            continue
                        self.data_well.dict_perforation.setdefault(plast, {}).setdefault(
                            "отрайбировано", False
                        )
                        self.data_well.dict_perforation.setdefault(plast, {}).setdefault(
                            "Прошаблонировано", False
                        )
                        self.data_well.dict_perforation.setdefault(plast, {}).setdefault("интервал", []).append(
                            (roof_num, sole_num)
                        )
                        self.data_well.dict_perforation_short.setdefault(plast, {}).setdefault(
                            "интервал", []
                        ).append((roof_num, sole_num))
                        self.data_well.dict_perforation.setdefault(plast, {}).setdefault("отключение", False)
                        self.data_well.dict_perforation_short.setdefault(plast, {}).setdefault(
                            "отключение", False
                        )

                if len(self.data_well.dict_perforation) != 0:
                    for plast_key in list(self.data_well.dict_perforation.keys()):
                        self.data_well.dict_perforation.setdefault(plast_key, {}).setdefault(
                            "отрайбировано", False
                        )
                        self.data_well.dict_perforation.setdefault(plast_key, {}).setdefault(
                            "Прошаблонировано", False
                        )
                        self.data_well.dict_perforation.setdefault(plast_key, {}).setdefault("отключение", False)
                        self.data_well.dict_perforation_short.setdefault(plast_key, {}).setdefault(
                            "отключение", False
                        )

            else:
                fluid = current_widget.fluid_edit.toPlainText().replace(',', '.')
                current_bottom = current_widget.current_bottom_edit.text()
                if current_bottom != '':
                    current_bottom = round_cell(current_bottom.replace(',', '.'))


                work_earlier = current_widget.work_edit.toPlainText()
                number_dp = current_widget.number_DP_Combo.currentText()

                template_depth_edit = current_widget.template_depth_edit.text()
                template_length_edit = current_widget.template_length_edit.text()
                if self.data_well.column_additional:
                    template_depth_addition_edit = current_widget.template_depth_addition_edit.text()
                    template_length_addition_edit = current_widget.template_length_addition_edit.text()
                skm_interval_edit = current_widget.skm_interval_edit.text()
                try:

                    if skm_interval_edit not in ['', 0, '0-0'] and '-' in skm_interval_edit:
                        if ',' in skm_interval_edit:
                            for skm_int in skm_interval_edit.split(','):
                                self.data_well.skm_interval.append(list(map(int, skm_int.split('-'))))
                        else:
                            self.data_well.skm_interval.append(list(map(int, skm_interval_edit.split('-'))))

                except Exception as e:
                    QMessageBox.warning(self, 'Ошибка',
                                        'в интервале скреперования отсутствует корректные интервалы скреперования')
                    return

                if current_bottom == '' or fluid == '' or work_earlier == '' or \
                        template_depth_edit == '' or template_length_edit == '':
                    # print(current_bottom, fluid, work_earlier)
                    QMessageBox.critical(self, 'Забой', 'не все значения введены')
                    return
                if template_length_edit == '0':
                    QMessageBox.critical(self, 'Длина шаблона',
                                         'Введите длину шаблонов которые были спущены в скважину')
                    return
                if float(template_depth_edit) > float(current_bottom):
                    QMessageBox.critical(self, 'Забой', 'Шаблонирование не может быть ниже текущего забоя')
                    return
                if number_dp != '':
                    self.data_well.number_dp = int(float(number_dp))

                if (0.87 <= float(fluid[:3].replace(',', '.')) <= 1.64) == False:
                    QMessageBox.critical(self, 'рабочая жидкость',
                                         'уд. вес рабочей жидкости не может быть меньше 0,87 и больше 1,64')
                    return

                # if float(current_bottom) > self.data_well.bottom_hole_drill.get_value:
                #     QMessageBox.critical(self, 'Забой', 'Текущий забой больше пробуренного забоя')
                #     return
                self.data_well.fluid_work, self.data_well.fluid_work_short = self.calc_work_fluid(fluid)

                self.data_well.template_depth = float(template_depth_edit)
                self.data_well.template_length = float(template_length_edit)
                if self.data_well.column_additional:
                    self.data_well.template_depth = float(template_depth_addition_edit)
                    self.data_well.template_length = float(template_length_addition_edit)

                work_list = self.work_list(work_earlier)
                # self.data_well.insert_index2 = self.insert_index
                self.populate_work_rows_with_remote_fallback(
                    "dop_plan_py", {}, self.table_widget, work_list, work_plan=self.work_plan
                )
                definition_plast_work(self)
        except Exception as e:
            logger.critical(e)

        data_list.pause = False
        self.close()
        self.close_modal_forcefully()

    @staticmethod
    def _insert_rows_into_excel_json(data: dict, row_heights: list, boundaries_dict: dict, insert_row: int,
                                    rows_values: list):
        """
        Вставляет строки (rows_values) в Excel-JSON `data` со сдвигом индексов, высот строк и merged-cells.
        `insert_row` — 1-based номер строки Excel, куда вставляем (до существующей строки).
        """
        if not isinstance(data, dict) or not rows_values:
            return data, row_heights, boundaries_dict

        # Сначала пытаемся найти строку "Порядок работы" и вставить ПЕРЕД ней
        order_row = None
        for k, row in data.items():
            if k == "image" or not str(k).isdigit():
                continue
            try:
                v = row[1]["value"] if len(row) > 1 else None
            except Exception:
                v = None
            if "Порядок работы" in str(v):
                order_row = int(k)
                break
        if order_row is not None:
            insert_row = order_row

        # Если insert_row неизвестен — пытаемся найти первый блок работ по "Наименование работ"
        if not isinstance(insert_row, int) or insert_row <= 0:
            insert_row = None
            for k, row in data.items():
                if k == "image":
                    continue
                try:
                    row2 = row[2]["value"] if len(row) > 2 else None
                except Exception:
                    row2 = None
                if "Наименование работ" in str(row2):
                    insert_row = int(k) + 1
                    break
            if insert_row is None:
                # fallback: в конец
                numeric_keys = [int(k) for k in data.keys() if str(k).isdigit()]
                insert_row = (max(numeric_keys) + 1) if numeric_keys else 1

        n_new = len(rows_values)

        # Подготовим шаблон строки (стили), чтобы новые строки не были "пустыми" по форматам.
        template_key = str(insert_row) if str(insert_row) in data else None
        if template_key is None:
            numeric_keys_sorted = sorted(int(k) for k in data.keys() if str(k).isdigit())
            # ближайшая существующая строка сверху
            prev = max([k for k in numeric_keys_sorted if k < insert_row], default=None)
            template_key = str(prev) if prev is not None else (str(numeric_keys_sorted[0]) if numeric_keys_sorted else None)

        template_row = copy.deepcopy(data.get(template_key, [])) if template_key else []

        def _make_row(values):
            if template_row:
                row_copy = copy.deepcopy(template_row)
                for idx, v in enumerate(values):
                    if idx < len(row_copy) and isinstance(row_copy[idx], dict):
                        row_copy[idx]["value"] = v
                return row_copy
            # Минимальный fallback: только значения
            return [{"value": v, "font": {"name": "Calibri", "size": 11.0, "bold": False, "italic": False, "color": "RGB(00000000)"},
                     "fill": {"color": "RGB(00000000)"},
                     "borders": {"left": None, "right": None, "top": None, "bottom": None},
                     "alignment": {"horizontal": None, "vertical": None, "wrap_text": None}} for v in values]

        # 1) Сдвигаем строки в data
        new_data = {}
        for k, row in data.items():
            if k == "image":
                new_data[k] = row
                continue
            if not str(k).isdigit():
                new_data[k] = row
                continue
            rk = int(k)
            if rk >= insert_row:
                new_data[str(rk + n_new)] = row
            else:
                new_data[str(rk)] = row

        # 2) Вставляем новые строки
        for i, values in enumerate(rows_values):
            new_data[str(insert_row + i)] = _make_row(values)

        # 3) Сдвигаем row_heights (если список)
        if isinstance(row_heights, list) and row_heights:
            idx0 = max(insert_row - 1, 0)
            default_h = row_heights[idx0 - 1] if idx0 - 1 >= 0 and idx0 - 1 < len(row_heights) else (row_heights[0] if row_heights else 15)
            row_heights = row_heights[:idx0] + [default_h] * n_new + row_heights[idx0:]

        # 4) Сдвигаем merged cells
        if isinstance(boundaries_dict, dict) and boundaries_dict:
            for key, val in boundaries_dict.items():
                try:
                    start_col, start_row, end_col, end_row = val
                except Exception:
                    continue
                if start_row >= insert_row:
                    start_row += n_new
                if end_row >= insert_row:
                    end_row += n_new
                boundaries_dict[key] = [start_col, start_row, end_col, end_row]

        def _next_merge_key(bdict: dict) -> str:
            try:
                keys_int = [int(k) for k in bdict.keys() if str(k).isdigit()]
                return str(max(keys_int) + 1) if keys_int else "1"
            except Exception:
                return "1"

        def _add_merge(bdict: dict, start_col: int, row_num: int, end_col: int):
            if row_num is None or row_num <= 0:
                return
            # Не добавляем дубль (иначе могут быть конфликты/накладки merge)
            try:
                for _k, v in bdict.items():
                    if not isinstance(v, (list, tuple)) or len(v) != 4:
                        continue
                    if int(v[0]) == start_col and int(v[1]) == row_num and int(v[2]) == end_col and int(v[3]) == row_num:
                        return
            except Exception:
                pass
            key = _next_merge_key(bdict)
            bdict[key] = [start_col, row_num, end_col, row_num]

        # 5) Добавляем merge для строки "Ранее проведенные работы" (B:L)
        if isinstance(boundaries_dict, dict):
            _add_merge(boundaries_dict, 2, insert_row, 12)

        # 6) Фиксируем merge заголовков:
        # - "Порядок работы" объединение B:K (2..11)
        # - "Наименование работ" объединение C:J (3..10)
        if isinstance(boundaries_dict, dict):
            order_row_new = None
            name_row_new = None
            for k, row in new_data.items():
                if k == "image" or not str(k).isdigit():
                    continue
                rk = int(k)
                # По договоренности: "Порядок работы" находится в B, "Наименование работ" — в C
                v_b = None
                v_c = None
                try:
                    if len(row) > 1 and isinstance(row[1], dict):
                        v_b = row[1].get("value")
                except Exception:
                    v_b = None
                try:
                    if len(row) > 2 and isinstance(row[2], dict):
                        v_c = row[2].get("value")
                except Exception:
                    v_c = None

                if order_row_new is None and "Порядок работы" in str(v_b):
                    order_row_new = rk
                if name_row_new is None and "Наименование работ" in str(v_c):
                    name_row_new = rk
                if order_row_new is not None and name_row_new is not None:
                    break

            _add_merge(boundaries_dict, 2, order_row_new, 11)
            _add_merge(boundaries_dict, 3, name_row_new, 10)

        return new_data, row_heights, boundaries_dict

    def add_work_excel(self, ws2, work_list, ind_ins):
        for i in range(1, len(work_list) + 1):  # Добавлением работ
            for j in range(1, 13):
                cell = ws2.cell(row=i, column=j)

                if cell and str(cell) != str(work_list[i - 1][j - 1]):

                    if i >= ind_ins:

                        if j != 1:
                            cell.border = data_list.thin_border
                        if j == 11:
                            cell.font = Font(name='Arial', size=11, bold=False)
                        # if j == 12:
                        #     cell.value = work_list[i - 1][j - 1]
                        else:
                            cell.font = Font(name='Arial', size=13, bold=False)
                        ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                        vertical='center')
                        ws2.cell(row=i, column=11).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')
                        ws2.cell(row=i, column=12).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')
                        ws2.cell(row=i, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                        vertical='center')

                        if 'порядок работы' in str(cell.value).lower() or \
                                'наименование работ' in str(cell.value).lower():
                            self.data_well.insert_index2 = i + 1
                            ws2.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=True)
                            ws2.cell(row=i, column=j).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                            vertical='center')

    def work_list(self, work_earlier):
        krs_begin = [[None,
                      f'Ранее проведенные работы: \n {work_earlier}',
                      None, None, None, None, None, None, None,
                      None, None, None],
                     [None, 'Порядок работы', None, None, None, None, None, None, None, None, None, None],
                     [None, 'п/п', 'Наименование работ', None, None, None, None, None, None, None,
                      'Ответственный',
                      'Нормы времени \n мин/час.']
                     ]

        return krs_begin


if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    # app.setStyleSheet()

    window = DopPlanWindow(1, 1, 1)
    window.show()
    sys.exit(app.exec_())
