# -*- coding: utf-8 -*-
import json
import os
import shutil

import data_list
import sys
import socket
import psutil
import win32com.client
import openpyxl
import re
import win32con
import property_excel.property_excel_pvr
import threading
import win32gui
from exceptions import UncaughtExceptions
from modal_dialogs import ModalDialog
from server_response import ApiClient
from openpyxl.reader.excel import load_workbook
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QMenu,
    QMenuBar,
    QAction,
    QTableWidget,
    QLineEdit,
    QFileDialog,
    QToolBar,
    QPushButton,
    QMessageBox,
    QInputDialog,
    QTabWidget,
    QTableWidgetItem,
    QSplashScreen,
    QDialog,
    QVBoxLayout,
)
from PyQt5 import QtCore, QtWidgets
from block_name import current_datetime

from datetime import datetime
from decrypt import decrypt
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font
from block_name import region_select

from log_files.log import logger
from openpyxl.drawing.image import Image
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QPixmap

from work_py.rationingKRS import lifting_nkt_norm, descentNKT_norm
from work_py.data_informations import dict_data_cdng


class MyMainWindow(QMainWindow):

    def __init__(self):
        super().__init__()

        self.table_schema = None
        self.table_widget = None
        self.table_title = None
        self.data_window = None
        self.perforation_correct_window2 = None
        self.work_plan = None
        self.gnkt_data = None
        self.threads = []
        self.repair_id = None

    def work_podpisant_list(self, region, contractor):
        with open(
                f"{data_list.path_image}podpisant.json", "r", encoding="utf-8"
        ) as file:
            podpis_dict = json.load(file)

        power_of_attorney = None
        expedition = ""
        if "Ойл" in contractor:
            chief_engineer_post = podpis_dict[data_list.contractor]["Руководство"][
                "chief_engineer"
            ]["post"]
            chief_engineer_surname = podpis_dict[data_list.contractor]["Руководство"][
                "chief_engineer"
            ]["surname"]
            chief_geologist_post = podpis_dict[data_list.contractor]["Руководство"][
                "chief_geologist"
            ]["post"]
            chief_geologist_surname = podpis_dict[data_list.contractor]["Руководство"][
                "chief_geologist"
            ]["surname"]
        elif "РН" in contractor:
            number_expedition = [
                number for number in data_list.user[0] if number.isdigit()
            ][0]
            if self.data_well.region == "ТГМ":
                expedition = f"Экспедиция № 3"
            elif self.data_well.region == "ЧГМ":
                expedition = f"Экспедиция № 2"
            elif self.data_well.region == "АГМ":
                expedition = f"Экспедиция № 5"
            elif self.data_well.region == "ИГМ":
                expedition = f"Экспедиция № 1"
            elif self.data_well.region == "КГМ":
                expedition = f"Экспедиция № 4"

            chief_engineer_post = podpis_dict[data_list.contractor]["Экспедиция"][
                expedition
            ]["chief_engineer"]["post"]
            chief_engineer_surname = podpis_dict[data_list.contractor]["Экспедиция"][
                expedition
            ]["chief_engineer"]["surname"]
            power_of_attorney = podpis_dict[data_list.contractor]["Экспедиция"][
                expedition
            ]["chief_engineer"]["power_of_attorney"]
            chief_geologist_post = podpis_dict[data_list.contractor]["Экспедиция"][
                expedition
            ]["chief_geologist"]["post"]
            chief_geologist_surname = podpis_dict[data_list.contractor]["Экспедиция"][
                expedition
            ]["chief_geologist"]["surname"]

            work_podpisant_list = [
                [
                    None, "СОГЛАСОВАНО:",
                    None, None, None, None, None,
                    "УТВЕРЖДАЕМ:", None, None, None, None,
                ],
                [
                    None,
                    podpis_dict[data_list.costumer][region]["gi"]["post"], None, None, None, None, None,
                    chief_engineer_post, None, None, None, None,
                ],
                [
                    None,
                    f'____________{podpis_dict[data_list.costumer][region]["gi"]["surname"]}', None, None, None, None,
                    None,
                    f"_____________{chief_engineer_surname}", None, None, None, None,
                ],
                [
                    None,
                    f'"____"_____________________{current_datetime.year}г.', None, None, None, None, None,
                    f'"____"_____________________{current_datetime.year}г.', None, None, None, None,
                ],
                [None, None, None, None, None, None, None,
                 power_of_attorney, None, None, None, None,
                 ],
                [
                    None,
                    podpis_dict[data_list.costumer][region]["gg"]["post"], None, None, None, None, None, None, None,
                    None, None, None,
                ],
                # [None, podpis_dict[data_list.costumer][region]['gg']['post'], None, None, None,
                #  None, None, f'{chief_geologist_post}', None, None, None, None],
                # [None, f'_____________{podpis_dict[data_list.costumer][region]["gg"]["surname"]}', None, None, None,
                #  None, None, f'_____________{chief_geologist_surname}', None, None, '',
                #  None],
                [
                    None,
                    f'_____________{podpis_dict[data_list.costumer][region]["gg"]["surname"]}', None, None, None, None,
                    None, None, None, None,
                    "", None,
                ],
                # [None, f'"____"_____________________{current_datetime.year}г.', None, None, '', None, None,
                #  f'"____"_____________________{current_datetime.year}г.', None, None, None, None],
                [
                    None,
                    f'"____"_____________________{current_datetime.year}г.', None, None,
                    "", None, None, None, None, None, None, None,
                ],
                [None, None, None, None, None, None, None, None, None, None, None, None,
                 ],
                [None, None, None, None, None, None, None, None, None, None, None, None,
                 ],
                [
                    None, None, None, None, None, None,
                    None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None, None, None, None,
                 ],
                [None, None, None, None, None, None, None, None, None, None, None, None,
                 ],
                [None, None, None, None, None, None, None, None, None, None, None, None,
                 ],
                [None, None, None, None, None, None, None, None, None, None, None, None,
                 ],
                [None, None, None, None, None, None, None, None, None, None, None, None,
                 ],
            ]

            if "3" in expedition or "2" in expedition:
                work_podpisant_list[5] = [
                    None,
                    podpis_dict[data_list.costumer][region]["gg"]["post"],
                    None, None, None, None, None, None, None, None, None, None,
                ]
                work_podpisant_list[6] = [
                    None,
                    f'_____________{podpis_dict[data_list.costumer][region]["gg"]["surname"]}',
                    None, None, None, None, None, None, None, None,
                    "", None,
                ]
                work_podpisant_list[7] = [
                    None,
                    f'"____"_____________________{current_datetime.year}г.', None, None,
                    "", None, None, None, None, None, None, None,
                ]

        if "prs" in self.data_well.work_plan:
            if "Ойл" in contractor:
                if (
                        region == "ЧГМ"
                        or region == "ТГМ"
                        or "gnkt" in self.data_well.work_plan
                ):
                    data_list.ctkrs = "ЦТКРС №1"
                elif region == "КГМ" or region == "АГМ":
                    data_list.ctkrs = "ЦТКРС №2"
                elif region == "ИГМ":
                    data_list.ctkrs = "ЦТКРС №4"

            nach_cdng_post = (
                    podpis_dict[data_list.costumer][self.data_well.region]["ЦДНГ"][
                        self.data_well.cdng.get_value
                    ]["Начальник"]["post"]
                    + " "
                    + self.data_well.cdng.get_value
            )
            nach_cdng_name = podpis_dict[data_list.costumer][self.data_well.region][
                "ЦДНГ"
            ][self.data_well.cdng.get_value]["Начальник"]["surname"]
            nach_cdng_name = nach_cdng_name.split(" ")
            nach_cdng_name = (
                f"{nach_cdng_name[0]} {nach_cdng_name[1][0]}.{nach_cdng_name[1][0]}."
            )
            technol_cdng_post = (
                    podpis_dict[data_list.costumer][self.data_well.region]["ЦДНГ"][
                        self.data_well.cdng.get_value
                    ]["Ведущий инженер-технолог"]["post"]
                    + " "
                    + self.data_well.cdng.get_value
            )
            technol_cdng_name = podpis_dict[data_list.costumer][self.data_well.region][
                "ЦДНГ"
            ][self.data_well.cdng.get_value]["Ведущий инженер-технолог"]["surname"]
            technol_cdng_name = technol_cdng_name.split(" ")
            technol_cdng_name = f"{technol_cdng_name[0]} {technol_cdng_name[1][0]}.{technol_cdng_name[1][0]}."
            geolog_cdng_post = (
                    podpis_dict[data_list.costumer][self.data_well.region]["ЦДНГ"][
                        self.data_well.cdng.get_value
                    ]["Ведущий геолог"]["post"]
                    + " "
                    + self.data_well.cdng.get_value
            )
            geolog_cdng_name = podpis_dict[data_list.costumer][self.data_well.region][
                "ЦДНГ"
            ][self.data_well.cdng.get_value]["Ведущий геолог"]["surname"]
            geolog_cdng_name = geolog_cdng_name.split(" ")
            geolog_cdng_name = f"{geolog_cdng_name[0]} {geolog_cdng_name[1][0]}.{geolog_cdng_name[1][0]}."
            nach_ctkrs_post = podpis_dict[data_list.contractor]["Экспедиция"][
                data_list.ctkrs
            ]["chief_engineer"]["post"]
            nach_ctkrs_name = podpis_dict[data_list.contractor]["Экспедиция"][
                data_list.ctkrs
            ]["chief_engineer"]["surname"]

            work_podpisant_list = [
                [None, 'СОГЛАСОВАНО:', None, None, None, None, None, 'УТВЕРЖДАЕМ:', None, None, None, None],
                [None, nach_cdng_post, None, None, None, None, None, nach_ctkrs_post, None, None, None, None],
                [None, f'____________{nach_cdng_name}',
                 None, None, None, None, None,
                 f'_____________{nach_ctkrs_name}', None, None, None, None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, None, None, None,
                 f'"____"_____________________{current_datetime.year}г.', None, None, None, None],
                [None, None, None, None, " ", None, None, None, None, None, None, None],
                [None,
                 technol_cdng_post,
                 None, None, None, None, None, None, None, None, None, None],
                [None,
                 f'____________{technol_cdng_name}',
                 None, None, None, None, None, None, None, None, None, None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, None, None, None,
                 None, None, None, None, None],
                [None,
                 geolog_cdng_post,
                 None, None, None, None, None, None, None, None, None, None],
                [None,
                 f'____________{geolog_cdng_name}',
                 None, None, None, None, None, None, None, None, None, None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, None, None, None,
                 None, None, None, None, None]]

        elif "krs" in self.data_well.work_plan and self.data_well.curator == "ВНС":
            work_podpisant_list = [
                [None, 'СОГЛАСОВАНО:', None, None, None, None, None, 'УТВЕРЖДАЕМ:', None, None, None, None],
                [None, "Первый заместитель генерального директора -\n главный инженер ООО 'Башнефть-Добыча'  ",
                 None, None, None, None, None,
                 chief_engineer_post, None, None, None, None],
                [None, f'_________________________Д.А.Чувакин', None, None, None,
                 None, None,
                 f'_____________{chief_engineer_surname}', None, None, None,
                 None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, None, None, None,
                 f'"____"_____________________{current_datetime.year}г.', None, None, None, None],
                [None, None, None, None, " ", None, None, None, None, None, None, None],
                [None, 'Заместитель генерального директора - \nглавный геолог  ООО "Башнефть-Добыча"  ', None, None,
                 None,
                 None, None, f'{chief_geologist_post}', None, None, None, None],
                [None, f'__________________________И.Р. Баширов ', None, None, None, None, None,
                 f'_____________{chief_geologist_surname}', None, None, '',
                 None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, '', None, None,
                 f'"____"_____________________{current_datetime.year}г.', None, None, None, None],
                [None, None, None, None, " ", None, None, None, None, None, None, None],
                [None, 'Начальник управления добычи нефти и газа ООО "Башнефть-Добыча" ', None, None, None,
                 None, None, None, None,
                 None, None, None],
                [None, f'__________________________М.А.Тенюнин', None, None, None,
                 None, None, None, None, None, None, None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, '', None, None,
                 None, None, None, None, None],
                [None, None, None, None, " ", None, None, None, None, None, None, None],
                [None, 'Начальник отдела-заместитель начальника Управления супервайзинга \nремонта скважин и '
                       'скважинных технологий ООО "Башнефть-Добыча"', None, None, None, None, None, None, None,
                 None, None, None],
                [None, f'__________________________А.Ю.Пензин', None, None, None,
                 None, None, None, None, None, None, None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, '', None, None,
                 None, None, None, None, None],
                [None, None, None, None, " ", None, None, None, None, None, None, None],
                [None, podpis_dict[data_list.costumer][region]['gi']['post'], None, None, None, None, None,
                 None, None, None, None, None],
                [None, f'____________{podpis_dict[data_list.costumer][region]["gi"]["surname"]}', None, None, None,
                 None, None, None, None, None, None, None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, None, None, None,
                 None, None, None, None, None],
                [None, None, None, None, None, None, None, power_of_attorney, None, None, None, None],
                [None, podpis_dict[data_list.costumer][region]['gg']['post'], None, None, None,
                 None, None, None, None, None, None, None],
                [None, f'_____________{podpis_dict[data_list.costumer][region]["gg"]["surname"]}', None, None, None,
                 None, None, None, None, None, None, None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, '', None, None,
                 None, None, None, None, None],
            ]

        else:
            work_podpisant_list = [
                [None, 'СОГЛАСОВАНО:', None, None, None, None, None, 'УТВЕРЖДАЕМ:', None, None, None, None],
                [None, podpis_dict[data_list.costumer][region]['gi']['post'], None, None, None, None, None,
                 chief_engineer_post, None, None, None, None],
                [None, f'____________{podpis_dict[data_list.costumer][region]["gi"]["surname"]}', None, None, None,
                 None, None,
                 f'__________________________{chief_engineer_surname}', None, None, None,
                 None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, None, None, None,
                 f'"____"_____________________{current_datetime.year}г.', None, None, None, None],
                [None, None, None, None, None, None, None, power_of_attorney, None, None, None, None],
                [None, podpis_dict[data_list.costumer][region]['gg']['post'], None, None, None,
                 None, None, f'{chief_geologist_post}', None, None, None, None],
                [None, f'_____________{podpis_dict[data_list.costumer][region]["gg"]["surname"]}', None, None, None,
                 None,
                 None,
                 f'__________________________{chief_geologist_surname}', None, None, '',
                 None],
                [None, f'"____"_____________________{current_datetime.year}г.', None, None, '', None, None,
                 f'"____"_____________________{current_datetime.year}г.', None, None, None, None],
            ]

        podp_bvo = [
            [None, None, None, None, " ", None, None, None, None, None, None, None],
            [None, 'Районный инженер Башкирского ', None, None, None, None, None, None, None, None, None, None],
            [None, 'военизированного отряда ', None, None, None, None, None, None, None, None, None, None],
            [None, '__________________________', None, None, None, None, None, None, None, None, None, None],
            [None, f'"____"_____________________{current_datetime.year}г.', None, None, None, None, None, None,
             None, None, None,
             None],
            [None, None, None, None, " ", None, None, None, None, None, None, None]]

        if data_list.data_in_base is False:
            if len(self.data_well.plast_work) != 0:

                try:
                    cat_P_1 = self.data_well.dict_category[
                        self.data_well.plast_work[0]
                    ]["по давлению"].category
                    category_h2s_list = self.data_well.dict_category[
                        self.data_well.plast_work[0]
                    ]["по сероводороду"].category
                    cat_gaz = self.data_well.dict_category[
                        self.data_well.plast_work[0]
                    ]["по газовому фактору"].category
                except:
                    cat_P_1 = self.data_well.category_pressure_well[0]
                    category_h2s_list = self.data_well.category_h2s_list[0]
                    cat_gaz = self.data_well.category_gaz_factor_percent[0]
            else:
                cat_P_1 = self.data_well.category_pressure_well[0]
                category_h2s_list = self.data_well.category_h2s_list[0]
                cat_gaz = self.data_well.category_gaz_factor_percent[0]
            try:
                cat_P_1_plan = self.data_well.dict_category[
                    self.data_well.plast_project[0]
                ]["по давлению"].category
                category_h2s_list_plan = self.data_well.dict_category[
                    self.data_well.plast_project[0]
                ]["по сероводороду"].category
                cat_gaz_plan = self.data_well.dict_category[
                    self.data_well.plast_project[0]
                ]["по газовому фактору"].category
            except:
                cat_P_1_plan = 3
                category_h2s_list_plan = 3
                cat_gaz_plan = 3

            if (
                    1
                    in [
                cat_P_1,
                cat_P_1_plan,
                category_h2s_list,
                cat_gaz,
                category_h2s_list_plan,
                cat_gaz_plan,
                self.data_well.category_pressure,
            ]
                    or "1"
                    in [
                cat_P_1,
                cat_P_1_plan,
                category_h2s_list,
                cat_gaz,
                category_h2s_list_plan,
                cat_gaz_plan,
                self.data_well.category_pressure,
            ]
                    or self.data_well.curator == "ВНС"
            ):
                work_podpisant_list.extend(podp_bvo)

        work_podpisant_list.extend([
            [None, None, None, None, " ", None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None]])

        return work_podpisant_list

    @staticmethod
    def close_process():
        for proc in psutil.process_iter():
            if proc.name() == "ZIMA.exe":
                proc.terminate()  # Принудительное завершение

    @staticmethod
    def show_info_message(data_well, message):
        from work_py.check_in_pz import CustomMessageBox

        dialog = CustomMessageBox(data_well, message)
        dialog.exec_()

    @staticmethod
    def check_if_none(value):
        if isinstance(value, int) or isinstance(value, float):
            return int(value)

        elif str(value).replace(".", "").replace(",", "").isdigit():
            if str(round(float(value.replace(",", ".")), 1))[-1] == 0:
                # print(str(round(float(value.replace(',', '.')), 1)))
                return int(float(value.replace(",", ".")))
            else:
                return float(value.replace(",", "."))
        else:
            return 0

    def definition_open_trunk_well(self):
        self.data_well.nkt_diam = 73 if self.data_well.column_diameter.get_value > 110 else 60
        self.data_well.nkt_template = 59.6 if self.data_well.column_diameter.get_value > 110 else 47.9

        if self.data_well.column_additional:
            if self.data_well.current_bottom > self.data_well.shoe_column_additional.get_value:
                self.data_well.open_trunk_well = True
            else:
                self.data_well.open_trunk_well = False
        else:
            if self.data_well.current_bottom > self.data_well.shoe_column.get_value:
                self.data_well.open_trunk_well = True
            else:
                self.data_well.open_trunk_well = False

    def normalization(self, current_depth, diameter_paker, gis_otz_after_true_quest):
        from work_py.alone_oreration import kot_select

        nkt_diam = self.data_well.nkt_diam

        normalization_list = [
            [
                f"Согласовать алгоритм нормализации письменной телефонограммой до H- {current_depth}м",
                None,
                f"Алгоритм работ согласовать с Заказчиком письменной телефонограммой: \n"
                f"В случае наличия ЗУМПФА менее 10м произвести работы по нормализации забоя,"
                f" в обратном случае согласовать с закачиком работы по нормализации не производить.\n",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Мастер КРС",
                None,
            ],
            [
                None,
                None,
                f"Спустить {kot_select(self, current_depth)} на НКТ{nkt_diam}мм до глубины текущего забоя"
                f" с замером, шаблонированием шаблоном {self.data_well.nkt_template}мм.",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "мастер КРС",
                descentNKT_norm(self.data_well.current_bottom, 1),
            ],
            [
                None,
                None,
                f"Произвести очистку забоя скважины до гл.{current_depth}м закачкой обратной промывкой тех жидкости"
                f" уд.весом {self.data_well.fluid_work}, по согласованию с Заказчиком",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "мастер КРС",
                0.4,
            ],
            [
                None,
                None,
                f"При необходимости согласовать закачку блок пачки по технологическому плану работ подрядчика",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "мастер КРС, предст. заказчика",
                None,
            ],
            [
                None,
                None,
                f"Поднять {kot_select(self, current_depth)} на НКТ{nkt_diam}мм c глубины {current_depth}м с "
                f"доливом скважины в "
                f"объеме {round(current_depth * 1.12 / 1000, 1)}м3 удельным весом {self.data_well.fluid_work}",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "мастер КРС",
                lifting_nkt_norm(current_depth, 1),
            ],
            [
                None,
                None,
                f"В случае наличия ЗУМПФа 10м и более продолжить работы с п. по отбивки забоя "
                f"В случае ЗУМПФа менее 10м: и не жесткая посадка компоновки СПО КОТ повторить. "
                f"В случае образование твердой корки (жесткой посадки): выполнить взрыхление ПМ с ВЗД"
                f" и повторить работы СПО КОТ.",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Мастер КРС",
                None,
            ],
            [
                None,
                None,
                f"Спустить компоновку с замером и шаблонированием НКТ:  долото Д="
                f"{diameter_paker + 2}мм, забойный двигатель,"
                f" НКТ - 20м, вставной фильтр, НКТмм до кровли проппантной пробки. "
                f"(При СПО первых десяти НКТ на спайдере дополнительно устанавливать элеватор ЭХЛ) ",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Мастер КРС",
                descentNKT_norm(current_depth, 1.2),
            ],
            [
                None,
                None,
                f"Подогнать рабочую трубу патрубками на заход 9-10м. Вызвать циркуляцию прямой промывкой. "
                f"Произвести допуск с прямой промывкой и рыхление проппантной пробки 10м с проработкой э/колонны по 10 раз. ",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Мастер КРС",
                0.9,
            ],
            [
                None,
                None,
                f"Поднять компоновку с глубины {current_depth}м с доливом скважины тех.жидкостью уд. весом"
                f" {self.data_well.fluid_work}  в объеме "
                f"{round(self.data_well.current_bottom * 1.12 / 1000, 1)}м3",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Мастер КРС",
                lifting_nkt_norm(current_depth, 1.2),
            ],
            [
                f"по согласованию с заказчиком: Отбивка забоя",
                None,
                f"по согласованию с заказчиком: \n"
                f'Вызвать геофизическую партию. Заявку оформить за 16 часов сутки через РИТС {data_list.contractor}". '
                f"Произвести монтаж ПАРТИИ ГИС согласно схемы №11 утвержденной главным инженером "
                f'{data_list.DICT_CONTRACTOR[data_list.contractor]["Дата ПВО"]}г. '
                f"ЗАДАЧА 2.8.2 Отбить забой по ГК и ЛМ",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Мастер КРС, подрядчик по ГИС",
                4,
            ],
        ]

        if gis_otz_after_true_quest == "Нет":
            normalization_list = normalization_list[:-1]
        self.data_well.current_bottom = current_depth

        return normalization_list

    def check_gpp_upa(self, table_widget):
        for row in range(table_widget.rowCount()):
            for column in range(table_widget.columnCount()):
                value = self.table_widget.item(row, column)
                if value is not None:
                    value = value.text()
                    if "Установить подъёмный агрегат на устье не менее 40т" in value:
                        new_value = QtWidgets.QTableWidgetItem(
                            f"Установить подъёмный агрегат на устье не менее 60т. "
                            f"Пусковой комиссией составить акт готовности подъемного "
                            f"агрегата и бригады для проведения ремонта скважины."
                        )
                        table_widget.setItem(row, column, new_value)

    @staticmethod
    def check_str_isdigit(string):
        # Паттерн для проверки: допустимы только цифры, точка и запятая
        pattern = r"^[\d.,]+$"
        # Проверка строки на соответствие паттерну
        if re.match(pattern, str(string)):
            return True
        else:
            return False

    @staticmethod
    def check_once_isdigit(string):
        # Паттерн для проверки: допустимы только цифры, точка и запятая
        string_exit = ""
        for i in str(string):
            if i in "1234567890,.":
                string_exit += i
        if string_exit.replace(",", "") == "":
            return None
        else:
            string_exit = string_exit.replace(",", ".")
            if string_exit.count(".") > 1:
                QMessageBox.warning(
                    None,
                    "ошибка в ячейке ПЗ",
                    f"ошибка в ячейке ПЗ c данными {string} "
                    f"приложение ожидает только одну точку",
                )
                return
            if string_exit.replace(".", "").isdigit():
                return float(string_exit)

    # def move_window(self, x, y):
    #     self.move(x, y)
    def add_window(self, window):
        if self.operation_window is None:
            data_list.pause = True
            self.operation_window = window(self.data_well, self.table_widget)
            self.operation_window.move(100, 100)
            self.set_modal_window(self.operation_window)
            self.operation_window = None
        else:
            self.operation_window.close()  # Close window.
            self.operation_window = None

    @staticmethod
    def insert_image(ws, file, coordinate, width=200, height=180):
        # Загружаем изображение с помощью библиотеки Pillow

        img = openpyxl.drawing.image.Image(file)
        img.width = width
        img.height = height
        img.anchor = coordinate
        ws.add_image(img, coordinate)

    def set_modal_window(self, window):
        self.data_well.modal_dialog = ModalDialog(window)
        self.data_well.modal_dialog.setModal(True)  # Установка модальности
        self.data_well.modal_dialog.show()

    def close_modal_forcefully(self):
        if self.data_well.modal_dialog:
            self.data_well.modal_dialog.close()  # Закрытие модального окна

    def read_excel_file(self):
        from find import FindIndexPZ
        from work_py.leakage_column import LeakageWindow

        from find import (
            WellNkt,
            WellPerforation,
            WellCondition,
            WellHistoryData,
            WellName,
            WellCategory,
            WellFondData,
            WellSuckerRod,
            WellExpectedPickUp,
            WellData,
        )

        # Запуск основного класса и всех дочерних классов в одной строке
        self.data_well = FindIndexPZ(self.ws, self.work_plan, self)
        if self.data_well is False:
            return
        self.data_well.work_plan = self.work_plan
        data_well_dict = {}
        self.data_well = WellName.read_well(
            self.data_well,
            self.data_well.cat_well_max.get_value,
            self.data_well.data_pvr_min.get_value,
        )
        if self.data_well is None:
            return None
        if self.data_well.cdng.get_value not in dict_data_cdng:
            return

        self.data_well.region = region_select(self.data_well.cdng.get_value)

        date_str2 = datetime.strptime("2024-09-19", "%Y-%m-%d")

        if self.work_plan == "dop_plan":

            from data_base.config_base import connection_to_database, WorkDatabaseWell

            number_list = list(map(str, range(1, 50)))
            self.data_well.number_dp, ok = QInputDialog.getItem(
                self,
                "Номер дополнительного плана работ",
                "Введите номер дополнительного плана работ",
                number_list,
                0,
                False,
            )

            db = connection_to_database(decrypt("DB_WELL_DATA"))
            data_well_base = WorkDatabaseWell(db, self.data_well)

            data_well = data_well_base.check_in_database_dp_data(
                self.data_well.well_number.get_value,
                self.data_well.well_area.get_value,
                f"ДП№{self.data_well.number_dp}",
            )

            if data_well:
                date_str1 = datetime.strptime(f"{data_well[1]}", "%Y-%m-%d")
                if date_str1 > date_str2:

                    change_work_work_plan = QMessageBox.question(
                        self,
                        "Наличие в базе данных",
                        "Проверка показала что данные по скважине есть в"
                        " базе данных, "
                        "загрузить с базы?",
                    )

                    if change_work_work_plan == QMessageBox.StandardButton.Yes:
                        self.data_well.type_kr = data_well[2]
                        self.data_well.work_plan = "dop_plan_in_base"
                        self.data_well.work_plan = "dop_plan_in_base"
                        data_list.data_in_base = True
                        from work_py.dop_plan_py import DopPlanWindow

                        self.rir_window = DopPlanWindow(
                            self.data_well.insert_index, None, self.data_well.work_plan
                        )
                        # self.rir_window.setGeometry(200, 400, 100, 200)
                        self.rir_window.show()
                        data_list.pause = True

                        return

        if data_list.data_well_is_True is False:
            # Сохранение изменений

            WellNkt.read_well(
                self.data_well,
                self.data_well.pipes_ind.get_value,
                self.data_well.condition_of_wells.get_value,
            )

            WellHistoryData.read_well(
                self.data_well,
                self.data_well.data_pvr_max.get_value,
                self.data_well.data_fond_min.get_value,
            )
            # Сохранение изменений

            if "prs" not in self.work_plan:
                WellCondition.read_well(
                    self.data_well,
                    self.data_well.condition_of_wells.get_value,
                    self.data_well.data_well_max.get_value,
                )
            else:
                WellCondition.read_well(
                    self.data_well,
                    self.data_well.data_pvr_max.get_value,
                    self.data_well.data_well_max.get_value,
                )

            WellExpectedPickUp.read_well(
                self.data_well,
                self.data_well.data_x_min.get_value,
                self.data_well.data_x_max.get_value,
            )
            if self.work_plan not in ["application_pvr", "application_gis"]:
                read_data = WellSuckerRod.read_well(
                    self.data_well,
                    self.data_well.sucker_rod_ind.get_value,
                    self.data_well.pipes_ind.get_value,
                )
                if read_data is None:
                    self.data_well = None
                    return

                read_data = WellFondData.read_well(
                    self.data_well,
                    self.data_well.data_fond_min.get_value,
                    self.data_well.pipes_ind.get_value,
                )
                if read_data is None:
                    self.data_well = None
                    return

            read_data = WellData.read_well(
                self.data_well,
                self.data_well.cat_well_max.get_value,
                self.data_well.data_pvr_min.get_value,
            )

            if self.data_window is None:
                from data_correct import DataWindow
                data_list.pause = True

                self.data_window = DataWindow(self.data_well)
                self.data_window.setWindowTitle("Сверка данных")
                self.data_window.setGeometry(100, 100, 300, 400)

                self.data_window.show()
                self.pause_app()
                data_list.pause = True
                self.data_window = None

            if data_list.operation_window is False:
                return

            WellPerforation.read_well(
                self.data_well,
                self.data_well.data_pvr_min.get_value,
                self.data_well.data_pvr_max.get_value + 1,
            )

            from perforation_correct import PerforationCorrect

            self.perforation_correct_window2 = PerforationCorrect(self.data_well)
            self.perforation_correct_window2.setWindowTitle("Сверка данных перфорации")
            # self.perforation_correct_window2.setGeometry(200, 400, 100, 400)

            self.set_modal_window(self.perforation_correct_window2)
            self.pause_app()
            data_list.pause = True
            self.perforation_correct_window2 = None
            # definition_plast_work(self)

            if data_list.operation_window is False or self.data_well is None:
                return

            self.data_well = WellCategory.read_well(
                self.data_well,
                self.data_well.cat_well_min.get_value,
                self.data_well.data_well_min.get_value,
            )
            if data_list.operation_window is False:
                return

            if self.data_well.leakiness is True:
                if WellCondition.leakage_window is None:
                    WellCondition.leakage_window = LeakageWindow(self.data_well)
                    WellCondition.leakage_window.setWindowTitle(
                        "Геофизические исследования"
                    )
                    self.set_modal_window(WellCondition.leakage_window)
                    # WellCondition.leakage_window.setGeometry(200, 400, 300, 400)
                    WellCondition.leakage_window.show()
                    # self.data_well.dict_leakiness = WellCondition.leakage_window.add_work()
                    self.pause_app()

                    # print(f'словарь нарушений {self.data_well.dict_leakiness}')
                    data_list.pause = True
                    WellCondition.leakage_window = None  # Discard reference.
            if data_list.operation_window is False:
                return
            if data_list.connect_in_base:
                params = {
                    "well_number": self.data_well.well_number.get_value,
                    "well_area": self.data_well.well_area.get_value,
                    "well_oilfield": self.data_well.well_oilfield.get_value,
                    "costumer": data_list.costumer,
                    "cdng": self.data_well.cdng.get_value,
                    "inventory_number": str(self.data_well.inventory_number.get_value),
                    "wellhead_fittings": str(self.data_well.wellhead_fittings),
                    "appointment": str(self.data_well.appointment_well.get_value),
                    "angle_data": {"инклинометрия": self.data_well.angle_data},
                    "column_direction": {
                        "diameter": self.data_well.column_direction_diameter.get_value,
                        "wall_thickness": self.data_well.column_direction_wall_thickness.get_value,
                        "head": 0,
                        "shoe": self.data_well.column_direction_length.get_value,
                        "level_cement": self.data_well.level_cement_direction.get_value,
                    },
                    "column_conductor": {
                        "diameter": self.data_well.column_conductor_diameter.get_value,
                        "wall_thickness": self.data_well.column_conductor_wall_thickness.get_value,
                        "head": 0,
                        "shoe": self.data_well.column_conductor_length.get_value,
                        "level_cement": self.data_well.level_cement_conductor.get_value,
                    },
                    "column_production": {
                        "diameter": self.data_well.column_diameter.get_value,
                        "wall_thickness": self.data_well.column_wall_thickness.get_value,
                        "head": self.data_well.head_column.get_value,
                        "shoe": self.data_well.shoe_column.get_value,
                        "level_cement": self.data_well.level_cement_column.get_value,
                    },
                    "column_additional": {
                        "diameter": self.data_well.column_additional_diameter.get_value,
                        "wall_thickness": self.data_well.column_additional_wall_thickness.get_value,
                        "head": self.data_well.head_column_additional.get_value,
                        "shoe": self.data_well.shoe_column_additional.get_value,
                        "level_cement": 0,
                    },
                    "bottom_hole_drill": float(
                        self.data_well.bottom_hole_drill.get_value
                    ),
                    "bottom_hole_artificial": float(
                        self.data_well.bottom_hole_artificial.get_value
                    ),
                    "max_angle": float(self.data_well.max_angle.get_value),
                    "distance_from_rotor_table": self.data_well.distance_from_well_to_sampling_point,
                    "max_angle_depth": float(self.data_well.max_angle_depth.get_value),
                    "max_expected_pressure": float(
                        self.data_well.max_expected_pressure.get_value
                    ),
                    "max_admissible_pressure": float(
                        self.data_well.max_admissible_pressure.get_value
                    ),
                    "rotor_altitude": 0.0,
                    "perforation": self.data_well.dict_perforation,
                    "equipment": self.data_well.data_well_dict["оборудование"],
                    "nkt_data": self.data_well.data_well_dict["НКТ"],
                    "sucker_pod": self.data_well.data_well_dict["штанги"],
                    "diameter_doloto_ek": float(
                        self.data_well.diameter_doloto_ek.get_value
                    ),
                    "last_pressure_date": self.data_well.result_pressure_date.get_value,
                    "date_commissioning": self.data_well.date_commissioning.get_value,
                    "date_drilling_run": self.data_well.date_drilling_run,
                    "date_drilling_finish": self.data_well.date_drilling_cancel,
                    "geolog": data_list.user[1],
                    "date_create": data_list.current_date,
                    "leakiness": self.data_well.dict_leakiness,
                    "contractor": data_list.contractor
                }

                response_find_data = ApiClient.find_wells(
                    self.data_well.well_number.get_value,
                    self.data_well.well_area.get_value,
                    ApiClient.find_wells_data_response_filter_well_number_well_area(),
                )
                if response_find_data:
                    mes = QMessageBox.question(
                        self,
                        "данные по скважине",
                        "Данные есть в базе данных, обновить?",
                    )
                    if mes == QMessageBox.StandardButton.Yes:
                        response = ApiClient.request_put_json(
                            ApiClient.update_wells_data_response(),
                            params,
                            {
                                "well_number": self.data_well.well_number.get_value,
                                "well_area": self.data_well.well_area.get_value
                            },
                            "json"
                        )
                    else:
                        response = True
                else:
                    response = ApiClient.request_post_json(
                        ApiClient.read_wells_data_response_for_add(),
                        params,
                        None,
                        "json",
                    )
                if response is None:
                    QMessageBox.warning(self, "ошибка", "скважина не добавлена в well_data")
                    return

            if self.data_well.emergency_well is True:
                emergency_quest = QMessageBox.question(
                    self,
                    "Аварийные работы ",
                    "Программа определила что в скважине"
                    f" авария - {self.data_well.emergency_count}, верно ли?",
                )
                if emergency_quest == QMessageBox.StandardButton.Yes:
                    self.data_well.emergency_well = True
                    self.data_well.emergency_bottom, ok = QInputDialog.getInt(
                        self,
                        "Аварийный забой",
                        "Введите глубину аварийного забоя",
                        0,
                        0,
                        int(self.data_well.bottom_hole_artificial.get_value),
                    )
                else:
                    self.data_well.emergency_well = False
            if self.data_well.problem_with_ek is True:
                problem_with_ek_quest = QMessageBox.question(
                    self,
                    "ВНИМАНИЕ НЕПРОХОД ",
                    f"Программа определила что в скважине "
                    f"сужение в ЭК -, верно ли?",
                )
                if problem_with_ek_quest == QMessageBox.StandardButton.Yes:
                    self.data_well.problem_with_ek = True
                    self.data_well.problem_with_ek_depth, ok = QInputDialog.getInt(
                        self,
                        "Глубина сужения",
                        "Введите глубину cсужения",
                        0,
                        0,
                        int(self.data_well.current_bottom),
                    )
                    self.data_well.problem_with_ek_diameter = QInputDialog.getInt(
                        self,
                        "диаметр внутренний cсужения",
                        "ВВедите внутренний диаметр cсужения",
                        0,
                        0,
                        int(float(self.data_well.current_bottom)),
                    )[0]
                else:
                    self.data_well.problem_with_ek = False

            if self.data_well.gips_in_well is True:
                gips_true_quest = QMessageBox.question(
                    self,
                    "Гипсовые отложения",
                    "Программа определила что скважина осложнена гипсовыми "
                    "отложениями "
                    "и требуется предварительно определить забой на НКТ, верно ли "
                    "это?",
                )

                if gips_true_quest == QMessageBox.StandardButton.Yes:
                    self.data_well.gips_in_well = True
                else:
                    self.data_well.gips_in_well = False

        if (
                self.data_well.inventory_number.get_value == "не корректно"
                or self.data_well.inventory_number is None
        ):
            QMessageBox.warning(
                self,
                "Инвентарный номер отсутствует",
                "Необходимо уточнить наличие инвентарного номера",
            )
            return

        return self.data_well

    def determination_injection_pressure(self):
        if self.data_well.region == "ЧГМ" and self.data_well.expected_pressure < 80:
            return 80
        else:
            return self.data_well.expected_pressure

    def privyazka_nkt(self):
        angle_text = ''
        priv_list = []
        if self.data_well.angle_data and self.data_well.max_angle.get_value > 60:
            angle_text = self.calculate_angle_grad(self.data_well.angle_data)
            priv_list = [
                [
                    None,
                    None,
                    angle_text,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Мастер КРС",
                    None,
                ]]

        priv_list.append(
            [
                f"ГИС Привязка по ГК и ЛМ",
                None,
                f"Вызвать геофизическую партию. Заявку оформить за 16 часов сутки через РИТС "
                f'{data_list.contractor}". '
                f"Произвести  монтаж ПАРТИИ ГИС согласно схемы №11 утвержденной главным инженером "
                f'{data_list.DICT_CONTRACTOR[data_list.contractor]["Дата ПВО"]}г. '
                f"ЗАДАЧА 2.8.1 Привязка технологического оборудования скважины",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Мастер КРС, подрядчик по ГИС",
                4,
            ]
        )
        return priv_list

    def open_read_excel_file_pz(self):
        from open_pz import CreatePZ
        from data_base.work_with_base import insert_data_new_excel_file
        from work_py.correct_plan import CorrectPlanWindow
        from work_py.dop_plan_py import DopPlanWindow
        from work_py.gnkt_grp import GnktOsvWindow
        from work_py.gnkt_frez import WorkWithGnkt
        from find import FindIndexPZ

        if self.work_plan in [
            "krs",
            "dop_plan",
            "gnkt_opz",
            "gnkt_after_grp",
            "gnkt_bopz",
            "gnkt_frez",
            "prs",
        ]:
            QMessageBox.information(
                self,
                "ВНИМАНИЕ",
                "Для корректного прочтения план заказа, план заказ должен быть "
                "пересохранен в формат .xlsx (КНИГА EXCEL, "
                "excel версия от 2010г и выше)",
            )
            self.fname, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "Выберите файл", ".", "Файлы Exсel (*.xlsx);;Файлы Exсel (*.xls)"
            )

            if self.fname:
                try:
                    self.read_pz(self.fname)
                    data_list.pause = True
                    self.data_well = self.read_excel_file()
                    if self.data_well is None:
                        return
                    read_pz = CreatePZ(self.data_well, self.ws, self)

                except FileNotFoundError as f:
                    QMessageBox.warning(
                        self, "Ошибка", f"Ошибка при прочтении файла {f}"
                    )
                    return

                self.wb2_prs = Workbook()
                self.ws2_prs = self.wb2_prs.active

                if self.work_plan in ["krs", "dop_plan"]:
                    if self.wb2_prs:
                        self.ws = read_pz.open_excel_file(
                            self.ws, self.work_plan, self.ws2_prs
                        )
                        self.copy_pz(self.ws, self.table_widget, self.work_plan)
                        if self.work_plan == "dop_plan":
                            self.rir_window = DopPlanWindow(
                                self.data_well, self.table_widget
                            )
                            self.set_modal_window(self.rir_window)

                            data_list.pause = True

                elif self.work_plan in ["gnkt_opz", "gnkt_after_grp", "gnkt_bopz"]:

                    self.gnkt_data = GnktOsvWindow(
                        self.ws,
                        self.table_title,
                        self.table_schema,
                        self.table_widget,
                        self.data_well,
                    )
                elif self.work_plan == "gnkt_frez":
                    self.gnkt_data = WorkWithGnkt(
                        self.ws,
                        self.table_title,
                        self.table_schema,
                        self.table_widget,
                        self.data_well,
                    )
                elif self.work_plan in ["prs"]:
                    if self.wb2_prs:
                        self.ws = read_pz.open_excel_file(
                            self.ws, self.work_plan, self.ws2_prs
                        )
                        self.copy_pz(self.ws, self.table_widget, self.work_plan, 15)

        elif self.work_plan in ["plan_change", "dop_plan_in_base"]:
            data_list.data_in_base = True
            self.data_well = FindIndexPZ

            if self.work_plan == "plan_change":
                self.data_well.work_plan = self.work_plan
                # Деактивируем кнопку до загрузки ID
                if hasattr(self, 'signPlanButton'):
                    self.signPlanButton.setEnabled(False)
                self.rir_window = CorrectPlanWindow(self.data_well, self.table_widget)

            elif self.work_plan == "dop_plan_in_base":
                self.data_well.work_plan = self.work_plan
                self.rir_window = DopPlanWindow(self.data_well, self.table_widget)

            self.rir_window.setGeometry(200, 400, 800, 200)
            self.set_modal_window(self.rir_window)
            data_list.pause = True
            self.pause_app()

            data_list.pause = True

            self.ws = insert_data_new_excel_file(
                self,
                data_list.data,
                data_list.row_heights,
                data_list.col_width,
                data_list.boundaries_dict,
            )

            self.copy_pz(self.ws, self.table_widget, self.work_plan)

        data_list.pause = True
        self.rir_window = None

        return self.ws

    def read_pz(self, fname):
        self.wb = load_workbook(fname, data_only=True)
        name_list = self.wb.sheetnames
        self.ws = self.wb.active

    def definition_filenames(self):
        contractor = "Ойл"
        if "РН" in data_list.contractor:
            contractor = "РН"
        elif "Ойл" in data_list.contractor:
            contractor = "Ойл"

        if self.data_well.work_plan in ["dop_plan", "dop_plan_in_base"]:
            string_work = f" ДП№ {self.data_well.number_dp}"
        elif self.data_well.work_plan == "krs":
            string_work = "ПР"
        elif self.data_well.work_plan == "plan_change":
            if self.data_well.work_plan_change == "krs":
                string_work = "ПР изм"
            else:
                string_work = f"ДП№{self.data_well.number_dp} изм "

        elif self.data_well.work_plan == "gnkt_bopz":
            string_work = "ГНКТ БОПЗ ВНС"
        elif self.data_well.work_plan == "gnkt_opz":
            string_work = "ГНКТ ОПЗ"
        elif self.data_well.work_plan == "gnkt_after_grp":
            string_work = "ГНКТ ОСВ ГРП"
        elif self.data_well.work_plan == "prs":
            string_work = "ПРС"
        else:
            string_work = "ГНКТ"

        filenames = (
            f"{self.data_well.well_number.get_value} {self.data_well.well_area.get_value} "
            f'{self.data_well.type_kr.split(" ")[0]} кат {self.data_well.category_pressure}'
            f" {string_work} {contractor}.xlsx"
        )
        return filenames

    @staticmethod
    def calculate_chemistry(type_chemistry, volume):
        if type_chemistry in ["HCl", "Нефтекислотка"]:
            data_list.DICT_VOLUME_CHEMISTRY["HCl"] += volume
        elif type_chemistry == "HF":
            data_list.DICT_VOLUME_CHEMISTRY["HF"] += volume
        elif type_chemistry == "ВТ":
            data_list.DICT_VOLUME_CHEMISTRY["ВТ СКО"] += volume
        elif type_chemistry == "Противогипсовая обработка":
            data_list.DICT_VOLUME_CHEMISTRY["NaOH"] += volume
        elif type_chemistry == "цемент":
            data_list.DICT_VOLUME_CHEMISTRY["цемент"] += volume
        elif type_chemistry == "Глина":
            data_list.DICT_VOLUME_CHEMISTRY["Глина"] += volume
        elif type_chemistry == "растворитель":
            data_list.DICT_VOLUME_CHEMISTRY["растворитель"] += volume
        elif type_chemistry == "песок":
            data_list.DICT_VOLUME_CHEMISTRY["песок"] += volume
        elif type_chemistry == "РПК":
            data_list.DICT_VOLUME_CHEMISTRY["РПК"] += 1
        elif type_chemistry == "РИР 2С":
            data_list.DICT_VOLUME_CHEMISTRY["РИР 2С"] += 1
        elif type_chemistry == "РИР ОВП":
            data_list.DICT_VOLUME_CHEMISTRY["РИР ОВП"] += 1
        elif type_chemistry == "гидрофабизатор":
            data_list.DICT_VOLUME_CHEMISTRY["гидрофабизатор"] += volume

    def save_file_dialog(self, wb2, full_path):
        while True:  # Начинаем бесконечный цикл
            try:
                file_name, _ = QFileDialog.getSaveFileName(
                    None, "Save excel-file", f"{full_path}", "Excel Files (*.xlsm)"
                )
                if file_name:  # Если имя файла не пустое
                    wb2.save(file_name)  # Пытаемся сохранить
                    break  # Если сохранение успешно, выходим из цикла

            except Exception as e:
                QMessageBox.critical(
                    None,
                    "Ошибка",
                    f"файл под таким именем открыт, закройте его: {type(e).__name__}\n  {str(e)}",
                )
        try:
            # Создаем объект Excel
            # Создаем объект Excel
            excel = win32com.client.Dispatch("Excel.Application")
            # win32api.SetFileAttributes(file_name, win32con.FILE_ATTRIBUTE_NORMAL)
            # Открываем файл
            workbook = excel.Workbooks.Open(file_name)
            # Выбираем активный лист
            worksheet = workbook.ActiveSheet

            # Назначаем область печати с колонок B до L
            worksheet.PageSetup.PrintArea = "B:L"

        except Exception as e:
            print(f"Ошибка при работе с Excel: {type(e).__name__}\n\n{str(e)}")

    @staticmethod
    def pause_app():
        while data_list.pause is True:
            QtCore.QCoreApplication.instance().processEvents()

    def check_pvo_schema(self, ws5, ind):
        global schema_path
        schema_pvo_set = set()
        for row in range(self.table_widget.rowCount()):
            if row > ind:
                for column in range(self.table_widget.columnCount()):
                    value = self.table_widget.item(row, column)
                    if value is not None:
                        value = value.text()
                        if (
                                "схеме №" in value
                                or "схемы №" in value
                                or "Схемы обвязки №" in value
                                or "схемы ПВО №" in value
                        ):
                            number_schema = value[
                                            value.index(" №") + 1: value.index(" №") + 4
                                            ].replace(" ", "")

                            schema_pvo_set.add(number_schema)
        # print(f'схема ПВО {schema_pvo_set}')

        n = 0
        if schema_pvo_set:
            for schema in list(schema_pvo_set):
                coordinate = f"{get_column_letter(2)}{1 + n}"
                if "Ойл" in data_list.contractor:
                    schema_path = (
                        f"{data_list.path_image}imageFiles/pvo/oil/схема {schema}.png"
                    )
                elif "РН" in data_list.contractor:
                    schema_path = (
                        f"{data_list.path_image}imageFiles/pvo/rn/схема {schema}.png"
                    )
                try:
                    img = openpyxl.drawing.image.Image(schema_path)
                    img.width = 750
                    img.height = 530
                    img.anchor = coordinate
                    ws5.add_image(img, coordinate)
                    n += 29
                except FileNotFoundError as f:
                    QMessageBox.warning(
                        None, "Ошибка", f"Схему {schema} не удалось вставилось:\n {f}"
                    )
            ws5.print_area = f"B1:M{n}"
            ws5.page_setup.fitToPage = True
            ws5.page_setup.fitToHeight = False
            ws5.page_setup.fitToWidth = True
            ws5.print_options.horizontalCentered = True
            # зададим размер листа
            ws5.page_setup.paperSize = ws5.PAPERSIZE_A4
            # содержимое по ширине страницы
            ws5.sheet_properties.pageSetUpPr.fitToPage = True
            ws5.page_setup.fitToHeight = False
            # Переместите второй лист перед первым

        return list(schema_pvo_set)

    def insert_data_in_database(self, row_number, row_max):

        dict_perforation_json = json.dumps(
            ApiClient.serialize_datetime(self.data_well.dict_perforation),
            default=str,
            ensure_ascii=False,
            indent=4,
        )
        # print(self.data_well.dict_leakiness)
        leakage_json = json.dumps(
            self.data_well.dict_leakiness, default=str, ensure_ascii=False, indent=4
        )
        plast_all_json = json.dumps(self.data_well.plast_all)
        plast_work_json = json.dumps(self.data_well.plast_work)
        skm_list_json = json.dumps(self.data_well.skm_interval)
        raiding_list_json = json.dumps(self.data_well.ribbing_interval)
        head_column = self.data_well.head_column.get_value

        template_ek = json.dumps(
            [
                self.data_well.template_depth,
                self.data_well.template_length,
                self.data_well.template_depth_addition,
                self.data_well.template_length_addition,
            ]
        )

        # Подготовленные данные для вставки (пример)
        data_values = [
            row_max,
            self.data_well.current_bottom,
            dict_perforation_json,
            plast_all_json,
            plast_work_json,
            leakage_json,
            self.data_well.column_additional,
            self.data_well.fluid_work,
            self.data_well.category_pressure,
            self.data_well.category_h2s,
            self.data_well.category_gas_factor,
            template_ek,
            skm_list_json,
            self.data_well.problem_with_ek_depth,
            self.data_well.problem_with_ek_diameter,
            raiding_list_json,
            head_column,
        ]

        if len(self.data_well.data_list) == 0:
            self.data_well.data_list.append(data_values)
        else:
            row_number = row_number - self.data_well.count_row_well
            self.data_well.data_list.insert(row_number, data_values)

    def check_depth_in_skm_interval(self, depth):

        check_true = False
        check_ribbing = False

        for interval in self.data_well.skm_interval:
            if float(interval[0]) <= float(depth) <= float(interval[1]):
                check_true = True
                return int(depth)

        for interval in self.data_well.ribbing_interval:
            if float(interval[0]) <= float(depth) <= float(interval[1]):
                check_ribbing = True

        if check_true is False or check_ribbing is True:
            false_question = QMessageBox.question(
                None,
                "Проверка посадки пакера в интервал скреперования",
                f"Проверка посадки показала, что пакер сажается не "
                f"в интервал скреперования {self.data_well.skm_interval}, "
                f"но сажается в интервал райбирования "
                f"райбирования {self.data_well.ribbing_interval} \n"
                f"Продолжить?",
            )
            if false_question == QMessageBox.StandardButton.No:
                return False

        if check_true is False and check_ribbing is False:
            QMessageBox.warning(
                None,
                "Проверка посадки пакера в интервал скреперования",
                f"Проверка посадки показала, что пакер сажается не "
                f"в интервал скреперования {self.data_well.skm_interval}, и "
                f"райбирования {self.data_well.ribbing_interval} \n"
                f"Нужно скорректировать интервалы скреперования или глубину посадки пакера",
            )
            return False

    def check_depth_paker_in_perforation(self, depth):

        check_true = False

        for plast in self.data_well.plast_all:
            if len(self.data_well.dict_perforation[plast]["интервал"]) >= 1:
                for interval in self.data_well.dict_perforation[plast]["интервал"]:
                    if float(interval[0]) < depth < float(interval[1]):
                        check_true = False
                        break
                    else:
                        check_true = True
            elif len(self.data_well.dict_perforation[plast]["интервал"]) == 0:
                check_true = True

        if check_true is False:
            QMessageBox.warning(
                None,
                f"Проверка посадки пакера в интервал перфорации {interval}",
                f"Проверка посадки показала пакер сажается в интервал перфорации, "
                f"необходимо изменить глубину посадки!!!",
            )

        return check_true

    def populate_row(self, insert_index, work_list, table_widget, work_plan="krs"):
        text_width_dict = {
            20: (0, 100),
            40: (101, 200),
            60: (201, 300),
            80: (301, 400),
            100: (401, 500),
            120: (501, 600),
            140: (601, 700),
            160: (701, 800),
            180: (801, 900),
            200: (901, 1200),
            250: (1201, 1350),
            300: (1351, 1500),
            350: (1501, 2000)
        }
        index_setSpan = 0
        if work_plan == "gnkt_frez":
            index_setSpan = 1
        row_max = table_widget.rowCount()
        # print(f'ДОП {work_plan}')

        for row in range(self.table_widget.rowCount()):
            for column in range(self.table_widget.columnCount()):
                value = self.table_widget.item(row, column)
                if value is not None:
                    value = value.text()
                    if "Категория скважины" in str(value):
                        self.data_well.cat_well_min = data_list.ProtectedIsDigit(row)


                    elif "Наименование работ" in str(value):
                        self.data_well.insert_index2 = row - 1

        if self.__class__.__name__ in ["GppWindow", "GrpWindow"]:
            if self.__class__.__name__ in ["GppWindow"]:
                podp_grp = [
                    [
                        None,
                        "Представитель подрядчика по ГРП",
                        None,
                        None,
                        None,
                        None,
                        None,
                        "Представитель OOO НТЦ ЗЭРС",
                        None,
                        None,
                        None,
                        None,
                    ],
                    [
                        None,
                        "_____________",
                        None,
                        None,
                        None,
                        None,
                        None,
                        "_____________",
                        None,
                        None,
                        None,
                        None,
                    ],
                    [
                        None,
                        f'"____"_____________________г.',
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        f'"____"_____________________г.',
                        None,
                        None,
                        None,
                    ],
                ]
            else:
                podp_grp = [
                    [
                        None,
                        "Представитель подрядчика по ГРП",
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    [
                        None,
                        "_____________",
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    [
                        None,
                        f'"____"_____________________г.',
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                ]

            for index_row, row in enumerate(podp_grp):
                for column, data in enumerate(row):
                    if not data is None:
                        item = QtWidgets.QTableWidgetItem(str(data))
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                        table_widget.setItem(
                            self.data_well.cat_well_min.get_value
                            - len(podp_grp)
                            + index_row,
                            column,
                            item,
                        )
                    else:
                        table_widget.setItem(
                            self.data_well.cat_well_min.get_value
                            - len(podp_grp)
                            + index_row,
                            column,
                            QtWidgets.QTableWidgetItem(str("")),
                        )

        for i, row_data in enumerate(work_list):
            if "prs" in self.data_well.work_plan:
                row_data.insert(-4, None)
                row_data.insert(-4, None)
                row_data.insert(-4, None)

            row = insert_index + i
            if work_plan not in [
                "application_pvr",
                "gnkt_frez",
                "gnkt_opz",
                "gnkt_bopz",
                "gnkt_after_grp",
                "application_gis",
            ]:
                self.insert_data_in_database(row, row_max + i)

            table_widget.insertRow(row)

            if len(str(row_data)[1]) > 3 and work_plan in "gnkt_frez":
                table_widget.setSpan(i + insert_index, 1, 1, 12)
            elif "prs" in self.data_well.work_plan:
                table_widget.setSpan(i + insert_index, 2, 1, 11)
            elif "Порядок работы" in row_data or "Ранее проведенные работы" in str(
                    row_data[1]
            ):
                table_widget.setSpan(i + insert_index, 1, 1, 12)
            elif (
                    len([row_str for row_str in row_data if row_str not in ["", None]]) > 6
            ):
                pass

            else:
                table_widget.setSpan(i + insert_index, 2, 1, 8 + index_setSpan)

            for column, data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(str(data))
                item.setFlags(item.flags() | Qt.ItemIsEditable)

                if not data is None:
                    table_widget.setItem(row, column, item)
                else:
                    table_widget.setItem(
                        row, column, QtWidgets.QTableWidgetItem(str(""))
                    )

                if column == 2:
                    if not data is None:
                        text = data
                        for key, value in text_width_dict.items():
                            if value[0] <= len(text) <= value[1]:
                                text_width = key + text.count("\n") * 4
                                table_widget.setRowHeight(row, int(text_width))
        if "gnkt" not in work_plan and self.data_well.insert_index2:
            for row in range(table_widget.rowCount()):
                if row >= self.data_well.insert_index2 + 2:
                    # Добавляем нумерацию в первую колонку
                    item_number = QtWidgets.QTableWidgetItem(
                        str(row - self.data_well.insert_index2 - 1)
                    )  # Номер строки + 1
                    table_widget.setItem(row, 1, item_number)

    def check_true_depth_template(self, depth):
        check = True
        if self.data_well.column_additional:

            if (
                    self.data_well.template_depth_addition < depth
                    and depth > self.data_well.head_column_additional.get_value
            ):
                check = False
                check_question = QMessageBox.question(
                    self,
                    "Проверка глубины пакера",
                    f"Проверка показало что пакер с глубиной {depth}м спускается ниже"
                    f" глубины  шаблонирования {self.data_well.template_depth_addition}",
                )
                if check_question == QMessageBox.StandardButton.Yes:
                    check = True
            if (
                    self.data_well.template_depth < depth
                    and depth < self.data_well.head_column_additional.get_value
            ):
                check = False
                check_question = QMessageBox.question(
                    self,
                    "Проверка глубины пакера",
                    f"Проверка показало что пакер с глубиной {depth}м спускается ниже "
                    f"глубины шаблонирования {self.data_well.template_depth}",
                )
                if check_question == QMessageBox.StandardButton.Yes:
                    check = True
        else:
            # print(f'глубина {self.data_well.template_depth, depth}')
            if self.data_well.template_depth < depth:
                check = False
                check_question = QMessageBox.question(
                    self,
                    "Проверка глубины пакера",
                    f"Проверка показало что пакер с глубиной {depth}м спускается ниже "
                    f"глубины шаблонирования {self.data_well.template_depth}",
                )
                if check_question == QMessageBox.StandardButton.Yes:
                    check = True

        return check

    def copy_pz(self, sheet, table_widget, work_plan="krs", count_col=12, list_page=1):
        from krs import GnoWindow

        if sheet:
            rows = sheet.max_row
            if work_plan in ["krs", "prs"]:
                self.data_well.insert_index2 = rows + 1

            merged_cells = sheet.merged_cells
            table_widget.setRowCount(rows)
            self.data_well.count_row_well = table_widget.rowCount()

            if work_plan == "plan_change":
                self.data_well.count_row_well = self.data_well.data_x_max.get_value

            border_styles = {}
            for row in sheet.iter_rows():
                for cell in row:
                    border_styles[(cell.row, cell.column)] = cell.border

            table_widget.setColumnCount(count_col)
            row_heights_exit = [
                (
                    sheet.row_dimensions[i + 1].height
                    if sheet.row_dimensions[i + 1].height is not None
                    else 18
                )
                for i in range(sheet.max_row)
            ]

            for row in range(1, rows + 2):
                if row > 1 and row < rows - 1:
                    table_widget.setRowHeight(row, int(row_heights_exit[row]))
                for col in range(1, count_col + 1):
                    if not sheet.cell(row=row, column=col).value is None:
                        if (
                                isinstance(sheet.cell(row=row, column=col).value, float)
                                and row > 25
                        ):
                            cell_value = str(
                                round(sheet.cell(row=row, column=col).value, 2)
                            )
                        elif isinstance(
                                sheet.cell(row=row, column=col).value, datetime
                        ):
                            cell_value = sheet.cell(row=row, column=col).value.strftime(
                                "%d.%m.%Y"
                            )
                        else:
                            cell_value = str(sheet.cell(row=row, column=col).value)

                        item = QtWidgets.QTableWidgetItem(str(cell_value))
                        table_widget.setItem(row - 1, col - 1, item)

                        # Проверяем, является ли текущая ячейка объединенной
                        for merged_cell in merged_cells:
                            range_row = range(
                                merged_cell.min_row, merged_cell.max_row + 1
                            )
                            range_col = range(
                                merged_cell.min_col, merged_cell.max_col + 1
                            )
                            if row in range_row and col in range_col:
                                # Устанавливаем количество объединяемых строк и столбцов для текущей ячейки
                                table_widget.setSpan(
                                    row - 1,
                                    col - 1,
                                    merged_cell.max_row - merged_cell.min_row + 1,
                                    merged_cell.max_col - merged_cell.min_col + 1,
                                )

                    else:
                        item = QTableWidgetItem("")

            if data_list.dop_work_list:
                self.populate_row(
                    table_widget.rowCount(),
                    data_list.dop_work_list,
                    self.table_widget,
                    self.work_plan,
                )
            if "gnkt" not in work_plan and self.data_well.insert_index2:
                for row in range(table_widget.rowCount()):
                    if row >= self.data_well.insert_index2 + 2:
                        # Добавляем нумерацию в первую колонку
                        item_number = QtWidgets.QTableWidgetItem(
                            str(row - self.data_well.insert_index2 - 1)
                        )  # Номер строки + 1
                        table_widget.setItem(row, 1, item_number)

                row_value_empty = (
                    True  # Флаг, указывающий, что все ячейки в строке пустые
                )
                # Проход по всем колонкам в текущей строке
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    # Проверка, является ли содержимое ячейки пустым
                    if item is not None and item.text() != "":
                        row_value_empty = (
                            False  # Если хотя бы одна ячейка не пустая, снимаем флаг
                        )
                        break

                # Если все ячейки в строке пустые, скрываем строку
                if row_value_empty:
                    table_widget.setRowHidden(row, True)
                else:
                    table_widget.setRowHidden(row, False)

            if work_plan in ["krs", "prs"]:
                self.work_window = GnoWindow(
                    table_widget.rowCount(), self.table_widget, self.data_well
                )
                self.set_modal_window(self.work_window)
                self.ws3 = self.wb.create_sheet("Расчет поглотителя сероводорода", 1)
                data_list.pause = True

                data_list.pause = True
                self.work_window = None
            check_str = ""
            if len(
                    self.data_well.check_data_in_pz
            ) != 0 and (self.data_well.work_plan in ["krs", "prs"] or
                        ("gnkt" in self.data_well.work_plan and list_page == 2)):

                for ind, check_data in enumerate(self.data_well.check_data_in_pz):
                    if check_data not in check_str:
                        check_str += f"{ind + 1}. {check_data} \n"
            if check_str != "":
                self.show_info_message(self.data_well, check_str)

        if work_plan in ["gnkt_frez"] and list_page == 2:
            col_width = [
                2.28515625,
                13.0,
                4.5703125,
                13.0,
                13.0,
                13.0,
                5.7109375,
                13.0,
                13.0,
                13.0,
                4.7109375,
                13.0,
                5.140625,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                4.7109375,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                13.0,
                5.42578125,
                13.0,
                4.5703125,
                2.28515625,
                10.28515625,
            ]
            for column in range(table_widget.columnCount()):
                table_widget.setColumnWidth(
                    column, int(col_width[column])
                )  # Здесь задайте требуемую ширину столбца

        elif (
                work_plan in ["gnkt_after_grp", "gnkt_opz", "gnkt_after_grp", "gnkt_bopz"]
                and list_page == 2
        ):
            col_width = property_excel.property_excel_pvr.col_width_gnkt_osv
            for column in range(table_widget.columnCount()):
                table_widget.setColumnWidth(
                    column, int(col_width[column] * 9)
                )  # Здесь задайте требуемую ширину столбца

        elif work_plan == "application_pvr":
            from property_excel import property_excel_pvr

            for column in range(table_widget.columnCount()):
                table_widget.setColumnWidth(
                    column, int(property_excel_pvr.col_width[column])
                )  # Здесь задайте требуемую ширину столбца
            data_list.pause = True


class MyWindow(MyMainWindow):

    def __init__(self):

        super().__init__()

        self.initUI()
        self.login_window = None
        self.copied_rows = None
        self.operation_window = None
        self.skm_depth = 0
        self.new_window = None
        self.raid_window = None
        self.leakage_window = None
        self.correct_window = None
        self.acid_windowPaker = None
        self.work_window = None
        self.signatures_window = None
        self.acid_windowPaker2 = None
        self.rir_window = None
        self.data_window = None
        self.filter_widgets = []
        self.table_class = None
        self.table_juming = None
        self.resize(1400, 800)
        self.showMaximized()

        # self.check_for_updates()

        self.perforation_correct_window2 = None
        self.ws = None
        self.insert_index = None
        self.perforation_list = []
        self.dict_perforation_project = {}

        self.insert_index_border = None
        self.work_plan = 0
        self.table_widget = None
        self.table_pvr = None
        self.data_well = None

        threading.Timer(2.0, self.close_splash).start()

        # self.log_widget = QPlainTextEditLogger(self)
        # logger.addHandler(self.log_widget)
        # self.setCentralWidget(self.log_widget.widget)

        # Обработка критических ошибок
        self.excepthook = UncaughtExceptions(self.data_well)
        self.excepthook._exception_caught.connect(self.excepthook.handle_uncaught_exception)

        try:

            print(f"Путь к изображению: {data_list.path_image}")

            # Загружаем изображение для заставки
            splash_pix = QPixmap(
                f"{data_list.path_image}imageFiles/icon/zima.png"
            )  # путь к  изображению

            splash = QSplashScreen(splash_pix, Qt.WindowStaysOnTopHint)
            splash.setMask(splash_pix.mask())
            splash.show()

            # Задержка на )5 секунд
            QTimer.singleShot(1000, splash.close)

            data_list.connect_in_base, self.db = connect_to_database(
                decrypt("DB_NAME_USER")
            )

            self.login_window = LoginWindow()
            self.login_window.setWindowModality(Qt.ApplicationModal)

            self.login_window.show()

            data_list.pause = False
        except Exception as e:
            logger.critical(e)
            QMessageBox.warning(
                None,
                "КРИТИЧЕСКАЯ ОШИБКА",
                f"Критическая ошибка, смотри в лог {type(e).__name__}\n\n{str(e)}",
            )

    def global_excepthook(self, exc_type, exc_value, exc_traceback):
        self.exception_handler.handle_exception(exc_type, exc_value, exc_traceback)

    def insert_data_in_chemistry(self):

        if self.data_well.work_plan in ["dop_plan", "dop_plan_in_base"]:
            string_work = f" ДП№ {self.data_well.number_dp}"
        elif self.data_well.work_plan == "krs":
            string_work = "ПР"
        elif self.data_well.work_plan == "plan_change":
            if self.data_well.work_plan_change == "krs":
                string_work = "ПР изм"
            else:
                string_work = f"ДП№{self.data_well.number_dp} изм "

        elif self.data_well.work_plan == "gnkt_bopz":
            string_work = "ГНКТ БОПЗ ВНС"
        elif self.data_well.work_plan == "gnkt_opz":
            string_work = "ГНКТ ОПЗ"
        elif self.data_well.work_plan == "gnkt_after_grp":
            string_work = "ГНКТ ОСВ ГРП"
        else:
            string_work = "ГНКТ"

        date_today = datetime.now()
        data_work = (
            self.data_well.well_number.get_value,
            self.data_well.well_area.get_value,
            self.data_well.region,
            data_list.costumer,
            data_list.contractor,
            string_work,
            self.data_well.type_kr.split(" ")[0],
            date_today,
            data_list.DICT_VOLUME_CHEMISTRY["цемент"],
            data_list.DICT_VOLUME_CHEMISTRY["HCl"],
            data_list.DICT_VOLUME_CHEMISTRY["HF"],
            data_list.DICT_VOLUME_CHEMISTRY["NaOH"],
            data_list.DICT_VOLUME_CHEMISTRY["ВТ СКО"],
            data_list.DICT_VOLUME_CHEMISTRY["Глина"],
            data_list.DICT_VOLUME_CHEMISTRY["песок"],
            data_list.DICT_VOLUME_CHEMISTRY["РПК"],
            data_list.DICT_VOLUME_CHEMISTRY["РПП"],
            data_list.DICT_VOLUME_CHEMISTRY["извлекаемый пакер"],
            data_list.DICT_VOLUME_CHEMISTRY["ЕЛАН"],
            data_list.DICT_VOLUME_CHEMISTRY["растворитель"],
            data_list.DICT_VOLUME_CHEMISTRY["РИР 2С"],
            data_list.DICT_VOLUME_CHEMISTRY["РИР ОВП"],
            data_list.DICT_VOLUME_CHEMISTRY["гидрофабизатор"],
            round(self.data_well.norm_of_time, 1),
            self.data_well.fluid,
        )

        query = (
            f"INSERT INTO chemistry "
            f"VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, "
            f"%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        )

        from data_base.work_with_base import ClassifierWell

        ClassifierWell.insert_database("well_data", data_work, query)

    @staticmethod
    def check_process():
        count_zima = 0
        for proc in psutil.process_iter():
            if proc.name() == "ZIMA.exe":
                count_zima += count_zima
        if count_zima > 1:
            return True  # Процесс найден

        return False  # Процесс не найден

    @staticmethod
    def show_confirmation():
        reply = QMessageBox.question(
            None,
            "Закрыть Zima?",
            "Приложение Zima.exe работает. Вы хотите Перезапустить его?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            MyWindow.close_process()

    @staticmethod
    def check_connection(host, port=5432):
        """Проверяет соединение с удаленным хостом по указанному порту."""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(2)  # Устанавливаем таймаут 2 секунды
            sock.connect((host, port))
            sock.close()
            return True
        except socket.error:
            return False

    # Остальная часть кода...

    def initUI(self):

        self.setWindowTitle("ZIMA")
        self.setGeometry(200, 100, 800, 800)

        self.create_menu_bar()
        self.le = QLineEdit()

        self.toolbar = QToolBar()
        self.addToolBar(self.toolbar)

        self.saveFileButton = QPushButton("Сохранить проект")
        self.saveFileButton.clicked.connect(self.save_to_excel)
        self.toolbar.addWidget(self.saveFileButton)

        self.correctDataButton = QPushButton("Скорректировать данные")
        self.correctDataButton.clicked.connect(self.correctData)
        self.toolbar.addWidget(self.correctDataButton)

        self.correctPVRButton = QPushButton("Скорректировать работающие ПВР")
        self.correctPVRButton.clicked.connect(self.correct_perforation)
        self.toolbar.addWidget(self.correctPVRButton)

        self.correctNEKButton = QPushButton("Скорректировать НЭК")
        self.correctNEKButton.clicked.connect(self.correct_nek)
        self.toolbar.addWidget(self.correctNEKButton)

        self.correct_curator_Button = QPushButton("Скорректировать куратора")
        self.correct_curator_Button.clicked.connect(self.correct_curator)
        self.toolbar.addWidget(self.correct_curator_Button)

        self.signPlanButton = QPushButton("План работ подписан")
        self.signPlanButton.clicked.connect(self.sign_work_plan)
        # Кнопка будет активна только для корректировки плана
        self.signPlanButton.setEnabled(False)
        self.toolbar.addWidget(self.signPlanButton)

        self.closeFileButton = QPushButton("Закрыть проект")
        self.closeFileButton.clicked.connect(self.close_file)
        self.toolbar.addWidget(self.closeFileButton)

    def create_menu_bar(self):
        self.menuBar = QMenuBar(self)
        self.setMenuBar(self.menuBar)
        self.fileMenu = QMenu("&Файл", self)
        self.application_geophysical = QMenu("&Заявка на ГИС", self)
        self.classifierMenu = QMenu("&Классификатор", self)
        self.signatories = QMenu("&Подписанты ", self)
        self.menuBar.addMenu(self.fileMenu)
        self.menuBar.addMenu(self.application_geophysical)
        self.menuBar.addMenu(self.classifierMenu)
        self.menuBar.addMenu(self.signatories)

        self.application_pvr = self.application_geophysical.addAction(
            "Заявка на ПВР", self.action_clicked
        )
        self.application_gis = self.application_geophysical.addAction(
            "Заявка на ГИС", self.action_clicked
        )

        self.create_file = self.fileMenu.addMenu("&Создать")

        self.create_KRS = self.create_file.addAction("План КРС", self.action_clicked)
        self.create_KRS_change = self.create_file.addAction(
            "Корректировка плана КРС", self.action_clicked
        )
        self.create_KRS_DP = self.create_file.addAction(
            "Дополнительный план КРС", self.action_clicked
        )
        self.create_KRS_DP_in_base = self.create_file.addAction(
            "Дополнительный план КРС из базы", self.action_clicked
        )
        self.create_GNKT = self.create_file.addMenu("&План ГНКТ")
        self.create_gnkt_opz = self.create_GNKT.addAction(
            " ГНКТ ОПЗ", self.action_clicked
        )
        self.create_GNKT_frez = self.create_GNKT.addAction(
            "ГНКТ Фрезерование", self.action_clicked
        )
        self.create_GNKT_GRP = self.create_GNKT.addAction(
            "Освоение после ГРП", self.action_clicked
        )
        self.create_GNKT_BOPZ = self.create_GNKT.addAction(
            "БОПЗ ГНКТ", self.action_clicked
        )
        self.create_PRS = self.create_file.addAction("План ПРС", self.action_clicked)
        # self.open_file = self.fileMenu.addAction('Открыть', self.action_clicked)
        # self.save_file = self.fileMenu.addAction('Сохранить', self.action_clicked)
        # self.save_file_as = self.fileMenu.addAction('Сохранить как', self.action_clicked)

        self.class_well = self.classifierMenu.addMenu("&ООО Башнефть-Добыча")
        self.costumer_class_well = self.class_well.addMenu("Классификатор")
        self.costumer_select = self.class_well.addMenu("Перечень скважин без глушения")

        self.class_well_TGM = self.costumer_class_well.addMenu("&Туймазинский регион")
        self.class_well_TGM_open = self.class_well_TGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.class_well_IGM = self.costumer_class_well.addMenu("&Ишимбайский регион")

        self.class_well_IGM_open = self.class_well_IGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.class_well_CHGM = self.costumer_class_well.addMenu(
            "&Чекмагушевский регион"
        )
        self.class_well_CHGM_open = self.class_well_CHGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.class_well_KGM = self.costumer_class_well.addMenu("&Краснохолмский регион")
        self.class_well_KGM_open = self.class_well_KGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.class_well_AGM = self.costumer_class_well.addMenu("&Арланский регион")
        self.class_well_AGM_open = self.class_well_AGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.without_jamming_TGM = self.costumer_select.addMenu("&Туймазинский регион")
        self.without_jamming_TGM_open = self.without_jamming_TGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.without_jamming_IGM = self.costumer_select.addMenu("&Ишимбайский регион")
        self.without_jamming_IGM_open = self.without_jamming_IGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.without_jamming_CHGM = self.costumer_select.addMenu(
            "&Чекмагушевский регион"
        )
        self.without_jamming_CHGM_open = self.without_jamming_CHGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.without_jamming_KGM = self.costumer_select.addMenu(
            "&Краснохолмский регион"
        )
        self.without_jamming_KGM_open = self.without_jamming_KGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.without_jamming_AGM = self.costumer_select.addMenu("&Арланский регион")
        self.without_jamming_AGM_open = self.without_jamming_AGM.addAction(
            "&открыть перечень", self.action_clicked
        )

        self.class_well_TGM_reload = self.class_well_TGM.addAction(
            "&обновить", self.action_clicked
        )
        self.class_well_IGM_reload = self.class_well_IGM.addAction(
            "&обновить", self.action_clicked
        )
        self.class_well_CHGM_reload = self.class_well_CHGM.addAction(
            "&обновить", self.action_clicked
        )
        self.class_well_KGM_reload = self.class_well_KGM.addAction(
            "&обновить", self.action_clicked
        )
        self.class_well_AGM_reload = self.class_well_AGM.addAction(
            "&обновить", self.action_clicked
        )
        self.without_jamming_TGM_reload = self.without_jamming_TGM.addAction(
            "&обновить", self.action_clicked
        )
        self.without_jamming_IGM_reload = self.without_jamming_IGM.addAction(
            "&обновить", self.action_clicked
        )
        self.without_jamming_CHGM_reload = self.without_jamming_CHGM.addAction(
            "&обновить", self.action_clicked
        )
        self.without_jamming_KGM_reload = self.without_jamming_KGM.addAction(
            "&обновить", self.action_clicked
        )
        self.without_jamming_AGM_reload = self.without_jamming_AGM.addAction(
            "&обновить", self.action_clicked
        )

        self.signatories_Bnd = self.signatories.addAction(
            "&БашНефть-Добыча", self.action_clicked
        )
        self.signatories_cdng = self.signatories.addAction("&ЦДНГ", self.action_clicked)
        self.signatories_contractor = self.signatories.addAction(
            "&Подрядчик КРС", self.action_clicked
        )

    @QtCore.pyqtSlot()
    def action_clicked(self):
        from data_correct_position_people import CorrectSignaturesWindow

        action = self.sender()
        if action == self.signatories_Bnd:
            if self.signatures_window is None:
                self.signatures_window = CorrectSignaturesWindow()
                self.signatures_window.setWindowTitle("Подписанты")
                # self.signatures_window.setGeometry(200, 400, 300, 400)
                self.signatures_window.show()
            else:
                self.signatures_window.close()
                self.signatures_window = None
        elif action == self.signatories_cdng:
            if self.signatures_window is None:
                from work_py.data_correct_position_cdng import CorrectSignaturesCdng

                self.signatures_window = CorrectSignaturesCdng()
                self.signatures_window.setWindowTitle("Подписанты")
                # self.signatures_window.setGeometry(200, 400, 300, 400)
                self.signatures_window.show()
            else:
                self.signatures_window.close()
                self.signatures_window = None
        elif action == self.signatories_contractor:
            if self.signatures_window is None:
                from work_py.data_correct_position_contractor import (
                    CorrectSignaturesContractor,
                )

                self.signatures_window = CorrectSignaturesContractor()
                self.signatures_window.setWindowTitle("Подписанты КРС")
                # self.signatures_window.setGeometry(200, 400, 300, 400)
                self.signatures_window.show()
            else:
                self.signatures_window.close()
                self.signatures_window = None

        elif action == self.without_jamming_TGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_without_damping(costumer, "ТГМ")
        elif action == self.without_jamming_IGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_without_damping(costumer, "ИГМ")
        elif action == self.without_jamming_CHGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_without_damping(costumer, "ЧГМ")
        elif action == self.without_jamming_KGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_without_damping(costumer, "КГМ")
        elif action == self.without_jamming_AGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_without_damping(costumer, "АГМ")
        elif action == self.class_well_TGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_class_well(costumer, "ТГМ")
        elif action == self.class_well_IGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_class_well(costumer, "ИГМ")
        elif action == self.class_well_CHGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_class_well(costumer, "ЧГМ")
        elif action == self.class_well_KGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_class_well(costumer, "КГМ")
        elif action == self.class_well_AGM_open:
            costumer = "ООО Башнефть-добыча"
            self.open_class_well(costumer, "АГМ")

        elif action == self.application_pvr:
            self.work_plan = "application_pvr"
            # self.table_widget_open_pvr()
            self.fname, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "Выберите файл", ".", "Файлы Exсel (*.xlsx);;Файлы Exсel (*.xls)"
            )
            if self.fname:
                self.open_pvr_application(self.fname)
        elif action == self.application_gis:
            self.work_plan = "application_gis"
            # self.table_widget_open_pvr()
            self.fname, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "Выберите файл", ".", "Файлы Exсel (*.xlsx);;Файлы Exсel (*.xls)"
            )
            if self.fname:
                self.open_gis_application(self.fname)
        elif action == self.application_geophysical:
            pass

        elif getattr(sys, "frozen", True):
            if action == self.class_well_TGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_class_well(costumer, "ТГМ")
            elif action == self.class_well_CHGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_class_well(costumer, "ЧГМ")
            elif action == self.class_well_KGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_class_well(costumer, "КГМ")
            elif action == self.class_well_AGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_class_well(costumer, "АГМ")
            elif action == self.without_jamming_TGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_without_damping(costumer, "ТГМ")
            elif action == self.without_jamming_CHGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_without_damping(costumer, "ЧГМ")
            elif action == self.without_jamming_KGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_without_damping(costumer, "КГМ")
            elif action == self.without_jamming_AGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_without_damping(costumer, "АГМ")
            elif action == self.without_jamming_IGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_without_damping(costumer, "ИГМ")
            elif action == self.class_well_IGM_reload:
                costumer = "ООО Башнефть-добыча"
                self.reload_class_well(costumer, "ИГМ")

        if not self.table_widget is None:
            mes = QMessageBox.question(
                self, "Информация", "Необходимо закрыть текущий проект, закрыть?"
            )
            if mes == QMessageBox.StandardButton.Yes:
                self.close_file()
            else:
                return

        if action == self.create_KRS and self.table_widget is None:
            self.work_plan = "krs"
            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()

        elif action == self.create_KRS_DP and self.table_widget is None:
            self.work_plan = "dop_plan"

            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()
        elif action == self.create_KRS_DP_in_base and self.table_widget is None:
            self.work_plan = "dop_plan_in_base"

            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()

        elif action == self.create_KRS_change and self.table_widget is None:
            self.work_plan = "plan_change"
            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()

        elif action == self.create_gnkt_opz and self.table_widget is None:
            self.work_plan = "gnkt_opz"
            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()

        elif action == self.create_GNKT_GRP and self.table_widget is None:
            self.work_plan = "gnkt_after_grp"
            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()

        elif action == self.create_GNKT_BOPZ and self.table_widget is None:
            self.work_plan = "gnkt_bopz"
            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()

        elif action == self.create_GNKT_frez and self.table_widget is None:
            self.work_plan = "gnkt_frez"
            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()

        elif action == self.create_PRS and self.table_widget is None:
            self.work_plan = "prs"
            self.table_widget_open(self.work_plan)
            self.ws = self.open_read_excel_file_pz()

    def reload_class_well(self, costumer, region):
        self.open_parent_class(costumer, region)
        self.classifier_well.export_to_database_class_well(self.fname)

    def close_splash(self):
        splash_hwnd = win32gui.FindWindow(
            None, "Splash Screen"
        )  # При необходимости измените название окна
        if splash_hwnd:
            win32gui.PostMessage(splash_hwnd, win32con.WM_CLOSE, 0, 0)

    def open_without_damping(self, costumer, region):
        from data_base.work_with_base import ClassifierWell

        if self.new_window is None:

            self.new_window = ClassifierWell(costumer, region, "damping")
            self.new_window.open_to_sqlite_without_juming()
            self.new_window.setWindowTitle("Перечень скважин без глушения")
            # self.new_window.setGeometry(200, 400, 300, 400)
            self.new_window.show()

        else:
            self.new_window.close()  # Close window.
            self.new_window = None  # Discard reference.

    def open_pvr_application(self, fname):
        from open_pz import CreatePZ
        from application_pvr import PvrApplication

        if fname:
            # try:
            self.read_pz(fname)
            data_list.pause = True
            self.data_well = self.read_excel_file()
            read_pz = CreatePZ(self.data_well, self.ws, self)
            self.ws = read_pz.open_excel_file(self.ws, self.work_plan)
            self.rir_window = PvrApplication(self.table_pvr, self.data_well)
            self.set_modal_window(self.rir_window)

            data_list.pause = False

    def open_gis_application(self, fname):
        from open_pz import CreatePZ
        from application_gis import GisApplication

        if fname:
            # try:
            self.read_pz(fname)
            data_list.pause = True
            self.data_well = self.read_excel_file()
            read_pz = CreatePZ(self.data_well, self.ws, self)
            self.ws = read_pz.open_excel_file(self.ws, self.work_plan)
            self.rir_window = GisApplication(self.table_pvr, self.data_well)
            self.set_modal_window(self.rir_window.show())

            data_list.pause = False
            # except:
            #     pass

    def open_class_well(self, costumer, region):
        from data_base.work_with_base import ClassifierWell

        if self.new_window is None:
            self.new_window = ClassifierWell(costumer, region, "ClassifierWell")
            self.new_window.open_to_sqlite_class_well()
            self.new_window.setWindowTitle("Классификатор")
            # self.new_window.setGeometry(200, 400, 300, 400)
            self.new_window.show()
        else:
            self.new_window.close()  # Close window.
            self.new_window = None  # Discard reference.

    def open_parent_class(self, costumer, region):

        from data_base.work_with_base import ClassifierWell

        self.fname, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Выберите файл", ".", "Файлы Exсel (*.xlsx);;Файлы Exсel (*.xls)"
        )
        if self.fname:
            self.classifier_well = ClassifierWell(costumer, region)

    def reload_without_damping(self, costumer, region):
        self.open_parent_class(costumer, region)
        self.classifier_well.export_to_sqlite_without_juming(self.fname)

    def table_widget_open(self, work_plan="krs"):

        if self.table_widget is None:

            # Создание объекта TabWidget
            self.tab_widget = QTabWidget()
            self.table_widget = QTableWidget()
            self.table_pvr = QTableWidget()

            self.table_widget.setContextMenuPolicy(
                QtCore.Qt.ContextMenuPolicy.CustomContextMenu
            )
            self.table_widget.customContextMenuRequested.connect(self.open_context_menu)
            self.setCentralWidget(self.tab_widget)
            self.model = self.table_widget.model()

            # Этот сигнал испускается всякий раз, когда ячейка в таблице нажата.
            # Указанная строка и столбец - это ячейка, которая была нажата.
            self.table_widget.cellPressed[int, int].connect(self.clickedRowColumn)
            if work_plan in ["gnkt_frez", "gnkt_after_grp", "gnkt_opz", "gnkt_bopz"]:
                self.table_title = QTableWidget()
                self.tab_widget.addTab(self.table_title, "Титульник")
                self.table_schema = QTableWidget()
                self.tab_widget.addTab(self.table_schema, "Схема скважины")
                self.tab_widget.addTab(self.table_widget, "Ход работ")

            else:
                self.tab_widget.addTab(self.table_widget, "Ход работ")

    def save_to_excel(self):
        from excel_saver import SaveInExcel
        excel_save_work = SaveInExcel(
            self.data_well,
            self.ws,
            self.table_widget,
            self.table_title,
            self.table_schema,
            self.gnkt_data,
        )

        if self.work_plan in ["gnkt_frez", "gnkt_after_grp", "gnkt_opz", "gnkt_bopz"]:
            excel_save_work.save_to_gnkt()
        else:
            excel_save_work.save_to_krs()

    def open_context_menu(self, position):
        context_menu = QMenu(self)

        action_menu = context_menu.addMenu("вид работ")
        geophysical = action_menu.addMenu("Геофизические работы")

        perforation_action = QAction("Перфорация", self)
        geophysical.addAction(perforation_action)
        perforation_action.triggered.connect(self.perforation_new_window)

        geophysical_action = QAction("Геофизические исследования", self)
        geophysical.addAction(geophysical_action)
        geophysical_action.triggered.connect(self.geophysical_new_window)

        rgd_menu = geophysical.addMenu("РГД")
        rgd_without_paker_action = QAction("РГД по колонне", self)
        rgd_menu.addAction(rgd_without_paker_action)
        rgd_without_paker_action.triggered.connect(self.rgd_without_paker_action)

        po_action = QAction("прихватоопределить", self)
        geophysical.addAction(po_action)
        po_action.triggered.connect(self.po_new_window)

        torpedo_action = QAction("Торпедирование и извлечение ЭК", self)
        geophysical.addAction(torpedo_action)
        torpedo_action.triggered.connect(self.torpedo_action_window)

        rgd_with_paker_action = QAction("РГД с пакером", self)
        rgd_menu.addAction(rgd_with_paker_action)
        rgd_with_paker_action.triggered.connect(self.rgd_with_paker_action)

        privyazka_action = QAction("Привязка НКТ", self)
        geophysical.addAction(privyazka_action)
        privyazka_action.triggered.connect(self.privyazka_nkt)

        definition_bottom_gklm_action = QAction("Отбивка забоя по ЭК", self)
        geophysical.addAction(definition_bottom_gklm_action)
        definition_bottom_gklm_action.triggered.connect(self.definition_bottom_gklm)

        vp_action = QAction("Установка ВП", self)
        geophysical.addAction(vp_action)
        vp_action.triggered.connect(self.vp_action)

        swibbing_action = QAction("Свабирование", self)
        geophysical.addAction(swibbing_action)
        swibbing_action.triggered.connect(self.swibbing_with_paker)

        kompressVoronka_action = QAction("Освоение компрессором с воронкой", self)
        geophysical.addAction(kompressVoronka_action)
        kompressVoronka_action.triggered.connect(self.kompress_with_voronka)

        kompressVoronka_action = QAction("Замер Рпл", self)
        geophysical.addAction(kompressVoronka_action)
        kompressVoronka_action.triggered.connect(self.pressure_gis)

        del_menu = context_menu.addMenu("удаление строки")
        empty_string_action = QAction("добавить пустую строку", self)
        del_menu.addAction(empty_string_action)
        empty_string_action.triggered.connect(self.empty_string)

        deleteString_action = QAction("Удалить строку", self)
        del_menu.addAction(deleteString_action)
        deleteString_action.triggered.connect(self.deleteString)

        copy_string_action = QAction("копировать строку", self)
        del_menu.addAction(copy_string_action)
        copy_string_action.triggered.connect(self.copy_string)

        cut_string_action = QAction("вырезать строку", self)
        del_menu.addAction(cut_string_action)
        cut_string_action.triggered.connect(self.cut_string)

        paste_string_action = QAction("Вставить строку", self)
        del_menu.addAction(paste_string_action)
        paste_string_action.triggered.connect(self.paste_string)

        descent_gno_action = QAction("Подьем ГНО", self)
        action_menu.addAction(descent_gno_action)
        descent_gno_action.triggered.connect(self.descent_gno_action)

        opressovka_action = QAction("Опрессовка колонны", self)
        action_menu.addAction(opressovka_action)
        opressovka_action.triggered.connect(self.pressure_test)

        template_with_skm = QAction("шаблон c СКМ", self)
        template_menu = action_menu.addMenu("Шаблоны")
        template_menu.addAction(template_with_skm)
        template_with_skm.triggered.connect(self.template_with_skm)

        sgm_work = QAction("СГМ", self)
        template_menu.addAction(sgm_work)
        sgm_work.triggered.connect(self.sgm_work)

        template_pero = QAction("проходимость по перу", self)
        template_menu.addAction(template_pero)
        template_pero.triggered.connect(self.template_pero)

        paker_aspo = QAction("очистка колонны от АСПО с пакером и заглушкой", self)
        template_menu.addAction(paker_aspo)
        paker_aspo.triggered.connect(self.paker_clear_aspo)

        ryber_action = QAction("Райбирование", self)
        action_menu.addAction(ryber_action)
        ryber_action.triggered.connect(self.ryber_add)

        drilling_menu = action_menu.addMenu("Бурение")

        drilling_action_nkt = QAction("бурение на НКТ", self)
        drilling_menu.addAction(drilling_action_nkt)
        drilling_action_nkt.triggered.connect(self.drilling_action_nkt)

        frezering_port_action = QAction("Фрезерование портов", self)
        drilling_menu.addAction(frezering_port_action)
        frezering_port_action.triggered.connect(self.frezering_port_action)

        template_without_skm = QAction("шаблон без СКМ", self)
        template_menu.addAction(template_without_skm)
        template_without_skm.triggered.connect(self.template_without_skm)

        emergency_menu = action_menu.addMenu("Аварийные работы")

        magnet_action = QAction("магнит", self)
        emergency_menu.addAction(magnet_action)
        magnet_action.triggered.connect(self.magnet_action)

        hook_action = QAction("Удочка-крючок", self)
        emergency_menu.addAction(hook_action)
        hook_action.triggered.connect(self.hook_action)

        emergency_sticking_action = QAction("Прихваченное оборудование", self)
        emergency_menu.addAction(emergency_sticking_action)
        emergency_sticking_action.triggered.connect(self.lar_po_action)

        emergency_print_work_action = QAction("СПО печати", self)
        emergency_menu.addAction(emergency_print_work_action)
        emergency_print_work_action.triggered.connect(self.emergency_print_work_action)

        lar_sbt_action = QAction("ловильные работы", self)
        emergency_menu.addAction(lar_sbt_action)
        lar_sbt_action.triggered.connect(self.lar_sbt_action)

        lapel_tubing_action = QAction("Отворот на СБТ левое", self)
        emergency_menu.addAction(lapel_tubing_action)
        lapel_tubing_action.triggered.connect(self.lapel_tubing_func)

        acid_menu = action_menu.addMenu("Кислотная обработка")
        acid_action1paker = QAction("Кислотная обработка", self)
        acid_menu.addAction(acid_action1paker)
        acid_action1paker.triggered.connect(self.acid_paker_new_window)

        acid_action_gons = QAction("ГОНС", self)
        acid_menu.addAction(acid_action_gons)
        acid_action_gons.triggered.connect(self.acid_action_gons)

        sand_menu = action_menu.addMenu("песчанный мост")
        filling_action = QAction("Отсыпка песком")
        sand_menu.addAction(filling_action)
        filling_action.triggered.connect(self.filling_sand)

        washing_action = QAction("вымыв песка")
        sand_menu.addAction(washing_action)
        washing_action.triggered.connect(self.washing_sand)

        grp_menu = action_menu.addMenu("ГРП")
        grp_with_paker_action = QAction("ГРП с одним пакером")
        grp_menu.addAction(grp_with_paker_action)
        grp_with_paker_action.triggered.connect(self.grp_with_paker)

        grp_with_gpp_action = QAction("ГРП с ГПП")
        grp_menu.addAction(grp_with_gpp_action)
        grp_with_gpp_action.triggered.connect(self.grp_with_gpp)

        alone_menu = action_menu.addMenu("одиночные операции")

        mkp_action = QAction("Ревизия МКП")
        alone_menu.addAction(mkp_action)
        mkp_action.triggered.connect(self.mkp_revision)

        resuscitation_work = QAction("Реанимация ШГН")
        alone_menu.addAction(resuscitation_work)
        resuscitation_work.triggered.connect(self.resuscitation_work)

        block_pack_action = QAction("Блок пачка")
        alone_menu.addAction(block_pack_action)
        block_pack_action.triggered.connect(self.block_pack)

        tubing_pressure_testing_action = QAction("Опрессовка поинтервальная НКТ")
        alone_menu.addAction(tubing_pressure_testing_action)
        tubing_pressure_testing_action.triggered.connect(self.tubing_pressure_testing)

        konte_action = QAction("Канатные технологии")
        alone_menu.addAction(konte_action)
        konte_action.triggered.connect(self.konte_action)

        definition_q_action = QAction("Определение приемитости по НКТ", self)
        alone_menu.addAction(definition_q_action)
        definition_q_action.triggered.connect(self.definition_q)

        definition_q_NEK_action = QAction("Определение приемитости по затрубу", self)
        alone_menu.addAction(definition_q_NEK_action)
        definition_q_NEK_action.triggered.connect(self.definition_q_nek)

        kot_action = QAction("Система обратных клапанов")
        alone_menu.addAction(kot_action)
        kot_action.triggered.connect(self.kot_work)

        fluid_change_action = QAction("Изменение удельного веса")
        alone_menu.addAction(fluid_change_action)
        fluid_change_action.triggered.connect(self.fluid_change_action)

        pvo_cat1_action = QAction("Монтаж первой категории")
        alone_menu.addAction(pvo_cat1_action)
        pvo_cat1_action.triggered.connect(self.pvo_cat1)

        rir_menu = action_menu.addMenu("РИР")

        paker_izvlek_action = QAction("извлекаемый пакер")
        rir_menu.addAction(paker_izvlek_action)
        paker_izvlek_action.triggered.connect(self.paker_izvlek_action)

        rir_action = QAction("РИР")
        rir_menu.addAction(rir_action)
        rir_action.triggered.connect(self.rir_action)

        clay_solision_action = QAction("Глинистый раствор в ЭК")
        rir_menu.addAction(clay_solision_action)
        clay_solision_action.triggered.connect(self.clay_solision)

        gno_menu = action_menu.addAction("Спуск фондового оборудования")
        gno_menu.triggered.connect(self.gno_bottom)

        context_menu.exec_(self.mapToGlobal(position))

    def copy_string(self):
        if self.insert_index > self.data_well.count_row_well:
            selected_rows = sorted(
                set(item.row() for item in self.table_widget.selectedItems())
            )
            if selected_rows:
                self.copied_rows = []
                self.merge_rows = {}
                self.data_copy = []
                count_col = 0
                rows_to_span = []
                for row in selected_rows:
                    row_data = [
                        self.table_widget.item(row, col).text()
                        for col in range(self.table_widget.columnCount())
                        if self.table_widget.item(row, col)
                    ]
                    for column in range(self.table_widget.columnCount()):
                        if (
                                self.table_widget.rowSpan(row, column) > 1
                                or self.table_widget.columnSpan(row, column) > 1
                        ):
                            start_row = row + 1
                            start_column = column + 1
                            end_row = (
                                    start_row + self.table_widget.rowSpan(row, column) - 1
                            )
                            end_column = (
                                    start_column
                                    + self.table_widget.columnSpan(row, column)
                                    - 1
                            )
                            if rows_to_span:
                                if start_row != rows_to_span[-1][0]:
                                    rows_to_span.append(
                                        (start_row, start_column, end_row, end_column)
                                    )
                                else:
                                    if (
                                            start_column > rows_to_span[-1][-1]
                                            and start_row == rows_to_span[-1][0]
                                    ):
                                        rows_to_span.append(
                                            (
                                                start_row,
                                                start_column,
                                                end_row,
                                                end_column,
                                            )
                                        )
                            else:
                                rows_to_span.append(
                                    (start_row, start_column, end_row, end_column)
                                )
                    index_row = row - self.data_well.count_row_well
                    self.data_copy.append(
                        self.data_well.data_list[row - self.data_well.count_row_well]
                    )

                    if rows_to_span:
                        self.merge_rows[row] = rows_to_span

                    self.copied_rows.append((row, row_data))
                return selected_rows

    def cut_string(self):
        if self.insert_index > self.data_well.count_row_well:
            selected_rows = self.copy_string()
            if selected_rows:
                for row in reversed(selected_rows):
                    self.table_widget.removeRow(row)
                    self.data_well.data_list.pop(row - self.data_well.count_row_well)
            else:
                QMessageBox.warning(self, "Ошибка", "Выберите строки для вырезания.")

    def paste_string(self):
        text_width_dict = {
            20: (0, 100),
            40: (101, 200),
            60: (201, 300),
            80: (301, 400),
            100: (401, 500),
            120: (501, 600),
            140: (601, 700),
            160: (701, 800),
            180: (801, 900),
            200: (901, 1500),
        }
        if self.insert_index > self.data_well.count_row_well:
            if self.copied_rows:
                currentRow = (
                    self.table_widget.currentRow() + 1
                    if self.table_widget.currentRow() >= 0
                    else self.table_widget.rowCount()
                )
                for original_row, row_data in self.copied_rows[::-1]:
                    self.table_widget.insertRow(currentRow)
                    for column in range(len(row_data)):
                        item = QTableWidgetItem(row_data[column])
                        self.table_widget.setItem(currentRow, column, item)
                        for key, value in text_width_dict.items():
                            if value[0] <= len(row_data[2]) <= value[1]:
                                self.table_widget.setRowHeight(currentRow, int(key))

                    index_row = currentRow - self.data_well.count_row_well
                    self.data_well.data_list.insert(index_row, self.data_copy[-1])
                    self.data_copy.pop(-1)
                if self.merge_rows:
                    for key, value in self.merge_rows.items():
                        for _, col, _, merge in value:
                            self.table_widget.setSpan(
                                currentRow, col - 1, 1, abs(merge - col) + 1
                            )

                        currentRow += 1

            else:
                QMessageBox.warning(self, "Ошибка", "Сначала вырежьте строки.")

    def clickedRowColumn(self, r, c):

        self.insert_index = r + 1
        self.data_well.insert_index = r + 1

        # print(r, self.data_well.count_row_well)
        if r > self.data_well.count_row_well and "gnkt" not in self.work_plan:
            data = self.read_clicked_mouse_data(r)

    def closeEvent(self, event):
        for thread in self.threads:
            if thread.isRunning():
                thread.requestInterruption()
                thread.quit()
                thread.wait()
        event.accept()

    def list_threads(self):
        print("Активные потоки:")
        asdawada = threading.enumerate()
        for thread in threading.enumerate():
            print(f"Поток: {thread.name}, идентификатор: {thread.ident}")

    def sign_work_plan(self):
        """Изменяет статус плана работ на 'подписан' (только для корректировки плана)"""
        from server_response import ApiClient
        
        # Проверяем, что это корректировка плана
        if self.data_well is None or self.data_well.work_plan != "plan_change":
            QMessageBox.warning(
                self,
                "Ошибка",
                "Функция доступна только для корректировки плана работ"
            )
            return
        
        # Проверяем наличие ID ремонта
        repair_id = self.repair_id
        if repair_id is None:
            QMessageBox.warning(
                self,
                "Ошибка",
                "ID ремонта не найден. Загрузите план работ из базы данных."
            )
            return
        
        # Подготавливаем данные для обновления статуса
        # StatusWorkPlan.PLAN_IS_SIGNED = "подписан"
        data = {
            "id": repair_id,
            "status_work_plan": "подписан"
        }
        
        # Вызываем API для обновления статуса
        try:
            result = ApiClient.request_put_json(
                ApiClient.update_plan_status_path(),
                data,
                answer="param"
            )
            
            if result:
                QMessageBox.information(
                    self,
                    "Успех",
                    f"Статус плана работ успешно изменен на 'подписан' для ID {repair_id}"
                )
            else:
                QMessageBox.warning(
                    self,
                    "Ошибка",
                    "Не удалось обновить статус плана работ"
                )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Произошла ошибка при обновлении статуса: {str(e)}"
            )

    def close_file(self):
        from find import ProtectedIsNonNone, ProtectedIsDigit

        temp_folder = r"C:\Windows\Temp"

        self.wb2_prs = None
        self.ws2_prs = None

        self.rir_window = None
        self.data_well = None
        self.repair_id = None  # Сбрасываем ID ремонта
        
        # Отключаем кнопку "План работ подписан"
        if hasattr(self, 'signPlanButton'):
            self.signPlanButton.setEnabled(False)

        if not self.table_widget is None:
            self.cdng = ProtectedIsNonNone("")
            self.table_widget.clear()
            self.work_window = None
            self.table_widget.resizeColumnsToContents()
            self.table_widget = None
            self.tab_widget = None

            data_list.lift_ecn_can = False
            data_list.pause = False
            data_list.diameter_length = 0
            self.operation_window = None

            data_list.pipe_mileage = 0
            data_list.pipe_fatigue = 0
            data_list.pvo = 0
            data_list.previous_well = 0
            data_list.b_plan = 0
            data_list.pipes_ind = ProtectedIsDigit(0)
            data_list.sucker_rod_ind = ProtectedIsDigit(0)

            data_list.plast_select = ""
            data_list.paker_layout = 0
            (
                data_list.data,
                data_list.row_heights,
                data_list.col_width,
                data_list.boundaries_dict,
            ) = ("", "", "", "")
            data_list.data_in_base = False
            data_list.DICT_VOLUME_CHEMISTRY = {
                "пункт": [],
                "цемент": 0.0,
                "HCl": 0.0,
                "HF": 0.0,
                "NaOH": 0.0,
                "ВТ СКО": 0.0,
                "Глина": 0.0,
                "растворитель": 0.0,
                "уд.вес": 0.0,
                "песок": 0.0,
                "РПК": 0.0,
                "РПП": 0.0,
                "извлекаемый пакер": 0.0,
                "ЕЛАН": 0.0,
                "РИР 2С": 0.0,
                "РИР ОВП": 0.0,
                "гидрофабизатор": 0.0,
            }

            data_list.work_perforations = []
            data_list.work_perforations_dict = {}

            data_list.values = []
            data_list.dop_work_list = None

            data_list.dict_pump = {"before": "0", "after": "0"}

            data_list.dict_pump_h = {"before": 0, "after": 0}

            data_list.len_work_podpisant_list = 0
            data_list.data_well_is_True = False
            data_list.countAcid = 0
            data_list.swab_type_comboIndex = 1
            data_list.swab_true_edit_type = 1
            data_list.pakerTwoSKO = False
            data_list.privyazkaSKO = 0
            data_list.lift_key = 0
            data_list.data_pvr_max = ProtectedIsNonNone(0)
            data_list.row_heights = []
            data_list.condition_of_wells = ProtectedIsNonNone(0)

            QMessageBox.information(self, "Обновление", "Данные обнулены")

        self.list_threads()

    def on_finished(self):
        print("Работа с файлом Excel завершена.")

    def tubing_pressure_testing(self):
        from work_py.tubing_pressuar_testing import TubingPressureWindow

        self.add_window(TubingPressureWindow)

    def read_clicked_mouse_data(self, row):
        from work_py.advanted_file import definition_plast_work
        from data_correct import DataWindow

        aswa = self.data_well.count_row_well
        row = row - self.data_well.count_row_well + 1
        # print(self.data_well.column_diameter.get_value)
        data = self.data_well.data_list
        if row <= len(data):
            row = row - 1
        self.data_well.current_bottom = data[row][1]
        self.data_well.dict_perforation = json.loads(data[row][2])
        self.data_well.plast_all = json.loads(data[row][3])
        self.data_well.plast_work = json.loads(data[row][4])
        self.data_well.dict_leakiness = json.loads(data[row][5])

        self.data_well.column_additional = data[row][6]

        self.data_well.fluid_work = data[row][7]
        (
            self.data_well.template_depth,
            self.data_well.template_length,
            self.data_well.template_depth_addition,
            self.data_well.template_length_addition,
        ) = json.loads(data[row][11])
        self.data_well.skm_interval = json.loads(data[row][12])

        self.data_well.problem_with_ek_depth = data[row][13]
        self.data_well.problem_with_ek_diameter = data[row][14]

        self.definition_open_trunk_well()

        definition_plast_work(self)
        try:
            self.data_well.ribbing_interval = json.loads(data[row][15])
        except:
            pass

        try:
            self.data_well.head_column = data_list.ProtectedIsDigit(data[row][16])
        except:
            print("отсутствуют данные по голове хвостовика")

        # print(self.data_well.skm_interval)

    def frezering_port_action(self):
        from work_py.drilling import DrillWindow

        drilling_work_list = DrillWindow.frezer_ports(self)
        self.populate_row(self.insert_index, drilling_work_list, self.table_widget)

    def drilling_action_nkt(self):
        from work_py.drilling import DrillWindow

        self.add_window(DrillWindow)

    def magnet_action(self):
        from work_py.emergency_magnit import EmergencyMagnit

        self.add_window(EmergencyMagnit)

    def emergency_sticking_action(self):
        from work_py.emergencyWork import emergency_sticking

        emergency_sticking_list = emergency_sticking(self)
        self.populate_row(self.insert_index, emergency_sticking_list, self.table_widget)

    def hook_action(self):
        from work_py.emergencyWork import emergency_hook

        hook_work_list = emergency_hook(self)
        self.populate_row(self.insert_index, hook_work_list, self.table_widget)

    def lapel_tubing_func(self):
        from work_py.emergencyWork import lapel_tubing

        emergency_sbt_list = lapel_tubing(self)
        self.populate_row(self.insert_index, emergency_sbt_list, self.table_widget)

    def lar_sbt_action(self):
        from work_py.emergency_lar import EmergencyLarWork

        self.add_window(EmergencyLarWork)

    def lar_po_action(self):
        from work_py.emergency_po import EmergencyPo

        self.add_window(EmergencyPo)

    def emergency_print_work_action(self):
        from work_py.emergency_printing import EmergencyPrintWork

        self.add_window(EmergencyPrintWork)

    def rgd_without_paker_action(self):
        from work_py.rgdVcht import rgd_without_paker

        rgd_without_paker_list = rgd_without_paker(self)
        self.populate_row(self.insert_index, rgd_without_paker_list, self.table_widget)

    def rgd_with_paker_action(self):
        from work_py.rgdVcht import rgd_with_paker

        rgd_with_paker_list = rgd_with_paker(self)
        self.populate_row(self.insert_index, rgd_with_paker_list, self.table_widget)

    def pressure_gis(self):
        from work_py.alone_oreration import pressure_gis

        pressure_gis_list = pressure_gis(self)
        self.populate_row(self.insert_index, pressure_gis_list, self.table_widget)

    def definition_bottom_gklm(self):
        from work_py.alone_oreration import definition_bottom_gklm

        definition_bottom_gklm_list = definition_bottom_gklm(self)
        self.populate_row(
            self.insert_index, definition_bottom_gklm_list, self.table_widget
        )

    def privyazka_nkt_work(self):

        priv_list = [
            [
                f"ГИС Привязка по ГК и ЛМ",
                None,
                f'Вызвать геофизическую партию. Заявку оформить за 16 часов сутки через РИТС {data_list.contractor}". '
                f"Произвести  монтаж ПАРТИИ ГИС согласно схемы  №11 утвержденной главным инженером "
                f'{data_list.DICT_CONTRACTOR[data_list.contractor]["Дата ПВО"]}г. '
                f"ЗАДАЧА 2.8.1 Привязка технологического оборудования скважины",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Мастер КРС, подрядчик по ГИС",
                4,
            ]
        ]
        return priv_list

    def privyazka_nkt(self):
        privyazka_nkt_list = self.privyazka_nkt_work()
        self.populate_row(self.insert_index, privyazka_nkt_list, self.table_widget)

    def definition_q(self):
        from work_py.alone_oreration import definition_q

        definition_q_list = definition_q(self)
        self.populate_row(self.insert_index, definition_q_list, self.table_widget)

    def definition_q_nek(self):
        from work_py.alone_oreration import definition_q_nek

        definition_q_list = definition_q_nek(self)
        self.populate_row(self.insert_index, definition_q_list, self.table_widget)

    def kot_work(self):
        from work_py.alone_oreration import kot_work

        kot_work_list = kot_work(self)
        self.populate_row(self.insert_index, kot_work_list, self.table_widget)

    def konte_action(self):
        from work_py.alone_oreration import konte

        konte_work_list = konte(self)
        self.populate_row(self.insert_index, konte_work_list, self.table_widget)

    def mkp_revision(self):
        from work_py.mkp import mkp_revision

        mkp_work_list = mkp_revision(self)
        self.populate_row(self.insert_index, mkp_work_list, self.table_widget)

    def resuscitation_work(self):
        from work_py.descent_gno import DescentParent

        mkp_work_list = DescentParent.append_note_6(self)
        mkp_work_list[0][2] = mkp_work_list[0][2].replace("ПРИМЕЧАНИЕ №6: ", "")
        self.populate_row(self.insert_index, mkp_work_list, self.table_widget)

    def acid_action_gons(self):
        from work_py.acids import GonsWindow

        self.add_window(GonsWindow)

    def paker_izvlek_action(self):
        from work_py.izv_paker import PakerIzvlek

        self.add_window(PakerIzvlek)

    def pvo_cat1(self):
        from work_py.alone_oreration import pvo_cat1

        pvo_cat1_work_list = pvo_cat1(self)
        self.populate_row(self.insert_index, pvo_cat1_work_list, self.table_widget)

    def fluid_change_action(self):
        from work_py.change_fluid import ChangeFluidWindow

        self.add_window(ChangeFluidWindow)

    def clay_solision(self):
        from work_py.claySolution import ClayWindow

        self.add_window(ClayWindow)

    def rir_action(self):
        from work_py.rir import RirWindow

        self.add_window(RirWindow)

    def grp_with_paker(self):
        from work_py.grp import GrpWindow

        self.add_window(GrpWindow)

    def grp_with_gpp(self):
        from work_py.gpp import GppWindow

        self.add_window(GppWindow)

    def filling_sand(self):
        from work_py.sand_filling import SandWindow

        self.add_window(SandWindow)

    def washing_sand(self):
        from work_py.sand_filling import SandWindow

        washing_work_list = SandWindow.sandWashing(self)
        self.populate_row(self.insert_index, washing_work_list, self.table_widget)

    def deleteString(self):
        selected_ranges = self.table_widget.selectedRanges()
        selected_rows = []
        if self.insert_index > self.data_well.count_row_well:
            # Получение индексов выбранных строк
            for selected_range in selected_ranges:
                top_row = selected_range.topRow()
                bottom_row = selected_range.bottomRow()

                for row in range(top_row, bottom_row + 1):
                    if row not in selected_rows:
                        selected_rows.append(row)

            # Удаление выбранных строк в обратном порядке
            selected_rows.sort(reverse=True)
            # print(selected_rows)
            for row in selected_rows:
                self.table_widget.removeRow(row)
                self.data_well.data_list.pop(row - self.data_well.count_row_well)

    def empty_string(self):
        if self.insert_index > self.data_well.count_row_well:
            ryber_work_list = [
                [None, None, None, None, None, None, None, None, None, None, None, None]
            ]
            self.populate_row(self.insert_index, ryber_work_list, self.table_widget)

    def vp_action(self):
        from work_py.vp_cm import VpWindow

        self.add_window(VpWindow)

    def swibbing_with_paker(self):
        from work_py.swabbing import SwabWindow

        self.add_window(SwabWindow)

    def kompress_with_voronka(self):
        from work_py.kompress import KompressWindow

        self.add_window(KompressWindow)

    def ryber_add(self):
        from work_py.raiding import Raid

        self.add_window(Raid)

    def gno_bottom(self):
        from work_py.descent_gno import GnoDescentWindow

        self.add_window(GnoDescentWindow)

    def descent_gno_action(self):
        from krs import GnoWindow

        self.work_window = GnoWindow(
            self.data_well.insert_index, self.table_widget, self.data_well
        )
        self.set_modal_window(self.work_window)

    def pressure_test(self):
        from work_py.opressovka import OpressovkaEK

        self.add_window(OpressovkaEK)

    def block_pack(self):
        from work_py.block_pack_work import BlockPackWindow

        self.add_window(BlockPackWindow)

    def template_pero(self):
        from work_py.pero_work import PeroWindow

        self.add_window(PeroWindow)

    def template_with_skm(self):
        from work_py.template_work import TemplateKrs

        self.add_window(TemplateKrs)

    def sgm_work(self):
        from work_py.sgm_work import TemplateKrs

        self.add_window(TemplateKrs)

    def paker_clear_aspo(self):
        from work_py.opressovka_aspo import PakerAspo

        self.add_window(PakerAspo)

    def template_without_skm(self):
        from work_py.template_without_skm import TemplateWithoutSkm

        self.add_window(TemplateWithoutSkm)

    def acid_paker_new_window(self):
        from work_py.acid_paker import AcidPakerWindow

        self.add_window(AcidPakerWindow)

    def geophysical_new_window(self):
        from work_py.geophysic import GeophysicWindow

        self.add_window(GeophysicWindow)

    def correct_perforation(self):
        from perforation_correct import PerforationCorrect

        # self.data_well.current_bottom, ok = QInputDialog.getDouble(
        #     self, "Необходимый забой", "Введите забой до которого нужно нормализовать"
        # )

        self.perforation_correct_window2 = PerforationCorrect(self.data_well)
        self.perforation_correct_window2.setWindowTitle("Сверка данных перфорации")
        self.set_modal_window(self.perforation_correct_window2)
        self.pause_app()
        data_list.pause = True
        if self.insert_index:
            if self.insert_index > self.data_well.count_row_well:
                assde = self.insert_index - self.data_well.count_row_well - 1
                self.data_well.data_list[self.insert_index - self.data_well.count_row_well][
                    1] = self.data_well.current_bottom

                self.data_well.data_list[self.insert_index - self.data_well.count_row_well][2] = json.dumps(
                    self.data_well.dict_perforation, default=str, ensure_ascii=False, indent=4
                )

    def correct_curator(self):
        from work_py.curators import SelectCurator

        self.new_window = SelectCurator(self.data_well)
        self.set_modal_window(self.new_window)

    def correct_nek(self):
        from find import WellCondition
        from work_py.leakage_column import LeakageWindow

        if self.leakage_window is None:
            self.leakage_window = LeakageWindow(self.data_well)
            self.leakage_window.setWindowTitle("Корректировка негерметичности")
            # WellCondition.leakage_window.setGeometry(200, 400, 300, 400)
            data_list.pause = False
            self.set_modal_window(self.leakage_window)

            # print(f'словарь нарушений {self.data_well.dict_leakiness}')
            data_list.pause = True
            # self.data_well.dict_leakiness = self.leakage_window.add_work()
            self.data_well.data_list[-1][5] = json.dumps(
                self.data_well.dict_leakiness, default=str, ensure_ascii=False, indent=4
            )
            # print(f'словарь нарушений {self.data_well.dict_leakiness}')

        else:
            data_list.pause = True
            self.leakage_window = None  # Discard reference.

    def correctData(self):
        from data_correct import DataWindow

        if self.data_well is None:
            QMessageBox.warning(
                self,
                "Внимание",
                "Сначала откройте или загрузите скважину (план работ)."
            )
            return

        if self.work_window is None:
            self.work_window = DataWindow(self.data_well)
            self.set_modal_window(self.work_window)

            data_list.pause = True
            self.work_window = None

    def torpedo_action_window(self):
        from work_py.torpedo import TorpedoWindow

        self.add_window(TorpedoWindow)

    def po_new_window(self):
        from work_py.emergencyWork import emergencyECN

        template_pero_list = emergencyECN(self)
        self.populate_row(self.insert_index, template_pero_list, self.table_widget)

    def perforation_new_window(self):
        from work_py.perforation import PerforationWindow

        if len(self.data_well.category_pressure_list) > 1:
            if (
                    self.data_well.category_pressure_list[1] == 1
                    and self.data_well.category_pvo != 1
            ):
                QMessageBox.information(
                    self, "Внимание", "Не произведен монтаж первой категории"
                )
                return
        self.add_window(PerforationWindow)

    def create_short_plan(self, wb2, plan_short):
        from work_py.descent_gno import TabPageGno

        ws4 = wb2.create_sheet("Sheet1")
        ws4.title = "Краткое содержание плана работ"
        ws4 = wb2["Краткое содержание плана работ"]

        for row in range(15):
            ws4.insert_rows(ws4.max_row)
        ws4.cell(row=1, column=1).value = self.data_well.well_number.get_value
        ws4.cell(row=2, column=1).value = self.data_well.well_area.get_value

        if (
                self.data_well.dict_pump_shgn["before"] != 0
                and self.data_well.dict_pump_ecn["before"] == 0
                and self.data_well.paker_before["before"] == 0
        ):
            ws4.cell(row=3, column=1).value = (
                f'{self.data_well.dict_pump_shgn["before"]} -на гл. '
                f'{self.data_well.dict_pump_shgn_depth["before"]}м'
            )
        elif (
                self.data_well.dict_pump_shgn["before"] == 0
                and self.data_well.dict_pump_ecn["before"] != 0
                and self.data_well.paker_before["before"] == 0
        ):
            ws4.cell(row=3, column=1).value = (
                f'{self.data_well.dict_pump_ecn["before"]} -на гл. '
                f'{self.data_well.dict_pump_ecn_depth["before"]}м'
            )
        elif (
                self.data_well.dict_pump_shgn["before"] == 0
                and self.data_well.dict_pump_ecn["before"] != 0
                and self.data_well.paker_before["before"] != 0
        ):
            ws4.cell(row=3, column=1).value = (
                f'{self.data_well.dict_pump_ecn["before"]} -на гл.'
                f' {self.data_well.dict_pump_ecn_depth["before"]}м \n'
                f'{self.data_well.paker_before["before"]} на '
                f'{self.data_well.depth_fond_paker_before["before"]}м'
            )
        elif (
                self.data_well.dict_pump_shgn["before"] != 0
                and self.data_well.dict_pump_ecn["before"] == 0
                and self.data_well.paker_before["before"] != 0
        ):
            ws4.cell(row=3, column=1).value = (
                f'{self.data_well.dict_pump_shgn["before"]} -на гл. '
                f'{self.data_well.dict_pump_shgn_depth["before"]}м \n'
                f'{self.data_well.paker_before["before"]} на '
                f'{self.data_well.depth_fond_paker_before["before"]}м'
            )
        elif (
                self.data_well.dict_pump_shgn["before"] == 0
                and self.data_well.dict_pump_ecn["before"] == 0
                and self.data_well.paker_before["before"] != 0
        ):
            ws4.cell(row=3, column=1).value = (
                f'{self.data_well.paker_before["before"]} на '
                f'{self.data_well.depth_fond_paker_before["before"]}м'
            )
        elif (
                self.data_well.dict_pump_shgn["before"] == 0
                and self.data_well.dict_pump_ecn["before"] == 0
                and self.data_well.paker_before["before"] == 0
        ):
            ws4.cell(row=3, column=1).value = " "
        elif (
                self.data_well.dict_pump_shgn["before"] != 0
                and self.data_well.dict_pump_ecn["before"] != 0
                and self.data_well.paker_before["before"] != 0
        ):
            ws4.cell(row=3, column=1).value = (
                f'{self.data_well.dict_pump_shgn["before"]} -на гл.'
                f' {self.data_well.dict_pump_shgn_depth["before"]}м \n'
                f'{self.data_well.dict_pump_ecn["before"]} -на гл. '
                f'{self.data_well.dict_pump_ecn_depth["before"]}м \n'
                f'{self.data_well.paker_before["before"]} на '
                f'{self.data_well.depth_fond_paker_before["before"]}м '
            )
        plast_str = ""
        pressur_set = set()
        adawdawd = self.data_well.data_list
        dict_perforation = json.loads(self.data_well.data_list[0][2])
        # print(f'После {self.data_well.dict_perforation_short}')
        for plast in list(dict_perforation.keys()):
            if (
                    dict_perforation[plast]["отключение"] is False
                    and plast in dict_perforation
            ):
                for interval in dict_perforation[plast][
                    "интервал"
                ]:
                    plast_str += f"{plast[:4]}: {interval[0]}- {interval[1]} \n"
            elif (
                    dict_perforation[plast]["отключение"]
                    and plast in dict_perforation
            ):
                for interval in dict_perforation[plast][
                    "интервал"
                ]:
                    plast_str += f"{plast[:4]} :{interval[0]}- {interval[1]} (изол)\n"
            try:
                filter_list_pressure = None
                if "давление" in dict_perforation[plast].keys():
                    filter_list_pressure = list(
                        filter(
                            lambda x: type(x) in [int, float],
                            list(
                                dict_perforation[plast]["давление"]
                            ),
                        )
                    )
                # print(f'фильтр -{filter_list_pressure}')
                if filter_list_pressure:
                    pressur_set.add(f"{plast[:4]} - {filter_list_pressure[0]}")
            except Exception as e:
                QMessageBox.warning(
                    None, "Ошибка", f"Ошибка вставки давления в краткое описание {e}"
                )

        ws4.cell(row=6, column=1).value = (
            f"НКТ: \n {TabPageGno.gno_nkt_opening(self.data_well.dict_nkt_before)}"
        )
        ws4.cell(row=7, column=1).value = f'Рпл: \n {" ".join(list(pressur_set))}атм'
        # ws4.cell(row=8, column=1).value = f'ЖГС = {self.data_well.fluid_work_short}г/см3'
        ws4.cell(row=9, column=1).value = (
            f"Нст- {self.data_well.static_level.get_value}м / Ндин -"
            f" {self.data_well.dinamic_level.get_value}м"
        )
        if self.data_well.curator == "ОР":
            ws4.cell(row=10, column=1).value = (
                f"Ожид {self.data_well.expected_pickup}м3/сут при Р-{self.data_well.expected_pressure}м3/сут"
            )
        else:
            ws4.cell(row=10, column=1).value = (
                f"Qн {self.data_well.expected_oil}т Qж- {self.data_well.water_cut}м3/сут"
            )
        ws4.cell(row=11, column=1).value = (
            f"макс угол {self.data_well.max_angle.get_value} на "
            f"{self.data_well.max_angle_depth.get_value}"
        )
        ws4.cell(row=1, column=2).value = self.data_well.cdng.get_value
        try:
            try:
                category_pressure = self.data_well.dict_category[
                    self.data_well.plast_work_short[0]
                ]["по давлению"].category
            except:
                category_pressure = self.data_well.category_pressure_second
            try:
                category_h2s = self.data_well.dict_category[
                    self.data_well.plast_work_short[0]
                ]["по сероводороду"].category
            except:
                category_h2s = self.data_well.category_h2s_second
            try:
                gaz_f_pr = self.data_well.gaz_factor_percent[0]
            except:
                gaz_f_pr = self.data_well.gaz_factor_pr_second

            ws4.cell(row=2, column=3).value = (
                f"Рпл - {category_pressure},"
                f" H2S -{category_h2s},"
                f" газ факт -{gaz_f_pr}т/м3"
            )
        except Exception as e:
            QMessageBox.warning(
                self,
                "ОШИБКА",
                f"Программа не смогла вставить данные в краткое содержание значения по "
                f"Рпл {type(e).__name__}\n\n{str(e)}",
            )
        column_well = (
            f"{self.data_well.column_diameter.get_value}х{self.data_well.column_wall_thickness.get_value} "
            f"в "
            f"инт 0 - {self.data_well.shoe_column.get_value}м "
            if self.data_well.column_additional is False
            else f"{self.data_well.column_diameter.get_value} х "
                 f"{self.data_well.column_wall_thickness.get_value} \n"
                 f"0 - {self.data_well.shoe_column.get_value}м/\n{self.data_well.column_additional_diameter.get_value}"
                 f" х {self.data_well.column_additional_wall_thickness.get_value} в инт "
                 f"{self.data_well.head_column_additional.get_value}-{self.data_well.head_column_additional.get_value}м"
        )
        ws4.cell(row=1, column=7).value = column_well
        ws4.cell(row=4, column=7).value = (
            f"Пробур забой {self.data_well.bottom_hole_drill.get_value}м"
        )
        ws4.cell(row=5, column=7).value = (
            f"Исскус забой {self.data_well.bottom_hole_artificial.get_value}м"
        )
        ws4.cell(row=6, column=7).value = f"Тек забой {self.data_well.bottom}м"

        ws4.cell(row=7, column=7).value = plast_str
        ws4.cell(row=11, column=7).value = (
            f"Рмакс {self.data_well.max_admissible_pressure.get_value}атм"
        )
        ws4.cell(row=3, column=2).value = plan_short
        nek_str = "НЭК "
        if len(self.data_well.leakiness_interval) != 0:

            for nek in self.data_well.leakiness_interval:
                nek_str += f"{nek} \n"

        ws4.cell(row=3, column=7).value = nek_str

        ws4.insert_rows(1, 2)
        ws4.insert_cols(1, 2)
        ws4.cell(row=2, column=3).value = "Краткое содержание плана работ"
        ws4.cell(row=2, column=3).font = Font(name="Arial", size=16, bold=True)

        # объединение ячеек
        ws4.merge_cells(
            start_row=2, start_column=3, end_row=2, end_column=9
        )  # Объединение оглавления
        ws4.merge_cells(
            start_row=5, start_column=3, end_row=7, end_column=3
        )  # Объединение строк ГНО
        ws4.merge_cells(
            start_row=4, start_column=5, end_row=4, end_column=6
        )  # объединение по класси
        ws4.merge_cells(
            start_row=3, start_column=9, end_row=4, end_column=9
        )  # Объединение строк данных по колонну
        ws4.merge_cells(start_row=9, start_column=9, end_row=12, end_column=9)
        ws4.merge_cells(start_row=5, start_column=4, end_row=13, end_column=8)

        for row_ind in range(3, 15):
            ws4.row_dimensions[row_ind].height = 80
            for col in range(3, 10):
                if row_ind == 3:
                    ws4.column_dimensions[get_column_letter(col)].width = 20

                ws4.cell(row=row_ind, column=col).border = data_list.thin_border
                ws4.cell(row=row_ind, column=col).font = Font(
                    name="Arial", size=13, bold=False
                )
                ws4.cell(row=row_ind, column=col).alignment = Alignment(
                    wrap_text=True, horizontal="left", vertical="center"
                )
        ws4.cell(row=5, column=4).font = Font(name="Arial", size=11, bold=False)
        ws4.hide = True
        ws4.page_setup.fitToPage = True
        ws4.page_setup.fitToHeight = False

        ws4.page_setup.fitToWidth = True
        ws4.print_area = "C2:I14"

    @staticmethod
    def delete_files():
        zip_path = os.path.dirname(os.path.abspath(__file__)) + "/ZIMA.zip"
        print(zip_path)
        destination_path = os.path.dirname(os.path.abspath(__file__)) + "/ZimaUpdate"
        a = os.path.exists(destination_path)
        if os.path.exists(destination_path):
            shutil.rmtree(destination_path)  # Удаляет папку ZimaUpdate
        if os.path.exists(zip_path):
            os.remove(zip_path)  # Удаляет файл Zima.zip


def show_splash_screen():
    # Создаем приложение
    app = QApplication(sys.argv)

    # Создаем главное окно
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    from data_base.config_base import connect_to_database
    from users.login_users import LoginWindow

    if MyWindow.check_process():
        MyWindow.show_confirmation()

    show_splash_screen()

    # window = MyWindow()
    # window.show()
