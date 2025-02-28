from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIntValidator, QDoubleValidator
from PyQt5.QtWidgets import QWidget, QLabel, QComboBox, QLineEdit, QGridLayout,  QPushButton, \
    QMessageBox, QApplication, QTableWidget, QTableWidgetItem, QVBoxLayout, QHeaderView
import openpyxl

import data_list


from work_py.parent_work import TabPageUnion, TabWidgetUnion, WindowUnion



class TabPageSoPvr(TabPageUnion):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.validator_int = QIntValidator(0, 600)
        self.validator_float = QDoubleValidator(0.87, 1.65, 2)

        self.number_brigada_label = QLabel('Номер бригады', self)
        self.number_brigada_combo = QComboBox(self)
        brigada_list = data_list.DICT_TELEPHONE
        self.number_brigada_combo.addItems(list(brigada_list.keys()))


        self.number_telephone_label = QLabel('номер телефона, self')
        self.number_telephone_edit = QLineEdit(self)


        self.date_new_label = QLabel('Дата заявки', self)
        self.date_new_edit = QLineEdit(self)
        self.date_new_edit.setText(f'{data_list.current_date.date()}')

        self.time_new_label = QLabel('Время заявки', self)
        self.time_new_edit = QLineEdit(self)
        self.time_new_edit.setText(f'12:00')

        self.work_label = QLabel("Ранее проведенные работы:", self)
        self.work_edit = QLineEdit(self)

        self.nkt_label = QLabel("Внутренный диаметр НКТ:", self)
        self.nkt_edit = QLineEdit(self)
        self.nkt_edit.setText('62')


        self.nkt_shoe_label = QLabel("Башмак НКТ:", self)
        self.nkt_shoe_edit = QLineEdit(self)

        self.nkt_com_label = QLabel("Компоновка НКТ (низ НКТ):", self)
        self.nkt_com_edit = QLineEdit(self)

        self.paker_type_label = QLabel("Тип пакера", self)
        self.paker_type= QLineEdit(self)

        self.pakerLabel = QLabel("глубина пакера", self)
        self.paker_depth = QLineEdit(self)

        self.fluid_label = QLabel("уд.вес жидкости глушения", self)
        self.fluid_edit = QLineEdit(self)


        self.labelType = QLabel("Кровля  перфорации", self)
        self.lineedit_type = QLineEdit(self)
        self.lineedit_type.setClearButtonEnabled(True)

        self.labelType2 = QLabel("Подошва  перфорации", self)
        self.lineedit_type2 = QLineEdit(self)
        self.lineedit_type2.setClearButtonEnabled(True)

        self.labelTypeCharges = QLabel("Тип зарядов", self)
        self.ComboBoxCharges = QComboBox(self)
        self.ComboBoxCharges.addItems(['ГП', 'БО'])
        # self.spinBox.setAlignment(QtCore.Qt.AlignCenter)
        # self.spinBox.setMinimum(1917)
        # self.spinBox.setMaximum(2060)
        self.ComboBoxCharges.setProperty("value", 'ГП')

        self.labelHolesMetr = QLabel("отверстий на 1п.м", self)
        self.lineEditHolesMetr = QComboBox(self)
        self.lineEditHolesMetr.addItems(['6', '8', '10', '16', '18', '20', '30'])
        self.lineEditHolesMetr.setProperty("value", '20')

        self.labelIndexFormation = QLabel("Индекс пласта", self)
        self.lineEditIndexFormation = QLineEdit(self)
        self.lineEditIndexFormation.setClearButtonEnabled(True)

        self.label_type_pvr = QLabel("Вид перфорации", self)
        self.combo_pvr_type = QComboBox(self)
        self.combo_pvr_type.addItems(["На кабеле", 'Трубная перфорация'])

        # self.grid = QGridLayout(self)

        self.grid.addWidget(self.number_brigada_label, 7, 2)
        self.grid.addWidget(self.number_brigada_combo, 8, 2)

        self.grid.addWidget(self.number_telephone_label, 7, 3)
        self.grid.addWidget(self.number_telephone_edit, 8, 3)

        self.grid.addWidget(self.date_new_label, 9, 3)
        self.grid.addWidget(self.date_new_edit, 10, 3)

        self.grid.addWidget(self.time_new_label, 9, 4)
        self.grid.addWidget(self.time_new_edit, 10, 4)

        self.grid.addWidget(self.work_label, 11, 3, 1, 2)
        self.grid.addWidget(self.work_edit, 12, 3, 2, 4)

        self.grid.addWidget(self.nkt_label, 15, 2)
        self.grid.addWidget(self.nkt_edit, 16, 2)

        self.grid.addWidget(self.nkt_shoe_label, 15, 3)
        self.grid.addWidget(self.nkt_shoe_edit, 16, 3)

        self.grid.addWidget(self.nkt_com_label, 15, 4)
        self.grid.addWidget(self.nkt_com_edit, 16, 4)


        self.grid.addWidget(self.paker_type_label, 17, 2)
        self.grid.addWidget(self.paker_type, 18, 2)

        self.grid.addWidget(self.pakerLabel, 17, 3)
        self.grid.addWidget(self.paker_depth, 18, 3)

        self.grid.addWidget(self.fluid_label, 19, 3)
        self.grid.addWidget(self.fluid_edit, 20, 3)

        self.grid.addWidget(self.labelType, 22, 2)
        self.grid.addWidget(self.lineedit_type, 23, 2)

        self.grid.addWidget(self.labelType2, 22, 3)
        self.grid.addWidget(self.lineedit_type2, 23, 3)

        self.grid.addWidget(self.labelTypeCharges, 22, 4)
        self.grid.addWidget(self.ComboBoxCharges, 23, 4)

        self.grid.addWidget(self.labelHolesMetr, 22, 5)
        self.grid.addWidget(self.lineEditHolesMetr, 23, 5)

        self.grid.addWidget(self.labelIndexFormation, 22, 6)
        self.grid.addWidget(self.lineEditIndexFormation, 23, 6)
        self.grid.addWidget(self.label_type_pvr, 22, 7)
        self.grid.addWidget(self.combo_pvr_type, 23, 7)

        self.number_brigada_combo.currentTextChanged.connect(self.update_brigade)

    def update_brigade(self, index):
        self.number_telephone_edit.setText(str(data_list.DICT_TELEPHONE[self.number_brigada_combo.currentText()]))


class TabWidget(TabWidgetUnion):
    def __init__(self, parent=None):
        super().__init__()
        self.addTab(TabPageSoPvr(parent), 'Заявка на ПВР')


class PvrApplication(WindowUnion):
    def __init__(self, table_pvr, parent=None):
        super().__init__(parent)
        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)
        self.table_pvr = table_pvr
        # self.model = self.table_pvr.model()
        # self.table_pvr.setColumnCount(42)
        # self.table_pvr.setRowCount(113)
        self.tabWidget = TabWidget(parent)


        self.tableWidget = QTableWidget(0, 7)
        self.tableWidget.setHorizontalHeaderLabels(
            ["Кровля перфорации", "Подошва Перфорации", "Тип заряда", "отв на 1 п.м.",
             "Количество отверстий", "Вскрываемые пласты", "доп информация"])
        for i in range(7):
            self.tableWidget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)


        # self.tableWidget.setSortingEnabled(True)
        # self.tableWidget.setAlternatingRowColors(True)

        self.buttonAdd = QPushButton('Добавить интервалы перфорации в таблицу')
        self.buttonAdd.clicked.connect(self.add_row_table)
        self.buttonDel = QPushButton('Удалить интервалы перфорации в таблице')
        self.buttonDel.clicked.connect(self.del_row_table)
        self.buttonadd_work = QPushButton('Создать заявку')
        self.buttonadd_work.clicked.connect(self.add_work, Qt.QueuedConnection)
        self.buttonAddProject = QPushButton('Добавить интервалы перфорации из плана')
        self.buttonAddProject.clicked.connect(self.addPerfProject)


        vbox = QGridLayout(self.centralWidget)
        vbox.addWidget(self.tabWidget, 0, 0, 1, 2)
        vbox.addWidget(self.tableWidget, 1, 0, 1, 2)
        vbox.addWidget(self.buttonAdd, 2, 0)
        vbox.addWidget(self.buttonDel, 2, 1)
        vbox.addWidget(self.buttonadd_work, 3, 0)
        vbox.addWidget(self.buttonAddProject, 3, 1)

    def addPerfProject(self):
        for row_ind, row in enumerate(self.data_well.ws.iter_rows(values_only=True)):
            for col_ind, col in enumerate(row):
                if col_ind in [3, 2]:
                    if 'кровля' in str(col).lower():
                        type_pvr = self.data_well.ws.cell(row=row_ind, column=3).value
                        index_row_pvr_begin = row_ind + 1
                    if 'произвести контрольную' in str(col).lower():
                        index_row_pvr_cancel = row_ind
                        if index_row_pvr_begin < index_row_pvr_cancel:
                            self.data_well.index_row_pvr_list.append(
                                (index_row_pvr_begin, index_row_pvr_cancel, type_pvr))
                            index_row_pvr_begin, index_row_pvr_cancel = 0, 0
        for pvr in self.data_well.index_row_pvr_list:
            for row in range(pvr[0], pvr[1]):
                row_list = []
                for col in range(2, 9):
                    row_list.append(str(self.data_well.ws.cell(row=row + 1, column=col + 1).value))
                self.data_well.pvr_row_list.append(row_list)
        if not self.data_well.pvr_row_list:

            QMessageBox.warning(self, 'Ошибка', 'Перфорация в плане работ не найдены')
            return
        rows = self.tableWidget.rowCount()
        for pvr in self.data_well.pvr_row_list:
            self.tableWidget.insertRow(rows)
            self.tableWidget.setItem(rows, 0, QTableWidgetItem(str(pvr[0])))
            self.tableWidget.setItem(rows, 1, QTableWidgetItem(str(pvr[2])))
            self.tableWidget.setItem(rows, 2, QTableWidgetItem(str(pvr[3])))
            self.tableWidget.setItem(rows, 3, QTableWidgetItem(str(pvr[4])))
            self.tableWidget.setItem(rows, 4, QTableWidgetItem(str(pvr[5])))
            self.tableWidget.setItem(rows, 5, QTableWidgetItem(str(pvr[6])))


    def add_row_table(self):

        edit_type = self.tabWidget.currentWidget().lineedit_type.text().replace(',', '.')
        edit_type2 = self.tabWidget.currentWidget().lineedit_type2.text().replace(',', '.')
        chargesx = str(self.tabWidget.currentWidget().ComboBoxCharges.currentText())
        editHolesMetr = self.tabWidget.currentWidget().lineEditHolesMetr.currentText()
        editIndexFormation = self.tabWidget.currentWidget().lineEditIndexFormation.text()
        dopInformation = self.tabWidget.currentWidget().lineEditDopInformation.text()
        if not edit_type or not edit_type2 or not chargesx or not editIndexFormation:
            QMessageBox.information(self, 'Внимание', 'Заполните все поля!')
            return
        if float(edit_type2.replace(',', '.')) >= float(self.data_well.current_bottom):
            QMessageBox.information(self, 'Внимание', 'Подошва интервала перфорации ниже текущего забоя')
            return

        # chargesx = PerforationWindow.charge(self, int(float(edit_type2)))[0][:-2] + chargesx)

        self.tableWidget.setSortingEnabled(False)
        rows = self.tableWidget.rowCount()
        self.tableWidget.insertRow(rows)
        self.tableWidget.setItem(rows, 0, QTableWidgetItem(edit_type))
        self.tableWidget.setItem(rows, 1, QTableWidgetItem(edit_type2))
        self.tableWidget.setItem(rows, 2, QTableWidgetItem(chargesx))
        self.tableWidget.setItem(rows, 3, QTableWidgetItem(editHolesMetr))
        self.tableWidget.setItem(rows, 4, QTableWidgetItem(str(int((float(edit_type2) - float(
            edit_type)) * int(editHolesMetr)))))
        self.tableWidget.setItem(rows, 5, QTableWidgetItem(editIndexFormation))
        self.tableWidget.setItem(rows, 6, QTableWidgetItem(dopInformation))
        self.tableWidget.setSortingEnabled(True)
        # print(edit_type, spinYearOfIssue, editSerialNumber, editSpecifications)

    def copy_pvr(self, ws, work_list):
        for row in range(len(work_list)):
            for col in range(42):
                if work_list[row][col]:

                    ws.cell(row=row + 1, column=col + 1).value = work_list[row][col]

        # Перебираем строки и скрываем те, у которых все значения равны None
        for row_ind, row in enumerate(ws.iter_rows(values_only=True)):
            if all(value is None for value in row[:42]):

                ws.row_dimensions[row_ind+1].hidden = True

    def add_work(self):
        

        wb = openpyxl.load_workbook(f'{data_list.path_image}property_excel/template_pvr.xlsx')
        # Выбираем активный лист
        self.ws_pvr = wb.active
        number_brigada = str(self.tabWidget.currentWidget().number_brigada_combo.currentText())
        number_telephone = self.tabWidget.currentWidget().number_telephone_edit.text()
        date_new_edit = self.tabWidget.currentWidget().date_new_edit.text()
        time_new_edit = self.tabWidget.currentWidget().time_new_edit.text()
        work_edit = self.tabWidget.currentWidget().work_edit.text()
        nkt_edit = self.tabWidget.currentWidget().nkt_edit.text()
        nkt_shoe_edit = self.tabWidget.currentWidget().nkt_shoe_edit.text()
        nkt_com_edit = self.tabWidget.currentWidget().nkt_com_edit.text()
        paker_type = self.tabWidget.currentWidget().paker_type.text()
        paker_depth = self.tabWidget.currentWidget().paker_depth.text()
        fluid = self.tabWidget.currentWidget().fluid_edit.text()
        type_pvr = self.tabWidget.currentWidget().combo_pvr_type.currentText()
        if type_pvr == "На кабеле":
            type_pvr_str = '2.9.1'
        else:
            type_pvr_str = '2.9.2'
        rows = self.tableWidget.rowCount()
        perf_list = []

        for row in range(rows):
            roof = self.tableWidget.item(row, 0).text()
            sole = self.tableWidget.item(row, 1).text()
            type_charge = self.tableWidget.item(row, 2).text()
            count_otv = self.tableWidget.item(row, 3).text()
            count_charge = self.tableWidget.item(row, 4).text()
            plast = self.tableWidget.item(row, 5).text()

            perf_list.append(
                [None, type_pvr_str, None, None, type_pvr, None, None, None, None, None, type_charge, None, None, None,
                 None,
                 None, None, count_otv, None, None, None, None, None, count_charge, None, None,
                 None, None, plast,
                 None, None, None, None, None,  roof, None,  None, None, sole, None,  None,
                 None,
                 None,
                 None])

        work_list = self.application_pvr_def(number_brigada, number_telephone, date_new_edit, time_new_edit, work_edit, nkt_edit,
                                             nkt_shoe_edit,  nkt_com_edit, paker_type, paker_depth, fluid)

        for index, row in enumerate(perf_list):
            work_list[15 + index] = row

        self.copy_pvr(self.ws_pvr, work_list)

        # MyWindow.copy_pz(self, self.ws_pvr, self.table_pvr, 'application_pvr', 42)

        data_list.pause = False
        self.close()
        self.ws_pvr.print_area = f'B1:AP{114}'

        filenames = f'{self.data_well.well_number.get_value} {self.data_well.well_area.get_value} ПВР {data_list.current_date.date()}.xlsx'
        path = 'D:\Documents\Desktop\ГТМ\заявки ГИС'
        full_path = path + "/" + filenames
        if wb:
            wb.close()
            self.save_file_dialog(wb, full_path)
            # wb2.save(full_path)
            print(f'Table data saved to Excel {full_path}')
        if wb:
            wb.close()


    def del_row_table(self):
        row = self.tableWidget.currentRow()
        if row == -1:
            QMessageBox.information(self, 'Внимание', 'Выберите строку для удаления')
            return
        self.tableWidget.removeRow(row)


    def application_pvr_def(self, number_brigada, number_telephone, date_new_edit, time_new_edit, work_edit, nkt_edit, nkt_shoe_edit,
                            nkt_com_edit, paker_type, paker_depth, fluid):

        column_data = f'{self.data_well.column_diameter.get_value}мм x {self.data_well.column_wall_thickness.get_value} в инт ' \
                      f'0-{self.data_well.shoe_column.get_value}м'
        if self.data_well.column_additional:
            column_data_add = f'{self.data_well.column_additional_diameter.get_value}мм x ' \
                              f'{self.data_well.column_additional_wall_thickness.get_value} в инт ' \
                          f'{self.data_well.head_column_additional.get_value}-{self.data_well.shoe_column_additional.get_value}м'
        else:
            column_data_add = ''
        pressure = self.data_well.dict_category[list(self.data_well.dict_category.keys())[0]]['по давлению'].data_pressure

        value_list = [
            ['З А Я В К А', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             '№', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None,
             None, None, None, None],
            ['на проведение прострелочно-взрывных работ', None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Исполнитель', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, 'по договору №', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Заказчик', None, None, None, f'{data_list.contractor}', None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, 'Цех', None, None, self.data_well.cdng.get_value, None, None, None, None, None, None, None, None,
             None,
             None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Уполномоченный представитель', None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, '№ скважины', None, None, None, None, self.data_well.well_number.get_value, None, None, None, None, None, None,
             'куст', None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Регион', None, None, None, self.data_well.region, None, None, None, None, None, None, None,
             'Месторождение',
             None, None, None, None, None, self.data_well.well_area.get_value, None, None, None, None, None, None, None, None, None,
             None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Дата', None, None, None, None, None, None, None, None, 'Время', None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, '(по регламенту за 16 часов)', None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Дата', None, None, date_new_edit, None,  None, None, None, None,
             'Время', None, None, time_new_edit, None,  None, None,
             None,
             None, None, None, None, None, None, None, '(готовности по факту)', None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Комплекс и интервал исследования:', None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None],
            [None, '№ Тех. \nкарты', None, None, 'Цель ГИРС', None, None, None, None, None, 'тип ПВА', None, None, None,
             None,
             None, None, 'плотность перф., \nотв. 1 п.м.', None, None, None, None, None, 'Объем ПВР', None, None, None,
             None,
             'Индекс пласта', None, None, None, None, None, 'интервал ПВР/ГИС, м.', None, None, None, None, None, None,
             None,
             None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 'от', None, None,
             None,
             'до', None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Примечание к комплексу исследования', None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Основные сведения по скважине:', None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Категория скважины по ГНВП', None, None, None, None, None, None, None, None, None, 'Рпл:', None,
             self.data_well.category_pressure,
             None, None, None, None, None, 'H2S:', None, self.data_well.category_h2s, None, None, None, None, None, None,
             'Газовый фактор:',
             None, None, None, None, None, None,self.data_well.category_gas_factor, None, None, None, None, None, None, None],
            [None, 'Пробуренный забой', None, None, None, None, None, None, self.data_well.bottom_hole_drill.get_value, None, None, None,
             None,
             'м.', None, 'Искусственный забой', None, None, None, None, None, None, None, self.data_well.bottom_hole_artificial.get_value, None,
             None,
             'м.', None, 'Текущий забой', None, None, None, None, None, self.data_well.current_bottom, None, None, None, 'м.',
             None,
             None, None, None],
            [None, 'Максимальный угол', None, None, None, None, None, None, self.data_well.max_angle.get_value, None, None, None, None,
             None, None, 'гр.', None, 'на глубине', None, None, None, self.data_well.max_angle_depth.get_value, None, None, None, None, 'м.',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Расстояние муфта-ротор', None, None, None, None, None, None, None, None, self.data_well.stol_rotor.get_value, None,
             None, None, 'м.', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'диаметр обсадной колонны, мм.', None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, column_data, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'доп колонна, мм.', None, None, None, None, None, None, None, None, None, None, None, None, None,
             column_data_add,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None],
            [None, 'Внутр. диаметр насосно-компрессорных труб, мм', None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, nkt_edit, None, None, None, None, 'Башмак НКТ, м.', None, None, None, None,
             None,
             None, f'{nkt_shoe_edit}м', None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Компоновка НКТ', None, None, None, None, None, nkt_com_edit, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None],
            [None, 'Пакер', None, None, paker_type, None, None, None, None, None, None, None, None, None, None, None, None,
             'Глубина спуска пакера', None, None, None, None, None, None, None, paker_depth, None, None, None, None, None,
             None, 'м.',
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Качество цементирования', None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, 'Высота подъема цемента за колонной, м.', None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, self.data_well.level_cement_column.get_value, None, None, None, None, None, None, None, None],
            [None, 'Устьевое оборудование скважины', None, None, None, None, None, None, None, None, None, None, None,
             f'ПШП-{self.data_well.column_diameter.get_value}',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Скважина заполнена:', None, None, None, None, None, None, None, 'Тип:', None, 'тех.вода ', None,
             None, None,
             None, None, None, None, None, None, None, None, None, 'Уровень, м', None, None, None,
             self.data_well.static_level.get_value, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Плотность, г./см3', None, None, None, None, None, fluid, None, None, None, None, None,
             'Вязкость, сек.', None, None, None, None, None, None, None, None, None, None, None, 'УЭС, Омм', None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Гидростатическое давление в интервале работ, атм', None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Ожидаемое пластовое давление, МПа', None, None, None, None, None, None, None, None, None, None,
             None, None,
             f'{pressure}атм', None, None, None, None, None, None, None, 'Газовый фактор, м3/т', None, None, None,
             None, None,
             None, None, self.data_well.gaz_factor_percent[0], None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Температура в интервале ПВР, С', None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Сведения о ранее проведенных ПВР', None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'интервал', None, None, None, None, None, 'тип перфоратора', None, None, None, None, None, None,
             'плотность',
             None, None, None, None, 'дата', None, None, None, None, None, 'индекс пласта', None, None, None, None,
             None, None,
             'примечание', None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Другие данные по скважине:', None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Наличие электроэнергии', None, None, None, None, None, None, None, None, 'нет', None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Способ эксплуатации', None, None, None, None, None, None, None, f'{self.data_well.dict_pump_ecn["before"]}', None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, 'глубина спуска, м', None, None, None, None, None,
             f'{self.data_well.dict_pump_ecn_depth["before"]}',
             None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Время остановки скважины', None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Текущий дебит, Приемистость', None, None, None, None, None, None, None, None, None, None, None,
             None, ',',
             None, None, 'т/сут, м3/сут', None, None, None, None, None, 'Обводненность', None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'наличие сужений(нет/да), интервал', None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'наличие уступов(нет/да), интервал', None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None],
            [None,
             'Описание работ проводимых непосредственно перед ГИРС (кислотные обработки, агрессивные растворы, '
             'продолжительность реакции обработки, время промывки после обработки реперфорация и др.):',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None],
            [None, work_edit, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Дополнительные сведения:', None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Максимально ожидаемое давление на устье скважины', None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, f'{self.data_well.max_admissible_pressure.get_value}', None, None, None, None, None, None,
             'атм.',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Расстояние до скважины', None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             'км.', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Ответственный представитель Заказчика на скважине во время ГИРС', None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             'телефон',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, number_brigada, None, None, None, None,
             number_telephone,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Заявку подал', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, 'телефон', None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None],
            [None, f'{data_list.user[0]} {data_list.user[1]}', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, 'Заявку согласовал', None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, 'телефон', None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             '№ заявки, присвоенный в ЦДС-Менеджер', None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Версия 1.02 от 08.04.2021', None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None],

        ]
        pvr_list = [[None, 'интервал', None, None, None, None, None, 'тип перфоратора', None, None, None, None, None, None,
             'плотность',
             None, None, None, None, 'дата', None, None, None, None, None, 'индекс пласта', None, None, None, None,
             None, None,
             'примечание', None, None, None, None, None, None, None, None, None, None]]


        for plast in self.data_well.plast_all:
            for interval in self.data_well.dict_perforation[plast]['интервал']:
                if self.data_well.dict_perforation[plast]['отключение']:
                    izol = 'Изолирован'
                else:
                    izol = 'рабочий'
                pvr_list.append(
                    [None, f'{interval[0]}-{interval[1]}', None, None, None, None, None, None, None, None, None, None,
                     None, None, None,
                     None, None, None, None, None, None, None, None, None, None, plast, None, None, None, None,
                     None, None,
                     izol, None, None, None, None, None, None,
                     None, None, None, None])
        col = 0
        for pvr in pvr_list:

            value_list[47 + col] = pvr
            col += 1


        return value_list



# app = QApplication(sys.argv)
# login_window = PvrApplication()
# login_window.show()
# sys.exit(app.exec_())