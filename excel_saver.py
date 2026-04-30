from datetime import datetime
import json
import os
from io import BytesIO
import base64

import win32com.client
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import get_column_letter
from PIL import Image
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QMainWindow

import data_list
from data_base.work_with_base import excel_in_json
from H2S import CalculateH2s
from block_name import pop_down
from excel_worker import ExcelWorker
from main import MyWindow, MyMainWindow
from open_pz import CreatePZ
from server_response import ApiClient
from work_py.alone_oreration import is_number
from work_py.progress_bar_save import ProgressBarWindow

from excel_date_normalize import normalize_workbook_iso_date_strings
from excel_post_save import (
    com_open_workbook_editable,
    ensure_excel_file_writable_on_disk,
)


# Я удалил handle_worker_error из этого блока

class SaveInExcel(MyWindow):
    def __init__(
            self,
            data_well,
            ws,
            table_widget,
            table_title=None,
            table_schema=None,
            gnkt_data=None,
    ):
        # ВАЖНО: SaveInExcel используется как вспомогательный класс при сохранении Excel.
        # Нам НЕ нужна полная инициализация MyWindow (она поднимает UI/логин и дергает connect_to_database).
        # Но для PyQt-объекта нужно инициализировать базовый QMainWindow, иначе будет RuntimeError.
        QMainWindow.__init__(self)
        self.threads = []
        self.data_well = data_well
        self.table_widget = table_widget
        self.table_title = table_title
        self.table_schema = table_schema
        self.ws = ws
        self.gnkt_data = gnkt_data

    @staticmethod
    def _sanitize_sheet_title(title):
        """Убирает из имени листа символы, недопустимые в Excel (\\ / : * ? [ ])."""
        invalid = '\\/:*?[]'
        for char in invalid:
            title = title.replace(char, ' ')
        return title.strip()[:31]

    def save_to_gnkt(self):
        from gnkt_data.gnkt_data import insert_data_base_gnkt

        if self.gnkt_data is None:
            return
        gnkt = self.gnkt_data
        sheets = ["Титульник", "СХЕМА", "Ход работ"]
        tables = [self.table_title, self.table_schema, self.table_widget]

        for i, sheet_name in enumerate(sheets):
            worksheet = gnkt.wb_gnkt[sheet_name]
            table = tables[i]

            work_list = []
            for row in range(table.rowCount()):
                row_lst = []
                # self.insert_index_border += 1
                for column in range(table.columnCount()):

                    item = table.item(row, column)
                    if not item is None:

                        row_lst.append(item.text())
                        # print(item.text())
                    else:
                        row_lst.append("")
                work_list.append(row_lst)
            gnkt.count_row_height(worksheet, work_list, sheet_name)
        if "СХЕМЫ КНК_38,1" not in gnkt.wb_gnkt.sheetnames:
            ws7 = gnkt.wb_gnkt.create_sheet(title="СХЕМЫ КНК_38,1")
            self.insert_image(
                ws7,
                f"{data_list.path_image}imageFiles/schema_well/СХЕМЫ КНК_38,1.png",
                "A1",
                550,
                900,
            )

        path = self.load_last_save_path()

        filenames = self.definition_filenames()
        full_path = (path + "/" + filenames) if path else filenames
        gnkt_data = gnkt.data_gnkt
        params = {
            "gnkt_number": gnkt_data.gnkt_number_combo,
            "well_number": self.data_well.well_number.get_value,
            "well_area": self.data_well.well_area.get_value,
            "contractor": data_list.contractor,
            "length_gnkt": int(gnkt_data.length_gnkt_edit),
            "diameter_gnkt": float(gnkt_data.diameter_length),
            "wear_gnkt": float(gnkt_data.iznos_gnkt_edit) * 1.014,
            "mileage_gnkt": int(gnkt_data.pipe_mileage_edit) + int(gnkt_data.current_bottom_edit * 1.1),
            "tubing_fatigue": gnkt_data.pipe_fatigue,
            "previous_well": gnkt_data.previous_well_combo,
            "date_repair": data_list.current_date,
            "pvo_number": int(gnkt_data.pvo_number)
        }
        response = ApiClient.request_post(ApiClient.add_data_gnkt(), params)
        if response is None:
            QMessageBox.warning(self, "ошибка", "скважина не добавлена в well_data")
            # return



        if self.data_well.bvo is True:
            ws5 = gnkt.wb_gnkt.create_sheet("Sheet1")
            ws5.title = "Схемы ПВО"
            ws5 = gnkt.wb_gnkt["Схемы ПВО"]
            gnkt.wb_gnkt.move_sheet(ws5, offset=-1)
            schema_list = self.check_pvo_schema(ws5, 2)

        if gnkt.wb_gnkt:
            self.save_file_dialog(gnkt.wb_gnkt, full_path)

            gnkt.wb_gnkt.close()
            print(f"Table data saved to Excel {full_path}")
        if gnkt.wb_gnkt:
            gnkt.wb_gnkt.close()

    def save_to_krs(self):

        from work_py.alone_oreration import is_number
        from data_base.work_with_base import excel_in_json

        if not self.table_widget is None:
            self.wb2 = Workbook()
            self.ws2 = self.wb2.active
            if self.ws2 is None:
                return
            self.ws2.title = "План работ"

            insert_index = (self.data_well.insert_index2 + 2) if self.data_well.insert_index2 is not None else 0

            merged_cells = []  # Список индексов объединения ячеек

            work_list = []
            for row in range(self.table_widget.rowCount()):
                row_lst = []
                # self.insert_index_border += 1
                for column in range(self.table_widget.columnCount()):
                    if (
                            self.table_widget.rowSpan(row, column) > 1
                            or self.table_widget.columnSpan(row, column) > 1
                    ):
                        merged_cells.append((row, column))
                    item = self.table_widget.item(row, column)
                    if not item is None:
                        if "Нормы времени" in item.text():
                            insert_index = row
                        if self.check_str_isdigit(item.text()):
                            row_lst.append(item.text().replace(",", "."))
                        else:
                            row_lst.append(item.text())
                    else:
                        row_lst.append("")

                work_list.append(row_lst)

            merged_cells_dict = {}
            # print(f' индекс объ {insert_index}')
            for row in merged_cells:
                if row[0] >= insert_index - 1:
                    merged_cells_dict.setdefault(row[0], []).append(row[1])
            plan_short = ""
            self.data_well.norm_of_time = 0
            number_index_norm = 11
            if "prs" in self.data_well.work_plan:
                number_index_norm = 14
            for i in range(1, len(work_list)):  # нумерация работ
                if i >= insert_index + 1:
                    if is_number(work_list[i][number_index_norm]) is True:
                        self.data_well.norm_of_time += round(
                            float(
                                str(work_list[i][number_index_norm]).replace(",", ".")
                            ),
                            1,
                        )
                    if work_list[i][0]:
                        plan_short += f"п.{work_list[i][1]} {work_list[i][0]} \n"

            self.count_row_height(
                self.wb2, self.ws, self.ws2, work_list, merged_cells_dict, insert_index
            )

            self.data_well.itog_ind_min = insert_index
            self.data_well.itog_ind_max = len(work_list)
            if "prs" not in self.data_well.work_plan:
                self.ws2_prs = None
            self.add_itog(
                self.ws2, self.table_widget.rowCount() + 1, self.data_well.work_plan
            )

            # try:
            for row_ind, row in enumerate(self.ws2.iter_rows(values_only=True)):
                if 15 < row_ind < 100 and (
                        self.data_well.data_pvr_max.get_value + 13 > row_ind
                        or row_ind > self.data_well.data_fond_min.get_value + 13
                ):
                    if all(cell in [None, ""] for cell in row) and (
                            "Интервалы темпа"
                            not in str(self.ws2.cell(row=row_ind, column=2).value)
                            and "Замечания к эксплуатационному периоду"
                            not in str(self.ws2.cell(row=row_ind, column=2).value)
                            and "Замечания к эксплуатационному периоду"
                            not in str(self.ws2.cell(row=row_ind - 2, column=2).value)
                    ):
                        # print(row_ind, ('Интервалы темпа' not in str(ws2.cell(row=row_ind, column=2).value)),
                        #       str(ws2.cell(row=row_ind, column=2).value))
                        self.ws2.row_dimensions[row_ind + 1].hidden = True
                for col, value in enumerate(row):
                    if "И.М. Зуфаров" in str(value):
                        coordinate = f"{get_column_letter(col + 1)}{row_ind - 1}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/Зуфаров.png",
                            coordinate,
                        )
                    elif "Зуфаров" in str(value):
                        coordinate = f"{get_column_letter(col - 3)}{row_ind - 2}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/Зуфаров.png",
                            coordinate,
                        )


                    elif " - коэффициент усадки песка" in str(value):
                        coordinate = f"I{row_ind + 1}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/schema_well/sanf_formular-Photoroom.png",
                            coordinate,
                            200,
                            80,
                        )
                    elif (
                            "Категория скважины:" in str(value)
                            and str(getattr(self.data_well, "category_h2s", "")) in ("1", "2")
                    ):
                        coordinate = f"I{max(1, row_ind - 5)}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/schema_well/сероводород.jpg",
                            coordinate,
                            400,
                            120,
                        )
                    elif "М.К.Алиев" in str(value):
                        coordinate = f"{get_column_letter(col - 2)}{row_ind - 2}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/Алиев махир.png",
                            coordinate,
                        )

                    elif "З.К. Алиев" in str(value):
                        coordinate = f"{get_column_letter(col - 2)}{row_ind - 2}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/Алиев Заур.png",
                            coordinate,
                        )
                    elif "Рахимьянов" in str(value):
                        coordinate = f"{get_column_letter(col - 3)}{row_ind - 2}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/рахимьянов.png",
                            coordinate,
                        )
                    elif "Галиев Р.Р." in str(value):
                        coordinate = f"{get_column_letter(col - 3)}{row_ind - 2}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/Галиев.png",
                            coordinate,
                        )
                        break
                    elif "Расчет жидкости глушения производится согласно МУ" in str(
                            value
                    ):
                        ind = 6
                        row_ind_ins = row_ind + 1
                        if "prs" in self.data_well.work_plan:
                            ind = 13
                            row_ind_ins = row_ind
                        coordinate = f"{get_column_letter(ind)}{row_ind_ins}"
                        self.insert_image(
                            self.ws2,
                            f"{data_list.path_image}imageFiles/schema_well/формула.png",
                            coordinate,
                            300,
                            120,
                        )
                        break
            if self.data_well.work_plan in ["krs", "plan_change", "dop_plan_in_base", "dop_plan"]:
                self.create_short_plan(self.wb2, plan_short)
                # Сохраняем plan_short в data_well для последующего сохранения в БД
                self.data_well.plan_short = plan_short
            #
            if "Ойл" in data_list.contractor and "prs" not in self.data_well.work_plan:
                self.insert_image(
                    self.ws2, f"{data_list.path_image}imageFiles/Хасаншин.png", "H1"
                )
                self.insert_image(self.ws2, f'{data_list.path_image}imageFiles/Котиков.png', 'H4')

            excel_data_dict = excel_in_json(self, self.ws2)

            # Отправка данных в API и обновление brief_work_plan в одном потоке
            import threading
            def send_and_update_brief_plan():
                # Отправляем данные в API
                repair_id = self._send_wells_repair_to_api(excel_data_dict)
                
                # Если есть plan_short и получен ID, обновляем brief_work_plan
                if repair_id and hasattr(self.data_well, 'plan_short') and self.data_well.plan_short:
                    self._update_brief_work_plan(repair_id, self.data_well.plan_short)
            
            # Запускаем в отдельном потоке, чтобы не блокировать основной процесс
            thread = threading.Thread(target=send_and_update_brief_plan)
            thread.daemon = True
            thread.start()
            self.thread_excel_insert = ExcelWorker(self.data_well)
            # response_answer = MyMainWindow.insert_data_in_database(self, excel_data_dict)
            self.threads.append(self.thread_excel_insert)

            

            if "prs" in self.data_well.work_plan:
                self.ws2.print_area = f"B1:O{self.ws2.max_row}"
            else:
                self.ws2.print_area = f"B1:L{self.ws2.max_row}"
            self.ws2.page_setup.fitToPage = True
            self.ws2.page_setup.fitToHeight = False
            self.ws2.page_setup.fitToWidth = True
            self.ws2.print_options.horizontalCentered = True
            # зададим размер листа
            self.ws2.page_setup.paperSize = self.ws2.PAPERSIZE_A4
            # содержимое по ширине страницы
            page_set_up_pr = getattr(self.ws2.sheet_properties, "pageSetUpPr", None)
            if page_set_up_pr is not None:
                page_set_up_pr.fitToPage = True
            self.ws2.page_setup.fitToHeight = False

            path = self.load_last_save_path()

            filenames = self.definition_filenames()
            if path:
                full_path = path + "/" + filenames
            else:
                full_path = filenames

            if self.data_well.work_plan not in ["dop_plan", "dop_plan_in_base"]:
                from H2S import CalculateH2s

                if self.data_well.bvo:
                    ws5 = self.wb2.create_sheet("Sheet1")
                    ws5.title = "Схемы ПВО"
                    ws5 = self.wb2["Схемы ПВО"]
                    self.wb2.move_sheet(ws5, offset=-1)
                    schema_list = self.check_pvo_schema(ws5, insert_index + 2)
                category_check_list = []
                for plast in self.data_well.dict_category:
                    if self.data_well.dict_category[plast] not in category_check_list:
                        if self.data_well.dict_category[plast][
                            "по сероводороду"
                        ].category in [1, 2]:
                            name_list = f"Расчет H2S {plast}"
                            name_list = self._sanitize_sheet_title(name_list)
                            self.ws3 = self.wb2.create_sheet(name_list, 1)
                            calculate = CalculateH2s(self.data_well)
                            calculate.calc_h2s(self.ws3, plast)
                            category_check_list.append(
                                self.data_well.dict_category[plast]
                            )

                            # # Скрываем лист
                            self.ws3.sheet_state = "hidden"

            # Перед сохранением установите режим расчета
            self.wb2.calculation.calcMode = "auto"

            if self.wb2:
                self.wb2.close()
                self.save_file_dialog(self.wb2, full_path)
            


    def _find_repair_id(self):
        """
        Находит ID записи ремонта по параметрам скважины.
        Возвращает ID или None, если запись не найдена.
        """
        try:
            # Проверяем, что данные доступны
            if not hasattr(self.data_well, 'well_number') or not hasattr(self.data_well, 'well_area'):
                return None

            # Важно: backend /wells_repair_router/find_well_id требует обязательный параметр wells_id,
            # поэтому сначала получаем wells_id через /wells_data_router/find_wells_data.
            wells_params = {
                "well_number": self.data_well.well_number.get_value,
                "well_area": self.data_well.well_area.get_value,
            }
            wells_response = ApiClient.request_params_get(
                ApiClient.find_wells_data_response_filter_well_number_well_area(),
                wells_params
            )
            if not wells_response or "id" not in wells_response:
                return None
            wells_id = wells_response["id"]
            
            # Подготавливаем данные для поиска записи ремонта
            params = {
                "well_number": self.data_well.well_number.get_value,
                "well_area": self.data_well.well_area.get_value,
                "wells_id": wells_id,
            }
            
            # Добавляем опциональные параметры, если они доступны
            if hasattr(self.data_well, 'type_kr') and self.data_well.type_kr:
                params["type_kr"] = self.data_well.type_kr.split()[0]
            
            if hasattr(self.data_well, 'work_plan') and self.data_well.work_plan:
                params["work_plan"] = self._work_plan_for_db(
                    self.data_well.work_plan,
                    getattr(self.data_well, "number_dp", 0),
                )
            
            if hasattr(data_list, 'current_date') and data_list.current_date:
                params["date_create"] = data_list.current_date.strftime("%Y-%m-%d")
            
            # Ищем запись ремонта
            response = ApiClient.request_params_get(
                ApiClient.find_wells_repair_well_by_id(),
                params
            )
            
            if response and "id" in response:
                return response["id"]
            return None
        except Exception as e:
            print(f"Ошибка при поиске ID записи ремонта: {e}")
            return None
    
    def _update_brief_work_plan(self, repair_id, plan_short):
        """
        Обновляет поле brief_work_plan для записи ремонта через API.
        """
        try:
            # Убеждаемся, что repair_id - это число
            if not isinstance(repair_id, int):
                repair_id = int(repair_id) if repair_id else None
                if repair_id is None:
                    print("Ошибка: repair_id не может быть None")
                    return
            
            # Подготавливаем данные для обновления brief_work_plan
            # Сохраняем plan_short как словарь с ключом "plan_short"
            if plan_short:
                brief_work_plan_data = {
                    "plan_short": str(plan_short)
                }
            else:
                brief_work_plan_data = None
            
            update_data = {
                "id": int(repair_id),
                "brief_work_plan": brief_work_plan_data
            }
            
            # Логируем данные для отладки
            print(f"Отправка данных для обновления brief_work_plan: {update_data}")
            print(f"Тип repair_id: {type(repair_id)}, значение: {repair_id}")
            print(f"Тип brief_work_plan_data: {type(brief_work_plan_data)}, значение: {brief_work_plan_data}")
            
            # Обновляем brief_work_plan через API
            # Передаем данные в теле запроса (JSON), а не в query параметрах
            result = ApiClient.request_put_json(
                ApiClient.update_brief_work_plan_path(),
                update_data,
                param=None,
                answer="json"
            )
            
            if result:
                print(f"Краткий план работ успешно сохранен для записи ID: {repair_id}")
            else:
                print(f"Не удалось сохранить краткий план работ для записи ID: {repair_id}")
        except Exception as e:
            print(f"Ошибка при сохранении краткого плана работ: {e}")
            import traceback
            traceback.print_exc()

    @staticmethod
    def _work_plan_for_db(work_plan: str, number_dp=None) -> str:
        """Преобразует work_plan для отправки в БД/API.

        plan_change → ПРизм
        dop_plan_in_base → ДП№<number_dp> (или просто "ДП", если number_dp неизвестен)
        """
        if not work_plan:
            return work_plan or ""
        if work_plan == "plan_change":
            return "ПРизм"
        if work_plan == "dop_plan_in_base":
            try:
                # number_dp в UI хранится как int (1..N), но безопасно обработаем строки.
                num = int(float(number_dp)) if number_dp not in (None, "", 0, "0") else 0
            except Exception:
                num = 0
            return f"ДП№{num}" if num > 0 else "ДП"
        return work_plan.replace("krs", "ПР").replace("dop_plan", "ДП").replace("prs", "ПР_ТРС")

    @staticmethod
    def _build_expected_data(data_well):
        """Собирает expected_data из полей data_well для отправки в API."""
        existing = getattr(data_well, 'expected_data', None)
        if existing and isinstance(existing, dict) and existing:
            return existing

        def _val(attr_name, default=0):
            v = getattr(data_well, attr_name, default)
            if v is None:
                return default
            if isinstance(v, tuple):
                v = v[0] if v else default
            if isinstance(v, (int, float, str)):
                return v
            if hasattr(v, 'get_value'):
                return getattr(v, 'get_value')
            return v

        return {
            "expected_oil": _val("expected_oil"),
            "water_cut": _val("water_cut"),
            "percent_water": _val("percent_water"),
            "expected_pressure": _val("expected_pressure"),
            "expected_pickup": _val("expected_pickup"),
        }

    def _send_wells_repair_to_api(self, excel_data_dict):
        """
        Отправляет данные ремонта скважины в API через ручку /wells_repair_router/add_wells_data.
        Возвращает ID созданной записи или None в случае ошибки.
        """
        try:
            # Проверяем, что необходимые данные доступны
            if not hasattr(self.data_well, 'well_number') or not hasattr(self.data_well, 'well_area'):
                print("Недостаточно данных для отправки в API: отсутствуют well_number или well_area")
                return None

            # Формируем data_change_paragraph: API ожидает dict с ключом "данные" (JSON-строка списка)
            dcp = getattr(self.data_well, 'data_change_paragraph', None)
            if dcp is None or dcp == "" or (isinstance(dcp, dict) and not dcp):
                data_list_paragraph = getattr(self.data_well, 'data_list', [])
                data_change_paragraph = {"данные": json.dumps(data_list_paragraph, ensure_ascii=False)}
            elif isinstance(dcp, str):
                data_change_paragraph = {"данные": dcp}
            else:
                data_change_paragraph = dcp if isinstance(dcp, dict) else {"данные": "[]"}

            # Формируем payload согласно схеме SWellsRepair
            payload = {
                "id": 0,  # Backend сам присвоит ID при создании
                "category_dict": getattr(self.data_well, 'dict_category', {}),
                "type_kr": getattr(self.data_well, 'type_kr', '').split()[0],
                "work_plan": self._work_plan_for_db(
                    getattr(self.data_well, "work_plan", ""),
                    getattr(self.data_well, "number_dp", 0),
                ),
                "excel_json": excel_data_dict,
                "data_change_paragraph": data_change_paragraph,
                "norms_time": getattr(self.data_well, 'norm_of_time', 0.0),
                "chemistry_need": getattr(self.data_well, 'chemistry_need', {}),
                "geolog_id": "",  # Backend сам получит из токена пользователя
                "date_create": data_list.current_date.strftime("%Y-%m-%d") if hasattr(data_list, 'current_date') else datetime.now().strftime("%Y-%m-%d"),
                "perforation_project": getattr(self.data_well, 'dict_perforation_project', {}),
                "type_absorbent": getattr(self.data_well, 'type_absorbent', ''),
                "static_level": self.data_well.static_level.get_value if hasattr(self.data_well, 'static_level') and self.data_well.static_level else None,
                "dinamic_level": self.data_well.dinamic_level.get_value if hasattr(self.data_well, 'dinamic_level') and self.data_well.dinamic_level else None,
                "expected_data": self._build_expected_data(self.data_well),
                "curator": getattr(self.data_well, 'curator', ''),
                "region": getattr(self.data_well, 'region', ''),
                "contractor": data_list.contractor if hasattr(data_list, 'contractor') else '',
                "brief_work_plan": getattr(self.data_well, "plan_short", '')
            }

            # Отправляем данные в API
            # Backend ожидает well_number и well_area в query параметрах
            api_url = ApiClient.read_wells_repair_response_for_add()
            params = {
                "well_number": self.data_well.well_number.get_value,
                "well_area": self.data_well.well_area.get_value,
            }
            # Используем request_post_json с параметрами
            result = ApiClient.request_post_json(api_url, payload, param=params, answer="json")
            
            if result and isinstance(result, dict) and result.get("status") == "success":
                data = result.get("data")
                repair_id = data.get("id") if isinstance(data, dict) else None
                if repair_id is not None:
                    print(f"Данные ремонта успешно отправлены в API. ID записи: {repair_id}")
                return repair_id
            elif result == 500:
                print("Ошибка при отправке данных в API: ошибка сервера")
                return None
            else:
                print(f"Данные ремонта отправлены в API. Ответ: {result}")
                # Пытаемся извлечь ID из ответа, даже если статус не "success"
                if isinstance(result, dict) and 'id' in result:
                    return result.get('id')
                return None
                
        except Exception as e:
            print(f"Ошибка при отправке данных ремонта в API: {e}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    def load_last_save_path():
        """Загрузить последний сохраненный путь из файла."""
        if os.path.exists(f"{data_list.path_image}work_py/last_save_path.txt"):
            with open(f"{data_list.path_image}work_py/last_save_path.txt", "r") as file:
                return file.read().strip()
        return None

    @staticmethod
    def save_last_save_path(path):
        """Сохранить последний сохраненный путь в файл."""
        with open(f"{data_list.path_image}work_py/last_save_path.txt", "w") as file:
            file.write(path)

    def save_file_dialog(self, wb2, full_path):

        while True:  # Начинаем бесконечный цикл
            try:
                file_name, _ = QFileDialog.getSaveFileName(
                    None, "Save excel-file", f"{full_path}", "Excel Files (*.xlsx)"
                )
                if file_name:  # Если имя файла не пустое
                    ensure_excel_file_writable_on_disk(file_name)
                    normalize_workbook_iso_date_strings(wb2)
                    wb2.save(file_name)  # Пытаемся сохранить
                    ensure_excel_file_writable_on_disk(file_name)
                    break  # Если сохранение успешно, выходим из цикла

            except Exception as e:
                QMessageBox.critical(
                    None,
                    "Ошибка",
                    f"файл под таким именем открыт, закройте его: {type(e).__name__}\n  {str(e)}",
                )

        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True  # Сделать Excel видимым
            workbook = com_open_workbook_editable(excel, file_name)
            # Выбираем активный лист
            worksheet = workbook.ActiveSheet

            # Назначаем область печати с колонок B до L
            worksheet.PageSetup.PrintArea = "B:L"

        except Exception as e:
            print(f"Ошибка при работе с Excel: {type(e).__name__}\n\n{str(e)}")

    @staticmethod
    def reformated_string_data(num):
        try:
            if isinstance(num, datetime):
                return num.strftime("%d.%m.%Y")
            elif str(round(float(num), 6))[-1] != 0:
                return round(float(num), 6)
            elif str(round(float(num), 5))[-1] != 0:
                return round(float(num), 5)
            elif str(round(float(num), 4))[-1] != 0:
                return round(float(num), 4)
            elif str(round(float(num), 3))[-1] != 0:
                return round(float(num), 3)
            elif str(round(float(num), 2))[-1] != 0:
                return round(float(num), 2)
            elif str(round(float(num), 1))[-1] != 0:
                return round(float(num), 1)
            elif str(round(float(num), 0))[-1] != 0:
                return int(float(num))
        except:
            return num

    def count_row_height(self, wb2, ws, ws2, work_list, merged_cells_dict, ind_ins):
        global cell_num
        from openpyxl.utils.cell import range_boundaries, get_column_letter
        from PIL import Image

        stop_str = len(work_list)
        self.progress_bar_window = ProgressBarWindow(stop_str)
        self.progress_bar_window.show()

        boundaries_dict = {}

        text_width_dict = {
            35: (0, 100),
            50: (101, 200),
            70: (201, 300),
            95: (301, 400),
            110: (401, 500),
            150: (501, 600),
            170: (601, 700),
            190: (701, 800),
            230: (801, 1000),
            270: (1000, 1200),
            300: (1200, 1500),
            330: (1501, 1700),
            350: (1701, 2000),
            380: (1701, 2000),
            400: (2001, 3000)
        }

        for ind, _range in enumerate(ws.merged_cells.ranges):
            boundaries_dict[ind] = range_boundaries(str(_range))

        row_heights1 = [ws.row_dimensions[i].height for i in range(ws.max_row)]
        col_width = [
            ws.column_dimensions[get_column_letter(i + 1)].width for i in range(0, 15)
        ]
        # print(col_width)
        for i, row_data in enumerate(work_list):
            for column, data in enumerate(row_data):
                if column == 2 and i > ind_ins:
                    if data is not None:
                        text = data
                        text_len = len(text)
                        for key, value in text_width_dict.items():
                            if value[0] <= text_len <= value[1]:
                                ws2.row_dimensions[i + 1].height = min([400, int(key) + text.count("\n") * 4])

        if "prs" not in self.data_well.work_plan:
            head = self.head_ind(0, ind_ins)
            merge_column = 10
            size_font = 12
            font_type = "Arial"
            ws2.column_dimensions[get_column_letter(7)].width = 20
        else:
            head = self.head_ind_prs(0, ind_ins)
            merge_column = 13
            size_font = 16
            font_type = "Times New Roman"

        self.copy_true_ws(self.data_well, ws, ws2, head)
        boundaries_dict_index = 1000
        stop_str = 1500
        row_center = 1
        col_center = 1
        for i in range(1, len(work_list) + 1):  # Добавлением работ
            self.progress_bar_window.start_loading(i)
            if "Наименование работ" in work_list[i - 1][2]:
                boundaries_dict_index = i + 1

            if "код площади" in work_list[i - 1] or "код площади :" in work_list[i - 1]:

                for j in range(1, 13):

                    cell = ws2.cell(row=i, column=j)

                    # cell.number_format = 'General'
                    cell.value = str(work_list[i - 1][j - 1])
                    if "инв. №" in str(work_list[i - 1][j - 1]).lower():
                        ws2.cell(row=i, column=j + 1).number_format = "0"
                    elif "код площади" in str(work_list[i - 1][j - 1]).lower():
                        ws2.cell(row=i, column=j + 2).number_format = "General"
            elif (
                    "по H2S" in work_list[i - 1]
                    or "по H2S :" in work_list[i - 1]
                    or "по Pпл :" in work_list[i - 1]
                    or "по Pпл" in work_list[i - 1]
                    or "по газовому фактору" in work_list[i - 1]
                    or "по ГФ" in work_list[i - 1]
            ):
                for j in range(3, 13):
                    cell = ws2.cell(row=i, column=j)
                    cell.number_format = "General"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.value = work_list[i - 1][j - 1]
            elif "ИТОГО:" in work_list[i - 1]:
                stop_str = i

            for j in range(1, len(work_list[i - 1]) + 1):
                cell = ws2.cell(row=i, column=j)

                if i < 15:
                    cell.font = Font(name=font_type, size=size_font, bold=True)

                if "Наименование работ" in work_list[i - 1][2]:
                    work_list[i - 1][1] = "п/п"
                if cell and str(cell) != str(work_list[i - 1][j - 1]):
                    # print(work_list[i - 1][j - 1])
                    cell.value = self.reformated_string_data(work_list[i - 1][j - 1])
                    if "Ранее проведенные работ" in str(cell.value):
                        cell.font = Font(name=font_type, size=size_font, bold=False)
                    if i >= ind_ins:
                        if abs(i - ind_ins - 1) >= 1 and stop_str > i:
                            if self.data_well.work_plan in [
                                "dop_plan",
                                "dop_plan_in_base",
                            ]:
                                adwd = str(ws2[f"C{i}"].value)
                                if "Ранее проведенные работ" not in str(
                                        ws2[f"B{i}"].value
                                ) and "Порядок работы" not in str(ws2[f"B{i}"].value):
                                    ws2[f"B{i}"].value = (
                                        f"=COUNTA($C${ind_ins + 2}:C{i})"
                                    )
                            else:
                                if i != ind_ins:
                                    ws2[f"B{i}"].value = (
                                        f"=COUNTA($C${ind_ins + 2}:C{i})"
                                    )

                        if j != 1:
                            cell.border = data_list.thin_border
                        if j == 11:
                            cell.font = Font(name=font_type, size=size_font, bold=False)
                        # if j == 12:
                        #     cell.value = work_list[i - 1][j - 1]

                        if work_list[i - 1][4]:
                            ws2.cell(row=i - 1, column=2).alignment = Alignment(
                                wrap_text=True, horizontal="center", vertical="center"
                            )
                            ws2.cell(row=i, column=j).alignment = Alignment(
                                wrap_text=True, horizontal="center", vertical="center"
                            )
                        else:
                            ws2.cell(row=i, column=2).alignment = Alignment(
                                wrap_text=True, horizontal="center", vertical="center"
                            )
                            ws2.cell(row=i, column=3).alignment = Alignment(
                                wrap_text=True, horizontal="left", vertical="center"
                            )
                            if "prs" not in self.data_well.work_plan:
                                ws2.cell(row=i, column=11).alignment = Alignment(
                                    wrap_text=True,
                                    horizontal="center",
                                    vertical="center",
                                )
                                ws2.cell(row=i, column=12).alignment = Alignment(
                                    wrap_text=True,
                                    horizontal="center",
                                    vertical="center",
                                )
                            else:
                                ws2.cell(row=i, column=14).alignment = Alignment(
                                    wrap_text=True,
                                    horizontal="center",
                                    vertical="center",
                                )
                                ws2.cell(row=i, column=15).alignment = Alignment(
                                    wrap_text=True,
                                    horizontal="center",
                                    vertical="center",
                                )
                        if (
                                "примечание" in str(cell.value).lower()
                                or "заявку оформить за 16 часов" in str(cell.value).lower()
                                or "ЗАДАЧА 2.9." in str(cell.value).upper()
                                or "ВСЕ ТЕХНОЛОГИЧЕСКИЕ ОПЕРАЦИИ" in str(cell.value).upper()
                                or "за 48 часов до спуска".upper()
                                in str(cell.value).upper()
                                or "Требования безопасности".upper()
                                in str(cell.value).upper()
                                or "Контроль воздушной среды проводится".upper()
                                in str(cell.value).upper()
                                or "ТЕХНОЛОГИЧЕСКИЕ ПРОЦЕССЫ".upper()
                                in str(cell.value).upper()
                                or "Требования безопасности при выполнении работ".upper()
                                in str(cell.value).upper()
                                or "за 48 часов до спуска".upper()
                                in str(cell.value).upper()
                                or "за 48 часов до спуска".upper()
                                in str(cell.value).upper()
                                or "за 48 часов до спуска".upper()
                                in str(cell.value).upper()
                                or "РИР" in str(cell.value).upper()
                                or "При отсутствии избыточного давления" in str(cell.value)
                        ):
                            # print('есть жирный')
                            asde = ws2.cell(row=i, column=j).value
                            ws2.cell(row=i, column=j).font = Font(
                                name=font_type, size=size_font, bold=True
                            )
                        elif (
                                "Порядок работы" == cell.value
                                or "Наименование работ" == cell.value
                        ):
                            row_center = i
                            col_center = j
                            ws2.cell(row=i, column=j).font = Font(
                                name=font_type, size=size_font, bold=True
                            )
                            ws2.cell(row=i, column=j).alignment = Alignment(
                                wrap_text=True, horizontal="center", vertical="center"
                            )
                            if "Наименование работ" == cell.value:
                                cell_num = i - 1
                            elif self.data_well.work_plan == "dop_plan" and "Планируемый объём работ" in str(cell.value).upper():
                                cell_num = i - 1


                        else:
                            ws2.cell(row=i, column=j).font = Font(
                                name=font_type, size=size_font, bold=False
                            )

        # print(merged_cells_dict)
        for row, col in merged_cells_dict.items():
            if len(col) != 2:
                ws2.merge_cells(
                    start_row=row + 1,
                    start_column=3,
                    end_row=row + 1,
                    end_column=merge_column,
                )
            if row == cell_num:
                ws2.unmerge_cells(
                    start_row=row, start_column=3, end_row=row, end_column=merge_column
                )
                ws2.merge_cells(
                    start_row=row,
                    start_column=2,
                    end_row=row,
                    end_column=merge_column + 2,
                )

        ws2.cell(row=row_center, column=col_center).alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        ws2.merge_cells(
            start_row=row_center,
            start_column=3,
            end_row=row_center,
            end_column=merge_column,
        )
        ws2.merge_cells(
            start_row=row_center - 1,
            start_column=3,
            end_row=row_center - 1,
            end_column=merge_column,
        )
        ws2.cell(row=row_center - 2, column=col_center).font = Font(
            name=font_type, size=size_font, bold=False
        )
        ws2.cell(row=row_center - 1, column=col_center).alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )

        ws2.cell(row=row_center - 2, column=col_center).alignment = Alignment(
            wrap_text=True, horizontal="left", vertical="top"
        )

        for key, value in boundaries_dict.items():
            if value[1] <= boundaries_dict_index - 3 and value[1] != row_center - 1:
                ws2.merge_cells(
                    start_column=value[0],
                    start_row=value[1],
                    end_column=value[2],
                    end_row=value[3],
                )

        if self.data_well.image_data:
            for image_info in self.data_well.image_data:
                coord = image_info["coord"]
                width = image_info["width"]
                height = image_info["height"]
                image_base64 = image_info["data"]

                try:
                    # Декодирование из Base64 и создание изображения:
                    decoded_image_data = base64.b64decode(image_base64)

                    # Создаем объект PIL Image из декодированных данных
                    image = Image.open(BytesIO(decoded_image_data))

                    # Проверка размеров изображения:
                    print(f"Размеры изображения: {image.size}")

                    self.insert_image(ws2, image, coord, width * 0.72, height * 0.48)

                except Exception as e:
                    print(
                        f"Ошибка при вставке изображения: {type(e).__name__}\n\n{str(e)}"
                    )

        for index_row, row in enumerate(ws2.iter_rows()):  # Копирование высоты строки
            if all([col is None for col in row]):
                ws2.row_dimensions[index_row].hidden = True
            try:
                if index_row < ind_ins:
                    ws2.row_dimensions[index_row].height = row_heights1[index_row]
            except:
                pass

        for col_ind in range(len(col_width)):
            ws2.column_dimensions[get_column_letter(col_ind + 1)].width = col_width[
                col_ind
            ]
        if "prs" in self.data_well.work_plan:
            ws2.column_dimensions[get_column_letter(14)].width = 20
            ws2.column_dimensions[get_column_letter(15)].width = 20

            ws2.column_dimensions[get_column_letter(6)].width = 18
        else:
            ws2.column_dimensions[get_column_letter(11)].width = 20
            ws2.column_dimensions[get_column_letter(12)].width = 20

            ws2.column_dimensions[get_column_letter(6)].width = 20
            ws2.column_dimensions[get_column_letter(7)].width = 20

        return "Высота изменена"

    @staticmethod
    def copy_true_ws(data_well, ws, ws2, head):
        from copy import copy

        for row_number, row in enumerate(ws[head]):
            for col_number, cell in enumerate(row):

                if cell.value:
                    if row_number == 0:
                        if col_number > 6:
                            break
                        ws2.cell(row_number + 1, col_number + 1, cell.value)

                if (
                        "катег" in str(cell.value).lower()
                        and "план" not in str(cell.value).lower()
                ):
                    if data_well.work_plan not in [
                        "krs",
                        "dop_plan",
                        "dop_plan_in_base",
                        "plan_change",
                    ]:
                        ws2.cell(
                            row=row_number + 1, column=col_number + 1
                        ).alignment = Alignment(
                            wrap_text=True, horizontal="left", vertical="center"
                        )
                if type(cell.value) == float:
                    ws2.cell(row_number + 1, col_number + 1, round(cell.value, 5))
                else:
                    ws2.cell(row_number + 1, col_number + 1, cell.value)

                if cell.has_style:
                    ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                    ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                    ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                    ws2.cell(row_number + 1, col_number + 1).number_format = copy(
                        cell.number_format
                    )
                    ws2.cell(row_number + 1, col_number + 1).protection = copy(
                        cell.protection
                    )
                    ws2.cell(row_number + 1, col_number + 1).alignment = copy(
                        cell.alignment
                    )
                    ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(
                        cell.quotePrefix
                    )
                    ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(
                        cell.pivotButton
                    )

    @staticmethod
    def head_ind_prs(start, finish):
        return f"A{start}:O{finish}"

    @staticmethod
    def head_ind(start, finish):
        return f"A{start}:L{finish}"

    def add_itog_string(self):
        number_col = "L"
        if "prs" in self.data_well.work_plan:
            number_col = "O"

        add_itog_list = [
            [
                None,
                "ИТОГО:",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                f"=ROUND(SUM({number_col}{self.data_well.itog_ind_min + 2}:{number_col}{self.data_well.itog_ind_max}),1)",
            ],
            [
                None,
                "Герметизация , разгерметизация  устья  скважины",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                f"=ROUND(SUM({number_col}{self.data_well.itog_ind_min + 2}:{number_col}{self.data_well.itog_ind_max - 1})/11.5*11/60 ,1)",
            ],
            [
                None,
                "Заправка ДВС",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                f"=ROUND(SUM({number_col}{self.data_well.itog_ind_min}:{number_col}{self.data_well.itog_ind_max - 1})/11.5*0.3    ,1)",
            ],
            [
                None,
                "ПЗР в начале и конце смены с заполнением вахтового журнала",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                f"=ROUND(SUM({number_col}{self.data_well.itog_ind_min}:{number_col}{self.data_well.itog_ind_max - 1})/11.5*0.3,1)",
            ],
            [
                None,
                "Непредвиденные  работы  : ",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                f"=ROUND(SUM({number_col}{self.data_well.itog_ind_min}:L{self.data_well.itog_ind_max + 2})*"
                f"{self.data_well.bottom_hole_artificial.get_value}/100*0.0004 ,1)",
            ],
            [
                None,
                "ВСЕГО  :",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                f"=ROUND({number_col}{self.data_well.itog_ind_max + 1} + {number_col}{self.data_well.itog_ind_max + 2} +"
                f" {number_col}{self.data_well.itog_ind_max + 3} + {number_col}{self.data_well.itog_ind_max + 4} +{number_col}{self.data_well.itog_ind_max + 5}, 1)",
            ],
            [
                None,
                "Примечания: В соответствии с регламентом на производство КРС – заблаговременно подавать заявки на "
                "необходимое оборудование, а так же вызывать представителя Заказчика на геофизические работы, ПВР, "
                "установку пакера, срыв планшайбы, опрессовку колонны и другие технологические операции, прием "
                "скважины в ремонт и сдача из ремонта.",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                None,
                "ПРИМЕЧАНИЕ:",
                None,
                " ",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                None,
                "При незначительных изменениях в плане работ (изменении компоновки подземного оборудования, "
                "объемов закачки и т.д.)  и доп. работах в виде единичных СПО, технол.операций и др. возможна "
                "работа без доп. плана - по письму Заказчика.   ",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                None,
                "поглощения жидкости не допускать",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                None,
                "Ответственный за соблюдением и создание безопасных условий работ – мастера КPС ",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ]
        return add_itog_list

    def add_itog(self, ws, insert_index, work_plan, ws2=None):
        if ws.merged_cells.ranges:
            merged_cells_copy = list(
                ws.merged_cells.ranges
            )  # Создаем копию множества объединенных ячеек
            for merged_cell in merged_cells_copy:
                if merged_cell.min_row > insert_index + 5:
                    try:
                        ws.unmerge_cells(str(merged_cell))
                    except:
                        pass

        if "prs" not in self.data_well.work_plan:
            merge_column = 11
            size_font = 12
            font_type = "Arial"
        else:
            merge_column = 14
            size_font = 14
            font_type = "Times New Roman"

        if work_plan not in [
            "gnkt_frez",
            "application_pvr",
            "gnkt_after_grp",
            "gnkt_opz",
            "gnkt_bopz",
        ]:
            itog_list = self.add_itog_string()
            j = 1
            for i in range(
                    insert_index, len(itog_list) + insert_index
            ):  # Добавлением итогов
                row_list = itog_list[i - insert_index]
                if "prs" in self.data_well.work_plan:
                    row_list.insert(-8, None)
                    row_list.insert(-8, None)
                    row_list.insert(-8, None)
                if i < insert_index + 6:
                    for j in range(1, len(itog_list[i - insert_index]) + 1):
                        awded = row_list[j - 1]
                        ws.cell(row=i, column=j).value = row_list[j - 1]
                        if j != 1:
                            ws.cell(row=i, column=j).border = data_list.thin_border
                            ws.cell(row=i, column=j).font = Font(
                                name=font_type, size=size_font, bold=False
                            )

                    ws.merge_cells(
                        start_row=i, start_column=2, end_row=i, end_column=merge_column
                    )
                    ws.cell(row=i, column=j).alignment = Alignment(
                        wrap_text=True, horizontal="left", vertical="center"
                    )
                else:
                    ws.row_dimensions[insert_index + 6].height = 50
                    ws.row_dimensions[insert_index + 8].height = 50
                    for j in range(1, 13):
                        ws.cell(row=i, column=j).value = row_list[j - 1]
                        ws.cell(row=i, column=j).border = data_list.thin_border
                        ws.cell(row=i, column=j).font = Font(
                            name=font_type, size=size_font, bold=False
                        )
                        ws.cell(row=i, column=j).alignment = Alignment(
                            wrap_text=True, horizontal="left", vertical="center"
                        )

                    ws.merge_cells(
                        start_row=i,
                        start_column=2,
                        end_row=i,
                        end_column=merge_column + 1,
                    )
                    ws.cell(row=i, column=j).alignment = Alignment(
                        wrap_text=False, horizontal="left", vertical="center"
                    )

            insert_index += len(itog_list) + 2

        curator_s = self.curator_sel(self.data_well.curator, self.data_well.region)
        # print(f'куратор {curator_sel, self.data_well.curator}')
        if curator_s is False:
            return
        from block_name import pop_down

        if "prs" not in self.data_well.work_plan:
            podp_down = pop_down(self, self.data_well.region, curator_s)
        else:
            podp_down = pop_down(self, self.data_well.region, curator_s)[:3]

        for i in range(1 + insert_index, 1 + insert_index + len(podp_down)):
            # Добавлением подписантов внизу
            for j in range(1, 11):
                asdd = i - 1 - insert_index
                ws.cell(row=i, column=j).value = podp_down[i - 1 - insert_index][j - 1]
                ws.cell(row=i, column=j).font = Font(
                    name=font_type, size=size_font, bold=False
                )

            if i in range(insert_index + 7, 1 + insert_index + 15):
                ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
                ws.cell(row=i, column=2).alignment = Alignment(
                    wrap_text=False, vertical="center", horizontal="left"
                )
            else:
                ws.cell(row=i, column=2).alignment = Alignment(
                    wrap_text=False, vertical="center", horizontal="left"
                )

            ws.row_dimensions[insert_index + 7].height = 30
            ws.row_dimensions[insert_index + 9].height = 25

        insert_index += len(podp_down)
        aaa = ws.max_row

        ws.delete_rows(insert_index, aaa - insert_index)
        if "prs" in self.data_well.work_plan:
            from open_pz import CreatePZ

            CreatePZ.copy_data_excel_in_excel(
                self.data_well.ws,
                ws,
                self.data_well.prs_copy_index.get_value,
                self.data_well.data_fond_min.get_value,
                1,
                17,
                ws.max_row + 1,
            )

            CreatePZ.copy_data_excel_in_excel(
                self.data_well.ws,
                ws,
                self.data_well.condition_of_wells.get_value,
                self.data_well.ws.max_row,
                1,
                17,
                ws.max_row + 1,
            )

    @staticmethod
    def curator_sel(curator, region):

        with open(
                f"{data_list.path_image}podpisant.json", "r", encoding="utf-8"
        ) as file:
            podpis_dict = json.load(file)
        if curator == "ОР":
            return (
                podpis_dict[data_list.costumer][region]["ruk_orm"]["post"],
                podpis_dict[data_list.costumer][region]["ruk_orm"]["surname"],
            )
        elif curator == "ГТМ":
            return (
                podpis_dict[data_list.costumer][region]["ruk_gtm"]["post"],
                podpis_dict[data_list.costumer][region]["ruk_gtm"]["surname"],
            )
        elif curator == "ГО":
            return (
                podpis_dict[data_list.costumer][region]["go"]["post"],
                podpis_dict[data_list.costumer][region]["go"]["surname"],
            )
        elif curator == "ВНС":
            return (
                podpis_dict[data_list.costumer][region]["go"]["post"],
                podpis_dict[data_list.costumer][region]["go"]["surname"],
            )
        elif curator == "ГРР":
            return (
                podpis_dict[data_list.costumer][region]["grr"]["post"],
                podpis_dict[data_list.costumer][region]["grr"]["surname"],
            )
        return False



