import block_name
import data_list
from datetime import datetime
from data_base.config_base import connection_to_database, GnktDatabaseWell
from PyQt5.QtWidgets import QInputDialog, QTabWidget, QWidget, QApplication, QLabel, \
    QLineEdit, QGridLayout, QComboBox, QPushButton, QMessageBox
from openpyxl.utils import get_column_letter

from decrypt import decrypt
from gnkt_data.gnkt_data import gnkt_1, gnkt_2, gnkt_dict, read_database_gnkt
from openpyxl.styles import PatternFill, Font, Alignment

from open_pz import CreatePZ
from work_py.acid_paker import CheckableComboBox
from work_py.alone_oreration import well_volume
from work_py.parent_work import TabWidgetUnion, WindowUnion, TabPageUnion
from PyQt5.QtGui import QIntValidator, QDoubleValidator


class TabPageGnkt(TabPageUnion):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.iznos_gnkt_edit = None
        self.gnkt = None
        self.iznos_gnkt_label = None
        self.length_gnkt_edit = None
        self.length_gnkt_label = None
        self.gnkt_number_combo = None
        self.gnkt_number_label = None
        self.validator_int = QIntValidator(0, 8000)
        self.validator_float = QDoubleValidator(0, 8000, 2)

        self.init_gnkt()

    def init_gnkt(self):
        self.gnkt_number_label = QLabel('Номер флота ГНКТ')
        self.gnkt_number_combo = QComboBox(self)
        self.gnkt_number_combo.addItems(gnkt_dict["Ойл-сервис"])
        if self.gnkt_number_combo.currentText() == 'ГНКТ №1':
            self.gnkt = gnkt_1
        elif self.gnkt_number_combo.currentText() == 'ГНКТ №2':
            self.gnkt = gnkt_2

        self.length_gnkt_label = QLabel('длина ГНКТ')
        self.length_gnkt_edit = QLineEdit(self)

        self.iznos_gnkt_label = QLabel('Износ трубы')
        self.iznos_gnkt_edit = QLineEdit(self)

        self.pvo_number_label = QLabel('Номер ПВО')
        self.pvo_number_edit = QLineEdit(self)

        self.pipe_mileage_label = QLabel('Пробег трубы')
        self.pipe_mileage_edit = QLineEdit(self)

        self.previous_well_label = QLabel('Предыдущая скважина')
        self.previous_well_combo = QComboBox(self)

        self.current_bottom_label = QLabel('необходимый текущий забой')
        self.current_bottom_edit = QLineEdit(self)
        self.current_bottom_edit.setText(f'{self.data_well.need_depth}')

        self.fluid_label = QLabel("уд.вес жидкости глушения", self)
        self.fluid_edit = QLineEdit(self)
        self.fluid_edit.setValidator(self.validator_float)
        if self.data_well.fluid:
            self.fluid_edit.setText(str(self.data_well.fluid))

        self.fluid_work_label = QLabel("уд.вес рабочей жидкости", self)
        self.fluid_work_edit = QLineEdit(self)
        if self.data_well.work_plan == 'gnkt_opz':
            self.fluid_work_edit.setText(f'{self.data_well.water_density.get_value:.2f}')
        else:
            self.fluid_work_edit.setText('1.01')

        self.distance_pntzh_label = QLabel('Расстояние до ПНТЖ')
        self.distance_pntzh_line = QLineEdit(self)
        self.distance_pntzh_line.setValidator(self.validator_int)
        self.distance_pntzh_line.setText(f'{self.data_well.distance_from_well_to_sampling_point}')

        # self.grid = QGridLayout(self)
        self.grid.addWidget(self.gnkt_number_label, 0, 2, 1, 5)
        self.grid.addWidget(self.gnkt_number_combo, 1, 2, 1, 5)
        self.grid.addWidget(self.length_gnkt_label, 2, 3)
        self.grid.addWidget(self.length_gnkt_edit, 3, 3)
        self.grid.addWidget(self.iznos_gnkt_label, 2, 4)
        self.grid.addWidget(self.iznos_gnkt_edit, 3, 4)

        self.grid.addWidget(self.pipe_mileage_label, 2, 5)
        self.grid.addWidget(self.pipe_mileage_edit, 3, 5)

        self.grid.addWidget(self.pvo_number_label, 2, 6)
        self.grid.addWidget(self.pvo_number_edit, 3, 6)

        self.grid.addWidget(self.previous_well_label, 2, 7)
        self.grid.addWidget(self.previous_well_combo, 3, 7)

        self.grid.addWidget(self.current_bottom_label, 4, 2)
        self.grid.addWidget(self.current_bottom_edit, 5, 2)
        self.grid.addWidget(self.fluid_label, 4, 3)
        self.grid.addWidget(self.fluid_edit, 5, 3)
        self.grid.addWidget(self.fluid_work_label, 4, 4)
        self.grid.addWidget(self.fluid_work_edit, 5, 4)
        self.grid.addWidget(self.distance_pntzh_label, 4, 6)
        self.grid.addWidget(self.distance_pntzh_line, 5, 6)

        self.acids_work_label = QLabel('Проведение ОПЗ')
        self.acids_work_combo = QComboBox(self)
        self.acids_work_combo.addItems(['Нет', 'Да'])

        if self.data_well.work_plan == 'gnkt_after_grp':
            self.osvoenie_label = QLabel('Необходимость освоения')
            self.osvoenie_combo = QComboBox(self)
            self.osvoenie_combo.addItems(['', 'Да', 'Нет'])

            self.grid.addWidget(self.osvoenie_label, 4, 5)
            self.grid.addWidget(self.osvoenie_combo, 5, 5)
            self.grid.addWidget(self.acids_work_label, 6, 2)
            self.grid.addWidget(self.acids_work_combo, 7, 2)

            self.acids_work_combo.setCurrentIndex(2)
        elif self.data_well.work_plan == 'gnkt_frez':
            self.need_frez_port_label = QLabel('Необходимость фрезерования портов')
            self.need_frez_port_combo = QComboBox(self)
            self.need_frez_port_combo.addItems(['', 'Нет', 'Да'])
            self.grid.addWidget(self.need_frez_port_label, 4, 5)
            self.grid.addWidget(self.need_frez_port_combo, 5, 5)
            self.grid.addWidget(self.acids_work_label, 6, 2)
            self.grid.addWidget(self.acids_work_combo, 7, 2)
        else:
            self.acids_work_label.setParent(None)
            self.acids_work_combo.setParent(None)

        self.acids_work_combo.currentTextChanged.connect(self.update_acids_work)
        self.gnkt_number_combo.currentTextChanged.connect(self.update_number_gnkt)
        self.previous_well_combo.currentTextChanged.connect(self.update_data_gnkt)

    def update_acids_work(self, index):
        if index == 'Нет':
            try:
                self.acids_type_label.setParent(None)
                self.acid_edit.setParent(None)
                self.acid_volume_label.setParent(None)
                self.acid_volume_edit.setParent(None)
                self.acid_proc_label.setParent(None)
                self.acid_proc_edit.setParent(None)
                self.pressure_Label.setParent(None)
                self.pressure_edit.setParent(None)
                self.roof_label.setParent(None)
                self.roof_edit.setParent(None)

                self.sole_label.setParent(None)
                self.sole_edit.setParent(None)

                self.plast_label.setParent(None)
                self.plast_combo.setParent(None)
            except:
                pass

        elif index == 'Да':
            self.acids_type_label = QLabel('Вид кислоты')
            self.acid_edit = QComboBox(self)
            self.acid_edit.addItems(['Лимонная кислота', 'HCl', 'HF', 'vt'])
            self.acid_volume_label = QLabel("Объем кислотной обработки", self)
            self.acid_volume_edit = QLineEdit(self)
            self.acid_volume_edit.setValidator(self.validator_float)
            self.acid_volume_edit.setText("3")
            self.acid_volume_edit.setClearButtonEnabled(True)
            self.acid_proc_label = QLabel("Концентрация кислоты", self)
            self.acid_proc_edit = QLineEdit(self)
            self.acid_proc_edit.setText('15')
            self.acid_proc_edit.setClearButtonEnabled(True)
            self.acid_proc_edit.setValidator(self.validator_int)
            self.pressure_Label = QLabel("Давление закачки", self)
            self.pressure_edit = QLineEdit(self)
            self.pressure_edit.setText(f'{self.data_well.max_admissible_pressure.get_value}')
            self.pressure_edit.setValidator(self.validator_int)

            self.roof_label = QLabel("кровля пласта", self)
            self.roof_edit = QLineEdit(self)
            self.roof_edit.setText(f'{self.data_well.perforation_roof}')
            self.roof_edit.setValidator(self.validator_float)

            self.sole_label = QLabel("подошва пласта", self)
            self.sole_edit = QLineEdit(self)
            self.sole_edit.setText(f'{self.data_well.perforation_sole}')
            self.sole_edit.setValidator(self.validator_float)
            plast_work = ['']
            plast_work.extend(self.data_well.plast_work)

            self.plast_label = QLabel("Выбор пласта", self)
            self.plast_combo = CheckableComboBox(self)
            self.plast_combo.combo_box.addItems(plast_work)

            self.grid.addWidget(self.plast_label, 8, 2)
            self.grid.addWidget(self.plast_combo, 9, 2)
            self.grid.addWidget(self.roof_label, 8, 3)
            self.grid.addWidget(self.roof_edit, 9, 3)
            self.grid.addWidget(self.sole_label, 8, 4)
            self.grid.addWidget(self.sole_edit, 9, 4)

            self.grid.addWidget(self.acids_type_label, 6, 3)
            self.grid.addWidget(self.acid_edit, 7, 3)
            self.grid.addWidget(self.acid_volume_label, 6, 4)
            self.grid.addWidget(self.acid_volume_edit, 7, 4)
            self.grid.addWidget(self.acid_proc_label, 6, 5)
            self.grid.addWidget(self.acid_proc_edit, 7, 5)
            self.grid.addWidget(self.pressure_Label, 6, 6)
            self.grid.addWidget(self.pressure_edit, 7, 6)

    def update_data_gnkt(self):
        previus_well = self.previous_well_combo.currentText()
        try:
            if previus_well:
                db = connection_to_database(decrypt("DB_NAME_GNKT"))
                data_gnkt = GnktDatabaseWell(db)

                if 'ойл-сервис' in data_list.contractor.lower():
                    contractor_in = 'oil_service'

                result_gnkt = data_gnkt.update_data_gnkt(contractor_in, previus_well)

                self.length_gnkt_edit.setText(str(result_gnkt[3]))
                self.iznos_gnkt_edit.setText(str(result_gnkt[5]))
                self.pipe_mileage_edit.setText(str(result_gnkt[6]))
                self.pvo_number_edit.setText(str(result_gnkt[10]))


        except Exception as e:
            print(f'Ошибка  с базой ГНКТ {e}')

    def update_number_gnkt(self, gnkt_number):
        if gnkt_number != '':
            well_previus_list = read_database_gnkt(data_list.contractor, gnkt_number)

            self.previous_well_combo.clear()
            self.previous_well_combo.addItems(list(map(str, well_previus_list)))


class TabWidget(TabWidgetUnion):
    def __init__(self, parent=None):
        super().__init__()
        self.addTab(TabPageGnkt(parent), 'Данные по ГНКТ')


class GnktModel(WindowUnion):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.fluid_work_edit = None
        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)

        self.tab_widget = TabWidget(self.data_well)

        self.buttonAdd = QPushButton('Добавить данные в план работ')
        self.buttonAdd.clicked.connect(self.add_work)
        vbox = QGridLayout(self.centralWidget)
        vbox.addWidget(self.tab_widget, 0, 0, 1, 2)
        vbox.addWidget(self.buttonAdd, 2, 0)

    def insert_image_schema(self, ws):

        if self.data_well.paker_before["before"] != 0:
            coordinate_nkt_with_paker = 'F6'
            self.insert_image(ws, f'{data_list.path_image}imageFiles/schema_well/НКТ с пакером.png',
                              coordinate_nkt_with_paker, 100, 510)
        elif self.data_well.dict_nkt_before:
            coordinate_nkt_with_voronka = 'F6'
            self.insert_image(ws, f'{data_list.path_image}imageFiles/schema_well/НКТ с воронкой.png',
                              coordinate_nkt_with_voronka, 70, 470)

        if self.work_plan in ['gnkt_bopz']:
            coordinate = 'F65'
            self.insert_image(ws, f'{data_list.path_image}imageFiles/schema_well/angle_well.png',
                              coordinate, 265, 430)

        elif self.work_plan in ['gnkt_after_grp', 'gnkt_opz']:
            coordinate_propant = 'F43'
            if self.work_plan in ['gnkt_after_grp']:
                self.insert_image(ws, f'{data_list.path_image}imageFiles/schema_well/пропант.png',
                                  coordinate_propant, 90, 500)

            n = 0
            m = 0
            k = 27
            count_insert = 0
            for plast in self.data_well.plast_work:
                k += m
                m = 0
                for roof_plast, sole_plast in self.data_well.dict_perforation[plast]['интервал']:
                    count_insert += 1
                    count_interval = self.data_well.dict_perforation[plast]['счет_объединение']



                    try:
                        if roof_plast > self.data_well.depth_fond_paker_before["before"] and \
                                roof_plast < self.data_well.current_bottom and count_insert <= 5:
                            interval_str = f'{plast} \n {roof_plast}-{sole_plast}'
                            coordinate_pvr = f'F{43 + n}'

                            ws.cell(row=43 + n, column=10).value = interval_str
                            ws.merge_cells(start_column=10, start_row=43 + n,
                                           end_column=12, end_row=43 + n + 1)
                            ws.cell(row=43 + n, column=10).font = Font(name='Arial', size=12, bold=True)
                            ws.cell(row=43 + n, column=10).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                                 vertical='center')
                            n += 3
                            self.insert_image(ws, f'{data_list.path_image}imageFiles/schema_well/ПВР.png',
                                              coordinate_pvr, 85, 50)
                    except:
                        QMessageBox.critical(self,
                                             'Ошибка', f'программа не смогла вставить интервал перфорации в схему'
                                                       f'{roof_plast}-{sole_plast}')
                ws.merge_cells(start_column=23, start_row=k + m,
                               end_column=23, end_row=k + count_interval - 1)
                ws.merge_cells(start_column=22, start_row=k + m,
                               end_column=22, end_row=k + count_interval - 1)
                ws.merge_cells(start_column=21, start_row=k + m,
                               end_column=21, end_row=k + count_interval - 1)
                m += count_interval

            coordinate_voln = f'E18'
            self.insert_image(ws,
                              f'{data_list.path_image}imageFiles/schema_well/переход.png',
                              coordinate_voln, 150, 60)

    def count_row_height(self, ws2, work_list, sheet_name):

        from openpyxl.utils.cell import range_boundaries, get_column_letter

        col_width = [2.85546875, 14.42578125, 16.140625, 22.85546875, 17.140625, 14.42578125, 13.0, 13.0, 17.0,
                    14.42578125, 13.0, 21, 12.140625, None]

        text_width_dict = {35: (0, 100), 50: (101, 200), 70: (201, 300), 110: (301, 400), 120: (401, 500),
                           130: (501, 600), 150: (601, 700), 170: (701, 800), 190: (801, 900), 230: (901, 1500)}

        boundaries_dict = {}

        for ind, _range in enumerate(ws2.merged_cells.ranges):
            boundaries_dict[ind] = range_boundaries(str(_range))

        for key, value in boundaries_dict.items():
            ws2.unmerge_cells(start_column=value[0], start_row=value[1],
                              end_column=value[2], end_row=value[3])

        insert_index = 1

        for i in range(1, len(work_list) + 1):  # Добавлением работ
            if sheet_name == 'Ход работ':
                if len(str(work_list[i - 1][1])) <= 3 and str(work_list[i - 1][1]) != '№' and \
                        str(work_list[i - 1][1]) != 'п/п':  # Нумерация
                    work_list[i - 1][1] = str(insert_index)
                    insert_index += 1
                elif str(work_list[i - 1][1]) == 'п/п':
                    insert_index = 1

            for j in range(1, 13):
                cell = ws2.cell(row=i, column=j)

                if cell and str(cell) != str(work_list[i - 1][j - 1]):
                    if str(work_list[i - 1][j - 1]).replace('.', '').isdigit() and \
                            str(work_list[i - 1][j - 1]).count('.') != 2:
                        cell.value = str(work_list[i - 1][j - 1]).replace('.', ',')

                    else:
                        cell.value = work_list[i - 1][j - 1]

        # print(merged_cells_dict)
        if sheet_name != 'Ход работ':
            for key, value in boundaries_dict.items():
                # print(value)
                ws2.merge_cells(start_column=value[0], start_row=value[1],
                                end_column=value[2], end_row=value[3])
            if sheet_name == "Титульник":
                for i, row_data in enumerate(work_list):
                    # print(f'gghhg {work_list[i][2]}')
                    for column, data in enumerate(row_data):
                        if i < 2:
                            ws2.cell(row=i + 1, column=column + 1).alignment = Alignment(horizontal='left',
                                                                                         vertical='center')

            elif sheet_name == 'СХЕМА' and self.work_plan != 'gnkt_frez':
                self.insert_image_schema(ws2)


        elif sheet_name == 'Ход работ':
            for i, row_data in enumerate(work_list):
                # print(f'gghhg {work_list[i][2]}')
                for column, data in enumerate(row_data):
                    if column == 2:
                        if not data is None:
                            text = data
                            for key, value in text_width_dict.items():
                                if value[0] <= len(text) <= value[1]:
                                    ws2.row_dimensions[i + 1].height = int(key)
                    elif column == 1:
                        if not data is None:
                            text = data
                            # print(text)
                            for key, value in text_width_dict.items():
                                if value[0] <= len(text) <= value[1]:
                                    ws2.row_dimensions[i + 1].height = int(key)
                    if column != 0:
                        ws2.cell(row=i + 1, column=column + 1).border = data_list.thin_border
                    if column == 1 or column == 11:
                        ws2.cell(row=i + 1, column=column + 1).alignment = Alignment(wrap_text=True,
                                                                                     horizontal='center',
                                                                                     vertical='center')
                        ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=False)
                    else:
                        ws2.cell(row=i + 1, column=column + 1).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                                     vertical='center')
                        ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=False)
                        if 'примечание' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'внимание' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'мероприятия' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'порядок работ' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'По доп.согласованию с Заказчиком' in str(
                            ws2.cell(row=i + 1, column=column + 1).value).lower():
                            # print('есть жирный')
                            ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=True)

                if len(work_list[i][1]) > 5:
                    ws2.merge_cells(start_column=2, start_row=i + 1, end_column=12, end_row=i + 1)
                    ws2.cell(row=i + 1, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                        vertical='center')
                    ws2.cell(row=i + 1, column=2).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1',
                                                                     fill_type='solid')
                    ws2.cell(row=i + 1, column=2).font = Font(name='Arial', size=13, bold=True)
                elif i <= 39:
                    ws2.merge_cells(start_column=3, start_row=i + 1, end_column=11, end_row=i + 1)
                    ws2.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                        vertical='center')
                    ws2.cell(row=i + 1, column=11).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')
                elif i > 39 and self.data_well.work_plan in ['gnkt_opz']:
                    ws2.merge_cells(start_column=3, start_row=i + 1, end_column=10, end_row=i + 1)
                    ws2.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                        vertical='center')
                    ws2.cell(row=i + 1, column=11).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')
                else:
                    ws2.merge_cells(start_column=3, start_row=i + 1, end_column=11, end_row=i + 1)
                    ws2.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                        vertical='center')
                    ws2.cell(row=i + 1, column=11).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')

            for col in range(13):
                ws2.column_dimensions[get_column_letter(col + 1)].width = col_width[col]
            ws2.column_dimensions[get_column_letter(11)].width = 30
            ws2.print_area = f'B1:L{self.table_widget.rowCount() + 45}'
            ws2.page_setup.fitToPage = True
            ws2.page_setup.fitToHeight = False
            ws2.page_setup.fitToWidth = True
            ws2.print_options.horizontalCentered = True
            # зададим размер листа
            ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
            # содержимое по ширине страницы
            ws2.sheet_properties.pageSetUpPr.fitToPage = True
            ws2.page_setup.fitToHeight = False

        for row_ind, row in enumerate(ws2.iter_rows(values_only=True)):
            for col, value in enumerate(row):
                if 'Ойл' in data_list.contractor:
                    if 'А.Р. Хасаншин' in str(value):
                        coordinate = f'{get_column_letter(col + 1)}{row_ind - 1}'
                        self.insert_image(ws2, f'{data_list.path_image}imageFiles/Хасаншин.png', coordinate)
                    elif 'Д.Д. Шамигулов' in str(value):
                        coordinate = f'{get_column_letter(col + 1)}{row_ind - 2}'
                        self.insert_image(ws2, f'{data_list.path_image}imageFiles/Шамигулов.png', coordinate)
                    elif 'Зуфаров' in str(value):
                        coordinate = f'{get_column_letter(col - 2)}{row_ind}'
                        self.insert_image(ws2, f'{data_list.path_image}imageFiles/Зуфаров.png', coordinate)
                    elif 'Закиев И.Э.' in str(value):
                        coordinate = f'{get_column_letter(col - 1)}{row_ind}'
                        self.insert_image(ws2, f'{data_list.path_image}imageFiles/Закиев.png', coordinate)
                    elif 'Котиков' in str(value) and 'И.А.' in str(value):
                        coordinate = f'{get_column_letter(col - 1)}{row_ind}'
                        self.insert_image(ws2, f'{data_list.path_image}imageFiles/Котиков.png', coordinate)
                    elif 'Рахимьянов' in str(value):
                        coordinate = f'{get_column_letter(col - 1)}{row_ind}'
                        self.insert_image(ws2, f'{data_list.path_image}imageFiles/рахимьянов.png', coordinate)
    def work_with_data_gnkt(self):
        if self.data_gnkt is None:

            self.data_gnkt = GnktOsvWindow2(self.data_well)
            self.data_gnkt.setWindowTitle("Данные по ГНКТ")
            self.data_gnkt.setGeometry(200, 400, 100, 400)
            self.set_modal_window(self.data_gnkt)
            self.pause_app()
            data_list.pause = True

            if self.data_well.work_plan in ['gnkt_after_grp', 'gnkt_opz', 'gnkt_bopz']:
                self.work_schema = self.data_gnkt.schema_well(self.data_gnkt.current_bottom_edit,
                                                              self.data_gnkt.fluid_edit,
                                                              self.data_gnkt.gnkt_number_combo,
                                                              self.data_gnkt.length_gnkt_edit,
                                                              self.data_gnkt.iznos_gnkt_edit,
                                                              self.data_gnkt.pvo_number,
                                                              self.data_gnkt.diameter_length,
                                                              self.data_gnkt.pipe_mileage_edit)

                self.copy_pvr(self.ws_schema, self.work_schema)
        else:
            self.data_gnkt.close()
            self.data_gnkt = None
        return self.data_gnkt

    def update_opz_data(self, current_widget):

        self.roof_plast = round(float(current_widget.roof_edit.text().replace(',', '.')), 1)
        self.sole_plast = round(float(current_widget.sole_edit.text().replace(',', '.')), 1)

        self.acid_edit = current_widget.acid_edit.currentText()

        self.acid_volume_edit = float(current_widget.acid_volume_edit.text().replace(',', '.'))
        self.acid_proc_edit = int(current_widget.acid_proc_edit.text().replace(',', '.'))
        self.pressure_edit = int(current_widget.pressure_edit.text())
        self.plast_combo = str(current_widget.plast_combo.combo_box.currentText())

    def add_work(self):
        self.current_widget = self.tab_widget.currentWidget()
        self.gnkt_number_combo = self.current_widget.gnkt_number_combo.currentText()
        self.length_gnkt_edit = self.current_widget.length_gnkt_edit.text()
        self.iznos_gnkt_edit = self.current_widget.iznos_gnkt_edit.text().replace(',', '.')
        self.pipe_mileage_edit = self.current_widget.pipe_mileage_edit.text()
        self.distance_pntzh = self.current_widget.distance_pntzh_line.text()
        self.pipe_fatigue = 0

        self.current_bottom_edit = self.current_widget.current_bottom_edit.text()
        if self.current_bottom_edit == '':
            QMessageBox.warning(self, 'Некорректные данные', f'не указан текущий забоя')
            return
        else:
            self.current_bottom_edit = float(self.current_bottom_edit.replace(',', '.'))
            self.data_well.current_bottom = self.current_bottom_edit
            if self.current_bottom_edit > float(self.data_well.bottom_hole_drill.get_value):
                QMessageBox.warning(self, 'Некорректные данные',
                                    f'Текущий забой ниже пробуренного забоя {self.data_well.bottom_hole_drill.get_value}')
                return

        self.fluid_edit = self.current_widget.fluid_edit.text().replace(',', '.')
        self.fluid_work_edit = self.current_widget.fluid_work_edit.text().replace(',', '.')

        if self.data_well.work_plan in ['gnkt_after_grp', 'gnkt_frez']:

            self.acids_work_combo = self.current_widget.acids_work_combo.currentText()
            if self.acids_work_combo == '':
                QMessageBox.critical(self, "Ошибка", "Нужно выбрать необходимость фрезерования")
                return

            if self.acids_work_combo == 'Да':
                self.update_opz_data(self.current_widget)
                if self.roof_plast in ['', None, 0] or self.sole_plast in ['', None, 0]:
                    QMessageBox.critical(self, "Ошибка", "Не введены данные по кровле и ли подошве обработки")
                    return
            if self.data_well.work_plan == 'gnkt_frez':
                self.need_frez_port = self.current_widget.need_frez_port_combo.currentText()
                if self.need_frez_port == '':
                    QMessageBox.critical(self, "Ошибка", "Нужно выбрать необходимость фрезерования")
                    return
            elif self.data_well.work_plan == 'gnkt_after_grp':
                self.osvoenie_combo_need = self.current_widget.osvoenie_combo.currentText()
                if self.osvoenie_combo_need == '':
                    QMessageBox.critical(self, "Ошибка", "Нужно выбрать необходимость освоения")
                    return

        self.pvo_number = self.current_widget.pvo_number_edit.text()

        self.previous_well_combo = self.current_widget.previous_well_combo.currentText()

        self.diameter_length = 38

        if '' in [self.gnkt_number_combo, self.length_gnkt_edit, self.iznos_gnkt_edit, self.fluid_edit,
                  self.pvo_number]:
            QMessageBox.warning(self, 'Некорректные данные', f'Не все данные заполнены')
            return

        fluid_question = QMessageBox.question(self, 'Удельный вес',
                                              f'Работы необходимо производить на тех воде {self.fluid_work_edit}г/см3?')
        if fluid_question == QMessageBox.StandardButton.No:
            return

        question = QMessageBox.question(self, 'Забой',
                                              f'Забой нужно нормализовать до глубины {self.current_bottom_edit}м?')
        if question == QMessageBox.StandardButton.No:
            return

        self.data_well.gnkt_number_combo = self.gnkt_number_combo
        self.data_well.length_gnkt_edit = float(self.length_gnkt_edit.replace(',', '.'))
        self.data_well.iznos_gnkt_edit = float(self.iznos_gnkt_edit.replace(',', '.'))
        self.data_well.fluid_edit = self.fluid_edit.replace(',', '.')
        self.data_well.fluid_work_edit = self.fluid_work_edit.replace(',', '.')
        self.data_well.pvo_number = self.pvo_number
        self.data_well.pipe_mileage_edit = self.pipe_mileage_edit.replace(',', '.')

        self.well_volume_ek, self.well_volume_dp = self.check_volume_well(self.data_well)

        data_list.pause = False
        self.close()
        self.close_modal_forcefully()
        # return work_list

    def select_text_acid(self, plast_combo, roof, sole, acid_edit, acid_proc_edit, acid_volume_edit):

        if acid_edit == 'HCl':
            acid_24 = round(acid_volume_edit * acid_proc_edit / 24 * 1.118, 1)
            acid_sel = f'Произвести  солянокислотную обработку {plast_combo} в объеме {acid_volume_edit}м3 ' \
                       f' ({acid_edit} - {acid_proc_edit} %) силами/' \
                       f' Крезол НС с протяжкой БДТ вдоль интервалов перфорации {roof}-{sole}м ' \
                       f'(снизу вверх) в ' \
                       f'присутствии представителя заказчика с составлением акта, не превышая давления' \
                       f' закачки не более Р={self.data_well.max_admissible_pressure.get_value}атм.\n' \
                       f' (для приготовления соляной кислоты в объеме {acid_volume_edit}м3 - ' \
                       f'{acid_proc_edit}% необходимо замешать {acid_24}т HCL 24% и пресной воды ' \
                       f'{round(acid_volume_edit - acid_24, 1)}м3)'
            acid_sel_short = f'СКО пласта {plast_combo} {roof}-{sole} в объеме  {acid_volume_edit}м3  ' \
                             f'({acid_edit} - {acid_proc_edit} %)'
        elif acid_edit == 'ВТ':
            acid_sel = f'Произвести кислотную обработку пласта f{roof}-{sole}м {acid_edit} силами Крезол ' \
                       f'НС с протяжкой БДТ вдоль интервалов перфорации {roof}-{sole}м (снизу вверх) в присутствии ' \
                       f'представителя ' \
                       f'Заказчика с составлением акта, не превышая давления закачки не более ' \
                       f'Р = {self.data_well.max_admissible_pressure.get_value}атм.'
            acid_sel_short = f'{acid_edit} пласта {plast_combo}  в объеме ' \
                             f'{acid_volume_edit}м3  ({acid_edit} - {acid_proc_edit} %)'
        elif acid_edit == 'HF':
            acid_sel = f'Произвести глинокислотную обработку пласта {plast_combo} в объеме ' \
                       f'{acid_volume_edit}м3 ' \
                       f'(концентрация в смеси HF 3% / HCl 13%) силами Крезол ' \
                       f'НС с протяжкой БДТ вдоль интервалов перфорации {roof}-{sole}м (снизу вверх) в' \
                       f' присутствии представителя ' \
                       f'Заказчика с составлением акта, не превышая давления закачки не более ' \
                       f'Р={self.data_well.max_admissible_pressure.get_value}атм.'
            acid_sel_short = f'ГКО пласта {plast_combo}  в объеме  {acid_volume_edit}м3'
        elif acid_edit == 'Лимонная кислота':
            acid_sel = f'Произвести лимонной кислотой пласта {plast_combo} в объеме ' \
                       f'{acid_volume_edit}м3 ' \
                       f'с протяжкой БДТ вдоль интервалов перфорации {roof}-{sole}м (снизу вверх) в присутствии' \
                       f' представителя ' \
                       f'Заказчика с составлением акта, не превышая давления закачки не более ' \
                       f'Р={self.data_well.max_admissible_pressure.get_value}атм.'
            acid_sel_short = f'КО лимонной кислотой пласта {plast_combo} в объеме {acid_volume_edit}м3'

        return acid_sel, acid_sel_short

    def work_opz_gnkt(self, acid_info):

        self.volume_gntk = round(float(self.data_well.length_gnkt_edit) * 0.74 / 1000, 1)

        depth_fond_paker_do = sum(map(int, list(self.data_well.dict_nkt_before.values())))
        if self.data_well.depth_fond_paker_before["before"] == 0:
            self.depth_fond_paker_do = sum(list(self.data_well.dict_nkt_before.values()))
            # print(depth_fond_paker_do)
            if self.depth_fond_paker_do >= self.data_well.current_bottom:
                depth_fond_paker_do, ok = QInputDialog.getDouble(self, 'глубина НКТ',
                                                                 'Введите Глубины башмака НКТ', 500,
                                                                 0, self.data_well.current_bottom)
        else:
            self.depth_fond_paker_do = self.data_well.depth_fond_paker_before["before"]

        opz = []
        volume_sko = 0
        for plast_combo, skv_como, roof, sole, acid_edit, acid_proc_edit, acid_volume_edit in acid_info:
            acid_sel, acid_sel_short = self.select_text_acid(plast_combo, roof, sole, acid_edit, acid_proc_edit,
                                                             acid_volume_edit)
            volume_sko += acid_volume_edit

            opz.extend([[f'Установить КНК до глубины {sole}м',
                         17, f'Установить КНК до глубины {sole}м',
                         None, None, None, None, None, None, None,
                         'Мастер ГНКТ, состав бригады', 0.2],
                        [acid_sel_short,
                         17, acid_sel,
                         None, None, None, None, None, None, None,
                         'Мастер ГНКТ, состав бригады, подрядчик по ОПЗ', round(acid_volume_edit * 0.2, 1)]]
                       )
        opz.insert(0, [None, None,
                       f'Исследовать скважину на приёмистость при Рзак={self.data_well.expected_pressure}атм с составлением акта в '
                       f' в присутствии представителя ЦДНГ с составлением соответствующего акта (для вызова представителя '
                       f'давать телефонограмму в ЦДНГ). Определение приёмистости производить после насыщения пласта не '
                       f'менее 6м3 или '
                       f'при установившемся давлении закачки, но не более 1 часов.',
                       None, None, None, None, None, None, None,
                       'Мастер ГНКТ, состав бригады, представитель Заказчика', None])
        opz.insert(0, [None,
                       f'КИСЛОТНАЯ ОБРАБОТКА в объеме {volume_sko}м3 {acid_edit} {acid_proc_edit}% ',
                       None, None, None, None, None, None, None, None,
                       'Мастер ГНКТ, состав бригады, представитель Заказчика', None])


        work_list = [[None,
                      18,
                      f'ПРИМЕЧАНИЕ:\n '
                      f'Закачку первого объема {self.volume_gntk}м3 кислоты производить при открытом малом '
                      f'затрубном пространстве на '
                      f'циркуляции. Закачку оставшейся '
                      f'кислоты в объеме {round(volume_sko - self.volume_gntk, 1)}м3 производить '
                      f'при закрытом малом затрубном '
                      f'пространстве. Составить Акт.',
                      None, None, None, None, None, None, None,
                      'Мастер ГНКТ, состав бригады', 2.88],
                     [None, 19, f'Продавить кислоту в пласт мин.водой уд.веса {self.data_well.fluid_work} '
                                f'в объёме '
                                f'{round(self.volume_gntk + 3, 1)}м3 при '
                                f'давлении не более '
                                f'{self.data_well.max_admissible_pressure.get_value}атм. Составить Акт',
                      None, None, None, None, None, None, None,
                      'Мастер ГНКТ, состав бригады', 1.11],
                     [None, 20,
                      f'Приподнять БДТ на {int(self.depth_fond_paker_do) - 20}м. Стоянка на реакции 2 часа. В СЛУЧАЕ'
                      f' ОТСУТСТВИЯ ДАВЛЕНИЯ '
                      f'ПРОДАВКИ ПРИ СКО, РАБОТЫ ПРОИЗВОДИМ БЕЗ РЕАГИРОВАНИЯ.СОСТАВИТЬ АКТ)',
                      None, None, None, None, None, None, None,
                      'Мастер ГНКТ, состав бригады', 2.16],
                     ['разрядку скважины для извлечения продуктов',
                      21,
                      'Произвести разрядку скважины для извлечения продуктов реакции кислоты в объёме не менее объёма '
                      'закаченной кислоты + объём малого затрубного пространства (из расчета 1,88л на 1 м пространства между '
                      '73мм колонной НКТ и БДТ;'
                      ' 0,46л между 60мм НКТ и БДТ; 3,38л между 89мм НКТ и БДТ) + 3м3. Разрядку производить до чистой'
                      ' промывочной '
                      'жидкости (без признаков продуктов реакции кислоты), но не более 2 часов. Зафиксировать '
                      'избыточное давление '
                      'на устье скважины, объём и описание скважинной жидкости на выходе с отражением их в акте,'
                      ' суточном рапорте '
                      'работы бригад. Составить Акт.',
                      None, None, None, None, None, None, None,
                      'Мастер ГНКТ, состав бригады', 1],
                     ['Допустить БДТ до забоя. Промыть скважину ',
                      16,
                      f'Допустить БДТ до забоя. Промыть скважину  мин.водой уд.веса {self.data_well.fluid_work}  с составлением '
                      f'соответствующего акта. При отсутствии циркуляции дальнейшие промывки исключить. Определить '
                      f'приемистость пласта в трубное пространство при давлении не более '
                      f'{self.data_well.max_admissible_pressure.get_value}атм'
                      f'  (перед определением приемистости произвести закачку тех.воды не менее 6м3 или при установившемся '
                      f'давлении закачки, но не более 1 часа). Установить БДТ на гл.{self.data_well.current_bottom}м.',
                      None, None, None, None, None, None, None,
                      'Мастер ГНКТ, состав бригады, представитель Заказчика', 1.33],
                     ]
        if self.data_well.region == 'ТГМ' and acid_edit == 'HF':
            work_list[2] = [None, 20,
                      'без реагирования',
                      None, None, None, None, None, None, None,
                      'Мастер ГНКТ, состав бригады', None]

        opz.extend(work_list)

        return opz

    def check_volume_well(self, data_well):
        if data_well.column_additional:
            well_volume_ek = well_volume(self, data_well.head_column_additional.get_value)
        else:
            well_volume_ek = well_volume(self, data_well.current_bottom)
        if abs(float(data_well.well_volume_in_pz[0]) - well_volume_ek) > 0.2:
            QMessageBox.warning(None, 'Некорректный объем скважины',
                                f'Объем скважины указанный в ПЗ -{data_well.well_volume_in_pz}м3 не совпадает '
                                f'с расчетным {well_volume_ek}м3')
            well_volume_ek, _ = QInputDialog.getDouble(None,
                                                       "корректный объем",
                                                       'Введите корректный объем',
                                                       data_well.well_volume_in_pz[0], 1,
                                                       80, 1)
            well_volume_dp = round(well_volume(self, data_well.current_bottom) - well_volume_ek, 1)
        else:
            well_volume_dp = well_volume(self, data_well.current_bottom) - well_volume_ek
        return well_volume_ek, well_volume_dp

    def schema_well(self, current_bottom_edit, fluid_edit, gnkt_number_combo,
                    gnkt_length, iznos_gnkt_edit, pvo_number, diameter_length, pipe_mileage_edit):
        self.gnkt = self.tab_widget.currentWidget()
        pressure = []
        vertikal = []
        koef_anomal = []

        for ind, plast_ind in enumerate(self.data_well.plast_work):
            if self.data_well.paker_before["before"] != 0:
                if self.data_well.dict_perforation[plast_ind]['подошва'] > \
                        self.data_well.depth_fond_paker_before["before"]:
                        if pressure != 0:
                            pressure.append(max(list(map(
                                float, self.data_well.dict_perforation[plast_ind]["давление"]))))
                            if "вертикаль" in list(self.data_well.dict_perforation[plast_ind].keys()):
                                vertikal.append(min(self.data_well.dict_perforation[plast_ind]["вертикаль"]))


            else:
                pressure.append(max(list(map(
                    float, self.data_well.dict_perforation[plast_ind]["давление"]))))
                vertikal.append(min(list(map(
                    float, self.data_well.dict_perforation[plast_ind]["вертикаль"]))))
        if len(pressure) == 0:
            QMessageBox.warning(self, 'ошибка', 'Приложение не смогло найти рабочие интервалы под пакером. '
                                                'Небходимо уточнить спущенную компоновку')
            return
        if len(pressure) != 0:
            self.pressure = max(pressure)
        else:
            self.pressure = max(self.data_well.dict_perforation[plast_ind]['давление'])

        vertikal = min(vertikal)
        koef_anomal.append(round(float(self.pressure) * 101325 / (float(vertikal) * 9.81 * 1000), 1))


        koef_anomal = max(koef_anomal)

        nkt_str = ''
        length_str = '0-'
        vn_str = ''
        nkt_widht_str = ''
        sorted_dict = dict(sorted(self.data_well.dict_nkt_before.items(), reverse=True))
        len_a = 0
        volume_vn_str = ''
        volume_str = ''
        nkt_length = 0
        for nkt_in, length in sorted_dict.items():
            nkt_length += length
            if '60' in str(nkt_in):
                nkt_in = 60
                nkt_str += f'{nkt_in}\n'
                nkt_widht = 5
                nkt_widht_str += f'{nkt_widht}\n'
                vn_str += f'{nkt_in - 2 * nkt_widht}\n'
                volume_vn = round((nkt_in - 2 * nkt_widht) ** 2 * 3.14 / 4 / 1000, 1)
                volume_vn_str += f'{volume_vn}\n'
                volume = round((volume_vn * length) / 1000, 1)
                volume_str += f'{volume}\n'

            elif '73' in str(nkt_in):
                nkt_in = 73
                nkt_str += f'{nkt_in}, \n'
                nkt_widht = 5.5
                nkt_widht_str += f'{nkt_widht}\n'
                vn_str += f'{nkt_in - 2 * nkt_widht}\n'
                volume_vn = round((nkt_in - 2 * nkt_widht) ** 2 * 3.14 / 4 / 1000, 1)
                volume_vn_str += f'{volume_vn}\n'
                volume = round((volume_vn * length) / 1000, 1)
                volume_str += f'{volume}\n'
            elif '89' in str(nkt_in):
                nkt_in = 89
                nkt_str += f'{nkt_in}\n'
                nkt_widht = 7.34
                nkt_widht_str += f'{nkt_widht}\n'
                vn_str += f'{nkt_in - 2 * nkt_widht}\n'
                volume_vn = round((nkt_in - 2 * nkt_widht) ** 2 * 3.14 / 4 / 1000, 1)
                volume_vn_str += f'{volume_vn}\n'
                volume = round((volume_vn * length) / 1000, 1)
                volume_str += f'{volume}\n'

            length_str += f'{length + len_a}, \n{length}-'
            len_a += length

        nkt_widht_str = nkt_widht_str[:-1]
        length_str = length_str.split('\n')
        length_str = length_str[:-1]
        volume_str = volume_str.split('\n')[:-1]
        volume_str = '\n'.join(volume_str)
        volume_vn_str = volume_vn_str.split('\n')[:-1]
        volume_vn_str = '\n'.join(volume_vn_str)
        nkt_str = nkt_str[:-1]
        vn_str = vn_str[:-1]

        length_nkt = '\n'.join(length_str)

        volume_pm_ek = round(
            3.14 * (self.data_well.column_diameter.get_value - 2 * self.data_well.column_wall_thickness.get_value) ** 2 / 4 / 1000, 2)
        volume_pm_dp = round(3.14 * (self.data_well.column_additional_diameter.get_value - 2 *
                                     self.data_well.column_additional_wall_thickness.get_value) ** 2 / 4 / 1000, 2)

        if self.data_well.column_additional:
            column_data_add_diam = self.data_well.column_additional_diameter.get_value
            column_data_add_wall_thickness = self.data_well.column_additional_wall_thickness.get_value

            column_data_add_vn_volume = round(
                self.data_well.column_additional_diameter.get_value - 2 * self.data_well.column_additional_wall_thickness.get_value, 1)
            column_add_head = self.data_well.head_column_additional.get_value
            column_add_shoe = self.data_well.shoe_column_additional.get_value

        else:
            column_data_add_diam = ''
            column_data_add_wall_thickness = ''

            column_data_add_vn_volume = ''
            column_add_head = ''
            column_add_shoe = ''
        if self.data_well.curator == 'ОР':
            expected_title = 'Ожидаемая приемистость скважины'
            expected_oil = f'{self.data_well.expected_pickup}м3/сут'
            water_cut = f'{self.data_well.expected_pressure}атм'
            proc_water = ''
        else:
            expected_title = 'Ожидаемый дебит скважины'
            expected_oil = f'{self.data_well.expected_oil}т/сут'
            water_cut = f'{self.data_well.water_cut}м3/сут'
            proc_water = f'{self.data_well.percent_water}%'

        nkt = list(self.data_well.dict_nkt_before.keys())
        voronka = sum(list(self.data_well.dict_nkt_before.values()))
        if self.data_well.curator == 'ОР' and self.data_well.region == 'ТГМ' and \
                self.data_well.depth_fond_paker_second_before["before"] != 0:
            length_paker = round(float(self.data_well.depth_fond_paker_second_before["before"]) - float(
                    self.data_well.depth_fond_paker_before["before"]), 1)
            voronka = round(nkt_length + length_paker, 1)
        voronka_str = ''
        nkt_schema = ''
        voronka_depth_str = ''
        if len(nkt) != 0:
            voronka_str = 'Воронка'
            voronka_depth_str = f'на гл.{voronka}м'
            nkt = nkt[0]
            nkt_schema = f'НКТ {nkt_str}мм'
        wellhead_fittings = self.data_well.wellhead_fittings
        if self.data_well.work_plan == 'gnkt_after_grp':

            if 'грп' in str(self.data_well.wellhead_fittings).lower():
                wellhead_fittings = self.data_well.wellhead_fittings
            else:
                wellhead_fittings = f'АУШГН-{self.data_well.column_diameter.get_value}/' \
                                    f'АУГРП {self.data_well.column_diameter.get_value}*14'

        elif self.data_well.work_plan == 'gnkt_bopz':

            list_gnkt_bopz = [
                None, None, None, None, None, None, None, None, None, None, None, None,
                None, None, None, None,
                f'{plast_ind}\n{self.data_well.dict_perforation[plast_ind]["кровля"]}-'
                f'{self.data_well.dict_perforation[plast_ind]["подошва"]}',
                None,
                None, None, None, f'Тек. забой: \n{self.data_well.current_bottom}м ', None]
        length_paker = 2.6
        if self.data_well.curator == 'ОР' and self.data_well.region == 'ТГМ' and \
                self.data_well.depth_fond_paker_second_before["before"] != 0:
            length_paker = self.data_well.depth_fond_paker_second_before["before"] - self.data_well.depth_fond_paker_before["before"]
            if length_paker > 4:
                QMessageBox.warning(self, 'Длина пакера', 'Длина пакера больше 4м')
                length_paker = 2.6

        schema_well_list = [
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'СХЕМА СКВАЖИНЫ', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Данные о размерности труб', None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тип ПВО', None, None,
             f'4-х секционный превентор БП 80-70.00.00.000 (700атм) К2 № {pvo_number}', None, None, None, None, None,
             None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тип ФА', None, None,
             wellhead_fittings, None, None, None,
             None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тип КГ', None, None,
             self.data_well.column_head_m, None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None,
             f'ЭК {self.data_well.column_diameter.get_value}мм', None,
             None,
             'Стол ротора', None, f'{self.data_well.stol_rotor.get_value}м',
             'Øнаруж мм', 'толщ, мм', 'Øвнут, мм', 'Интервал спуска, м', None, 'ВПЦ.\nДлина', 'Объем', None],
            [None, None, None, None, None, None, None, None, None, f'0-{self.data_well.shoe_column.get_value}м',
             None, None,
             'Ø канавки', None, f'{self.data_well.groove_diameter}', None, None,
             None, None, None, None, 'л/п.м.', 'м3'],

            [None, None, None, None, None, None, None, None, None, nkt_schema, None, None, 'Направление', None,
             None,
             f'{self.data_well.column_direction_diameter.get_value}',
             self.data_well.column_direction_wall_thickness.get_value,
             round(self.data_well.column_direction_diameter.get_value - 2 * self.data_well.column_direction_wall_thickness.get_value, 1),
             f'0-', self.data_well.column_direction_length.get_value,
             f'{self.data_well.level_cement_direction.get_value}-{self.data_well.column_direction_length.get_value}',
             None,
             None],
            [None, None, None, None, None, None, None, None, None, f'{length_nkt}м', None, None, 'Кондуктор',
             None, None, self.data_well.column_conductor_diameter.get_value,
             self.data_well.column_conductor_wall_thickness.get_value,
             f'{round(self.data_well.column_conductor_diameter.get_value - 2 * self.data_well.column_conductor_wall_thickness.get_value)}',
             f'0-', self.data_well.column_conductor_length.get_value,
             f'{self.data_well.level_cement_conductor.get_value}-{self.data_well.column_conductor_length.get_value}',
             None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Экспл. колонна', None, None,
             f'{self.data_well.column_diameter.get_value}',
             f'{self.data_well.column_wall_thickness.get_value}',
             f'{round(float(self.data_well.column_diameter.get_value - 2 * self.data_well.column_wall_thickness.get_value), 1)}',
             f'0-', self.data_well.shoe_column.get_value, f'{self.data_well.level_cement_column.get_value}-{self.data_well.shoe_column.get_value}',
             volume_pm_ek,
             self.well_volume_ek],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, "", ""],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, "", ""],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, "", ""],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тех.колонна', None, None,
             column_data_add_diam,
             column_data_add_wall_thickness, column_data_add_vn_volume, column_add_head, column_add_shoe, None,
             volume_pm_dp, self.well_volume_dp],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'колонна НКТ', None, None, nkt_str,
             nkt_widht_str, vn_str, length_nkt, None, None, volume_vn_str, volume_str],
            [None, None, None, None, None, None, None, None, None, None, None, None,
             f'{self.data_well.paker_before["before"]}',
             None, None, None, None,
             50, self.data_well.depth_fond_paker_before["before"],
             round(self.data_well.depth_fond_paker_before["before"] + length_paker, 1),
             length_paker, None, None],
            [None, None, None, None, None, None, None, None, None, 'пакер', None, None, 'без патрубка', None, None,
             None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None,
             f'на гл {self.data_well.depth_fond_paker_before["before"]}м',
             None, None,
             voronka_str, None, None, nkt, None,
             None, voronka, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Данные о перфорации', None, None,
             None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, voronka_str, None, None, 'Пласт\nгоризонт', None,
             'Глубина пласта по вертикали', None, 'Интервал перфорации', None, None, None, 'вскрытия/\nотключения',
             'Рпл. атм', None],
            [None, None, None, None, None, None, None, None, None, voronka_depth_str, None, None, None, None, None,
             None, 'от', None, 'до', None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тек. забой по ПЗ ', None, None,
             None, None, None, None, None, None, self.data_well.bottom, None],
            [None, None, None, None, None, None, None, None, None, None, None, None,
             'необходимый текущий забой ', None, None, None, None, None, None, None, None, current_bottom_edit, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Искусственный забой  ', None,
             None, None, None, None, None, None, None, self.data_well.bottom_hole_artificial.get_value, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Дополнительная информация', None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Категория скважины', None, None,
             None, None, self.data_well.category_pressure, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Содержание H2S, мг/л', None, None,
             None,
             None, 'отсут' if self.data_well.value_h2s_mg[0] in [None, 0] else self.data_well.value_h2s_mg[0], None,
             None, None, None, None],
            [None, None, None, None, None, None, None, None,
             None, None, None, None, 'Газовый фактор', None, None, None,
             None, self.data_well.gaz_factor_percent[0], None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Коэффициент аномальности', None,
             None, None, None, koef_anomal, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Расчетная плотность жидкости глушения',
             None, None, None, None, fluid_edit, None, 'в объеме', None, self.well_volume_ek, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Плотность рабочей жидкости',
             None, None, None, None, self.fluid_work_edit, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, expected_title, None,
             None, None, None, expected_oil, None, water_cut, None, proc_water, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Максимальный угол наклона', None,
             None, None, None, self.data_well.max_angle.get_value, None, 'на глубине', None,
             self.data_well.max_angle_depth.get_value,
             None],
            [None, None, None, None, None, None, None, None, None, None, None, None,
             'Интервалы темпа набора кривизны более 1,5°  на 10 м', None,
             None, None, None, self.data_well.interval_temp.get_value, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Дата начало / окончания бурения',
             None, None, None, None, self.date_dmy(self.data_well.date_drilling_run), None,
             self.date_dmy(self.data_well.date_drilling_cancel),
             None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Дата ввода в эксплуатацию', None,
             None, None, None, f'{self.data_well.date_commissioning.get_value}', None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Р в межколонном пространстве',
             None, None, None, None, self.data_well.pressure_mkp.get_value, None, ' ', None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Первоначальное Р опр-ки ЭК', None,
             None, None, None, self.data_well.first_pressure.get_value, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Результат предыдущей опрес-и ЭК',
             None, None, None, None, self.data_well.result_pressure.get_value, None, '', None, 'гермет.', None],
            [None, None, None, None, None, None, None, None,
             'Тек.забой' if self.data_well.work_plan != 'gnkt_bopz' else '',
             None, None, None,
             'Макс.допустимое Р опр-ки ЭК', None, None, None, None,
             self.data_well.max_admissible_pressure.get_value,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None,
             current_bottom_edit if self.data_well.work_plan != 'gnkt_bopz' else '', None, None, None,
             'Макс. ожидаемое Р на устье ',
             None, None, None, None, self.data_well.max_expected_pressure.get_value, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, f'флот {gnkt_number_combo}',
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, 'длина трубы', 'Øнаруж мм', 'толщ, мм',
             'Øвнут, мм', 'Объем', None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 'л/п.м.', None,
             'м3', None, '%', None, 'м', None, None],
            [None, None, None, None, None, None, None, None, None, None, None, gnkt_length,
             diameter_length, 3.68, '=M69-2*N69',
             '=ROUND(O69*O69*3.14/4/1000,2)', None, '=ROUND(L69*P69/1000, 1)', None, iznos_gnkt_edit, None,
             pipe_mileage_edit, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, 'объем наземной линии, м3', None,
             'объем до границы перфорации, м3', None, 'Объем продавки, м3 ', None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, 1, None, '=L69*P69/1000', None,
             '=L72+N72', None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, 'Вид и категория ремонта, его шифр',
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None,
             f'{self.data_well.type_kr}', None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None,
             None, f'{self.data_well.bur_rastvor}', None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
        ]
        if self.data_well.column_direction_mine_true:
            schema_well_list.insert(10, [None, None, None, None, None, None, None, None, None, None, None, None, 'Шахтное направление', None,
             None,
             f'{self.data_well.column_direction_mine_diameter.get_value}',
             self.data_well.column_direction_mine_wall_thickness.get_value,
             round(
                 float(self.data_well.column_direction_mine_diameter.get_value) - 2 * float(self.data_well.column_direction_mine_wall_thickness.get_value),
                 1),
             f'0-', self.data_well.column_direction_mine_length.get_value,
             f'{self.data_well.level_cement_direction_mine.get_value}-{self.data_well.column_direction_mine_length.get_value}',
             None,
             None])
        else:
            schema_well_list.insert(10, [None, None, None, None, None, None, None, None, None, None, None, None,
                                        'Шахтное направление', None, None, '-', '-', '-', None, None, '-', '-', None])

        if self.data_well.work_plan == 'gnkt_bopz':
            schema_well_list.append(list_gnkt_bopz)
        if self.data_well.paker_before["before"] == 0:
            schema_well_list[21] = [None, None, None, None, None, None, None, None, None, None,
                                    None, None,
                                    'воронка', None, None, nkt, None,
                                    None, self.data_well.depth_fond_paker_before["before"], None, None, None, None]
            schema_well_list[20] = [None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                                    None,
                                    None, None, None, None, None, None, None, None]
            schema_well_list[19] = [None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                                    None,
                                    None, None, None, None, None, None, None, None]

        try:
            pvr_list = []
            self.data_well.img_pvr_list = []
            for plast in sorted(self.data_well.plast_all, key=lambda x: self.get_start_depth(
                    self.data_well.dict_perforation[x]['интервал'][0])):
                count_interval = 0
                for index, interval in enumerate(self.data_well.dict_perforation[plast]['интервал']):
                    count_interval += 1
                    if self.data_well.dict_perforation[plast]['отключение']:
                        izol = 'Изолирован'
                    else:
                        self.data_well.img_pvr_list = \
                            [(plast, self.data_well.dict_perforation[plast]['интервал'])]
                        izol = 'рабочий'
                    if self.data_well.paker_before["before"] != 0:
                        if self.data_well.dict_perforation[plast]['кровля'] < \
                                self.data_well.depth_fond_paker_before["before"]:
                            izol = 'над пакером'
                    vertikal_1 = None
                    if 'вертикаль' in self.data_well.dict_perforation[plast].keys():
                        if len(self.data_well.dict_perforation[plast]['вертикаль']) > 2:
                            vertikal_1 = sorted(self.data_well.dict_perforation[plast]['вертикаль'])[count_interval-1]
                        else:
                            vertikal_1 = sorted(self.data_well.dict_perforation[plast]['вертикаль'])[0]
                    pressure_1 = ''
                    if 'давление' in list(self.data_well.dict_perforation[plast].keys()):
                        pressure_1 = list(filter(lambda  x: x != 0, self.data_well.dict_perforation[plast]['давление']))
                        if len(pressure_1) != 0:
                            pressure_1 = pressure_1[0]
                    zamer_1 = ''
                    if 'замер' in list(self.data_well.dict_perforation[plast].keys()):
                        asdawda = self.data_well.dict_perforation[plast]['замер']
                        if len(self.data_well.dict_perforation[plast]['замер']) != 0:
                            zamer_1 = list(filter(lambda x: x != 0, self.data_well.dict_perforation[plast]['замер']))
                            if len(zamer_1) != 0:
                                zamer_1 = zamer_1[0]

                    pvr_list.append(
                        [None, None, None, None, None, None, None, None, None, None, None, None, plast, None, vertikal_1,
                         None, interval[0],
                         None, interval[1], None, izol, pressure_1,
                         zamer_1, None])
                self.data_well.dict_perforation[plast]['счет_объединение'] = count_interval

            for index, pvr in enumerate(pvr_list):
                schema_well_list[26 + index] = pvr

            self.data_well.current_bottom = round(float(current_bottom_edit), 1)
        except Exception as e:
            QMessageBox.warning(self, 'ошибка', f'Ошибка обработки интервала перфорации, необходимо изменить '
                                                f'изначальные данные {e}')
            return
        return schema_well_list

    def create_title_list(self, ws2):

        # print(f'цднг {self.data_well.cdng.get_value}')
        self.data_well.region = block_name.region_select(self.data_well.cdng.get_value)

        self.region = self.data_well.region

        title_list = [
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'ЗАКАЗЧИК:', None, None, None, None, None, None, None, None, None, None],
            [None, 'ООО «Башнефть-Добыча»', None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'ПЛАН РАБОТ НА СКВАЖИНЕ С ПОМОЩЬЮ УСТАНОВКИ С ГИБКОЙ ТРУБОЙ', None, None, None,
             None, None, None, None, None, None, None],

            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, '№ скважины:', f'{self.data_well.well_number.get_value}', 'куст:', None,
             'Месторождение:', None, None,
             self.data_well.well_oilfield.get_value, None, None],
            [None, None, 'инв. №:', self.data_well.inventory_number.get_value, None, None, None, None, 'Площадь: ',
             self.data_well.well_area.get_value,
             None,
             1],
            [None, None, None, None, None, None, None, 'цех:', f'{self.data_well.cdng.get_value}',
             None, None, None]]

        razdel = CreatePZ.work_podpisant_list(self, self.data_well.region, data_list.contractor)

        for row in razdel:  # Добавлением работ
            title_list.append(row)

        for row in [
            [None, None, None, None, None, "дата", datetime.now().strftime('%d.%m.%Y'), None, None, None, None]]:
            title_list.insert(-1, row)
        # print(title_list)
        index_insert = 11
        ws2.cell(row=1, column=2).alignment = Alignment(wrap_text=False, horizontal='left',
                                                        vertical='center')
        ws2.cell(row=1, column=4).alignment = Alignment(wrap_text=False, horizontal='left',
                                                        vertical='center')
        # ws2.column_dimensions[get_column_letter(1)].width = 15
        # ws2.column_dimensions[get_column_letter(2)].width = 20
        # ws2.column_dimensions[get_column_letter(3)].width = 20
        a = None
        for row in range(len(title_list)):  # Добавлением работ
            if row not in range(8, 13):
                ws2.row_dimensions[row].height = 35
            for col in range(1, 12):
                ws2.column_dimensions[get_column_letter(col)].width = 15
                cell = ws2.cell(row=row + index_insert, column=col)
                # print(f' Х {title_list[i ][col - 1]}')
                if title_list[row - 1][col - 1] is not None:
                    ws2.cell(row=row + index_insert, column=col).value = str(title_list[row - 1][col - 1])
                ws2.cell(row=row + index_insert, column=col).font = Font(name='Arial', size=11, bold=False)
                ws2.cell(row=row + index_insert, column=col).alignment = Alignment(wrap_text=False, horizontal='left',
                                                                                   vertical='center')
                if 'ПЛАН РАБОТ' in str(title_list[row - 1][col - 1]):
                    ws2.merge_cells(start_row=row + index_insert, start_column=2, end_row=row + index_insert,
                                    end_column=11)
                    ws2.merge_cells(start_row=row - 4 + index_insert, start_column=2, end_row=row - 4 + index_insert,
                                    end_column=4)
                    ws2.cell(row=row + index_insert, column=col).font = Font(name='Arial', size=14, bold=True)
                    ws2.cell(row=row + index_insert, column=col).alignment = Alignment(wrap_text=False,
                                                                                       horizontal='center',
                                                                                       vertical='center')
                if 'СОГЛАСОВАНО:' in str(title_list[row - 1][col - 1]):
                    a = row + index_insert

            # for row in range(len(title_list)):  # Добавлением работ
            if a:
                if a > row:
                    # print(f'сссооссоссо {row + index_insert}')
                    ws2.merge_cells(start_row=row + index_insert, start_column=2, end_row=row + index_insert,
                                    end_column=6)
                    ws2.merge_cells(start_row=row + index_insert, start_column=8, end_row=row + index_insert,
                                    end_column=11)

        ws2.print_area = f'B1:K{44}'
        ws2.page_setup.fitToPage = True
        ws2.page_setup.fitToHeight = False
        ws2.page_setup.fitToWidth = True
        ws2.print_options.horizontalCentered = True
        # зададим размер листа
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4

    def date_dmy(self, date_str):

        if '-' in str(date_str):
            return date_str
        elif type(date_str) is datetime:
            date_obj = datetime.strftime('%d.%m.%Y')
        else:
            date_obj = date_str

        return date_obj

    # Функция для получения глубины начала интервала
    def get_start_depth(self, interval):
        return interval[0]

    def calc_fluid(self):

        fluid_list = []
        try:

            fluid_p = 0.83
            for plast in self.data_well.plast_work:
                if float(list(self.data_well.dict_perforation[plast]['рабочая жидкость'])[0]) > fluid_p:
                    fluid_p = list(self.data_well.dict_perforation[plast]['рабочая жидкость'])[0]
            fluid_list.append(fluid_p)

            fluid_work_insert, ok = QInputDialog.getDouble(self, 'Рабочая жидкость',
                                                           'Введите расчетный удельный вес жидкости глушения в '
                                                           'конце жидкости',
                                                           max(fluid_list), 0.87, 2, 2)
        except:
            fluid_work_insert, ok = QInputDialog.getDouble(self, 'Рабочая жидкость',
                                                           'Введите удельный вес рабочей жидкости',
                                                           0, 0.87, 2, 2)
        return fluid_work_insert


class GnktOsvWindow2(GnktModel):
    def __init__(self, parent=None):
        super().__init__(parent)


if __name__ == '__main__':
    app = QApplication([])
    window = GnktOsvWindow2()
    window.show()
    app.exec_()
