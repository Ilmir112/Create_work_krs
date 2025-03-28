import json
import data_list
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import QWidget, QLabel, QComboBox, QLineEdit, QGridLayout, QPushButton
from PyQt5.QtCore import Qt
from data_base.config_base import connection_to_database, WorkDatabaseWell
from decrypt import decrypt
from main import ExcelWorker

from work_py.parent_work import TabPageUnion, WindowUnion, TabWidgetUnion


class TabPageDp(TabPageUnion):
    def __init__(self, data_well, tableWidget, old_index):
        super().__init__(data_well)

        self.tableWidget = tableWidget
        self.old_index = old_index

        self.work_plan = data_well.work_plan

        self.well_number_label = QLabel('номер скважины')
        self.well_number_edit = QLineEdit(self)
        # self.well_number_edit.setValidator(self.validator_int)

        self.well_area_label = QLabel('площадь скважины')
        self.well_area_edit = QLineEdit(self)

        self.table_name = ''

        # self.grid = QGridLayout(self)

        self.grid.addWidget(self.well_number_label, 2, 1)
        self.grid.addWidget(self.well_number_edit, 3, 1)

        self.grid.addWidget(self.well_area_label, 2, 2)
        self.grid.addWidget(self.well_area_edit, 3, 2)

        self.grid.addWidget(self.well_area_label, 2, 3)
        self.grid.addWidget(self.well_area_edit, 3, 3)

        # self.well_area_edit.setText(f'{self.data_well.well_area.value}')
        # self.well_area_edit.textChanged.connect(self.update_well)
        self.well_number_edit.editingFinished.connect(self.update_well)
        # self.change_pvr_combo.currentTextChanged.connect(self.update_change_pvr)
        # self.change_pvr_combo.setCurrentIndex(1)
        # self.change_pvr_combo.setCurrentIndex(0)
        # if self.work_plan not in ['dop_plan_in_base']:
        #     self.well_number_edit.setText(f'{self.data_well.well_number.value}')

        if data_list.data_in_base:
            self.well_data_label = QLabel('файл excel')
            self.well_data_in_base_combo = QComboBox()
            self.well_data_in_base_combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)

            self.grid.addWidget(self.well_data_label, 2, 6)
            self.grid.addWidget(self.well_data_in_base_combo, 3, 6)

            self.well_data_in_base_combo.currentIndexChanged.connect(self.update_area)

    def update_area(self):

        table_in_base_combo = self.well_data_in_base_combo.currentText()
        well_area = table_in_base_combo.split(" ")[1]
        self.well_area_edit.setText(well_area)
        self.well_number_edit.setText(table_in_base_combo.split(" ")[0])

    def update_well(self):
        from work_py.dop_plan_py import TabPageDp

        if data_list.data_in_base:

            well_list = TabPageDp.check_in_database_well_data2(self, self.well_number_edit.text())

            if well_list:
                self.well_data_in_base_combo.clear()
                self.well_data_in_base_combo.addItems(well_list)


class TabWidget(TabWidgetUnion):
    def __init__(self, work_plan, tableWidget=0, old_index=0):
        super().__init__()
        self.addTab(TabPageDp(work_plan, tableWidget, old_index), 'Корректировка плана работ')


class CorrectPlanWindow(WindowUnion):
    def __init__(self, data_well, table_widget):
        super().__init__(data_well)

        self.current_widget = None
        self.data_well.insert_index2 = 0
        self.tab_widget = TabWidget(self.data_well)

        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)

        self.table_widget = table_widget
        self.work_plan = self.data_well.work_plan
        self.dict_perforation = []

        self.data, self.row_heights, self.col_width, self.boundaries_dict = None, None, None, None
        self.target_row_index = None
        self.target_row_index_cancel = None
        self.old_index = 0

        self.buttonadd_work = QPushButton('Загрузить план работ')
        self.buttonadd_work.clicked.connect(self.add_work, Qt.QueuedConnection)

        vbox = QGridLayout(self.centralWidget)
        vbox.addWidget(self.tab_widget, 0, 0, 1, 2)

        vbox.addWidget(self.buttonadd_work, 3, 0, 1, 2)

    def add_work(self):
        from data_base.work_with_base import insert_data_well_dop_plan
        from work_py.dop_plan_py import DopPlanWindow

        from data_list import ProtectedIsNonNone
        self.current_widget = self.tab_widget.currentWidget()

        well_number = self.current_widget.well_number_edit.text()
        well_area = self.current_widget.well_area_edit.text()

        if data_list.data_in_base:
            data_well_data_in_base_combo, data_table_in_base_combo = '', ''
            well_data_in_base_combo = self.current_widget.well_data_in_base_combo.currentText()

            if ' от' in well_data_in_base_combo:
                data_well_data_in_base_combo = well_data_in_base_combo.split(' ')[-1]
                well_data_in_base = well_data_in_base_combo.split(' ')[3]
                if 'ДП' in well_data_in_base:
                    self.data_well.number_dp = ''.join(filter(str.isdigit, well_data_in_base))
                    self.data_well.work_plan_change = 'dop_plan'
                else:
                    self.data_well.work_plan_change = 'krs'

            db = connection_to_database(decrypt("DB_WELL_DATA"))
            data_well_base = WorkDatabaseWell(db, self.data_well)

            data_well = data_well_base.check_in_database_well_data(
                well_number, well_area, well_data_in_base, data_well_data_in_base_combo)

            if data_well:
                self.data_well.type_kr = data_well[2]
                if data_well[3]:

                    self.data_well.well_oilfield = ProtectedIsNonNone(data_well[4])
                    self.data_well.appointment_well = ProtectedIsNonNone(data_well[5])
                    self.data_well.inventory_number = ProtectedIsNonNone(data_well[6])
                    self.data_well.wellhead_fittings = data_well[7]
                    self.data_well.emergency_well = False
                if data_well[8]:
                    self.data_well.angle_data = json.loads(data_well[8])
                else:
                    self.data_well.angle_data = []

                insert_data_well_dop_plan(self, data_well[0])

            self.extraction_data(well_data_in_base_combo)

            DopPlanWindow.work_with_excel(self, well_number, well_area, well_data_in_base, self.data_well.type_kr)

            data_list.data, data_list.row_heights, data_list.col_width, data_list.boundaries_dict = \
                DopPlanWindow.change_pvr_in_bottom(self, self.data, self.row_heights, self.col_width,
                                                   self.boundaries_dict)

            if well_number != '' and well_area != '':
                self.data_well.well_number, self.data_well.well_area = \
                    ProtectedIsNonNone(well_number), ProtectedIsNonNone(well_area)

            self.thread_excel = ExcelWorker(self)

            self.without_damping, stop_app = self.thread_excel.check_well_existence(
                self.data_well.well_number.get_value, self.data_well.well_area.get_value,
                self.data_well.region)

            data_list.pause = False
            if well_number != '' and well_area != '':
                self.close()
                self.close_modal_forcefully()


    def add_work_excel(self, ws2, work_list, ind_ins):
        from data_list import ProtectedIsDigit
        for i in range(1, len(work_list) + 1):  # Добавлением работ
            for j in range(1, 13):
                cell = ws2.cell(row=i, column=j)

                if cell and str(cell) != str(work_list[i - 1][j - 1]):

                    if i >= ind_ins:

                        if j != 1:
                            cell.border = data_list.thin_border
                        if j == 11:
                            cell.font = Font(name='Arial', size=11, bold=False)
                        # if j == 12:
                        #     cell.value = work_list[i - 1][j - 1]
                        else:
                            cell.font = Font(name='Arial', size=13, bold=False)
                        ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                        vertical='center')
                        ws2.cell(row=i, column=11).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')
                        ws2.cell(row=i, column=12).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')
                        ws2.cell(row=i, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                        vertical='center')

                        if 'порядок работы' in str(cell.value).lower() or \
                                'наименование работ' in str(cell.value).lower():
                            self.data_well.insert_index2 = i + 1
                            self.data_well.data_x_max = ProtectedIsDigit(i + 2)

                            ws2.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=True)
                            ws2.cell(row=i, column=j).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                            vertical='center')
    @staticmethod
    def work_list(work_earlier):
        krs_begin = [
            [None, None, f' Ранее проведенные работы: \n {work_earlier}', None, None, None, None, None, None, None,
                      'Мастер КРС', None]]

        return krs_begin
