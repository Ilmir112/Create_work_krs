import json
import os
import re
import sqlite3

import openpyxl
import psycopg2
import data_list

from datetime import datetime

from PyQt5 import QtWidgets

from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem, QLineEdit, QHeaderView, QVBoxLayout, QWidget, \
    QTableWidget

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils import get_column_letter, range_boundaries
from data_base.config_base import connect_to_database, connection_to_database, CheckWellExistence
from decrypt import decrypt

from main import MyMainWindow, ExcelWorker
from server_response import ResponseWork, ApiClient
from work_py.alone_oreration import well_volume
from work_py.progress_bar_save import ProgressBarWindow


class ClassifierWell(MyMainWindow):
    number_well = None

    def __init__(self, costumer, region, parent=None):
        super().__init__()

        self.classification_well = None
        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)
        self.table_class = QTableWidget()
        self.region = region
        self.costumer = costumer
        self.number_well = None
        self.db = connection_to_database(decrypt("DB_NAME_USER"))
        # if self.well_number:
        #     self.number_well = self.data_well.well_number.get_value
        self.api_client = ApiClient
        self.setCentralWidget(self.table_class)
        self.model = self.table_class.model()
        if ClassifierWell == 'ClassifierWell':
            self.open_to_sqlite_class_well(costumer, region)
        elif ClassifierWell == 'damping':
            self.open_to_sqlite_without_juming(costumer, region)

    def get_data_from_db(self, region):
        if data_list.connect_in_base:
            region_json = {"region": region}
            data_response = ApiClient.request_post(ApiClient.read_wells_silencing_response_all(), region_json)
            data = []

            for num in data_response:
                wells_data = []
                for key, value in num.items():
                    if key in ["well_number", "deposit_area", "today"]:
                        wells_data.append(value)
                data.append(wells_data)

        else:
            well_classification = CheckWellExistence(self.db)
            data = well_classification.get_data_from_db(region)

        return data

    def open_to_sqlite_without_juming(self):
        layout = QVBoxLayout()
        self.edit_well_number = QLineEdit()
        self.edit_well_number.setPlaceholderText("Ввести номер скважины для фильтрации")

        self.edit_well_number.textChanged.connect(self.filter)
        self.edit_well_number.setText(self.number_well)
        layout.addWidget(self.edit_well_number)

        data = self.get_data_from_db(self.region)
        if data:
            self.table_class.setColumnCount(len(data[0]))
            self.table_class.setRowCount(len(data))
            self.table_class.setCellWidget(0, 0, self.edit_well_number)
            for row in range(len(data)):
                for col in range(len(data[row])):
                    item = QTableWidgetItem(str(data[row][col]))
                    self.table_class.setItem(row + 1, col, item)

        self.table_class.setHorizontalHeaderLabels(['номер скважины', 'площадь', 'Текущий квартал'])
        self.table_class.horizontalHeader().setStretchLastSection(True)
        self.table_class.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # layout.addWidget(table)
        self.setLayout(layout)

    def closeEvent(self, event):
        if self.sender() is None:  # Проверяем вызывающий объект
            # Закрываем основное окно при закрытии окна входа
            self.new_window = None
            event.accept()  # Принимаем событие закрытия

    def get_data_from_class_well_db(self, region):
        if data_list.connect_in_base:
            region_json = {"region": region[:3]}
            data_response = ApiClient.request_post(ApiClient.read_wells_classifier_response_all(), region_json)
            data = []
            for num in data_response:
                wells_data = [num["cdng"], num["well_number"], num["deposit_area"], num["oilfield"],
                              num['category_pressure'], num['pressure_gst'], num["pressure_ppl"],
                              num['date_measurement'], num['category_h2s'], num['h2s_pr'], num['h2s_mg_l'],
                              num['h2s_mg_m'], num['category_gf'], num['gas_factor'], num["today"]]

                data.append(wells_data)
        else:
            well_classification = CheckWellExistence(self.db)
            data = well_classification.get_data_from_class_well_db(region)
        return data

    def open_to_sqlite_class_well(self):
        layout = QVBoxLayout()
        self.edit_well_number = QLineEdit()
        self.edit_well_number.setPlaceholderText("Ввести номер скважины для фильтрации")

        self.edit_well_number.textChanged.connect(self.filter_class)
        self.edit_well_number.setText(self.number_well)
        layout.addWidget(self.edit_well_number)

        self.edit_well_area = QLineEdit()
        self.edit_well_area.setPlaceholderText("Ввести площадь для фильтрации")
        self.edit_well_area.textChanged.connect(self.filter_class_area)
        layout.addWidget(self.edit_well_area)
        region = f'{self.region}_классификатор'
        # print(region)
        data = self.get_data_from_class_well_db(region)
        if data:
            # print(data)

            self.table_class.setColumnCount(len(data[0]))
            self.table_class.setRowCount(len(data))
            self.table_class.setCellWidget(0, 1, self.edit_well_number)
            self.table_class.setCellWidget(0, 2, self.edit_well_area)
            for row in range(len(data)):
                for col in range(len(data[row])):
                    if col in [5, 6, 10, 11, 13]:
                        if str(data[row][col]).replace('.', '').isdigit() and str(data[row][col]).count('.') < 2:
                            item = QTableWidgetItem(str(round(float(data[row][col]), 1)))
                        else:
                            item = QTableWidgetItem(str(data[row][col]))
                    elif col in [9]:
                        if str(data[row][col]).replace('.', '').isdigit() and str(data[row][col]).count('.') < 2:
                            item = QTableWidgetItem(str(round(float(data[row][col]), 7)))
                        else:
                            item = QTableWidgetItem(str(data[row][col]))

                    else:
                        item = QTableWidgetItem(str(data[row][col]))
                    self.table_class.setItem(row + 1, col, item)

        self.table_class.setHorizontalHeaderLabels(
            ['ЦДНГ', 'номер скважины', 'площадь', 'Месторождение', 'Категория \n по Рпл',
             'Ргд', 'Рпл', 'Дата замера', 'категория \nH2S', 'H2S-%', "H2S-мг/л",
             "H2S-мг/м3", 'Категория по газу', "Газовый фактор", "версия от"])
        self.table_class.horizontalHeader().setStretchLastSection(True)
        self.table_class.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # layout.addWidget(table)
        self.setLayout(layout)

    def insert_database(self, data_base, data_work, query):
        if data_list.connect_in_base:

            # Параметры подключения к PostgreSQL
            try:
                # Создание подключения к базе данных PostgreSQL
                conn = connect_to_database(data_base)
                param = '%s'

            except psycopg2.Error as e:
                QMessageBox.warning(None, 'Ошибка', f'Ошибка подключения к базе данных: {type(e).__name__}\n\n{str(e)}')

        else:
            try:
                db_path = self.connect_to_db('data_list.db', 'data_base_well')
                # Создание подключения к базе данных SQLite
                conn = sqlite3.connect(db_path)
                query.replace('%s', '?')
                param = '?'


            except sqlite3.Error as e:
                QMessageBox.warning(None, 'Ошибка', f'Ошибка подключения к базе данных {type(e).__name__}\n\n{str(e)}')
        # Выполнение SQL-запроса для получения данных

        cursor = conn.cursor()

        cursor.execute(f"""SELECT * FROM chemistry
                           WHERE well_number={param} AND well_area={param} AND region={param}
                            AND costumer={param} AND contractor={param} AND work_plan={param} AND type_kr={param}
                           """, (data_work[:7]))

        row_exists = cursor.fetchone()

        if row_exists:
            query2 = f'''UPDATE chemistry
            SET today={param}, cement={param}, HCl={param}, HF={param}, NaOH={param}, VT_SKO={param},
                clay={param}, sand={param}, RPK={param}, RPP={param}, RKI={param},
                 ELAN={param},ASPO={param}, RIR_2C={param}, RIR_OVP={param}, gidrofabizator={param}, 
                 norm_time={param}, fluid={param}
            WHERE well_number={param} AND well_area={param} AND region={param} AND
                costumer={param} AND contractor={param} AND work_plan={param} AND type_kr={param}'''
            data_work2 = data_work[7:] + data_work[:7]

            cursor.execute(query2, data_work2)

        else:
            cursor.execute(query, data_work)

        # Сохранение изменений
        conn.commit()

        # Закрыть курсор и соединение
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    def export_to_sqlite_without_juming(self, fname):
        # Загрузка файла Excel
        wb = load_workbook(fname)
        ws = wb.active
        check_param = ''
        self.progress_bar_window = ProgressBarWindow(ws.max_row - 1)
        self.progress_bar_window.show()
        # Получение данных из Excel и запись их в базу данных
        for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            for col, value in enumerate(row):
                if not value is None and col <= 18:
                    # print(value)
                    if 'туймазин' in str(value).lower():
                        check_param = 'ТГМ'
                    elif 'ишимбай' in str(value).lower():
                        check_param = 'ИГМ'
                    elif 'чекмагуш' in str(value).lower():
                        check_param = 'ЧГМ'
                    elif 'красно' in str(value).lower():
                        check_param = 'КГМ'
                    elif 'арлан' in str(value).lower():
                        check_param = 'АГМ'
                    if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                            value) or '01.10.' in str(value):

                        version_year = re.findall(r'[0-9.]', str(value))
                        version_year = ''.join(version_year)
                        if version_year[-1] == '.':
                            version_year = version_year[:-1]
            if index_row > 18:
                break

        try:
            # Подключение к базе данных

            # self.classification_well = CheckWellExistence(self.db)

            REGION_LIST = ['ЧГМ', 'АГМ', 'ТГМ', 'ИГМ', 'КГМ', ]

            for region_name in REGION_LIST:
                if region_name == self.region:
                    if check_param not in region_name:
                        QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                            f'регион выбран некорректно  {region_name}')
                    elif check_param in region_name:
                        # self.classification_well.create_table_without_juming(region_name)
                        # path = self.api_client.read_wells_silencing_response_for_delete_well()
                        # responce = self.api_client.delete_wells_by_region(region_name, path)
                        try:
                            params_list = []
                            # Получение данных из Excel и запись их в базу данных
                            for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                if 'ПЕРЕЧЕНЬ' in row:
                                    check_file = True
                                if 'Скважина' in row:
                                    area_row = index_row + 2
                                    for col, value in enumerate(row):
                                        if not value is None and col <= 20:
                                            if 'Скважина' == value:
                                                well_column = col
                                            elif 'Площадь' == value:
                                                area_column = col

                            for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                self.progress_bar_window.start_loading(index_row + 1)
                                if index_row > area_row:
                                    well_number = row[well_column]
                                    area_well = row[area_column]
                                    params = {}
                                    if well_number and len(str(well_number)) <= 5:

                                        params['well_number'] = str(well_number).lstrip().rstrip()
                                        params['deposit_area'] = str(area_well).lstrip().rstrip().replace(" ", "_")
                                        params['costumer'] = str(self.costumer)
                                        params['today'] = str(version_year)
                                        params['region'] = str(region_name)
                                        if data_list.connect_in_base:
                                            params_list.append(params)
                                            # self.classification_well.insert_data_in_table_without_juming(
                                            #     str(well_number), area_well, version_year, region_name, self.costumer)
                                        else:
                                            self.classification_well.insert_data_in_table_without_juming(
                                                str(well_number), area_well, version_year, region_name, self.costumer)
                            if data_list.connect_in_base:
                                self.api_client.add_well_in_database(params_list,
                                    self.api_client.read_wells_silencing_response_for_add_well())

                            QMessageBox.information(self, 'данные обновлены', 'Данные обновлены')
                        except Exception as e:
                            QMessageBox.warning(self, 'ОШИБКА', f'Выбран файл с не корректными данными {e}')

                    else:
                        QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                            f'в Данном перечне отсутствую скважины {region_name}')


        except psycopg2.Error as e:
            # Выведите сообщение об ошибке
            QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')

        except sqlite3.Error as e:
            # Выведите сообщение об ошибке
            QMessageBox.warning(None, 'Ошибка',
                                f'Ошибка подключения к базе данных: /n {type(e).__name__}\n\n{str(e)}')

    def filter(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(0, 1, 2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break

    def filter_class(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(1, 2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break

    def filter_class_area(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break

    def export_to_database_class_well(self, fname):

        REGION_LIST = ['ЧГМ_классификатор', 'АГМ_классификатор', 'ТГМ_классификатор', 'ИГМ_классификатор',
                       'КГМ_классификатор']

        # Загрузка файла Excel
        params_list = []
        wb = load_workbook(fname)
        ws = wb.active
        self.progress_bar_window = ProgressBarWindow(ws.max_row - 1)
        self.progress_bar_window.show()

        try:
            self.classification_well = CheckWellExistence(self.db)

            check_param, well_column, cdng, area_column, oilfield, categoty_pressure, pressure_Gst, \
            date_measurement, pressure_Ppl, categoty_h2s, h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, \
            gas_factor, area_row, check_file = self.classification_well.read_excel_file_classification(ws)

            for region_name in REGION_LIST:
                if self.region in region_name:
                    if check_param not in region_name:
                        QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                            f'регион выбрано некорректно  {region_name}')
                    if check_param in region_name:
                        # if data_list.connect_in_base is False:
                            # self.classification_well.create_table_classification(region_name)
                        # path = self.api_client.read_wells_classifier_response_for_delete_well()
                        # response = self.api_client.delete_wells_by_region(self.region, path)
                        try:
                            # Вставка данных в таблицу
                            for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True, max_col=25)):
                                if index_row < area_row and check_file:
                                    for col, value in enumerate(row):
                                        if not value is None and col <= 18:
                                            if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                                                    value) or '01.10.' in str(value):
                                                version_year = re.findall(r'[0-9.]', str(value))
                                                version_year = ''.join(version_year)
                                                date_obj = datetime.strptime(version_year, "%d.%m.%Y")
                                                formatted_date = date_obj.strftime("%Y-%m-%d")
                                                if version_year[-1] == '.':
                                                    version_year = version_year[:-1]

                                elif index_row > area_row and check_file:
                                    well_number = row[well_column]
                                    area_well = row[area_column]
                                    oilfield_str = row[oilfield]
                                    if well_number and len(str(well_number)) < 10 and row[categoty_pressure] and data_list.connect_in_base:
                                        self.progress_bar_window.start_loading(index_row + 1)
                                        params = {
                                                "cdng": str(row[cdng]),
                                                "well_number": str(well_number).lstrip().rstrip(),
                                                "deposit_area": str(area_well).lstrip().rstrip().replace(" ", "_"),
                                                "oilfield": f"{oilfield_str}",
                                                "category_pressure": f"{row[categoty_pressure]}",
                                                "pressure_ppl": f"{row[pressure_Ppl]}",
                                                "pressure_gst": f"{row[pressure_Gst]}",
                                                "date_measurement": f'{ApiClient.serialize_datetime(row[date_measurement])}',
                                                "category_h2s": f"{row[categoty_h2s]}",
                                                "h2s_pr": f"{row[h2s_pr]}",
                                                "h2s_mg_l": f"{row[h2s_mg_l]}",
                                                "h2s_mg_m": f"{row[h2s_mg_m]}",
                                                "category_gf": f"{row[categoty_gf]}",
                                                "gas_factor": f"{row[gas_factor]}",
                                                "today": formatted_date,
                                                "region": self.region,
                                                "costumer": self.costumer}

                                        params_list.append(params)

                                    # else:
                                    #     self.classification_well.insert_data_in_classification(
                                    #         region_name, row[cdng], well_number, area_well, oilfield_str,
                                    #         row[categoty_pressure],
                                    #         row[pressure_Ppl], row[pressure_Gst], row[date_measurement],
                                    #         row[categoty_h2s],
                                    #         row[h2s_pr], row[h2s_mg_l], row[h2s_mg_m], row[categoty_gf],
                                    #         row[gas_factor],
                                    #         version_year, self.region, self.costumer
                                    #     )
                                    if well_number is None and area_well is None and oilfield is None and categoty_pressure is None:
                                        break
                            self.progress_bar_window.start_loading(ws.max_row - 1)
                            self.api_client.add_well_in_database(
                                params_list,
                                self.api_client.read_wells_classifier_response_for_add_well())

                        except Exception as e:
                            QMessageBox.warning(self, 'ОШИБКА', f'Выбран файл с не корректными данными {e}')



                    else:
                        QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                            f'в Данном перечне отсутствую скважины {region_name}')

            QMessageBox.information(self, 'Успешно', 'Классификатор успешно обновлен')

        except (psycopg2.Error, Exception) as e:
            # Выведите сообщение об ошибке
            QMessageBox.warning(self, 'Ошибка', f'Ошибка подключения к базе данных {e}')

    @staticmethod
    def connect_to_db(name_base, folder_base):
        # Получаем текущий каталог приложения
        current_dir = os.path.dirname(__file__)

        # Определяем путь к папке с базой данных
        db_folder = os.path.join(current_dir, folder_base)

        # Формируем полный путь к файлу базы данных
        db_path = os.path.join(db_folder, name_base)

        return db_path


def excel_in_json(self, sheet):
    index_end_copy = 46

    data = {}
    for row_index, row in enumerate(sheet.iter_rows()):

        if all(cell is None for cell in row[:32]) is False:
            if any([cell.value == "ИТОГО:" for cell in row[:4]]):
                index_end_copy = row_index

                break
            for cell in row[:32]:
                if any([cell.value == "ИТОГО:" for cell in row[:4]]):
                    index_end_copy = row_index

                    break
                row_data = []
                if all(cell is None for cell in row[:32]) is False:
                    if any([cell.value == "ИТОГО:" for cell in row[:4]]):
                        index_end_copy = row_index
                        break
                    for cell in row[:32]:
                        # Получение значения и стилей
                        value = cell.value

                        font = cell.font
                        if font.color:
                            # Преобразуем RGB в строковый формат
                            rgb_string = f"RGB({font.color.rgb})"
                        else:
                            rgb_string = None

                        fill = cell.fill
                        # Преобразуем RGB в строковый формат
                        rgb_string_fill = f"RGB({fill.fgColor.rgb})"

                        borders = cell.border

                        borders_style_left = None
                        borders_style_right = None
                        borders_style_top = None
                        borders_style_bottom = None
                        if borders.left.style:
                            borders_style_left = borders.left.style
                        if borders.left.color:
                            left_border = f"RGB({borders.left.color.rgb})"
                            if left_border == "RGB(Values must be of type <class 'str'>)":
                                left_border = None
                        else:
                            left_border = None
                        if borders.right.style:
                            borders_style_right = borders.right.style
                        if borders.right.color:
                            right_border = f"RGB({borders.right.color.rgb})"
                            if right_border == "RGB(Values must be of type <class 'str'>)":
                                right_border = None
                        else:
                            right_border = None
                        if borders.top.style:
                            borders_style_top = borders.top.style
                        if borders.top.color:
                            top_border = f"RGB({borders.top.color.rgb})"
                            if top_border == "RGB(Values must be of type <class 'str'>)":
                                top_border = None
                        else:
                            top_border = None

                        if borders.bottom.style:
                            borders_style_bottom = borders.bottom.style
                        if borders.bottom.color:
                            bottom_border = f"RGB({borders.bottom.color.rgb})"
                            if bottom_border == "RGB(Values must be of type <class 'str'>)":
                                bottom_border = None
                        else:
                            bottom_border = None
                        alignment = cell.alignment

                        row_data.append({
                            'value': value,
                            'font': {
                                'name': font.name,
                                'size': font.size,
                                'bold': font.bold,
                                'italic': font.italic,
                                'color': rgb_string
                            },
                            'fill': {
                                'color': rgb_string_fill
                            },
                            'borders': {
                                'left': {
                                    'style': borders_style_left,
                                    'color': left_border
                                },
                                'right': {
                                    'style': borders_style_right,
                                    'color': right_border
                                },
                                'top': {
                                    'style': borders_style_top,
                                    'color': top_border
                                },
                                'bottom': {
                                    'style': borders_style_bottom,
                                    'color': bottom_border
                                },
                            },
                            'alignment': {
                                'horizontal': alignment.horizontal,
                                'vertical': alignment.vertical,
                                'wrap_text': alignment.wrap_text
                            },
                        })
                        data[row[0].row] = row_data

    data['image'] = self.data_well.image_data
    row_heights = [sheet.row_dimensions[i + 1].height for i in range(sheet.max_row) if i <= index_end_copy]
    # row_heights = [sheet.row_dimensions[i + 1].height for i in range(sheet.max_row)if i <= 46]
    col_width = [sheet.column_dimensions[get_column_letter(i + 1)].width for i in range(0, 80)] + [None]
    boundaries_dict = {}

    for ind, _range in enumerate(sheet.merged_cells.ranges):
        if range_boundaries(str(_range))[1] <= index_end_copy:
            boundaries_dict[ind] = range_boundaries(str(_range))

    data_excel = {'data': data, 'rowHeights': row_heights, 'col_width': col_width, 'merged_cells': boundaries_dict}

    return data_excel


def insert_data_well_dop_plan(self, data_well):
    from data_list import ProtectedIsDigit, ProtectedIsNonNone

    well_data_dict = json.loads(data_well)
    self.data_well.column_direction_true = False
    self.data_well.column_direction_diameter = ProtectedIsDigit(well_data_dict["направление"]["диаметр"])
    self.data_well.column_direction_wall_thickness = ProtectedIsDigit(well_data_dict["направление"]["толщина стенки"])
    self.data_well.column_direction_length = ProtectedIsDigit(well_data_dict["направление"]["башмак"])
    self.data_well.level_cement_direction = ProtectedIsDigit(well_data_dict["направление"]["цемент"])
    self.data_well.column_conductor_diameter = ProtectedIsDigit(well_data_dict["кондуктор"]["диаметр"])
    self.data_well.column_conductor_wall_thickness = ProtectedIsDigit(well_data_dict["кондуктор"]["толщина стенки"])
    if self.data_well.column_conductor_diameter.get_value != '0':
        self.data_well.column_direction_true = True

    self.data_well.level_cement_column = ProtectedIsDigit(well_data_dict["ЭК"]["цемент"])
    if '-' in str(self.data_well.level_cement_column.get_value):
        self.data_well.level_cement_column = ProtectedIsDigit(
            self.data_well.level_cement_column.get_value.split('-')[1])

    self.data_well.column_conductor_length = ProtectedIsDigit(well_data_dict["кондуктор"]["башмак"])
    self.data_well.level_cement_conductor = ProtectedIsDigit(well_data_dict["кондуктор"]["цемент"])
    self.data_well.column_diameter = ProtectedIsDigit(well_data_dict["ЭК"]["диаметр"])
    self.data_well.column_wall_thickness = ProtectedIsDigit(well_data_dict["ЭК"]["толщина стенки"])
    self.data_well.shoe_column = ProtectedIsDigit(well_data_dict["ЭК"]["башмак"])
    self.data_well.column_additional = well_data_dict["допколонна"]["наличие"]
    self.data_well.column_additional_diameter = ProtectedIsDigit(well_data_dict["допколонна"]["диаметр"])
    self.data_well.column_additional_wall_thickness = ProtectedIsDigit(well_data_dict["допколонна"]["толщина стенки"])
    self.data_well.shoe_column_additional = ProtectedIsDigit(well_data_dict["допколонна"]["башмак"])
    self.data_well.head_column_additional = ProtectedIsDigit(well_data_dict["допколонна"]["голова"])
    self.data_well.curator = well_data_dict["куратор"]
    self.data_well.dict_pump_shgn = well_data_dict["оборудование"]["ШГН"]["тип"]
    if 'do' in list(self.data_well.dict_pump_shgn.keys()):
        self.data_well.dict_pump_shgn['before'] = self.data_well.dict_pump_shgn['do']
        self.data_well.dict_pump_shgn['after'] = self.data_well.dict_pump_shgn['posle']
    self.data_well.dict_pump_shgn_depth = well_data_dict["оборудование"]["ШГН"]["глубина "]
    if 'do' in list(self.data_well.dict_pump_shgn_depth.keys()):
        self.data_well.dict_pump_shgn_depth['before'] = self.data_well.dict_pump_shgn_depth['do']
        self.data_well.dict_pump_shgn_depth['after'] = self.data_well.dict_pump_shgn_depth['posle']
    self.data_well.dict_pump_ecn = well_data_dict["оборудование"]["ЭЦН"]["тип"]
    if 'do' in list(self.data_well.dict_pump_ecn.keys()):
        self.data_well.dict_pump_ecn['before'] = self.data_well.dict_pump_ecn['do']
        self.data_well.dict_pump_ecn['after'] = self.data_well.dict_pump_ecn['posle']
    self.data_well.dict_pump_ecn_depth = well_data_dict["оборудование"]["ЭЦН"]["глубина "]
    if 'do' in list(self.data_well.dict_pump_ecn_depth.keys()):
        self.data_well.dict_pump_ecn_depth['before'] = self.data_well.dict_pump_ecn_depth['do']
        self.data_well.dict_pump_ecn_depth['after'] = self.data_well.dict_pump_ecn_depth['posle']
    self.data_well.paker_before = well_data_dict["оборудование"]["пакер"]["тип"]
    if 'do' in list(self.data_well.paker_before.keys()):
        self.data_well.paker_before['before'] = self.data_well.paker_before['do']
        self.data_well.paker_before['after'] = self.data_well.paker_before['posle']
    self.data_well.depth_fond_paker_before = well_data_dict["оборудование"]["пакер"]["глубина "]
    if 'do' in list(self.data_well.depth_fond_paker_before.keys()):
        self.data_well.depth_fond_paker_before['before'] = self.data_well.depth_fond_paker_before['do']
        self.data_well.depth_fond_paker_before['after'] = self.data_well.depth_fond_paker_before['posle']

    self.data_well.paker_second_before = well_data_dict["оборудование"]["пакер2"]["тип"]
    if 'do' in list(self.data_well.paker_second_before.keys()):
        self.data_well.paker_second_before['before'] = self.data_well.paker_second_before['do']
        self.data_well.paker_second_before['after'] = self.data_well.paker_second_before['posle']
    self.data_well.depth_fond_paker_second_before = well_data_dict["оборудование"]["пакер2"]["глубина "]
    if 'do' in list(self.data_well.depth_fond_paker_second_before.keys()):
        self.data_well.depth_fond_paker_second_before['before'] = self.data_well.depth_fond_paker_second_before['do']
        self.data_well.depth_fond_paker_second_before['after'] = self.data_well.depth_fond_paker_second_before['posle']
    self.data_well.static_level = ProtectedIsDigit(well_data_dict["статика"])
    self.data_well.dinamic_level = ProtectedIsDigit(well_data_dict["динамика"])
    if 'После' in list(well_data_dict["НКТ"].keys()):
        self.data_well.dict_nkt_after = well_data_dict["НКТ"]['После']
        self.data_well.dict_nkt_before = well_data_dict["НКТ"]["До"]
        self.data_well.dict_sucker_rod_after = well_data_dict["штанги"]["После"]
        self.data_well.dict_sucker_rod = well_data_dict["штанги"]["До"]
    else:
        self.data_well.dict_nkt_after = well_data_dict["НКТ"]
        self.data_well.dict_nkt_before = well_data_dict["НКТ"]
        self.data_well.dict_sucker_rod_after = well_data_dict["штанги"]
        self.data_well.dict_sucker_rod = well_data_dict["штанги"]
    self.data_well.expected_oil = well_data_dict['ожидаемые']['нефть']
    self.data_well.water_cut = well_data_dict['ожидаемые']['вода']
    self.data_well.percent_water = well_data_dict['ожидаемые']['обводненность']
    self.data_well.expected_pressure = well_data_dict['ожидаемые']['давление']
    self.data_well.expected_pickup = well_data_dict['ожидаемые']['приемистость']

    self.data_well.bottom_hole_drill = ProtectedIsDigit(well_data_dict['данные']['пробуренный забой'])
    self.data_well.bottom_hole_artificial = ProtectedIsDigit(well_data_dict['данные']['искусственный забой'])
    self.data_well.max_angle = ProtectedIsDigit(well_data_dict['данные']['максимальный угол'])
    self.data_well.max_angle_depth = ProtectedIsDigit(well_data_dict['данные']['глубина'])
    self.data_well.max_expected_pressure = ProtectedIsDigit(well_data_dict['данные']['максимальное ожидаемое давление'])
    self.data_well.max_admissible_pressure = ProtectedIsDigit(
        well_data_dict['данные']['максимальное допустимое давление'])

    self.data_well.curator = well_data_dict['куратор']
    self.data_well.region = well_data_dict['регион']
    self.data_well.cdng = ProtectedIsNonNone(well_data_dict['ЦДНГ'])
    if 'голова ЭК' in list(well_data_dict['ЭК'].keys()):
        self.data_well.head_column = ProtectedIsDigit(well_data_dict['ЭК']['голова ЭК'])
        self.data_well.diameter_doloto_ek = ProtectedIsDigit(well_data_dict['данные']['диаметр долото при бурении'])
    else:
        self.data_well.head_column = ProtectedIsDigit(0)
        self.data_well.diameter_doloto_ek = ProtectedIsDigit(0)
        print(f'отсутствуют данные в базе по голове хвостовика и диаметру долото')
    if 'дата ввода в эксплуатацию' in list(well_data_dict.keys()):
        self.data_well.date_commissioning = data_list.ProtectedIsNonNone(well_data_dict['дата ввода в эксплуатацию'])

    else:
        self.data_well.date_commissioning = data_list.ProtectedIsNonNone('01.01.2000')

    if 'дата опрессовки' in list(well_data_dict.keys()):
        self.data_well.result_pressure_date = data_list.ProtectedIsNonNone(well_data_dict['дата опрессовки'])
    else:
        self.data_well.result_pressure_date = data_list.ProtectedIsNonNone('01.01.2000')

    if 'ПВР план' in list(well_data_dict.keys()):
        self.data_well.dict_perforation_project = well_data_dict['ПВР план']
    else:
        self.data_well.dict_perforation_project = ''

    self.data_well.data_well_dict = well_data_dict

    self.data_well.well_volume_in_pz = [well_volume(self, self.data_well.head_column_additional.get_value)]
    self.data_well.check_data_in_pz = []
    self.data_well.without_damping = False
    self.thread_excel = ExcelWorker(self)

    # QMessageBox.information(None, 'Данные с базы', "Данные вставлены из базы данных")

    # definition_plast_work(self)


def round_cell(data):
    try:
        if data is None or type(data) is datetime:
            return data
        float_item = float(data)
        if float_item.is_integer():
            return int(float_item)
        else:
            return round(float_item, 4)
    except ValueError:
        return data


def insert_data_new_excel_file(self, data, row_heights, col_width, boundaries_dict):
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active

    for row_index, row_data in data.items():
        if row_index != 'image':
            for col_index, cell_data in enumerate(row_data, 1):
                cell = sheet_new.cell(row=int(row_index), column=int(col_index))
                if cell_data:
                    cell.value = round_cell(cell_data['value'])

    row_max = sheet_new.max_row

    for key, value in boundaries_dict.items():
        if value[1] <= row_max:
            sheet_new.merge_cells(start_column=value[0], start_row=value[1],
                                  end_column=value[2], end_row=value[3])

    # Восстановление данных и стилей из словаря
    for row_index, row_data in data.items():
        if row_index != 'image':
            for col_index, cell_data in enumerate(row_data, 1):
                cell = sheet_new.cell(row=int(row_index), column=int(col_index))

                try:
                    # Получение строки RGB из JSON
                    rgb_string = cell_data['font']['color']
                    hex_color = cell_data['fill']['color'][4:-1]
                    cell.font = Font(name=cell_data['font']['name'], size=cell_data['font']['size'],
                                     bold=cell_data['font']['bold'], italic=cell_data['font']['italic'])
                    if 'color' in list(cell_data['font'].keys()):
                        if hex_color != '00000000':
                            color = Color(rgb=hex_color)
                            cell.fill = PatternFill(patternType='solid', fgColor=color)

                        color_font = change_rgb_to_hex(rgb_string)

                        cell.font = Font(name=cell_data['font']['name'], size=cell_data['font']['size'],
                                         bold=cell_data['font']['bold'], italic=cell_data['font']['italic'],
                                         color=color_font)
                    else:
                        # Извлекаем шестнадцатеричный код цвета
                        hex_color = rgb_string[4:-1]

                        cell.font = Font(name=cell_data['font']['name'], size=cell_data['font']['size'],
                                         bold=cell_data['font']['bold'], italic=cell_data['font']['italic'])

                        if hex_color != '00000000':

                            try:
                                color = Color(rgb=hex_color)

                                # Создание объекта заливки
                                fill = PatternFill(patternType='solid', fgColor=color)
                                cell.fill = fill
                            except Exception:
                                pass

                except:
                    pass

                if type(cell_data['borders']['left']) is dict:

                    rgb_string_left = cell_data['borders']['left']['color']

                    color_font_left = change_rgb_to_hex(rgb_string_left)

                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(
                            style=cell_data['borders']['left']['style'],
                            color=color_font_left),
                        right=openpyxl.styles.Side(
                            style=cell_data['borders']['right']['style'],
                            color=change_rgb_to_hex(cell_data['borders']['right']['color'])),
                        top=openpyxl.styles.Side(
                            style=cell_data['borders']['top']['style'],
                            color=change_rgb_to_hex(cell_data['borders']['top']['color'])),
                        bottom=openpyxl.styles.Side(
                            style=cell_data['borders']['bottom']['style'],
                            color=change_rgb_to_hex(cell_data['borders']['bottom']['color'])),
                    )
                else:
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style=cell_data['borders']['left']),
                        right=openpyxl.styles.Side(style=cell_data['borders']['right']),
                        top=openpyxl.styles.Side(style=cell_data['borders']['top']),
                        bottom=openpyxl.styles.Side(style=cell_data['borders']['bottom'])
                    )

                wrap_true = cell_data['alignment']['wrap_text']

                cell.alignment = openpyxl.styles.Alignment(horizontal=cell_data['alignment']['horizontal'],
                                                           vertical=cell_data['alignment']['vertical'],
                                                           wrap_text=wrap_true)

    try:
        self.data_well.image_data = data['image']
        # Добавьте обработку ошибки, например, пропуск изображения или запись информации об ошибке в лог
    except ValueError as e:
        print(f"Ошибка при вставке изображения: {type(e).__name__}\n\n{str(e)}")

    for col in range(13):
        sheet_new.column_dimensions[get_column_letter(col + 1)].width = col_width[col]
    index_delete = 47

    for index_row, row in enumerate(sheet_new.iter_rows()):
        # Копирование высоты строки
        if any(['Наименование работ' in str(col.value) for col in row[:13]]) and self.data_well.work_plan not in [
            'plan_change']:
            index_delete = index_row
            data_list.gns_ind2 = index_row

        elif any(['ПЛАН РАБОТ' in str(col.value).upper() for col in row[:4]]) and self.data_well.work_plan not in [
            'plan_change']:
            sheet_new.cell(row=index_row + 1,
                           column=2).value = f'ДОПОЛНИТЕЛЬНЫЙ ПЛАН РАБОТ № {self.data_well.number_dp}'

        elif any(['ИТОГО:' in str(col.value).upper() for col in row[:4]]) and self.data_well.work_plan in [
            'plan_change']:
            index_delete = index_row + 2
            data_list.gns_ind2 = index_row + 2

        elif any(['КАТЕГОРИЯ СКВАЖИНЫ' in str(col.value).upper() for col in row[:4]]):
            self.data_well.cat_well_min = data_list.ProtectedIsDigit(index_row)

        elif all([col is None for col in row[:13]]):
            sheet_new.row_dimensions[index_row].hidden = True
        try:
            sheet_new.row_dimensions[index_row].height = row_heights[index_row - 1]
        except Exception:
            pass

    if self.data_well.work_plan not in ['plan_change']:
        sheet_new.delete_rows(index_delete, sheet_new.max_row - index_delete + 1)

        razdel = self.work_podpisant_list(self.data_well.region, data_list.contractor)
        for index_row, row in enumerate(sheet_new.iter_rows(max_row=self.data_well.cat_well_min.get_value, max_col=12)):
            if len(razdel) > index_row:
                for index_col, column in enumerate(row):
                    if sheet_new.cell(row=index_row + 1, column=index_col + 1).value:
                        sheet_new.cell(row=index_row + 1, column=index_col + 1).value = razdel[index_row][index_col]
            else:
                for index_col, column in enumerate(row):
                    if sheet_new.cell(row=index_row + 1, column=index_col + 1).value:
                        sheet_new.cell(row=index_row + 1, column=index_col + 1).value = None

    return sheet_new


def change_rgb_to_hex(rgb_string_font_color):
    if rgb_string_font_color is None or rgb_string_font_color == "RGB(Values must be of type <class 'str'>)":
        rgb_string_font_color = 'RGB(00000000)'

    # Извлекаем шестнадцатеричный код цвета
    hex_color_font = rgb_string_font_color[4:-1]

    color_font = Color(rgb=hex_color_font)
    return color_font


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    # app.setStyleSheet()
    window = ClassifierWell()
    window.show()
    sys.exit(app.exec_())
