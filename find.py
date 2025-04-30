import re
from io import BytesIO

import data_list
import base64
from datetime import datetime

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QInputDialog, QMessageBox
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl_image_loader import SheetImageLoader
from work_py.data_informations import dict_data_cdng

from decrypt import decrypt
from main import ExcelWorker, MyMainWindow, MyWindow

from data_list import ProtectedIsDigit, ProtectedIsNonNone


class FindIndexPZ(MyMainWindow):
    wb_pvr = Workbook()

    def __init__(self, ws, work_plan, parent=None):
        super().__init__()

        self.end_index = None

        self.insert_index2 = None
        self.need_depth = None
        self.count_row_well = None
        self.prs_copy_index = ProtectedIsDigit(0)
        self.perforation_sole = 5000
        self.number_dp = 0
        self.skm_interval = []
        self.head_column = 0
        self.modal_dialog = None
        self.image_loader = None
        self.water_density = ProtectedIsDigit(1.18)
        self.result_pressure_date = ProtectedIsNonNone('01.01.2000')
        self.column_direction_mine_diameter, self.column_direction_mine_wall_thickness, \
        self.column_direction_mine_length, self.level_cement_direction_mine = None, None, None, None
        self.fluid_work = None
        self.nkt_template = None
        self.nkt_diam = None
        self.perforation_roof = 5000
        self.old_version = False
        self.region = None
        self.fluid = None
        self.insert_index = 0
        self.curator = None
        self.data_well_max = ProtectedIsDigit(0)
        self.plan_correct_index = ProtectedIsDigit(0)
        self.condition_of_wells = ProtectedIsDigit(0)
        self.plan_correct_index = ProtectedIsDigit(0)
        self.type_absorbent = ''
        self.sucker_rod_none = False
        self.fluid_work_short = '0.87г/см3'

        self.ws = ws
        self.wb = parent.wb

        self.category_h2s = 0
        self.category_gas_factor = 0

        self.thread_excel = None

        self.without_damping = False
        self.insert_index = 0
        self.bottom = 5000
        self.work_plan = work_plan
        self.row_expected = []
        self.column_direction_true = False
        self.category_h2s_list = []
        self.value_h2s_mg = []
        self.category_gaz_factor_percent = []
        self.gaz_factor_percent = []
        self.bur_rastvor = ''
        self.bottom_hole_drill = ProtectedIsNonNone('не корректно')
        self.bottom_hole_artificial = ProtectedIsNonNone(5000)
        self.max_angle = ProtectedIsNonNone('не корректно')
        self.well_number = ProtectedIsNonNone('не корректно')
        self.well_area = ProtectedIsNonNone('не корректно')
        self.max_angle_depth = ProtectedIsNonNone('не корректно')
        self.stol_rotor = ProtectedIsNonNone('не корректно')
        self.column_conductor_diameter = ProtectedIsNonNone('не корректно')
        self.column_conductor_wall_thickness = ProtectedIsNonNone('не корректно')
        self.column_conductor_length = ProtectedIsNonNone('не корректно')
        self.level_cement_direction = ProtectedIsNonNone('не корректно')
        self.column_direction_diameter = ProtectedIsNonNone('отсут')
        self.column_direction_wall_thickness = ProtectedIsNonNone('отсут')
        self.column_direction_length = ProtectedIsNonNone('отсут')
        self.level_cement_conductor = ProtectedIsNonNone('отсут')
        self.column_diameter = ProtectedIsNonNone('не корректно')
        self.column_wall_thickness = ProtectedIsNonNone('не корректно')
        self.shoe_column = ProtectedIsNonNone('не корректно')
        self.level_cement_column = ProtectedIsNonNone('не корректно')
        self.pressure_mkp = ProtectedIsNonNone('0')
        self.column_additional_diameter = ProtectedIsNonNone('не корректно')
        self.column_additional_wall_thickness = ProtectedIsNonNone('не корректно')
        self.head_column_additional = ProtectedIsNonNone('не корректно')
        self.shoe_column_additional = ProtectedIsNonNone('не корректно')
        self.interval_temp = ProtectedIsNonNone('не корректно')

        self.leakiness_count = 0
        self.emergency_count = 0
        self.distance_from_well_to_sampling_point = 0

        self.date_drilling_cancel = ''
        self.date_drilling_run = ''

        self.max_expected_pressure = ProtectedIsNonNone('не корректно')
        self.max_admissible_pressure = ProtectedIsNonNone('не корректно')
        self.result_pressure = ProtectedIsNonNone('не корректно')
        self.first_pressure = ProtectedIsNonNone('не корректно')
        self.static_level = ProtectedIsNonNone('не корректно')
        self.dinamic_level = ProtectedIsNonNone('не корректно')
        self.well_volume_in_pz = []
        self.well_fluid_in_pz = []
        self.pressure_mkp = ProtectedIsNonNone('не корректно')
        self.column_direction_true = False
        self.column_additional = False
        self.data_window = None
        self.perforation_correct_window2 = None
        self.for_paker_list = False
        self.grp_plan = False
        self.angle_data = []
        self.expected_oil = 0
        self.water_cut = 0
        self.date_commissioning = ProtectedIsNonNone('01.01.2000')
        self.expected_pressure = 0
        self.appointment_well = ProtectedIsNonNone('')
        self.expected_pickup = 0
        self.sucker_rod_ind = ProtectedIsDigit(0)
        self.percent_water = 0
        self.expected_pick_up = {}
        self.ribbing_interval = []
        self.drilling_interval = []
        self.nkt_opress_true = False
        self.stabilizator_need = False
        self.norm_of_time = 0
        self.leakiness = False
        self.pipes_ind = ProtectedIsDigit(0)
        self.data_x_min = ProtectedIsDigit(0)
        self.data_x_max = ProtectedIsDigit(0)
        self.data_pvr_min = ProtectedIsDigit(0)
        self.data_fond_min = ProtectedIsDigit(0)
        self.data_pvr_min = ProtectedIsDigit(0)
        self.pipes_ind = ProtectedIsDigit(0)
        self.data_pvr_max = ProtectedIsDigit(0)
        self.well_area = ProtectedIsNonNone("")
        self.well_oilfield = ProtectedIsNonNone("")
        self.inventory_number = ProtectedIsNonNone("")
        self.cdng = ProtectedIsNonNone("")
        self.paker_before = {"before": 0, "after": 0}
        self.depth_fond_paker_before = {"before": 0, "after": 0}
        self.paker_second_before = {"before": 0, "after": 0}
        self.depth_fond_paker_second_before = {"before": 0, "after": 0}
        self.dict_pump_shgn = {"before": 0, "after": 0}
        self.dict_pump_ecn = {"before": 0, "after": 0}
        self.dict_pump_shgn_depth = {"before": 0, "after": 0}
        self.dict_pump_ecn_depth = {"before": 0, "after": 0}
        self.dict_sucker_rod = {}

        self.pvr_row_list = []
        self.dict_nkt_before = {}
        self.dict_nkt_after = {}
        self.dict_sucker_rod_after = {}
        self.column_direction_mine_true = False
        self.dict_pump = {"before": 0, "after": 0}
        self.column_head_m = ''
        self.wellhead_fittings = ''
        self.groove_diameter = ''
        self.image_list = []
        self.image_data = []
        self.dict_leakiness = {}
        self.data_list = []
        self.open_trunk_well = False
        self.count_template = 0
        self.template_depth = 0
        self.leakiness_interval = []
        self.category_pressure_second = ''
        self.category_h2s_second = ''
        self.category_pressure_well = []
        self.gaz_factor_pr_second = ''
        self.current_bottom_second = 0
        self.current_bottom = 5000
        self.template_length = 0
        self.template_depth_addition = 0
        self.template_length_addition = 0

        self.type_kr = ''
        self.konte_true = False
        self.bvo = False
        self.category_pvo = 2

        self.problem_with_ek = False
        self.problem_with_ek_diameter = 220
        self.problem_with_ek_depth = 10000
        self.dict_perforation = {}
        self.dict_perforation_short = {}
        self.dict_perforation_project = {}
        self.category_pressure_list = []
        self.category_pressure = None
        self.skm_interval = []
        self.leakiness = False

        self.emergency_well = False
        self.problem_with_ek = False
        self.gips_in_well = False
        self.index_row_pvr_list = []
        self.plast_all = []

        self.value_h2s_percent = []
        self.value_h2s_mg = []
        self.value_h2s_mg_m3 = []
        self.dict_category = {}

        self.plast_project = []
        self.bcu_level = False
        self.plast_work = []
        self.category_p_p = []
        self.image_data = []
        self.check_data_in_pz = ['Рассмотрев план заказ были выявлены следующие нарушения, '
                                 'прошу дополнить и внести изменения:\n']

        self.gis_list = []
        self.index_row_pvr_list = []

        self.cat_well_min = ProtectedIsDigit(0)
        self.cat_well_max = ProtectedIsDigit(0)
        self.data_well_min = ProtectedIsDigit(0)


        self.data_pvr_min = ProtectedIsDigit(0)

        if self.work_plan not in ['prs']:
            self.read_pz()
        else:
            self.read_pz_prs()



    def delete_rows_pz(self, ws, cat_well_min, data_well_max, data_x_max):
        boundaries_dict = {}

        for ind, _range in enumerate(ws.merged_cells.ranges):
            boundaries_dict[ind] = range_boundaries(str(_range))

        # row_heights_top = [None, 18.0, 18, 18,None, 18.0, 18, 18,None, 18.0, 18, 18, 18.0, 18, 18, 18.0, 18, 18, 18.0, 18, 18]
        row_heights1 = [ws.row_dimensions[i + 1].height for i in range(cat_well_min.get_value, ws.max_row)]
        for key, value in boundaries_dict.items():
            ws.unmerge_cells(start_column=value[0], start_row=value[1],
                             end_column=value[2], end_row=value[3])

        # print(f'индекс удаления {1, self.cat_well_min - 1} , {data_well_max + 2, ws.max_row - data_well_max}')

        if 'prs' not in self.work_plan:
            ws.delete_rows(data_x_max.get_value, ws.max_row - data_x_max.get_value)

        ws.delete_rows(1, cat_well_min.get_value - 1)

        # print(sorted(boundaries_dict))
        data_list.row_heights = row_heights1

        for _ in range(16):
            ws.insert_rows(1, 1)

        for key, value in boundaries_dict.items():
            if value[1] <= data_well_max.get_value + 1 and value[1] >= cat_well_min.get_value:
                ws.merge_cells(start_column=value[0], start_row=value[1] + 16 - cat_well_min.get_value + 1,
                               end_column=value[2], end_row=value[3] + 16 - cat_well_min.get_value + 1)

        # print(f'{ws.max_row, len(data_list.prow_heights)}dd')
        for index_row, row in enumerate(ws.iter_rows()):  # Копирование высоты строки
            ws.row_dimensions[index_row + 17].height = data_list.row_heights[index_row - 1]

    @staticmethod
    def check_text_in_row(text, row):
        return any([text.lower() in str(col).lower() for col in row])

    def read_pz(self):
        cat_well_min = []
        self.image_loader = None
        try:
            # Копирование изображения
            self.image_loader = SheetImageLoader(self.ws)

        except Exception as e:
            QMessageBox.warning(None, 'Ошибка', f'Ошибка в копировании изображений {e}')

        for row_ind, row in enumerate(self.ws.iter_rows(values_only=True, max_row=300, max_col=20)):
            self.ws.row_dimensions[row_ind].hidden = False
            if self.cat_well_min.get_value != 0:
                coord_image = list(map(
                    lambda x: int("".join([y for y in x if y.isdigit()])), list(self.image_loader._images.keys())))
                if self.data_x_max.get_value == 0 and row_ind in coord_image:
                    self.work_with_img(self.image_loader, row_ind)

            if 'Категория скважины' in row:
                cat_well_min.append(row_ind + 1)
                self.cat_well_min = ProtectedIsDigit(min(cat_well_min))  # индекс начала категории

            elif any(['план-заказ' in str(col).lower() or 'план работ' in str(col).lower() for col in row]) \
                    and row_ind < 50:
                self.cat_well_max = ProtectedIsDigit(row_ind)
                self.data_well_min = ProtectedIsDigit(row_ind + 1)
            elif any(['стабилизатор' in str(col).lower() and 'желез' in str(col).lower() for col in row]):
                self.stabilizator_need = True

            elif any(['Ожидаемые показатели после' in str(col) for col in row]):
                self.data_x_min = ProtectedIsDigit(row_ind + 1)
                # print(f' индекс Ожидаемые показатели {self.data_x_min}')
            elif any(['эксплуатационные горизонты и интервалы перфорации' in str(col).lower() for col in row]):
                self.data_pvr_min = ProtectedIsDigit(row_ind)
            elif 'Оборудование скважины ' in row:
                self.data_fond_min = ProtectedIsDigit(row_ind)

            elif any(['VIII. Вид и категория ремонта, его шифр' in str(col) for col in row]):
                type_kr = self.ws.cell(row=row_ind + 2, column=1).value
                n = 1
                while type_kr is None and n != 8:
                    type_kr = self.ws.cell(row=row_ind + 2, column=1 + n).value
                    n += 1
                self.type_kr = type_kr

            elif any(['IX. Мероприятия по предотвращению' in str(col) for col in row]) or \
                    any(['IX. Мероприятия по предотвращению аварий, инцидентов и осложнений::' in str(col) for col in
                         row]):

                self.data_well_max = ProtectedIsDigit(row_ind)

            elif 'НКТ' == str(row[1]).upper():
                self.pipes_ind = ProtectedIsDigit(row_ind + 1)

            elif 'ШТАНГИ' == str(row[1]).upper():
                self.sucker_rod_ind = ProtectedIsDigit(row_ind + 1)

            elif (self.check_text_in_row('Планируемый объём работ', row) or \
                  self.check_text_in_row('Планируемый объём работ', row) or \
                  self.check_text_in_row('Порядок работы', row) \
                  or self.check_text_in_row('Ранее проведенные работ', row)) and \
                    self.data_x_max.get_value == 0:
                self.data_x_max = ProtectedIsDigit(row_ind)
                if self.check_text_in_row('Ранее проведенные работ', row):
                    self.data_x_max = ProtectedIsDigit(row_ind - 2)
            elif any(['Должность' == str(col) for col in row]):
                self.end_index = ProtectedIsDigit(row_ind - 2)
                break
            elif any(['II. История эксплуатации скважины' in str(col) for col in row]):
                self.data_pvr_max = ProtectedIsDigit(row_ind)

            elif 'III. Состояние скважины к началу ремонта ' in row:
                self.condition_of_wells = ProtectedIsDigit(row_ind)
            elif any([('безопасный' in str(col).lower() and ('урове' in str(col).lower() or 'Нст' in str(col).lower()))
                      or 'бсу' in str(col).lower() for col in row]):
                self.bcu_level = True
            for col, value in enumerate(row):
                if value is not None and col <= 12:
                    if 'сужен' in str(value).lower() or 'не проход' in str(value).lower() or \
                            'дорн' in str(value).lower() or 'пластырь' in str(value).lower():
                        self.problem_with_ek = True
                        self.problem_with_ek = True

                    if 'гипс' in str(value).lower() or 'гидратн' in str(value).lower():
                        self.gips_in_well = True

        if self.data_x_max.get_value == 0:
            QMessageBox.warning(self, 'Ошибка', 'Не корректный файл excel, либо отсутствует строка с '
                                                'текстом Порядок работ')
            # self.pause_app()
            return

        if self.cat_well_max.get_value == 0:
            QMessageBox.warning(self, 'Ошибка', 'Не корректный файл excel, либо отсутствует строка с '
                                                'текстом ПЛАН-ЗАКАЗ или ПЛАН-РАБОТ')
            # self.pause_app()
            return

        if self.cat_well_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала копирования',
                                'Программа не смогла определить строку начала копирования, нужно '
                                'добавить "Категория скважины" в ПЗ для определения начала копирования')
            # self.pause_app()
            return
        if self.data_well_max.get_value == 0:
            QMessageBox.warning(self, 'индекс окончания копирования',
                                'Программа не смогла определить строку с IX. Мероприятия по предотвращению аварий '
                                'нужно добавить "IX. Мероприятия по предотвращению аварии" в ПЗ')
            # self.pause_app()
            return
        if self.data_well_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала строки после план заказ',
                                'Программа не смогла найти начала строку с названием "План работ" или "план заказ"')
            # self.pause_app()
            return

        if self.sucker_rod_none:
            if self.sucker_rod_ind.get_value == 0:
                sucker_mes = QMessageBox.question(self, 'ШТАНГИ', 'Программа определелила, что в скважине '
                                                                  'отсутствуют штанги, корректно ли это?')
                if sucker_mes == QMessageBox.StandardButton.Yes:
                    self.sucker_rod_ind = ProtectedIsDigit(0)
                else:
                    QMessageBox.information(self, 'ШТАНГИ', 'Нужно добавить "ШТАНГИ" в таблицу?')
                    # self.pause_app()
                    return

        if self.data_x_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала копирования ожидаемых показателей',
                                'Программа не смогла определить строку начала копирования ожидаемых показателей')
            # self.pause_app()
            return

        if self.data_pvr_max.get_value == 0:
            QMessageBox.warning(self, 'индекс историю',
                                'Программа не смогла найти "II. История эксплуатации скважины"')
            # self.pause_app()
            return

        if self.pipes_ind.get_value == 0:
            QMessageBox.warning(self, 'индекс начала строки с НКТ',
                                'Программа не смогла найти строку с НКТ, необходимо проверить столбец В')
            # self.pause_app()
            return
        if self.data_pvr_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала начала ПВР', 'Программа не смогла найти индекс начала ПВР')
            # self.pause_app()
            return
        if self.data_fond_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала строки с таблицей фондового оборудования',
                                'Программа не смогла найти строку с таблицей фондового оборудования')
            # self.pause_app()
            return
        if self.type_kr == '':
            QMessageBox.information(self, 'Вид ГТМ', 'Приложение не смогло найти тип КР, '
                                                     'необходимо внести вручную')
        if self.condition_of_wells.get_value == 0:
            QMessageBox.warning(
                self, 'индекс копирования',
                'Программа не смогла определить строку n\ III. Состояние скважины к началу ремонта ')
            # self.pause_app()
            return
        if self.type_kr in ['', None]:
            self.check_data_in_pz.append('Не указан Вид и категория ремонта, его шифр\n')

        # if self.work_plan != 'plan_change':
        #     self.row_expected = []
        #     for j in range(self.data_x_min.get_value,
        #                    self.data_x_max.get_value):  # Ожидаемые показатели после ремонта
        #         lst = []
        #         for i in range(0, 12):
        #             lst.append(self.ws.cell(row=j + 1, column=i + 1).value)
        #         self.row_expected.append(lst)

        if self.bcu_level is False:
            QMessageBox.warning(self, 'безопасный статический уровень',
                                'В план заказе не указан безопасный статический уровень')
            self.check_data_in_pz.append('В план заказе не указан безопасный статический уровень\n'
                                         'Нарушен п. 9.1.9 инструкции БНД по предупреждению '
                                         'ГНВП №ПЗ-05 И-102089 ЮЛ-305 ')

    def read_pz_prs(self):

        cat_well_min = []
        self.image_loader = None
        try:
            # Копирование изображения
            self.image_loader = SheetImageLoader(self.ws)

        except Exception as e:
            QMessageBox.warning(None, 'Ошибка', f'Ошибка в копировании изображений {e}')
        if any([self.ws.cell(row=i, column=1).value != None for i in range(1, 20)]):
            QMessageBox.warning(self, 'Ошибка', 'Для корректной работы приложения нужно сместить колонку "А" в право')
            # self.pause_app()
            return

        for row_ind, row in enumerate(self.ws.iter_rows(values_only=True, max_row=300, max_col=20)):
            self.ws.row_dimensions[row_ind].hidden = False
            if self.cat_well_min.get_value != 0:
                coord_image = list(map(
                    lambda x: int("".join([y for y in x if y.isdigit()])), list(self.image_loader._images.keys())))
                if self.data_x_max.get_value == 0 and row_ind in coord_image:
                    self.work_with_img(self.image_loader, row_ind)

            if 'Категория скважины' in row:
                cat_well_min.append(row_ind + 1)
                self.cat_well_min = ProtectedIsDigit(min(cat_well_min))  # индекс начала категории

            elif any(['план-заказ' in str(col).lower() or 'план работ' in str(col).lower() for col in row]) \
                    and row_ind < 50:
                self.ws.cell(row=row_ind + 1, column=2).value = 'План работ'
                self.cat_well_max = ProtectedIsDigit(row_ind)
                self.data_well_min = ProtectedIsDigit(row_ind + 1)
            elif any(['стабилизатор' in str(col).lower() and 'желез' in str(col).lower() for col in row]):
                self.data_well.stabilizator_need = True
            elif self.check_text_in_row('XI. ПЛАН РАБОТ:', row):
                self.prs_copy_index = ProtectedIsDigit(row_ind + 1)
                self.ws.cell(row=row_ind + 1, column=2).value = 'XIa ТЕХНИЧЕСКОЕ ЗАДАНИЕ от ЦДНГ:'
            elif self.check_text_in_row('Ожидаемые показатели после', row):
                self.data_x_min = ProtectedIsDigit(row_ind)
                # print(f' индекс Ожидаемые показатели {self.data_x_min}')
            elif self.check_text_in_row('9. Эксплуатационные горизонты и интервалы перфорации:', row):
                self.data_pvr_min = ProtectedIsDigit(row_ind)
            elif 'Оборудование скважины ' in row:
                self.data_fond_min = ProtectedIsDigit(row_ind)

            elif self.check_text_in_row('VIII. Вид и категория ремонта, его шифр', row):
                type_kr = self.ws.cell(row=row_ind + 2, column=1).value
                n = 1
                while type_kr is None and n != 8:
                    type_kr = self.ws.cell(row=row_ind + 2, column=1 + n).value
                    n += 1
                self.type_kr = type_kr

            elif any(['IX. Мероприятия по предотвращению' in str(col) for col in row]) or \
                    any(['IX. Мероприятия по предотвращению аварий, инцидентов и осложнений::' in str(col) for col in
                         row]):

                self.data_well_max = ProtectedIsDigit(row_ind)

            elif 'НКТ' == str(row[1]).upper():
                self.pipes_ind = ProtectedIsDigit(row_ind + 1)

            elif 'ШТАНГИ' == str(row[1]).upper():
                self.sucker_rod_ind = ProtectedIsDigit(row_ind + 1)

            elif self.check_text_in_row('Перемещение НКТ, ШН:', row) or \
                    self.check_text_in_row('Порядок работы', row) and self.data_x_max.get_value == 0:
                self.data_x_max = ProtectedIsDigit(row_ind)
                self.condition_of_wells = ProtectedIsDigit(row_ind)
                if 'prs' not in self.work_plan:
                    break

            elif self.check_text_in_row('Запуск скважины из ремонта:', row):
                self.data_x_max_prs = ProtectedIsDigit(row_ind + 1)
                break

            elif any(['II. История эксплуатации скважины' in str(col) for col in row]):
                self.data_pvr_max = ProtectedIsDigit(row_ind)

            elif 'III. Состояние скважины к началу ремонта ' in row or 'Перемещение НКТ, ШН:' in row:
                self.condition_of_wells = ProtectedIsDigit(row_ind)
            elif 'Герметизация , разгерметизация  устья  скважины' in row:
                self.plan_correct_index = ProtectedIsDigit(row_ind)

            for col, value in enumerate(row):
                if value is not None and col <= 12:
                    if 'сужен' in str(value).lower() or 'не проход' in str(value).lower() or \
                            'дорн' in str(value).lower() or 'пластырь' in str(value).lower():
                        self.problem_with_ek = True
                        self.problem_with_ek = True

                    if 'гипс' in str(value).lower() or 'гидратн' in str(value).lower():
                        self.gips_in_well = True

        if self.data_x_max_prs.get_value == 0:
            QMessageBox.warning(self, 'Ошибка', 'Не корректный файл excel, либо отсутствует строка с '
                                                'текстом "Запуск скважины из ремонта:"')
            # self.pause_app()
            return

        if self.prs_copy_index.get_value == 0:
            QMessageBox.warning(self, 'Ошибка', 'Не корректный файл excel, либо отсутствует строка с '
                                                'текстом "XI. ПЛАН РАБОТ"')
            # self.pause_app()
            return

        if self.cat_well_max.get_value == 0:
            QMessageBox.warning(self, 'Ошибка', 'Не корректный файл excel, либо отсутствует строка с '
                                                'текстом ПЛАН-ЗАКАЗ или ПЛАН-РАБОТ')
            # self.pause_app()
            return

        if self.cat_well_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала копирования',
                                'Программа не смогла определить строку начала копирования, нужно '
                                'добавить "Категория скважины" в ПЗ для определения начала копирования')
            # self.pause_app()
            return
        if self.data_well_max.get_value == 0:
            QMessageBox.warning(self, 'индекс окончания копирования',
                                'Программа не смогла определить строку с IX. Мероприятия по предотвращению аварий '
                                'нужно добавить "IX. Мероприятия по предотвращению аварии" в ПЗ')
            # self.pause_app()
            return
        if self.data_well_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала строки после план заказ',
                                'Программа не смогла найти начала строку с названием "План работ" или "план заказ"')
            # self.pause_app()
            return

        if self.sucker_rod_none:
            if self.sucker_rod_ind.get_value == 0:
                sucker_mes = QMessageBox.question(self, 'ШТАНГИ', 'Программа определелила, что в скважине '
                                                                  'отсутствуют штанги, корректно ли это?')
                if sucker_mes == QMessageBox.StandardButton.Yes:
                    self.sucker_rod_ind = ProtectedIsDigit(0)
                else:
                    QMessageBox.information(self, 'ШТАНГИ', 'Ключевое слово для поиска индекса строк в п ПЗ это \n'
                                                            'ШТАНГИ,\n до ремонта, \nпосле ремонта \nплан. '
                                                            'При необходимости нужно исправить ПЗ в '
                                                            'соответствии с этими данными')
                    data_list.pause = False
                    self.pause_app()
                    return

        if self.data_x_max.get_value == 0:
            QMessageBox.warning(self, 'индекс окончания копирования ожидаемых показателей',
                                'Программа не смогла определить строку окончания копирования'
                                ' ожидаемых показателей "ХI Планируемый объём работ"')
            # self.pause_app()
            return

        if self.data_x_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала копирования ожидаемых показателей',
                                'Программа не смогла определить строку начала копирования ожидаемых показателей')
            # self.pause_app()
            return

        if self.data_pvr_max.get_value == 0:
            QMessageBox.warning(self, 'индекс историю',
                                'Программа не смогла найти "II. История эксплуатации скважины"')
            # self.pause_app()
            return

        if self.pipes_ind.get_value == 0:
            QMessageBox.warning(self, 'индекс начала строки с НКТ',
                                'Программа не смогла найти строку с НКТ, необходимо проверить столбец В')
            # self.pause_app()
            return
        if self.data_pvr_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала начала ПВР', 'Программа не смогла найти индекс начала ПВР')
            # self.pause_app()
            return
        if self.data_fond_min.get_value == 0:
            QMessageBox.warning(self, 'индекс начала строки с таблицей фондового оборудования',
                                'Программа не смогла найти строку с таблицей фондового оборудования')
            # self.pause_app()
            return
        if self.type_kr == '':
            QMessageBox.information(self, 'Вид ГТМ', 'Приложение не смогло найти тип КР, '
                                                     'необходимо внести вручную')
        if self.condition_of_wells.get_value == 0:
            QMessageBox.warning(
                self, 'индекс копирования',
                'Программа не смогла определить строку n\ III. Состояние скважины к началу ремонта ')
            # self.pause_app()
            return
        if self.type_kr in ['', None]:
            self.check_data_in_pz.append('Не указан Вид и категория ремонта, его шифр\n')

        # if self.work_plan != 'plan_change':
        #     self.row_expected = []
        #     for j in range(self.data_x_min.get_value,
        #                    self.data_x_max.get_value):  # Ожидаемые показатели после ремонта
        #         lst = []
        #         for i in range(0, 12):
        #             lst.append(self.ws.cell(row=j + 1, column=i + 1))
        #         self.row_expected.append(lst)

    def work_with_img(self, image_loader, row):
        for col in range(1, 12):
            coord = f'{get_column_letter(col)}{row}'
            if image_loader.image_in(coord):
                # Загружаем изображение из текущей ячейки
                try:
                    image_file = image_loader.get(coord)
                except:
                    image_file = None
                if image_file:
                    coord = f'{get_column_letter(col)}{row + 17 - self.cat_well_min.get_value}'

                    # Создаем байтовый поток
                    image_bytes = BytesIO()
                    # Сохраняем изображение в байтовом потоке
                    image_file.save(image_bytes, format='PNG')  # Замените 'PNG' на другой формат, если необходимо
                    image_bytes.seek(0)  # Возвращаемся в начало байтового потока
                    image_base64 = base64.b64encode(image_bytes.read()).decode("utf-8")

                    # Создание словаря для изображения
                    image_info = {
                        "coord": coord,
                        "width": image_file.width,
                        "height": image_file.height,
                        "data": image_base64
                    }

                    # Добавление информации в список
                    self.image_data.append(image_info)

    def read_work_data(self):
        need_depth = None
        for row_ind, row in enumerate(self.ws.iter_rows(values_only=True, min_row=self.data_x_max.get_value,
                                                        max_row=self.end_index.get_value)):  # словарь количества НКТ и метраж
            for col_index, col in enumerate(row):
                if ("до гл" in str(col) or (
                        'восстано' in str(col).lower() or 'нормал' in str(col).lower())) or "до забо" in str(col):
                    text = list(filter(lambda x: "до гл" in x or "до забо" in x, col.split('.')))
                    if text:
                        need_depth = self.check_once_isdigit(text[0][text[0].index("до"):])

            if need_depth or row_ind > 6:
                break

        if need_depth and need_depth > self.current_bottom:
            self.need_depth = need_depth
        if self.current_bottom > self.perforation_sole:
            self.need_depth = self.bottom_hole_artificial.get_value

    def check_str_none(self, string):

        try:
            if MyWindow.check_str_isdigit(str(string)) is True:
                if str(round(float(str(string).replace(',', '.')), 1))[-1] == "0":
                    return int(float(str(string).replace(',', '.')))
                else:
                    return round(float(str(string).replace(',', '.')), 4)
            elif str(string).replace(' ', '') == '-' or 'отсут' in str(string).lower() or \
                    str(string).strip() == '' or string is None:
                return '0'
            elif '(мм)' in string and '(м)' in string:
                return string
            elif len(str(string).split('/')) == 2:
                lst = []
                for i in str(string).split('/'):
                    b = ''
                    for j in i:
                        if j in '0123456789.x':
                            b = str(b) + j
                        elif j == ',':
                            b = str(b) + '.'
                    lst.append(float(b))
                return lst
            elif len(str(string).split('-')) == 2:
                lst = []
                for i in str(string).split('-'):
                    # print(i)
                    lst.append(float(i.replace(',', '.').strip()))
                return lst

            else:
                b = 0
                for i in str(string.strip()):
                    i.replace(',', '.')
                    if i in '0123456789,.x':
                        b = str(b) + i
                b = float(b)
                return b
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка',
                                f'Ошибка в прочтении файла в строке {string} {e}, Проверьте excel файл')

    def definition_is_none(self, data, row, col, step, m=12):
        if data.__class__ == ProtectedIsNonNone:
            data = data.get_value
            while (data is None) and (step < m):
                data = self.ws.cell(row=row, column=col + step).value
                step += 1
            return ProtectedIsNonNone(data)
        elif data.__class__ == ProtectedIsDigit:
            data = data.get_value
            while (data is None) and (step < m):
                data = self.check_once_isdigit(self.ws.cell(row=row, column=col + step).value)
                step += 1
            return ProtectedIsDigit(data)
        else:
            while (data is None) and (step < m):
                data = self.ws.cell(row=row, column=col + step).value
                step += 1
            return data


def insert_column_direction(text):
    try:
        column_direction_data = text.split('(мм),')
        try:
            diameter = ProtectedIsDigit(
                column_direction_data[0].replace(' ', ''))
        except Exception:
            diameter = ProtectedIsNonNone('не корректно')

        try:
            wall_thickness = ProtectedIsDigit(
                column_direction_data[1].replace(' ', ''))
        except Exception:
            wall_thickness = ProtectedIsNonNone(
                'не корректно')
        try:
            try:
                length = ProtectedIsDigit(
                    column_direction_data[2].split('-')[1].replace('(м)', '').replace(" ", ""))
            except Exception:
                length = ProtectedIsDigit(
                    column_direction_data[2].replace('(м)', '').replace(" ", ""))

        except Exception:
            length = ProtectedIsNonNone('не корректно')
    except Exception:
        diameter = ProtectedIsNonNone('не корректно')
        wall_thickness = ProtectedIsNonNone('не корректно')
        length = ProtectedIsNonNone('не корректно')
    return diameter, wall_thickness, length


class WellNkt(FindIndexPZ):
    def __init__(self):
        super().__init__()
        # self.read_well(self.ws, data_list.pipes_ind.get_value, data_list.condition_of_wells.get_value)

    def read_well(self, begin_index, cancel_index):
        dict_nkt = {}
        dict_nkt_po = {}

        a_plan = 0
        data_list.nkt_mistake = False
        self.column_index_lenght_nkt = None
        self.column_index_column_nkt = None
        for row_ind, row in enumerate(self.ws.iter_rows(values_only=True, min_row=begin_index - 1,
                                                        max_row=cancel_index)):  # словарь количества НКТ и метраж
            if self.check_text_in_row('план', row) or self.check_text_in_row('карта спуска (планируемое)', row):
                a_plan = row_ind
            if row_ind < 2:
                for col_index, col in enumerate(row):
                    if 'диаметр' in str(col).lower() and 'мм' in str(col):
                        self.column_index_diametr_nkt = col_index
                    if 'кол-во' in str(col).lower() and 'шт' in str(col):
                        self.column_index_column_nkt = col_index
                    if 'длина' in str(col).lower() and 'м' in str(col):
                        self.column_index_lenght_nkt = col_index
        if self.column_index_lenght_nkt is None:
            QMessageBox.warning(self, 'Ошибка', 'Ошибка в поиске индекса длины НКТ')
            self.pause_app()
            return

        if a_plan == 0:
            QMessageBox.warning(self, 'Индекс планового НКТ',
                                'Программа не могла определить начала строку с ПЗ НКТ - план')
            self.pause_app()
            return

        for row_ind, row in enumerate(
                self.ws.iter_rows(values_only=True, min_row=begin_index - 1, max_row=cancel_index)):
            if row_ind >= 1:
                key = str(row[self.column_index_diametr_nkt]).strip()
                if key != str(None) and key != '-' and "диам" not in key.lower():
                    value = row[self.column_index_lenght_nkt]
                    if value:
                        if row_ind < a_plan:
                            dict_nkt[key] = dict_nkt.get(
                                key, 0) + round(self.check_str_none(value), 1)
                        elif row_ind >= a_plan:
                            dict_nkt_po[key] = dict_nkt_po.get(
                                key, 0) + round(self.check_str_none(value), 1)
        self.dict_nkt_before = dict_nkt
        self.dict_nkt_after = dict_nkt_po


class WellSuckerRod(FindIndexPZ):
    def __init__(self):
        super().__init__()

    # self.read_well(self.ws, data_list.sucker_rod_ind.get_value, data_list.pipes_ind.get_value)

    def read_well(self, begin_index, cancel_index):
        self.dict_sucker_rod = {}
        self.dict_sucker_rod_after = {}

        # try:
        dict_sucker_rod = {}
        dict_sucker_rod_po = {}

        b_plan = 0
        if self.sucker_rod_ind.get_value != 0:
            for row_ind, row in enumerate(
                    self.ws.iter_rows(values_only=True, min_row=begin_index - 1, max_row=cancel_index, max_col=20)):
                if self.check_text_in_row('план', row) or \
                        self.check_text_in_row('карта спуска (планируемое)', row):
                    b_plan = row_ind
                if row_ind <= 1:
                    for col_index, col in enumerate(row):
                        if 'диаметр' in str(col).lower() and 'мм' in str(col):
                            self.column_index_diametr_nkt = col_index
                        if 'кол-во' in str(col).lower() and 'шт' in str(col):
                            self.column_index_column_nkt = col_index
                        if 'длина' in str(col).lower() and 'м' in str(col):
                            self.column_index_lenght_nkt = col_index
            asd = self.sucker_rod_none

            if b_plan == 0:
                sucker_rod_question = QMessageBox.question(self,
                                                           'отсутствие штанг',
                                                           'Программа определило что штанг в '
                                                           'скважине нет, корректно?')
                if sucker_rod_question == QMessageBox.StandardButton.Yes:
                    self.sucker_rod_none = False
                else:
                    self.sucker_rod_none = True

                if self.sucker_rod_none is True:
                    QMessageBox.warning(self, 'Индекс планового НКТ',
                                        'Программа не могла определить начала строку с ПЗ'
                                        ' штанги - план.\n'
                                        'Ключевое слово для поиска индекса строк в п ПЗ это \n'
                                        '"ШТАНГИ",\n "до ремонта", \n"после ремонта" \nплан. '
                                        'При необходимости нужно исправить ПЗ в '
                                        'соответствии с этими данными. Приложение нужно перезапустить')
                    self.pause_app()
                    return
            # print(f'б {b_plan}')

            for row_ind, row in enumerate(
                    self.ws.iter_rows(values_only=True, min_row=begin_index - 1, max_row=cancel_index - 1)):
                if row_ind >= 1:
                    key = str(row[self.column_index_diametr_nkt]).replace(' ', '')
                    value = row[self.column_index_lenght_nkt]
                    if key != str(None) and key != '-' and key != '' and 'отсут' not in str(
                            key).lower() and 'диам' not in str(key).lower():
                        # print(key, value)
                        if key is not None and row_ind < b_plan:
                            try:
                                dict_sucker_rod[key] = dict_sucker_rod.get(key, 0) + int(
                                    float(str(value).replace(',', '.'))) + 1
                            except Exception:
                                QMessageBox.warning(self, 'Ошибка', 'Ошибка в определении длины штанг до ремонта, '
                                                                    'скорректируйте план заказ')
                                # self.pause_app()
                                break

                                return
                        if key is not None and row_ind >= b_plan:
                            try:
                                dict_sucker_rod_po[key] = dict_sucker_rod_po.get(key, 0) + int(
                                    float(str(value).replace(',', '.')))
                            except Exception:
                                QMessageBox.warning(self, 'Ошибка', 'Ошибка в определении длины штанг до ремонта, '
                                                                    'скорректируйте план заказ')
                                # self.pause_app()
                                break

                                return
        self.dict_sucker_rod = dict_sucker_rod
        self.dict_sucker_rod_after = dict_sucker_rod_po

        return True


class WellFondData(FindIndexPZ):

    def __init__(self):
        super().__init__()
        # self.read_well(self.ws, data_list.data_fond_min.get_value, data_list.condition_of_wells.get_value)

    def read_well(self, begin_index, cancel_index):

        paker_do = {"before": 0, "after": 0}
        depth_fond_paker_do = {"before": 0, "after": 0}
        paker2_do = {"before": 0, "after": 0}
        depth_fond_paker2_do = {"before": 0, "after": 0}
        dict_pump_shgn = {"before": '0', "after": '0'}
        dict_pump_ecn = {"before": '0', "after": '0'}
        dict_pump_shgn_h = {"before": '0', "after": '0'}
        dict_pump_ecn_h = {"before": '0', "after": '0'}
        dict_pump = {"before": '0', "after": '0'}
        data_list.old_index = 1
        wellhead_fittings = ''
        column_head_m = ''
        groove_diameter = ''
        for row_index, row in enumerate(self.ws.iter_rows(values_only=True, min_row=begin_index, max_row=cancel_index)):
            row_index += begin_index
            for col, cell in enumerate(row):
                value = cell
                if value:
                    if 'карта спуска' in str(value).lower():
                        col_plan = col
                    if 'до ремонта' in str(value).lower() and row_index < 6 + begin_index:
                        col_do = col
                    if 'колонная головка' in str(value) and 'типоразмер' in str(row[col + 2]):
                        column_head_m = row[col_do]
                    if 'Арматура устьевая' in str(value) and 'типоразмер' in str(row[col + 2]):
                        wellhead_fittings = row[col_do]

                    if 'диаметр канавки' in str(value).lower():
                        groove_diameter = row[col_do]
                        if groove_diameter is None:
                            groove_diameter = ''

                    if 'Пакер' in str(value) and 'типоразмер' in str(row[col + 2]):
                        if '/' in str(row[col_do]):
                            paker_do["before"] = str(row[col_do]).split('/')[0]
                            paker2_do["before"] = str(row[col_do]).split('/')[1]
                        else:
                            paker_do["before"] = row[col_do]

                        if '/' in str(row[col_plan]):
                            paker_do["after"] = str(row[col_plan]).split('/')[0]
                            paker2_do["after"] = str(row[col_plan]).split('/')[1]
                        else:
                            paker_do["after"] = row[col_plan]

                    elif value == 'Насос' and row[col + 2] == 'типоразмер':
                        if row[col_do]:
                            if 'ЭЦН' in str(row[col_do]).upper() or 'ВНН' in str(row[col_do]).upper():
                                dict_pump_ecn["before"] = row[col_do]
                                if '/' in str(row[col_do]):
                                    dict_pump_ecn["before"] = [ecn for ecn in row[col_do].split('/')
                                                               if 'ЭЦН' in ecn or 'ВНН' in ecn][0]
                            if ('НВ' in str(row[col_do]).upper() or 'ШГН' in str(row[col_do]).upper() or
                                'НН' in str(row[col_do]).upper()) or 'RH' in str(row[col_do]).upper():
                                dict_pump_shgn["before"] = row[col_do]
                                if '/' in str(row[col_do]):
                                    dict_pump_shgn["before"] = [ecn for ecn in row[col_do].split('/')
                                                                if 'НВ' in ecn or 'НН' in ecn or
                                                                'ШГН' in ecn or 'RH' in ecn][0]

                                # print(dict_pump_ecn["before"])

                        if row[col_plan]:
                            if 'ЭЦН' in str(row[col_plan]).upper() or 'ВНН' in str(row[col_plan]).upper():
                                dict_pump_ecn["after"] = row[col_plan]
                                if '/' in str(row[col_plan]):
                                    dict_pump_ecn["after"] = [ecn for ecn in row[col_plan].split('/')
                                                              if 'ЭЦН' in ecn or 'ВНН' in ecn][0]

                            if 'НВ' in str(row[col_plan]).upper() or \
                                    'ШГН' in str(row[col_plan]).upper() or \
                                    'НН' in str(row[col_plan]).upper() or \
                                    'RHAM' in str(row[col_plan]).upper():
                                dict_pump_shgn["after"] = row[col_plan]
                                if '/' in str(row[col_plan]):
                                    dict_pump_shgn["after"] = [ecn for ecn in row[col_plan].split('/')
                                                               if 'НВ' in ecn or 'НН' in ecn or
                                                               'ШГН' in ecn or 'RHAM' in ecn][0]

                        if dict_pump_ecn["before"] != 0:
                            dict_pump_ecn_h["before"] = self.check_str_none(
                                self.ws.cell(row=row_index + 4,
                                             column=col_do + 1).value)
                            if '/' in str(self.ws.cell(row=row_index + 4, column=col_do + 1)):
                                dict_pump_ecn_h["before"] = max(self.check_str_none(self.ws.cell(
                                    row=row_index + 4, column=col_do + 1)))
                        if dict_pump_shgn["before"] != 0:
                            dict_pump_shgn_h["before"] = self.check_str_none(
                                self.ws.cell(row=row_index + 4,
                                             column=col_do + 1).value)
                            if '/' in str(self.ws.cell(row=row_index + 4, column=col_do + 1)):
                                dict_pump_shgn_h["before"] = min(self.check_str_none(self.ws.cell(
                                    row=row_index + 4, column=col_do + 1).value))
                        if dict_pump_ecn["after"] != 0:
                            dict_pump_ecn_h["after"] = self.check_str_none(self.ws.cell(row=row_index + 4,
                                                                                        column=col_plan + 1).value)
                            if '/' in str(self.ws.cell(row=row_index + 4, column=col_plan + 1).value):
                                dict_pump_ecn_h["after"] = max(
                                    self.check_str_none(self.ws.cell(row=row_index + 4,
                                                                     column=col_plan + 1).value))
                        if dict_pump_shgn["after"] != 0:
                            dict_pump_shgn_h["after"] = self.check_str_none(
                                self.ws.cell(row=row_index + 4,
                                             column=col_plan + 1).value)
                            if '/' in str(self.ws.cell(row=row_index + 4, column=col_plan + 1).value):
                                dict_pump_shgn_h["after"] = min(
                                    self.check_str_none(self.ws.cell(row=row_index + 4,
                                                                     column=col_plan + 1).value))

                    elif value == 'Н посадки, м':
                        try:
                            if paker_do["before"] != 0:
                                depth_fond_paker_do["before"] = \
                                    self.check_str_none(row[col_do])[0]
                                depth_fond_paker2_do["before"] = \
                                    self.check_str_none(row[col_do])[1]
                        except Exception:
                            if paker_do["before"] != 0:
                                depth_fond_paker_do["before"] = row[col_do]
                        try:
                            if paker_do["after"] != 0:
                                depth_fond_paker_do["after"] = \
                                    self.check_str_none(row[col_plan])[0]
                                depth_fond_paker2_do["after"] = \
                                    self.check_str_none(row[col_plan])[1]
                        except Exception as e:
                            if paker_do["after"] != 0:
                                depth_fond_paker_do["after"] = row[col_plan]

        if self.depth_fond_paker_before["after"] <= 900 and self.paker_before['after'] != 0 and \
                '89' not in list(self.dict_nkt_after.keys()) and self.dict_pump_ecn['after'] == 0 and \
                self.dict_pump_shgn['after'] == 0:
            QMessageBox.warning(self, 'НКТ89мм', f'При глубине спуска ф.пакера до глубины 900м '
                                                 f'необходимо использование НКТ89мм, а не {list(self.dict_nkt_after.keys())}')
            self.check_data_in_pz.append(
                f'Согласно мероприятий по недопущению разгерметизации системы НКТ- пакер от 20.04.2021:\n'
                f'При глубине спуска ф.пакера до глубины 900м '
                f'необходимо использование НКТ89мм, а не {list(self.dict_nkt_after.keys())}')

        if wellhead_fittings in [None, '']:
            self.check_data_in_pz.append('Не указан тип устьевой арматуры\n '
                                         'Нарушен п. 9.1.9 инструкции БНД по предупреждению ГНВП №ПЗ-05 И-102089 ЮЛ-305')
        if column_head_m in [None, '']:
            self.check_data_in_pz.append('Не указан тип Колонной головки или завод-изготовитель \n'
                                         'Нарушен п. 8.1.9 инструкции БНД по предупреждению ГНВП №ПЗ-05 И-102089. ЮЛ-305')
        if groove_diameter in [None, '']:
            self.check_data_in_pz.append(
                'Не указан Диаметр канавки устьевой арматуры или тип резьбы\n ')

        self.paker_before = paker_do
        self.depth_fond_paker_before = depth_fond_paker_do
        self.paker_second_before = paker2_do
        self.depth_fond_paker_second_before = depth_fond_paker2_do
        self.dict_pump_shgn = dict_pump_shgn
        self.dict_pump_ecn = dict_pump_ecn
        self.dict_pump_shgn_depth = dict_pump_shgn_h
        self.dict_pump_ecn_depth = dict_pump_ecn_h
        self.dict_pump = dict_pump
        self.column_head_m = column_head_m
        self.wellhead_fittings = wellhead_fittings
        self.groove_diameter = groove_diameter

        return True


class WellHistoryData(FindIndexPZ):

    def __init__(self):
        super().__init__()
        self.leakage_window = None

        # self.read_well(self.ws, data_list.data_pvr_max.get_value, data_list.data_fond_min.get_value)

    def read_well(self, begin_index, cancel_index):

        for row_index, row in enumerate(
                self.ws.iter_rows(values_only=True, min_row=begin_index, max_row=cancel_index, max_col=20)):
            for col, cell in enumerate(row):
                value = cell
                if value:
                    if 'нэк' in str(value).lower() or 'негерм' in str(
                            value).lower() or 'нарушение э' in str(
                        value).lower() or \
                            'нарушение г' in str(value).lower():
                        self.leakiness_count += 1

                    if ('авар' in str(
                            value).lower() or 'расхаж' in str(
                        value).lower() or 'лар' in str(value).lower()) \
                            and 'акт о расследовании аварии прилагается' not in str(value).lower():
                        self.emergency_well = True
                        self.emergency_well = True
                        self.emergency_count += 1
                        self.emergency_count += 1

                    if 'Начало бурения' in str(value):
                        # self.date_drilling_run = row[col + 2]
                        self.date_drilling_run = row[col + 2]

                    elif 'Конец бурения' in str(value):
                        self.date_drilling_cancel = row[col + 1]

                        self.date_drilling_cancel = \
                            self.definition_is_none(self.date_drilling_cancel,
                                                    row_index + begin_index, col + 1, 1)
                    elif 'Дата ввода в экспл' in str(value):
                        self.date_commissioning = ProtectedIsNonNone(row[col + 2])
                        if type(self.date_commissioning.get_value) is datetime:
                            self.date_commissioning = ProtectedIsNonNone(self.date_commissioning.get_value.strftime(
                                '%d.%m.%Y'))

                    elif 'ствол скважины' in str(row[col]).lower() and 'буров' in str(row[col]).lower():
                        self.bur_rastvor = row[col]

                    elif 'Максимально ожидаемое давление на устье' in str(value):
                        aaa = row[col + 1]
                        self.max_expected_pressure = ProtectedIsDigit(row[col + 1])
                        self.max_expected_pressure = self.definition_is_none(self.max_expected_pressure,
                                                                             row_index + begin_index, col + 1, 1)

                    elif 'Результат предыдущей ' in str(value):
                        self.result_pressure = ProtectedIsDigit(row[col + 1])
                        self.result_pressure = self.definition_is_none(
                            self.result_pressure,
                            row_index + begin_index,
                            col + 1, 1)
                    elif 'дата опрессовки' in str(value).lower():
                        self.result_pressure_date = ProtectedIsDigit(row[col + 2])
                        if type(self.result_pressure_date.get_value) is datetime:
                            self.result_pressure_date = ProtectedIsDigit(self.result_pressure_date.get_value.strftime(
                                '%d.%m.%Y'))

                    elif 'Первоначальное давление опрессовки э/колонны' in str(value):
                        self.first_pressure = ProtectedIsDigit(row[col + 3])
                    elif 'максимально допустимое давление' in str(value).lower():
                        self.max_admissible_pressure = ProtectedIsDigit(row[col + 1])
                        self.max_admissible_pressure = \
                            self.definition_is_none(self.max_admissible_pressure, row_index + begin_index, col + 1, 1)
        if self.date_drilling_run == '':
            self.check_data_in_pz.append('не указано начало бурения\n')
        if self.date_drilling_cancel == '':
            self.check_data_in_pz.append('не указано окончание бурения\n')

        if self.date_commissioning.get_value == '':
            self.check_data_in_pz.append('не указано дата ввода\n')
        if self.max_expected_pressure.get_value:
            if self.max_expected_pressure.get_value < 30:
                QMessageBox.warning(None, 'допустимое давление ',
                                    'максимально ожидаемое давление на устье слишком маленькое')
                self.max_expected_pressure = ProtectedIsDigit(30)
        if self.max_expected_pressure.get_value in ['', 0, '0']:
            self.check_data_in_pz.append('не указано максимально ожидаемое давление на устье\n')
        if self.max_admissible_pressure.get_value in ['', 0, '0']:
            self.check_data_in_pz.append('не указано максимально допустимое давление на устье\n')
        if float(self.max_admissible_pressure.get_value) < 30:
            QMessageBox.warning(None, 'допустимое давление ',
                                'максимально допустимое давление на устье слишком маленькое')
            self.check_data_in_pz.append(f'максимально допустимое давление на устье слишком маленькое'
                                         f' {self.max_admissible_pressure}\n')


class WellCondition(FindIndexPZ):
    leakage_window = None

    def __init__(self):
        super().__init__()

        # self.read_well(self.ws, data_list.condition_of_wells.get_value, self.data_well.data_well_max.get_value)

    def read_well(self, begin_index, cancel_index):

        for row_index, row in enumerate(
                self.ws.iter_rows(values_only=True, min_row=begin_index, max_row=cancel_index, max_col=20)):
            row_index += begin_index
            for col, cell in enumerate(row):
                value = cell
                if value is not None:
                    if 'нэк' in str(value).lower() or 'негерм' in str(value).lower() or \
                            'нарушение э' in str(value).lower() or 'нарушение г' in str(value).lower():
                        self.leakiness_count += 1

                    if ('авар' in str(value).lower() or 'расхаж' in str(value).lower() or 'лар' in str(value)) \
                            and 'акт о расследовании аварии прилагается' not in str(value):
                        self.emergency_well = True
                        self.emergency_count += 1
                    if value:
                        if "Hст " in str(value):
                            if '/' in str(row[col + 1]):
                                self.static_level = ProtectedIsDigit(row[col + 1].split('/')[0])
                            else:
                                self.static_level = ProtectedIsDigit(row[col + 1])
                        elif 'плотн.воды' in str(value):
                            self.water_density = ProtectedIsDigit(row[col + 1])
                        elif 'Рмкп ' in str(value):
                            self.pressure_mkp = ProtectedIsNonNone(row[col + 3])

                        elif "грп" in str(value).lower():
                            self.grp_plan = True

                        elif "Ндин " in str(value):
                            self.dinamic_level = ProtectedIsDigit(row[col + 1])
                        elif "% воды " in str(value):
                            self.percent_water = str(row[col + 1]).strip().replace('%', '')
                            self.percent_water = FindIndexPZ.definition_is_none(
                                self, self.percent_water, row_index,
                                col + 1, 1)
                        elif 'расстояние от скважин' in str(value).lower():
                            self.distance_from_well_to_sampling_point = str(row[col + 2]).replace(',', '.')

                        elif 'плотность жидкости ' in str(value).lower():

                            if 'prs' in self.work_plan:
                                well_volume_in_pz = str(row[col + 4]).replace(',', '.')
                            else:
                                well_volume_in_pz = str(row[col + 5]).replace(',', '.')
                            step = 0
                            while self.check_str_isdigit(well_volume_in_pz) is False:
                                for step in range(4, 8):
                                    well_volume_in_pz = str(row[col + step]).replace(',', '.')
                                if self.check_str_isdigit(well_volume_in_pz) or step == 7:
                                    break
                            if self.check_str_isdigit(well_volume_in_pz) is False:
                                well_volume_in_pz, _ = QInputDialog.getDouble(self, 'Объем глушения',
                                                                              f'Введите объем глушения согласно ПЗ',
                                                                              50, 1, 70)
                            self.well_volume_in_pz.append(round(float(well_volume_in_pz), 1))

                            if 'prs' in self.work_plan:
                                well_fluid_in_pz = str(row[col + 2]).replace(',', '.')
                            else:
                                well_fluid_in_pz = str(row[col + 2]).replace(',', '.')
                            step = 0
                            while self.check_str_isdigit(well_fluid_in_pz) is False:
                                for step in range(2, 5):
                                    well_fluid_in_pz = str(row[col + step]).replace(',', '.')
                                if self.check_str_isdigit(well_fluid_in_pz) or step == 4:
                                    break

                            if self.check_str_isdigit(well_fluid_in_pz):
                                self.well_fluid_in_pz.append(round(float(well_fluid_in_pz), 2))
                            else:
                                QMessageBox.warning(self, 'Ошибка', f'Не корректно прочитан удельный '
                                                                    f'вес {well_fluid_in_pz}')

        if self.static_level.get_value == 'не корректно':
            self.check_data_in_pz.append('не указан статический уровень')
        if self.pressure_mkp.get_value in [None, 'не корректно', '-', 'нет', 'отсут']:
            asde = self.pressure_mkp.get_value
            self.check_data_in_pz.append(
                'не указано наличие наличие устройство замера давления и наличие давления в МКП \n'
                'Нарушен п. 9.1.9 инструкции БНД по предупреждению ГНВП №ПЗ-05 И-102089 ЮЛ-305')

        if self.leakiness_count != 0:
            leakiness_quest = QMessageBox.question(self, 'нарушение колонны',
                                                   f'Программа определила что в скважине'
                                                   f' есть нарушение - {self.leakiness_count}, '
                                                   f'верно ли?')
            if leakiness_quest == QMessageBox.StandardButton.Yes:
                self.leakiness = True
        return True


class WellExpectedPickUp(FindIndexPZ):
    def __init__(self):
        super().__init__()
        # self.read_well(self.ws, self.data_well.data_x_min.get_value, self.data_well.data_x_max.get_value)

    def read_well(self, begin_index, cancel_index):

        for row_index, row in enumerate(
                self.ws.iter_rows(values_only=True, min_row=begin_index, max_row=cancel_index, max_col=20)):
            row_index += begin_index
            # print(row_index)
            for col, cell in enumerate(row[0:15]):
                # print(row_index)
                value = cell
                if value:

                    if 'прием' in str(value).lower() or 'qж' in str(value).lower():
                        self.expected_pickup = row[col + 1]

                        self.expected_pickup = self.definition_is_none(self.expected_pickup, row_index,
                                                                       col + 1, 1)

                    if 'зак' in str(value).lower() or 'давл' in str(value).lower() or 'p' in str(value).lower():
                        self.expected_pressure = row[col + 1]
                        self.expected_pressure = self.definition_is_none(self.expected_pressure, row_index,
                                                                         col + 1, 1)

                    if 'qж' in str(value).lower():
                        self.water_cut = str(row[col + 1]).strip().replace(' ', '').replace(
                            'м3/сут', '')
                        self.water_cut = self.definition_is_none(
                            self.water_cut,
                            row_index, col + 1, 1)

                    if 'qн' in str(value).lower():
                        self.expected_oil = str(row[col + 1]).replace(' ', '').replace('т/сут', '')
                        self.expected_oil = self.definition_is_none(self.expected_oil,
                                                                    row_index, col + 1, 1)
                    if 'воды' in str(value).lower() and "%" in str(value).lower():
                        try:
                            proc_water = str(row[col + 1]).replace(' ', '').replace('%', '')

                            proc_water = self.definition_is_none(proc_water, row_index, col + 1, 1)
                            self.percent_water = int(float(proc_water)) if float(
                                proc_water) > 1 else round(
                                float(proc_water) * 100,
                                0)
                        except Exception:
                            print(f'ошибка в определение')

            try:
                if self.expected_pickup and self.expected_pressure:
                    self.expected_pick_up[self.expected_pickup] = self.expected_pressure
            except Exception as e:
                print(f'Ошибка в определении ожидаемых показателей {e}')

        return self


class WellName(FindIndexPZ):
    def __init__(self):
        super().__init__(self, parent=None)
        # self.read_well(self.ws, self.cat_well_max.get_value, data_list.data_pvr_min.get_value)

    def read_well(self, begin_index, cancel_index):
        for row_index, row in enumerate(
                self.ws.iter_rows(values_only=True, min_row=begin_index, max_row=cancel_index, max_col=20)):
            row_index += begin_index

            for col, cell in enumerate(row):
                value = cell
                if value:
                    if '№скв.:' in str(value):
                        self.well_number = ProtectedIsNonNone(str(row[col + 1]))

                    if 'площадь' in str(value):
                        if self.work_plan not in ['prs']:
                            self.well_number = ProtectedIsNonNone(str(row[col - 1]).lstrip().rstrip())
                        self.well_area = ProtectedIsNonNone(str(row[col + 1]).lstrip().rstrip().replace(" ", "_"))
                        # self.well_number = ProtectedIsNonNone(row[col - 1])
                        # self.well_area = ProtectedIsNonNone(row[col + 1])

                    elif 'месторождение' in str(value):  # определение номера скважины
                        self.well_oilfield = ProtectedIsNonNone(str(row[col + 2]).lstrip().rstrip())
                        # self.well_oilfield = ProtectedIsNonNone(row[col + 2])
                    elif 'инв. №' in str(value).lower():
                        self.inventory_number = ProtectedIsNonNone(row[col + 1])
                        # self.inventory_number = ProtectedIsNonNone(row[col + 1])
                    elif 'цех' in str(value):
                        self.cdng = ProtectedIsDigit(row[col + 1])
                        # self.cdng = ProtectedIsDigit(row[col + 1])
                    elif 'назначение' in str(value):
                        if self.work_plan in ['prs']:
                            a = row[col + 3], row[col + 2], row[col + 1]
                            self.appointment_well = ProtectedIsNonNone(row[col + 2])
                        else:
                            a = row[col + 3], row[col + 2], row[col + 1]
                            self.appointment_well = ProtectedIsNonNone(row[col + 2])
                        # well_data.appointment_well = ProtectedIsDigit(row[col + 1])
                        # print(f' ЦДНГ {self.cdng.get_value}')

        if self.cdng.get_value not in list(dict_data_cdng.keys()):
            QMessageBox.warning(self, 'Ошибка', f'ЦДНГ - {self.cdng.get_value} отсутствует в списках, '
                                                f'нужно уточнить промысел')
            return
        return self


class WellData(FindIndexPZ):

    def __init__(self):
        super().__init__()

        # self.read_well(self.ws, self.cat_well_max.get_value, data_list.data_pvr_min.get_value)

    def read_well(self, begin_index, cancel_index):
        for row_index, row in enumerate(
                self.ws.iter_rows(values_only=True, min_row=begin_index, max_row=cancel_index, max_col=20)):
            row_index += begin_index

            for col, cell in enumerate(row[:15]):
                value = cell
                if value:
                    if 'пробуренный забой' in str(value).lower() or 'пробуренный:' in str(value).lower():
                        self.bottom_hole_drill = ProtectedIsDigit(row[col + 1])
                        self.bottom_hole_drill = self.definition_is_none(
                            self.bottom_hole_drill,
                            row_index, col, 2)

                        self.bottom_hole_artificial = ProtectedIsDigit(row[col + 4])

                        self.bottom_hole_artificial = \
                            self.definition_is_none(self.bottom_hole_artificial,
                                                    row_index, col, 5)
                        # print(f'пробуренный забой {self.bottom_hole_artificial}')
                    if 'иск. забой' in str(value).lower():
                        self.bottom_hole_artificial = ProtectedIsDigit(row[col + 1])

                    elif 'интервалы темпа набора кривизны ' in str(value).lower():
                        self.interval_temp = ProtectedIsNonNone(row[col + 2])
                        self.interval_temp = self.definition_is_none(
                            self.interval_temp, row_index, col + 2, 1)

                    elif 'зенитный угол' in str(value).lower():
                        self.max_angle = ProtectedIsDigit(self.check_once_isdigit(row[col + 4]))
                        for index, col1 in enumerate(row[:14]):
                            if 'на глубине' in str(col1):
                                self.max_angle_depth = ProtectedIsDigit(row[index + 1])
                            if index > 10:
                                break

                    elif 'тек. забой:' in str(value).lower() and len(value) < 15:
                        self.current_bottom = self.check_once_isdigit(row[col + 1])

                    elif 'текущий забой' in str(value).lower() and len(value) < 15:
                        self.current_bottom = self.check_once_isdigit(row[col + 2])
                        self.current_bottom = \
                            FindIndexPZ.definition_is_none(
                                self, self.current_bottom, row_index, col, 2)

                        self.bottom = self.current_bottom
                    elif 'Расстояние от стола ротора ' in str(value):
                        self.stol_rotor = FindIndexPZ.definition_is_none(
                            self, ProtectedIsDigit(row[col + 5]), row_index, col + 1, 1)
                    elif 'Шахтное направление' in str(value):
                        asawawq = row[col + 3] not in ['-', None, '0', 0, '', 'отсутствует', '(мм), (мм), -(м)',
                                                       'отсут'], 'отсут' not in str(row[col + 3]).lower()
                        if row[col + 3] not in ['-', None, '0', 0, '', 'отсутствует', '(мм), (мм), -(м)', 'отсут'] and \
                                'отсут' not in str(row[col + 3]).lower():
                            self.column_direction_mine_true = True
                            column_direction_mine_data = row[col + 3]
                            column_direction_mine_data = FindIndexPZ.definition_is_none(self,
                                                                                        column_direction_mine_data,
                                                                                        row_index, col, 2)
                            self.column_direction_mine_diameter, self.column_direction_mine_wall_thickness, \
                            self.column_direction_mine_length = insert_column_direction(column_direction_mine_data)
                            self.level_cement_direction_mine = ProtectedIsNonNone(row[col + 9])
                        else:
                            self.column_direction_mine_true = False

                    elif 'Направление (диаметр наружный(мм)' in str(value):
                        self.column_direction_true = True
                        if self.column_direction_true:
                            column_direction_data = row[col + 3]
                            column_direction_data = FindIndexPZ.definition_is_none(self, column_direction_data,
                                                                                   row_index, col, 2)
                            self.column_direction_diameter, self.column_direction_wall_thickness, \
                            self.column_direction_length = insert_column_direction(column_direction_data)
                            self.level_cement_direction = ProtectedIsNonNone(0)
                    elif 'Кондуктор (диаметр наружный(мм), ' in str(value):
                        self.column_conductor_true = True
                        if self.column_conductor_true:
                            column_conductor_data = row[col + 3]
                            column_conductor_data = FindIndexPZ.definition_is_none(self, column_conductor_data,
                                                                                   row_index, col, 2)
                            self.column_conductor_diameter, self.column_conductor_wall_thickness, \
                            self.column_conductor_length = insert_column_direction(column_conductor_data)
                            self.level_conductor_direction = ProtectedIsNonNone(0)

                    elif 'Направление' in str(value) and 'Шахтное направление' not in str(value) and \
                            self.ws.cell(row=row_index + 1, column=col + 1) is not None and \
                            self.check_str_none(row[col + 3]) != '0':
                        self.column_direction_true = True
                        column_direction_data = row[col + 3]
                        if self.column_direction_true:
                            cell = None
                            for col1, cell in enumerate(row):
                                if 'Уровень цемента' in str(cell):
                                    n = 1
                                    while row[col1 + n] is None or n > 6:
                                        if 'уст' in str(row[col1 + 2]).lower() or str(
                                                row[col1 + 2]).isdigit():
                                            self.level_cement_direction = ProtectedIsDigit(0)
                                        else:
                                            if '-' in str(row[col1 + 2]):
                                                self.level_cement_direction = ProtectedIsDigit(
                                                    str(row[col1 + 2].split('-')[0]).replace(" ", ""))
                                        n += 1
                        else:
                            self.level_cement_direction = ProtectedIsNonNone('отсут')
                        self.column_direction_diameter, self.column_direction_wall_thickness, \
                        self.column_direction_length = insert_column_direction(column_direction_data)

                    elif 'Кондуктор' in str(value) and \
                            self.check_str_none(row[col + 3]) != '0':
                        cell = None
                        for col1, cell in enumerate(row):
                            if 'Уровень цемента' in str(cell):
                                try:
                                    if 'уст' in str(row[col1 + 2]).lower() or str(row[col1 + 2]).isdigit():
                                        self.level_cement_conductor = ProtectedIsDigit(0)
                                    else:
                                        self.level_cement_conductor = ProtectedIsDigit(
                                            str(row[col1 + 2].split('-')[0]).replace(' ', ''))
                                except Exception:
                                    self.level_cement_conductor = ProtectedIsNonNone('не корректно')
                        try:
                            column_conductor_data = str(row[col + 3])
                            self.column_conductor_diameter, self.column_conductor_wall_thickness, \
                            self.column_conductor_length = insert_column_direction(column_conductor_data)

                        except Exception:
                            self.column_conductor_diameter = ProtectedIsNonNone('не корректно')
                            self.column_conductor_wall_thickness = ProtectedIsNonNone('не корректно')
                            self.column_conductor_length = ProtectedIsNonNone('не корректно')
                    elif self.check_text_in_row('3.Эксплуатационная колонна (диаметр наружный(мм)', row):
                        data_main_production_string = self.ws.cell(row=row_index + 1,
                                                                   column=2).value
                        self.column_diameter, self.column_wall_thickness, \
                        self.shoe_column = insert_column_direction(data_main_production_string)
                    elif 'Эксплуатационная колонна (' in str(value):

                        data_main_production_string = str(
                            self.ws.cell(row=row_index + 1, column=col + 1).value)
                        self.column_diameter, self.column_wall_thickness, \
                        self.shoe_column = insert_column_direction(data_main_production_string)


                    elif 'Уровень цемента за колонной' in str(value):

                        self.level_cement_column = ProtectedIsDigit(str(row[col + 2]).strip())
                        self.level_cement_column = self.definition_is_none(
                            self.level_cement_column, row_index, col, 1)

                    elif 'онструкция хвостовика' in str(value):

                        data_column_additional = self.check_str_none(self.ws.cell(row=row_index + 2,
                                                                                  column=col + 2).value)

                        if data_column_additional != '0':
                            self.column_additional = True
                        if self.column_additional is True:
                            try:
                                self.head_column_additional = ProtectedIsDigit(
                                    data_column_additional[0])
                            except Exception:
                                self.head_column_additional = ProtectedIsNonNone('не корректно')
                            try:
                                self.shoe_column_additional = ProtectedIsDigit(
                                    data_column_additional[1])
                            except Exception:
                                self.shoe_column_additional = ProtectedIsNonNone('не корректно')

                            try:
                                try:
                                    data_add_column = self.check_str_none(
                                        self.ws.cell(row=row_index + 2,
                                                     column=col + 4).value)
                                    # print(f' доп колонна {data_add_column}')
                                    self.column_additional_diameter = ProtectedIsDigit(
                                        data_add_column[0])

                                except Exception:
                                    self.column_additional_diameter = ProtectedIsDigit(
                                        self.check_str_none(
                                            self.ws.cell(row=row_index + 2,
                                                         column=col + 4).value))

                                try:
                                    data_add_column = self.check_str_none(
                                        self.ws.cell(row=row_index + 2,
                                                     column=col + 4).value)
                                    self.column_additional_wall_thickness = ProtectedIsDigit(
                                        data_add_column[1])
                                except Exception:

                                    self.column_additional_wall_thickness = ProtectedIsDigit(
                                        self.check_str_none(
                                            self.ws.cell(row=row_index + 2,
                                                         column=col + 6).value))

                            except Exception:
                                self.column_additional_wall_thickness = ProtectedIsNonNone(
                                    'не корректно')
                                self.column_additional_diameter = ProtectedIsNonNone('не корректно')
                        else:
                            self.column_additional_diameter = ProtectedIsNonNone('отсут')
                            self.column_additional_wall_thickness = ProtectedIsNonNone('отсут')
                            self.head_column_additional = ProtectedIsNonNone('отсут')
                            self.shoe_column_additional = ProtectedIsNonNone('отсут')
        try:
            if self.stol_rotor.get_value in ['не корректно', None, '']:
                self.check_data_in_pz.append('не указано Стол ротора \n')
            if self.max_angle.get_value in ['не корректно', None, '']:
                self.check_data_in_pz.append('не указано максимальный угол \n')
            if self.max_angle_depth.get_value in ['не корректно', None, '']:
                self.check_data_in_pz.append('не указано глубина максимального угла\n')
            if self.level_cement_column.get_value in ['не корректно', None, '']:
                self.check_data_in_pz.append('не указан уровень цемент за колонной\n')

            if self.dict_pump_shgn['before'] not in ['0', 0] and self.dict_pump_shgn_depth['before'] not in ['0', 0]:
                adwdr = abs(sum(list(self.dict_sucker_rod.values())) - self.dict_pump_shgn_depth['before'])
                if self.dict_sucker_rod:
                    if abs(sum(list(self.dict_sucker_rod.values())) - self.dict_pump_shgn_depth['before']) > 10:
                        QMessageBox.warning(self, 'Ошибка', f'Длина штанг {sum(list(self.dict_sucker_rod.values()))}м '
                                                            f'до ремонта не равно глубине насоса '
                                                            f'{self.dict_pump_shgn_depth["before"]}м \n')
                        self.check_data_in_pz.append(
                            f'Ошибка в карте спуска: Длина штанг {sum(list(self.dict_sucker_rod.values()))}м '
                            f'до ремонта не равно глубине насоса '
                            f'{self.dict_pump_shgn_depth["before"]}м \n')
                if self.dict_nkt_before:
                    if sum(list((self.dict_nkt_before.values()))) - self.dict_pump_shgn_depth["before"] < 0:
                        QMessageBox.warning(self, 'Ошибка',
                                            f'Длина НКТ {sum(list(self.dict_nkt_after.values()))}м '
                                            f'после ремонта меньше глубины насоса'
                                            f'{self.dict_pump_shgn_depth["after"]}м')
                        self.check_data_in_pz.append(
                            f'Ошибка в карте спуска: \n Длина НКТ {sum(list(self.dict_nkt_after.values()))}м '
                            f'после ремонта не равно глубине насоса '
                            f'{self.dict_pump_shgn_depth["after"]}м')
            if self.distance_from_well_to_sampling_point == 0:
                QMessageBox.warning(self, 'Ошибка', f'Не указано расстояние до пункта налива')
                self.check_data_in_pz.append(f'Не указано расстояние до пункта налива')

            if self.dict_pump_shgn['after'] not in ['0', 0] and self.dict_pump_shgn_depth['after'] not in ['0', 0]:
                if self.dict_sucker_rod_after:
                    if abs(sum(list(self.dict_sucker_rod_after.values())) - self.dict_pump_shgn_depth['after']) > 10:
                        QMessageBox.warning(self, 'Ошибка',
                                            f'Длина штанг {sum(list(self.dict_sucker_rod_after.values()))}м '
                                            f'после ремонта не равно глубине насоса '
                                            f'{self.dict_pump_shgn_depth["after"]}м')
                        self.check_data_in_pz.append(
                            f'Ошибка в карте спуска: \nОшибка в карте спуска: Длина штанг '
                            f'{sum(list(self.dict_sucker_rod_after.values()))}м '
                            f'после ремонта не равно глубине насоса '
                            f'{self.dict_pump_shgn_depth["after"]}м')
                if self.dict_nkt_after:
                    if sum(list((self.dict_nkt_after.values()))) - self.dict_pump_shgn_depth["after"] < 0:
                        QMessageBox.warning(self, 'Ошибка',
                                            f'Длина НКТ {sum(list(self.dict_nkt_after.values()))}м '
                                            f'после ремонта не равно глубине насоса '
                                            f'{self.dict_pump_shgn_depth["after"]}м')

                        self.check_data_in_pz.append(
                            f'Ошибка в карте спуска: \n Длина НКТ {sum(list(self.dict_nkt_after.values()))}м '
                            f'после ремонта не равно глубине насоса '
                            f'{self.dict_pump_shgn_depth["after"]}м')
            if self.dict_nkt_before:
                if sum(list(self.dict_nkt_before.values())) > self.current_bottom and \
                        '48' not in list(self.dict_nkt_before.keys()):
                    QMessageBox.warning(self, 'Ошибка', f'Длина НКТ {sum(list(self.dict_nkt_before.values()))}м '
                                                        f'до ремонта больше текущего забоя {self.current_bottom}м')
                    self.check_data_in_pz.append(
                        f'Ошибка в карте спуска: Длина НКТ {sum(list(self.dict_nkt_before.values()))}м '
                        f'до ремонта больше текущего забоя {self.current_bottom}м')

            if self.max_angle.get_value > 45 or 'gnkt' in self.work_plan:
                angle_true_question = QMessageBox.question(self,
                                                           'Зенитный угол',
                                                           'Зенитный угол больше 45 градусов, '
                                                           'для корректной работы необходимо '
                                                           'загрузить данные инклинометрии. Загрузить?')
                if angle_true_question == QMessageBox.StandardButton.Yes:
                    self.angle_data = WellData.read_angle_well()
                    if self.angle_data is None:
                        # self.pause_app()
                        return
            if self.dict_pump_ecn["before"] != '0' and self.dict_pump_shgn["before"] != '0':
                if self.paker_before["before"] in ['0', None, 0]:
                    self.check_data_in_pz.append(
                        f'В план заказе не указано посадка пакера при спущенной компоновке ОРД ')
            if self.dict_pump_ecn["before"] != '0' and \
                    self.dict_pump_shgn["before"] != '0':
                if self.paker_before["before"] in ['0', None, 0]:
                    self.check_data_in_pz.append(
                        f'В план заказе не указано посадка пакера при cпуске ОРД ')

            if str(self.well_number.get_value) in ['1871', '1906', '1600', '3129', '2166', '1352', '1678']:
                QMessageBox.warning(self, 'Канатные технологии', f'Скважина согласована на канатные технологии')
                self.konte_true = True

            if '0' != str(self.dict_pump_ecn['before']):
                if sum(list((self.dict_nkt_before.values()))) - self.dict_pump_ecn_depth['before'] < 10 and \
                        '48' not in list(self.dict_nkt_before.keys()):
                    QMessageBox.warning(self, 'Ошибка',
                                        f'Длина НКТ {sum(list(self.dict_nkt_before.values()))}м '
                                        f'до ремонта меньше глубины насоса '
                                        f'{self.dict_pump_ecn_depth["before"]}м')
                    self.check_data_in_pz.append(
                        f'Ошибка в карте спуска: \n Длина НКТ {sum(list(self.dict_nkt_before.values()))}м '
                        f'до ремонта меньше глубины насоса'
                        f'{self.dict_pump_ecn_depth["before"]}м')
            if '0' != str(self.dict_pump_ecn['after']):
                if sum(list((self.dict_nkt_after.values()))) - self.dict_pump_ecn_depth['after'] < 10 and \
                        '48' not in list(self.dict_nkt_before.keys()):
                    QMessageBox.warning(self, 'Ошибка',
                                        f'Длина НКТ {sum(list(self.dict_nkt_after.values()))}м '
                                        f' меньше глубины насоса {self.dict_pump_ecn_depth["after"]}м')

                    self.check_data_in_pz.append(
                        f'Ошибка в карте спуска: \n Длина НКТ {sum(list(self.dict_nkt_after.values()))}м '
                        f'до ремонта меньше глубины насоса'
                        f'{self.dict_pump_shgn_depth["after"]}м')

        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', f'Ошибка в расчетах {e}')

        if self.data_window is None:
            from data_correct import DataWindow
            self.data_window = DataWindow(self)
            self.data_window.setWindowTitle("Сверка данных")
            self.data_window.setGeometry(100, 100, 300, 400)

            self.data_window.show()
            self.pause_app()
            data_list.pause = True
            self.data_window = None

        if self.work_plan in ['krs', 'prs']:
            from data_base.config_base import connection_to_database
            from data_base.config_base import WorkDatabaseWell
            db = connection_to_database(decrypt("DB_WELL_DATA"))
            check_in_base = WorkDatabaseWell(db, self)
            tables_filter = check_in_base.get_tables_starting_with(self.well_number.get_value,
                                                                   self.well_area.get_value, 'ПР',
                                                                   self.type_kr.split(' ')[0])
            if tables_filter:
                mes = QMessageBox.question(None, 'Наличие в базе',
                                           f'В базе имеется план работ по скважине:\n {" ".join(tables_filter)}. '
                                           f'При продолжении план пересохранится, продолжить?')
                if mes == QMessageBox.StandardButton.No:
                    # self.pause_app()
                    return
        return True

    @staticmethod
    def read_angle_well():
        fname = None
        while fname is None:
            fname, _ = QtWidgets.QFileDialog.getOpenFileName(None, 'Выберите файл', '.',
                                                             "Файлы Exсel (*.xlsx);;Файлы Exсel (*.xls)")
            # Загрузка файла Excel
            wb = load_workbook(fname)
            sheet_angle = wb.active
            angle_data = []
            depth_column = ''
            row_data = ''
            angle_column = ''
            curvature_column = ''
            for index_row, row in enumerate(sheet_angle.iter_rows(min_row=1, max_row=10, values_only=True)):
                for col, value in enumerate(row):
                    if value is not None:
                        if 'глубина' in str(value).lower() and col <= 7:
                            depth_column = col
                            row_data = index_row

                        elif ('Угол, гpад' in str(value) and col <= 7) or 'зенитный' in str(value).lower():
                            angle_column = col

                        elif 'кривизна' in str(value).lower() or 'гр./10' in str(value).lower():
                            curvature_column = col

            if depth_column == '' or row_data == '' or angle_column == '' or curvature_column == '':
                mes = QMessageBox.question(None, 'Ошибка', 'Файл не корректный, необходимо загрузить корректный файл')
                if mes == QMessageBox.StandardButton.No:
                    fname = 'без загрузки'
                else:
                    fname = None
        if fname not in ['без загрузки']:
            # Вставка данных в таблицу
            for index_row, row in enumerate(sheet_angle.iter_rows(min_row=row_data, values_only=True)):
                if str(row[depth_column]).replace(',', '').replace('.', '').isdigit() \
                        and row[depth_column] and row[angle_column] and row[curvature_column]:
                    angle_data.append((float(str(row[depth_column]).replace(',', '.')),
                                       float(str(row[angle_column]).replace(',', '.')),
                                       float(str(row[curvature_column]).replace(',', '.'))))
        else:
            return []
        return angle_data


class WellPerforation(FindIndexPZ):
    def __init__(self):
        super().__init__()

    def read_well(self, begin_index, cancel_index):
        from work_py.alone_oreration import is_number, calculation_fluid_work

        self.old_version = False
        col_old_open_index = 0
        bokov_stvol = False
        osnov_stvol = False
        col_plast_index = 0
        col_vert_index = 0
        col_roof_index = 0
        col_sole_index = 0
        col_open_index = 0
        col_close_index = 0
        col_udlin_index = 0
        col_pressure_index = 0
        col_date_pressure_index = 0

        if len(self.dict_perforation) == 0:
            for row in self.ws.iter_rows(values_only=True, min_row=begin_index + 1, max_row=begin_index + 3):
                # print(row)
                for col_index, column in enumerate(row[:20]):
                    if 'по вертикали'.lower() in str(column).lower():
                        col_vert_index = col_index
                    # print(f'вер {col_index}')

                    elif 'оризонт' in str(column).lower() or 'пласт' in str(column).lower():
                        col_plast_index = col_index

                    elif 'кровля'.lower() in str(column).lower():
                        col_roof_index = col_index
                        # print(f'кров {col_index}')
                    elif 'подошва'.lower() in str(column).lower():
                        # print(f'подо {col_index}')
                        col_sole_index = col_index
                    elif 'вскрытия'.lower() in str(column).lower():
                        # print(f'вскр {col_index}')
                        col_open_index = col_index

                    elif 'удлине'.lower() in str(column).lower():
                        # print(f'удл {col_index}')
                        col_udlin_index = col_index
                    elif 'Рпл' in str(column) and 'атм' in str(column):
                        col_pressure_index = col_index
                    elif 'замера' in str(column).lower():
                        col_date_pressure_index = col_index
                    if 'отключен'.lower() in str(column).lower() and col_index < 8:
                        # print(f'октл {col_index}')
                        col_close_index = col_index
                    if 'вскрыт'.lower() in str(column).lower() and 'откл'.lower() in str(column).lower():
                        self.old_version = True
                        col_close_index = col_index
                    if "сновной" in str(column).lower():
                        osnov_stvol = True
                    if "боков" in str(column).lower():
                        bokov_stvol = True
            column_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
            if col_date_pressure_index == 0:
                col_date_pressure_index = column_index_from_string(QInputDialog.getItem(
                    self, 'Ошибка', 'Программа не смогла определить колонку в таблице ПВР где указано в дата замера',
                    column_list, 11)[0]) - 2

            if col_pressure_index == 0:
                col_pressure_index = column_index_from_string(
                    QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                         'в таблице ПВР где указано в Рпл',
                                         column_list, 10)[0]) - 2
            if col_udlin_index == 0:
                col_udlin_index = column_index_from_string(
                    QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                         'в таблице ПВР где указано в удлинение',
                                         column_list, 9)[0]) - 2
            if col_close_index == 0:
                col_close_index = column_index_from_string(
                    QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                         'в таблице ПВР где указано дата отключения',
                                         column_list, 6)[0]) - 2
            if col_open_index == 0:
                col_open_index = column_index_from_string(
                    QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                         'в таблице ПВР где указано дата вскрытия',
                                         column_list, 5)[0]) - 2
            if col_sole_index == 0:
                col_sole_index = column_index_from_string(
                    QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                         'в таблице ПВР где указано Подошва ИП',
                                         column_list, 4)[0]) - 2
            if col_roof_index == 0:
                col_roof_index = column_index_from_string(
                    QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                         'в таблице ПВР где указано кровля ИП',
                                         column_list, 3)[0]) - 2
            if col_vert_index == 0:
                col_vert_index = column_index_from_string(
                    QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                         'в таблице ПВР где указано вертикаль',
                                         column_list, 2)[0]) - 2
            if col_plast_index == 0:
                col_plast_index = column_index_from_string(
                    QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                         'в таблице ПВР где указано пласт',
                                         column_list, 1)[0]) - 2

            if osnov_stvol is True and bokov_stvol is True:
                QMessageBox.warning(self, 'ОШИБКА', 'Ошибка в определении рабочий интервалов перфорации')
                begin_index = QInputDialog.getInt(self, 'Индекс начала', 'ВВедите индекс начала рабочих интервалов ПВР',
                                                  0, 0, 300)[0] - 3

                cancel_index = QInputDialog.getInt(self, 'Индекс начала',
                                                   'ВВедите индекс окончания рабочих интервалов ПВР', 0, 0, 300)[0] - 2

            perforations_intervals = []
            try:
                row_index = ''
                for row_index, row in enumerate(
                        self.ws.iter_rows(values_only=True, min_row=begin_index + 3, max_row=cancel_index + 2)):

                    aswa = row[col_roof_index], row[col_sole_index]
                    if str(row[col_roof_index]).replace('.', '').replace(',', '').isdigit() and \
                            str(row[col_sole_index]).replace('.', '').replace(',', '').isdigit():
                        perforations_intervals.append(row)
            except Exception as e:
                QMessageBox.warning(self, 'ОШИБКА',
                                    f'Приложение не смогло определить индекс пласта в строке {e}')
                data_list.pause = True
                # self.pause_app()
                return

            for ind, row in enumerate(perforations_intervals):
                if row[col_plast_index]:
                    plast = row[col_plast_index].strip()

                if any(['проект' in str((i)).lower() or 'не пер' in str((i)).lower() for i in row]) is False and all(
                        [str(i).strip() is None for i in row]) is False and is_number(row[col_roof_index]) is True \
                        and is_number(row[col_sole_index]) is True:
                    # print(f'5 {row}')
                    if self.check_str_isdigit(row[col_vert_index]):
                        self.dict_perforation.setdefault(plast, {}).setdefault('вертикаль', []).append(
                            float(str(row[col_vert_index]).replace(',', '.')))
                    if any(['фильтр' in str(i).lower() for i in row]):
                        self.dict_perforation.setdefault(plast, {}).setdefault('отрайбировано', True)
                    else:
                        self.dict_perforation.setdefault(plast, {}).setdefault('отрайбировано', False)
                    self.dict_perforation.setdefault(plast, {}).setdefault('Прошаблонировано', False)
                    roof_int = round(float(str(row[col_roof_index]).replace(',', '.')), 1)
                    sole_int = round(float(str(row[col_sole_index]).replace(',', '.')), 1)
                    self.dict_perforation.setdefault(plast, {}).setdefault('интервал', []).append(
                        (roof_int, sole_int))
                    self.dict_perforation_short.setdefault(plast, {}).setdefault('интервал',
                                                                                 []).append((roof_int, sole_int))
                    # for interval in list(self.dict_perforation[plast]["интервал"]):
                    # print(interval)

                    self.dict_perforation.setdefault(plast, {}).setdefault('вскрытие', []).append(
                        row[col_open_index])

                    if self.old_version is False:
                        if row[col_close_index] is None or row[col_close_index] == '-':
                            self.dict_perforation.setdefault(plast, {}).setdefault('отключение',
                                                                                   False)
                            self.dict_perforation_short.setdefault(plast, {}).setdefault('отключение',
                                                                                         False)
                        else:
                            self.dict_perforation.setdefault(plast, {}).setdefault('отключение', True)
                            self.dict_perforation_short.setdefault(plast, {}).setdefault('отключение',
                                                                                         True)
                    else:
                        asdef = row[col_close_index], type(row[col_close_index])
                        if isinstance(row[col_close_index], datetime) or '/' not in str(row[col_close_index]):
                            self.dict_perforation.setdefault(plast, {}).setdefault('отключение',
                                                                                   False)
                            self.dict_perforation_short.setdefault(plast, {}).setdefault('отключение',
                                                                                         False)
                        else:
                            self.dict_perforation.setdefault(plast, {}).setdefault('отключение', True)
                            self.dict_perforation_short.setdefault(plast, {}).setdefault('отключение',
                                                                                         True)

                    if str(row[col_pressure_index]).replace(',', '').replace('.', '').isdigit() and row[col_vert_index]:
                        data_p = float(str(row[col_pressure_index]).replace(',', '.'))
                        self.dict_perforation.setdefault(plast, {}).setdefault('давление',
                                                                               []).append(
                            round(data_p, 1))
                        self.dict_perforation_short.setdefault(plast, {}).setdefault('давление',
                                                                                     []).append(
                            round(data_p, 1))
                    else:
                        self.dict_perforation.setdefault(plast, {}).setdefault('давление',
                                                                               []).append(
                            round(0, 1))
                        self.dict_perforation_short.setdefault(plast, {}).setdefault('давление',
                                                                                     []).append(
                            round(0, 1))
                    if 'давление' in self.dict_perforation[plast]:
                        self.dict_perforation[plast]['давление'] = [max(self.dict_perforation[plast]['давление'])]

                    if row[col_date_pressure_index]:
                        self.dict_perforation.setdefault(
                            plast, {}).setdefault('замер', []).append(row[col_date_pressure_index])
                    else:
                        self.dict_perforation.setdefault(
                            plast, {}).setdefault('замер', []).append(0)

                    self.dict_perforation.setdefault(plast, {}).setdefault('рабочая жидкость',
                                                                           []).append(
                        calculation_fluid_work(self, row[col_vert_index], row[col_pressure_index]))

                elif any([str((i)).lower() == 'проект' for i in row]) is True and all(
                        [str(i).strip() is None for i in row]) is False and is_number(row[col_roof_index]) is True \
                        and is_number(
                    float(
                        str(row[col_roof_index]).replace(',',
                                                         '.'))) is True:  # Определение проектных интервалов перфорации
                    roof_int = round(float(str(row[col_roof_index]).replace(',', '.')), 1)
                    sole_int = round(float(str(row[col_sole_index]).replace(',', '.')), 1)

                    if len(perforations_intervals) > ind + 1 and perforations_intervals[ind + 1][4] is None:
                        perforations_intervals[ind + 1][4] = 'проект'

                    if row[col_vert_index] is not None:
                        self.dict_perforation_project.setdefault(
                            plast, {}).setdefault('вертикаль', []).append(
                            round(float(str(row[col_vert_index]).replace(',', '.')), 1))
                        self.dict_perforation_project.setdefault(
                            plast, {}).setdefault('интервал', []).append((roof_int, sole_int))

                    if row[col_pressure_index] is not None:
                        self.dict_perforation_project.setdefault(plast, {}).setdefault('давление',
                                                                                       []).append(
                            round(self.check_str_none(row[col_pressure_index]), 1))
                    self.dict_perforation_project.setdefault(plast, {}).setdefault('рабочая жидкость',
                                                                                   []).append(
                        calculation_fluid_work(self, row[col_vert_index], row[col_pressure_index]))



            # объединение интервалов перфорации если они пересекаются
            for plast, value in self.dict_perforation.items():
                intervals = value['интервал']
                merged_segments = list()
                for roof_int, sole_int in sorted(list(intervals), key=lambda x: x[0]):

                    if not merged_segments or roof_int > merged_segments[-1][1]:
                        merged_segments.append((roof_int, sole_int))
                    else:
                        merged_segments[-1] = [merged_segments[-1][0], max(sole_int, merged_segments[-1][1])]

                self.dict_perforation[plast]['интервал'] = merged_segments

        for plast, data in self.dict_perforation.items():
            try:
                if data["вертикаль"] and data["давление"][0] != 0 and data['рабочая жидкость']:
                    bsu_data = int(float(min(data["вертикаль"])) - float(
                        max(data["давление"])) * 100 / max(data['рабочая жидкость']) / 9.81)
                    self.dict_perforation.setdefault(plast, {}).setdefault('БСУ', bsu_data)
                else:
                    self.dict_perforation.setdefault(plast, {}).setdefault('БСУ', 89)
            except:
                self.dict_perforation.setdefault(plast, {}).setdefault('БСУ', 89)
                print('Ошибка БСУ')

        self.fluid = max([max(data['рабочая жидкость']) for plast, data in self.dict_perforation.items() if data['отключение'] is False])

        if self.perforation_correct_window2 is None:
            from perforation_correct import PerforationCorrect
            self.perforation_correct_window2 = PerforationCorrect(self)
            self.perforation_correct_window2.setWindowTitle("Сверка данных перфорации")
            # self.perforation_correct_window2.setGeometry(200, 400, 100, 400)

            self.perforation_correct_window2.show()
            self.pause_app()
            data_list.pause = True
            self.perforation_correct_window2 = None
            # definition_plast_work(self)
        else:
            self.perforation_correct_window2.close()
            self.perforation_correct_window2 = None

        if len(self.dict_perforation_project) != 0:
            self.plast_project = list(self.dict_perforation_project.keys())


class WellCategory(FindIndexPZ):

    def read_well(self, begin_index, cancel_index):

        if self.end_index:
            self.read_work_data()

        if data_list.data_in_base is False:
            try:
                for row in range(begin_index, cancel_index):
                    for col in range(1, 13):
                        cell = self.ws.cell(row=row, column=col).value
                        if cell:
                            if str(cell).strip() in ['атм'] and self.ws.cell(row=row, column=col - 2).value:

                                self.category_pressure_list.append(self.ws.cell(row=row, column=col - 2).value)
                                self.category_pressure_well.append(self.ws.cell(row=row, column=col - 1).value)

                            elif str(cell).strip() in ['%', 'мг/л', 'мг/дм3', 'мг/м3', 'мг/дм', 'мгдм3']:
                                if str(cell).strip() == '%':
                                    if self.ws.cell(row=row, column=col - 2).value is None:
                                        self.category_h2s_list.append(
                                            self.ws.cell(row=row - 1, column=col - 2).value)
                                    else:
                                        self.category_h2s_list.append(
                                            self.ws.cell(row=row, column=col - 2).value)
                                    if str(self.ws.cell(row=row, column=col - 1).value).strip() in ['', '-', '0',
                                                                                                    'None'] or \
                                            'отс' in str(self.ws.cell(row=row, column=col - 1).value).lower():
                                        self.value_h2s_percent.append(0)
                                        if self.ws.cell(row=row - 1, column=col - 2).value not in ['3', 3]:
                                            self.check_data_in_pz.append(
                                                'Не указано значение сероводорода в процентах. \nСогласно п.4 '
                                                'Распоряжения от 11.04.2022г об утверждении методики расчета '
                                                'расходной нормы нейтрализатора сероводорода необходимо обеспечить '
                                                'в план-заказах на ТиКРС двух параметров по содержанию сероводороду '
                                                'объемного (в %) и массового в мг/дм3 (мг/л) \n')
                                    else:
                                        self.value_h2s_percent.append(
                                            float(str(self.ws.cell(row=row, column=col - 1).value).replace(',', '.')))
                                if str(cell).strip() in ['мг/л', 'мг/дм3', 'мг/дм', 'мгдм3']:
                                    if str(self.ws.cell(row=row, column=col - 1).value).strip() in ['', '-', '0',
                                                                                                    'None'] or \
                                            'отс' in str(self.ws.cell(row=row, column=col - 1).value).lower():
                                        self.value_h2s_mg.append(0)

                                        if self.ws.cell(row=row, column=col - 2).value not in ['3', 3]:
                                            self.check_data_in_pz.append(
                                                'Не указано значение сероводорода в мг/л \n'
                                                'Согласно п.4 '
                                                'Распоряжения от 11.04.2022г об утверждении методики расчета '
                                                'расходной нормы нейтрализатора сероводорода необходимо обеспечить '
                                                'в план-заказах на ТиКРС двух параметров по содержанию сероводороду '
                                                'объемного (в %) и массового в мг/дм3 (мг/л)')

                                    else:
                                        self.value_h2s_mg.append(
                                            float(str(self.ws.cell(row=row, column=col - 1).value).replace(',', '.')))

                                if str(cell).strip() in ['мг/м3'] and self.ws.cell(row=row - 1,
                                                                                   column=col - 1).value not in \
                                        ['мг/л', 'мг/дм3', 'мг/дм', 'мгдм3'] and \
                                        self.ws.cell(row=row + 1, column=col - 1).value not in \
                                        ['мг/л', 'мг/дм3', 'мг/дм', 'мгдм3']:
                                    if str(self.ws.cell(
                                            row=row, column=col - 1).value).strip() in ['', '-', '0', 'None'] or \
                                            'отс' in str(self.ws.cell(row=row, column=col - 1).value).lower():
                                        self.value_h2s_mg.append(0)

                                    else:
                                        self.value_h2s_mg.append(float(str(
                                            self.check_str_none(
                                                str(self.ws.cell(row=row,
                                                                 column=col - 1).value).replace(
                                                    ',', '.')))) / 1000)

                            elif str(cell).strip() == 'м3/т':

                                self.category_gaz_factor_percent.append(self.ws.cell(row=row, column=col - 2).value)
                                if 'отс' in str(self.ws.cell(row=row, column=col - 1).value) or \
                                        'None' in str(self.ws.cell(row=row, column=col - 1).value) or \
                                        '-' in str(self.ws.cell(row=row, column=col - 1).value):
                                    self.gaz_factor_percent.append(3)
                                else:
                                    self.gaz_factor_percent.append(float(
                                        str(self.ws.cell(row=row, column=col - 1).value).replace(',', '.')))
            except Exception as e:
                QMessageBox.warning(self, 'Ошибка', f'Ошибка обработки данных по категориям {e}')

            if len(self.category_h2s_list) == 0:
                QMessageBox.warning(self, 'ОШИБКА', 'Приложение не смогла найти значение '
                                                    'сероводорода в процентах')
                data_list.pause = True
                # self.pause_app()
                return

            if self.data_window is None:
                from category_correct import CategoryWindow
                self.data_window = CategoryWindow(self)
                self.data_window.setWindowTitle("Сверка данных")
                # self.data_window.setGeometry(200, 200, 200, 200)
                self.data_window.show()
                self.pause_app()
                data_list.pause = True
            else:
                self.data_window.close()
                self.data_window = None

            if len(self.value_h2s_percent) == 0:
                QMessageBox.warning(self, 'Ошибка', 'Программа не смогла найти данные по содержания '
                                                    'сероводорода в процентах')
                h2s_pr, _ = QInputDialog.getDouble(self, 'сероводород в процентах',
                                                   'Введите значение сероводорода в процентах', 0, 0, 100, 5)

                self.value_h2s_percent.append(h2s_pr)

            self.category_pressure = self.category_pressure_list[0]
            # print(f'категория по давлению {self.category_pressure}')
            self.category_h2s = self.category_h2s_list[0]
            self.category_gas_factor = self.category_gaz_factor_percent[0]

            self.thread_excel = ExcelWorker(self)

            self.without_damping, stop_app = self.thread_excel.check_well_existence(
                self.well_number.get_value, self.well_area.get_value,
                self.region)
            if stop_app:
                # self.pause_app()
                return

            try:
                category_pressure_well, categoty_h2s_well, categoty_gf, data = self.thread_excel.check_category(
                    self.well_number, self.well_area, self.region)

                if category_pressure_well:
                    if str(category_pressure_well) != str(self.category_pressure):
                        QMessageBox.warning(None, 'Некорректная категория давления',
                                            f'согласно классификатора от {data} категория скважина '
                                            f'по давлению {category_pressure_well} категории')
                        self.check_data_in_pz.append(
                            f'согласно классификатора от {data} категория скважины '
                            f'по давлению {category_pressure_well} категории\n')
                if categoty_h2s_well:
                    if str(self.category_h2s_list[0]) != str(self.category_h2s):
                        # print(str(self.category_h2s_list[0]), self.category_h2s)
                        #
                        QMessageBox.warning(None, 'Некорректная категория давления',
                                            f'согласно классификатора от {data} категория скважина '
                                            f'по сероводороду {categoty_h2s_well} категории')
                        self.check_data_in_pz.append(
                            f'согласно классификатора от {data} категория скважина '
                            f'по сероводороду {categoty_h2s_well} категории\n')

                if categoty_gf:
                    if str(categoty_gf) != str(self.category_gas_factor):
                        QMessageBox.warning(None, 'Некорректная категория давления',
                                            f'согласно классификатора от {data} категория скважина '
                                            f'по газовому фактору {categoty_gf} категории')
                        self.check_data_in_pz.append(f'согласно классификатора от {data} категория скважина '
                                                     f'по газовому фактору {categoty_gf} категории\n')
            except Exception as e:
                QMessageBox.warning(self, 'Ошибка',
                                    f'Скважина {self.well_number.get_value} '
                                    f'{self.well_area.get_value} не найдена в классификаторе. '
                                    f'Необходимо проверить соответствие данных с классификатором')

        for plast in self.dict_perforation:
            if self.dict_perforation[plast]['отключение'] is False:
                zamer_str = None
                zamer = list(filter(lambda x: x != 0, self.dict_perforation[plast]['замер']))
                if zamer:
                    zamer = zamer[0]
                    if type(zamer) != datetime:
                        if zamer:
                            zamer = str(zamer).replace(' ', '')
                        date = re.search(r'\d{2}\.\d{2}\.\d{4}', zamer)
                        if date is None:
                            date = re.search(r'\d{2}\.\d{2}\.\d{2}', zamer)
                        if date:
                            extracted_date = date.group()
                            if '202' not in extracted_date:
                                extracted_date = extracted_date[:-2] + "20" + extracted_date[6:]
                            zamer_str = datetime.strptime(extracted_date, '%d.%m.%Y').date()
                    else:
                        zamer_str = zamer.date()
                    date_now = datetime.now().date()
                    if zamer_str != None:
                        # Вычитаем даты, получая timedelta (разницу в днях)
                        difference = date_now - zamer_str

                        if self.category_pressure in [3, '3']:
                            if difference.days > 90:
                                self.check_data_in_pz.append(
                                    'Согласно требований инструкций БНД № П3-05 И-102089 ЮЛ-305 версия 2 '
                                    f'замер по пласту {plast} не соответствует регламенту '
                                    f'для скважин 3-й категории не более 3 месяцев до '
                                    f'начала ремонта')
                        elif self.category_pressure in [2, '2']:
                            if difference.days > 30:
                                self.check_data_in_pz.append(
                                    f'Согласно требований инструкций БНД № П3-05 И-102089 ЮЛ-305 версия 2  замер по '
                                    f'пласту {plast} не соответствует регламенту '
                                    f'для скважин 2-й категории не более 1 месяца до '
                                    f'начала ремонта')
                        elif self.category_pressure in [1, '1']:
                            if difference.days > 3:
                                self.check_data_in_pz.append(
                                    f'Согласно требований инструкций БНД № П3-05 И-102089 ЮЛ-305 версия 2  замер по '
                                    f'пласту {plast} не соответствует регламенту '
                                    f'для скважин 1-й категории не более 3 дней до '
                                    f'начала ремонта')

        if self.work_plan not in ['gnkt_frez', 'application_pvr',
                                  'application_gis', 'gnkt_after_grp', 'gnkt_opz', 'gnkt_bopz', 'plan_change', 'prs']:
            ase = self.cat_well_min.get_value, self.data_well_max.get_value, self.data_x_max.get_value
            # if self.work_plan == 'prs':
            #     self.data_well_max = self.data_x_max_prs

            # self.delete_rows_pz(self.ws, self.cat_well_min, self.data_well_max, self.data_x_max)
            #
            self.insert_index = self.data_well_max.get_value

        return self
