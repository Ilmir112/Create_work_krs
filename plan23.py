import openpyxl as op


def open_excel_file(fname):
    wb = op.load_workbook(fname, data_only=True)
    ws = wb['план работ']
    max_rows = ws.max_row

    lst2 = []
    for j in range(1, max_rows):
        lst3 = []
        for i in range(1, 13):
            lst3.append(ws.cell(row=j, column = i).value)
        lst2.append(lst3)

    return lst2

print(open_excel_file('ГНКТ ОПЗ.xlsx'))