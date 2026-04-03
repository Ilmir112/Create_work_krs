import sys

file_path = "main.py"

# Read the content of the file
with open(file_path, "r", encoding="utf-8") as f:
    lines = f.readlines()

new_lines = []
in_pz_worker_section = False
in_run_method = False

new_run_method_content = '''    def run(self):
        try:
            from open_pz import CreatePZ
            from data_base.work_with_base import insert_data_new_excel_file
            from work_py.correct_plan import CorrectPlanWindow
            from work_py.dop_plan_py import DopPlanWindow
            from work_py.gnkt_grp import GnktOsvWindow
            from work_py.gnkt_frez import WorkWithGnkt
            from find import FindIndexPZ
            from openpyxl import load_workbook, Workbook

            try:
                self.wb = load_workbook(self.fname, data_only=True)
                self.ws = self.wb.active

                excel_worker_instance = ExcelWorker(self.ws, self.work_plan)
                excel_worker_instance.run()
                self.data_well = excel_worker_instance.data_well
                if self.data_well is False or self.data_well is None:
                    self.error.emit("Ошибка при чтении данных скважины из Excel.")
                    return

                read_pz_instance = CreatePZ(self.data_well, self.ws, None) # No parent for now

                self.wb2_prs = Workbook()
                self.ws2_prs = self.wb2_prs.active

                if self.work_plan in ["krs", "dop_plan"]:
                    if self.wb2_prs:
                        self.ws = read_pz_instance.open_excel_file(
                            self.ws, self.work_plan, self.ws2_prs
                        )

                elif self.work_plan in ["gnkt_opz", "gnkt_after_grp", "gnkt_bopz"]:
                    pass # Placeholder: UI interactions will be in main thread
                elif self.work_plan == "gnkt_frez":
                    pass # Placeholder: UI interactions will be in main thread
                elif self.work_plan in ["prs"]:
                    if self.wb2_prs:
                        self.ws = read_pz_instance.open_excel_file(
                            self.ws, self.work_plan, self.ws2_prs
                        )

                elif self.work_plan in ["plan_change", "dop_plan_in_base"]:
                    # self.data_well is already populated by ExcelWorker
                    if self.work_plan == "plan_change":
                        self.data_well.work_plan = self.work_plan

                    elif self.work_plan == "dop_plan_in_base":
                        self.data_well.work_plan = self.work_plan

                    pass # Placeholder: UI interactions will be in main thread

                self.finished.emit(self.ws, self.data_well, self.wb2_prs, self.ws2_prs)

            except FileNotFoundError as f:
                self.error.emit(f"Ошибка при прочтении файла: {f}")
            except Exception as e:
                self.error.emit(str(e))

        except Exception as e:
            self.error.emit(str(e))
'''

for line in lines:
    if "class PZWorker(QObject):" in line:
        in_pz_worker_section = True
        new_lines.append(line)
        continue

    if in_pz_worker_section and "def run(self):" in line:
        in_run_method = True
        new_lines.append(new_run_method_content)
        continue

    if in_run_method and "except Exception as e:" in line: # This is the end of the original run method
        in_run_method = False
        continue # Skip the original except block as it's part of new_run_method_content

    if in_pz_worker_section and not in_run_method and "def " in line and "self):" in line:
        # We are inside PZWorker, but after the run method and before the next method
        # or end of class. We should skip original run method lines.
        continue

    if in_pz_worker_section and "class MyMainWindow(QMainWindow):" in line:
        in_pz_worker_section = False
        new_lines.append(line)
        continue

    if not in_run_method: # Only append lines if we are not inside the original run method
        new_lines.append(line)

with open(file_path, "w", encoding="utf-8") as f:
    f.writelines(new_lines)


