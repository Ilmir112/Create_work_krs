from copy import copy

from openpyxl.styles import Alignment
from openpyxl.utils.cell import range_boundaries, get_column_letter

import data_list


def delete_rows_pz(self, ws, cat_well_min, data_well_max, data_x_max):
    boundaries_dict = {}

    for ind, _range in enumerate(ws.merged_cells.ranges):
        boundaries_dict[ind] = range_boundaries(str(_range))

    # row_heights_top = [None, 18.0, 18, 18,None, 18.0, 18, 18,None, 18.0, 18, 18, 18.0, 18, 18, 18.0, 18, 18, 18.0, 18, 18]
    row_heights1 = [ws.row_dimensions[i + 1].height for i in range(cat_well_min.get_value, ws.max_row)]
    for key, value in boundaries_dict.items():
        ws.unmerge_cells(start_column=value[0], start_row=value[1],
                         end_column=value[2], end_row=value[3])

    # print(f'индекс удаления {1, self.cat_well_min - 1} , {data_well_max + 2, ws.max_row - data_well_max}')
    if 'prs' not in self.work_plan:
        ws.delete_rows(data_x_max.get_value, ws.max_row - data_x_max.get_value)


    ws.delete_rows(1, cat_well_min.get_value - 1)



    # print(sorted(boundaries_dict))
    data_list.row_heights = row_heights1

    for _ in range(16):
        ws.insert_rows(1, 1)

    for key, value in boundaries_dict.items():
        if value[1] <= data_well_max.get_value + 1 and value[1] >= cat_well_min.get_value:
            ws.merge_cells(start_column=value[0], start_row=value[1] + 16 - cat_well_min.get_value + 1,
                           end_column=value[2], end_row=value[3] + 16 - cat_well_min.get_value + 1)

    # print(f'{ws.max_row, len(data_list.prow_heights)}dd')
    for index_row, row in enumerate(ws.iter_rows()):  # Копирование высоты строки
        ws.row_dimensions[index_row + 17].height = data_list.row_heights[index_row - 1]




def head_ind(start, finish):
    return f'A{start}:L{finish}'

def head_ind_prs(start, finish):
    return f'A{start}:O{finish}'


def copy_true_ws(data_well, ws, ws2, head):

    for row_number, row in enumerate(ws[head]):
        for col_number, cell in enumerate(row):

            if cell.value:
                if row_number == 0:
                    if col_number > 6:
                        break
                    ws2.cell(row_number + 1, col_number + 1, cell.value)

            if 'катег' in str(cell.value).lower() and 'план' not in str(cell.value).lower():
                if data_well.work_plan not in ['krs', 'dop_plan', 'dop_plan_in_base', 'plan_change']:
                    ws2.cell(row=row_number + 1, column=col_number + 1).alignment = Alignment(wrap_text=True,
                                                                                              horizontal='left',
                                                                                              vertical='center')
            if type(cell.value) == float:
                ws2.cell(row_number + 1, col_number + 1, round(cell.value, 5))
            else:
                ws2.cell(row_number + 1, col_number + 1, cell.value)

            if cell.has_style:
                ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                ws2.cell(row_number + 1, col_number + 1).number_format = copy(
                    cell.number_format)
                ws2.cell(row_number + 1, col_number + 1).protection = copy(
                    cell.protection)
                ws2.cell(row_number + 1, col_number + 1).alignment = copy(
                    cell.alignment)
                ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(
                    cell.quotePrefix)
                ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(
                    cell.pivotButton)