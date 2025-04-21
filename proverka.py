# import openpyxl
# from openpyxl.utils.cell import coordinate_from_string
# import json
# from data_base.work_with_base import ClassifierWell
#
# def excel_to_html(excel_file):
#     """Преобразует Excel-файл с объединенными ячейками в HTML-код."""
#
#     workbook = openpyxl.load_workbook(excel_file)
#     worksheet = workbook.active
#
#     # Получаем список всех объединенных диапазонов
#     merged_ranges = worksheet.merged_cells.ranges
#
#     # Создаем HTML-код для таблицы
#     html = '''
#     <!DOCTYPE html>
#     <html>
#     <head>
#         <title>Таблица из Excel</title>
#         <style>
#             table {
#                 border-collapse: collapse;
#                 width: 100%;
#             }
#
#             th, td {
#                 border: 1px solid black;
#                 padding: 8px;
#                 text-align: center;
#             }
#
#             th {
#                 background-color: #f2f2f2;
#             }
#
#             .merged {
#                 text-align: center;
#             }
#         </style>
#     </head>
#     <body>
#         <table>
#             <thead>
#     '''
#
#     # Заголовки столбцов
#     for col_num in range(1, worksheet.max_column + 1):
#         html += f'<th>{worksheet.cell(row=1, column=col_num).value}</th>'
#
#     html += '''
#             </thead>
#             <tbody>
#     '''
#
#     # Данные
#     for row_num in range(2, worksheet.max_row + 1):
#         html += '<tr>'
#
#         for col_num in range(1, worksheet.max_column + 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#
#             # Проверяем, является ли ячейка частью объединенной
#             is_merged = False
#             for merged_range in merged_ranges:
#                 if cell.coordinate in merged_range:
#                     is_merged = True
#                     break
#
#             if is_merged:
#                 # Получаем координаты первой ячейки объединенного диапазона
#                 first_cell = coordinate_from_string(merged_range.coord[0])
#                 first_row = first_cell[1]
#                 first_col = first_cell[0]
#
#                 # Вычисляем colspan и rowspan
#                 colspan = merged_range.coord[1].split(':')[1][0] - merged_range.coord[0].split(':')[0][0] + 1
#                 rowspan = int(merged_range.coord[1].split(':')[1][1:]) - int(merged_range.coord[0].split(':')[0][1:]) + 1
#
#                 # Добавляем объединенную ячейку
#                 html += f'<td colspan="{colspan}" rowspan="{rowspan}" class="merged">{worksheet.cell(row=first_row, column=first_col).value}</td>'
#             else:
#                 # Добавляем обычную ячейку
#                 html += f'<td>{cell.value}</td>'
#
#         html += '</tr>'
#
#     html += '''
#             </tbody>
#         </table>
#     </body>
#     </html>
#     '''
#
#     return html
import openpyxl

ade='15.СОГЛАСНО ПИСЬМА ООО "БАШНЕФТЬ-ДОБЫЧА" № 02-03-05/0134 ОТ 30.09.2024. ПРИ ОСВОЕНИИ (СВАБИРОВАНИИ ' \
    'ИЛИ КОМПРЕССИРОВАНИИ) СКВАЖИН, СОДЕРЖАЩИХ СЕРОВОДОРОД ПРИМЕНЯТЬ МЕРЫ БЕЗОПАСНОСТИ: \n'\
             '1. Персонал должен находиться с наветренной стороны от устья скважины и от емкости, в которую поступает '\
             'жидкость и, по возможности, на возвышенном месте. \n'\
             '2. При выходе скважинной жидкости в технологическую емкость контроль воздушной среды проводить '\
             'работниками бригады ТКРС с применением СИЗОД каждые 30 мин. \n'\
             '3. Обеспечить добавление нейтрализатора сероводорода в приемную емкость до начала отбора жидкости. '\
             'Объем нейтрализатора должен соответствовать максимальному объему заполнения емкости. \n'\
             '4. Обвязка емкости с выкидной линией должна быть выполнена таким образом, чтобы скважинная жидкость '\
             'поступала в емкость под уровень жидкости, обработанной нейтрализатором сероводорода. Емкость должна '\
             'быть с закрытым верхом для исключения объемного выхода сероводорода в окружающую среду. \n'\
             '5. Перед откачкой технологической емкости провести дополнительный контроль воздушной среды в СИЗОД. \n'\
             '6. Работниками геофизической партии должен быть обеспечен контроль воздушной среды в воздухе рабочей '\
             'зоны (внутри ПКС) каждые 30 мин с фиксацией в журнале ГВС. '
print(len(ade))

wb = openpyxl.load_workbook('шаблон обновленная отсыпка песка.xlsx')
ws = wb.active

dict_tel = {}
colii2 = []
for row_ind, row in enumerate(ws.iter_rows(values_only=True, min_row=1, max_row=35, max_col=12)):
    colii = []
    for col in row:
        colii.append(col)
        # if col == 'Начальник':
        #     dict_tel.setdefault(
        #         'ОАО "Башнефть"', {}).setdefault(
        #         "регион", {}).setdefault(
        #         row[0], {}).setdefault("ЦДНГ", {}).setdefault(row[1], {}).setdefault("Начальник", {
        #         "post": row[2], "surname": row[3], "telephone": row[4], "email": row[5]})
        # if col == 'Заместитель начальника':
        #     dict_tel.setdefault(
        #         'ОАО "Башнефть"', {}).setdefault(
        #         "регион", {}).setdefault(
        #         row[0], {}).setdefault("ЦДНГ", {}).setdefault(row[1], {}).setdefault("Заместитель начальника", {
        #         "post": row[2], "surname": row[3], "telephone": row[4], "email": row[5]})
        # if col == 'Ведущий геолог':
        #     dict_tel.setdefault(
        #         'ОАО "Башнефть"', {}).setdefault(
        #         "регион", {}).setdefault(
        #         row[0], {}).setdefault("ЦДНГ", {}).setdefault(row[1], {}).setdefault("Ведущий геолог", {
        #         "post": row[2], "surname": row[3], "telephone": row[4], "email": row[5]})
        # if col == 'Ведущий инженер-технолог':
        #     dict_tel.setdefault(
        #         'ОАО "Башнефть"', {}).setdefault(
        #         "регион", {}).setdefault(
        #         row[0], {}).setdefault("ЦДНГ", {}).setdefault(row[1], {}).setdefault("Ведущий инженер-технолог", {
        #         "post": row[2], "surname": row[3], "telephone": row[4], "email": row[5]})
    colii2.append(colii)

# for row in colii2:
#     print(row)
print(colii2, sep='\n')




