
from openpyxl.styles import Border, Side, Font,  Alignment
from work_py.calculate_work_parametrs import volume_work, volume_well_pod_nkt_calculate


from find import FindIndexPZ


class CalculateH2s:
    def __init__(self, data_well: FindIndexPZ):
        self.data_well = data_well

    def calc_h2s(self, ws3, plast):
        # print(self.data_well.dict_nkt_before)
        nkt_1 = list(self.data_well.dict_nkt_before.keys())[0]
        nkt_1_l = self.data_well.dict_nkt_before[nkt_1]

        try:
            nkt_2 = list(self.data_well.dict_nkt_before.keys())[1]
            nkt_2_l = self.data_well.dict_nkt_before[nkt_2]
        except Exception:
            nkt_2 = 0
            nkt_2_l = 0

        try:
            nkt_3 = list(self.data_well.dict_nkt_before.keys())[2]
        except Exception:
            nkt_3 = 0
        try:
            sucker_rod_l_25 = self.data_well.dict_sucker_rod['25']
        except Exception:
            sucker_rod_l_25 = 0
        try:
            sucker_rod_l_22 = self.data_well.dict_sucker_rod['22']
        except Exception:
            sucker_rod_l_22 = 0
        try:
            sucker_rod_l_19 = self.data_well.dict_sucker_rod['19']
        except Exception:
            sucker_rod_l_19 = 0
        self.data_well.fluid_work = 'EVASORB'
        if 'EVASORB' in self.data_well.fluid_work:
            self.data_well.type_absorbent = 'EVASORB марки 121'
        elif 'ХИМТЕХНО' in self.data_well.fluid_work:
            self.data_well.type_absorbent = 'ХИМТЕХНО 101 Марка А'
        elif 'СНПХ-1200' in self.data_well.fluid_work:
            self.data_well.type_absorbent = 'СНПХ-1200'


        if self.data_well.type_absorbent == 'EVASORB марки 121':
            koeff_zapas = 1.05
        else:
            koeff_zapas = 1
        if self.data_well.column_additional:
            shoe_column = self.data_well.head_column_additional.get_value
        else:
            shoe_column = self.data_well.shoe_column.get_value
        SNPKH = [
            [None, f'Расчет расхода нейтрализатора сероводорода на модификацию раствора глушения* по пласту {plast}',
             None, None, None, None],
            [None, 'Масса нейтрализатора сероводорода - по результатам расчета*', None, None, None, None],
            [None, 'Объем раствора глушения - по объему скважины от устья до забоя', None, None, None, None],
            [None, None, None, None, None, None],
            [None, '№', 'Параметр', None, 'Результат расчета', None],
            [None, 1, 'Параметры скважины',
             None, f'{self.data_well.well_number.get_value} {self.data_well.well_area.get_value}', None],
            [None, 1.1, 'Забой скважины', 'м',
             round(float(self.data_well.bottom_hole_artificial.get_value), 1), 'формула'],
            [None, 1.2, 'текущий забой', 'м',
             round(float(self.data_well.bottom), 1), 'ввод'],
            [None, 1.3, 'Диаметр ЭК (ступень 1 верхняя)', 'мм', int(self.data_well.column_diameter.get_value), 'ввод'],
            [None, '1.3.1.', 'Толщина стенки ЭК (ступень 1 верхняя)', 'мм',
             round(float(self.data_well.column_wall_thickness.get_value), 1), 'ввод'],
            [None, '1.3.2.',
             'Длина подвески ЭК (ступень 1 верхняя)', 'м', int(shoe_column), 'ввод'],
            [None, '1.3.3.',
             'Диаметр ЭК (ступень 2 хвостовик)', 'мм', int(self.data_well.column_additional_diameter.get_value),
             'ввод'],
            [None, '1.3.4.', 'Толщина стенки ЭК (ступень 2 хвостовик)', 'м',
             float(self.data_well.column_additional_wall_thickness.get_value),
             'ввод'],
            [None, '1.3.5.', 'Длина подвески ЭК (ступень 2 хвостовик)', 'м',
             abs(int(self.data_well.head_column_additional.get_value) - int(self.data_well.shoe_column_additional.get_value)), 'ввод'],
            [None, '1.3.6.',
             'Глубина "головы" (ступень 2 хвостовик)', 'м', int(self.data_well.head_column_additional.get_value),
             'ввод',
             ],
            [None, '1.3.7.',
             'Глубина "башмака" (ступень 2 хвостовик)', 'м', int(self.data_well.shoe_column_additional.get_value),
             'формула'],
            [None, None, None, None, None, None],
            [None, 2, 'Параметры ГНО', None, None, None, None],
            [None, '2.1.', 'Общая глубина подвески НКТ', 'м', '=E22+E25', 'формула'],
            [None, '1.4.1', 'Толщина стенки подвески НКТ (ступень 1 верхняя)', 'мм', 5.5, 'ввод'],
            [None, '1.4.2', 'Внешний диаметр подвески НКТ (ступень 1 верхняя)', 'мм', nkt_1, 'ввод'],
            [None, '1.4.3', 'Длина подвески НКТ (ступень 1 верхняя)', 'м', nkt_1_l, 'ввод'],
            [None, '1.4.4', 'Толщина стенки подвески НКТ (ступень 2 нижняя)', 'мм', 0, 'ввод'],
            [None, '1.4.5', 'Внешний диаметр подвески НКТ (ступень 2 нижняя)', 'мм', nkt_2, 'ввод'],
            [None, '1.4.6', 'Длина подвески НКТ (ступень 2 нижняя)', 'м', nkt_2_l, 'ввод'],
            [None, '2.2', 'Общая глубина подвески штанг', 'м', '=E27+E28+E29', 'формула'],
            [None, '2.2.1.', 'Длина штанг 25 мм', 'м', sucker_rod_l_25, 'ввод'],
            [None, '2.2.2.', 'Длина штанг 22 мм', 'м', sucker_rod_l_22, 'ввод'],
            [None, '2.2.3.', 'Длина штанг 19 мм', 'м', sucker_rod_l_19, 'ввод'],
            [None, 3, 'Расчеты емкости', None, None, None, None],
            [None, 3.1,
             'Удельный  внутренний объем ЭК', 'дм3/м', "=ROUND(10*3.14*((E9-E10*2)*0.01)^2/4, 2)", 'формула'],
            [None, 3.2, 'Удельный  внутренний объем хвостовика', 'дм3/м', '=ROUND(10*3.14*((E12-E13*2)*0.01)^2/4,2)',
             'формула'],
            [None, 3.3, 'Объем жидкости под ГНО, в т.ч.:', 'м3',
             '=ROUND(IF(E14=0,3.14*(E9-E10*2)^2/4/1000*(E8-E19)/1000,IF(E19<E11,(3.14*(E12-E13*2)^2/4/1000*(E8-E15)/1000)+(3.14*(E9-E10*2)^2/4/1000*(E11-E19)/1000),(3.14*(E12-E13*2)^2/4/1000*(E8-E19)/1000))),2)',
             'формула'],
            [None, '3.3.1.', 'Объем скважины', 'м3',
             '=ROUND(if(E14=0,3.14*(E9-E10*2)^2/4/1000*(E8)/1000,(3.14*(E12-E13*2)^2/4/1000*(E8-E15)/1000)+(3.14*(E9-E10*2)^2/4/1000*(E15)/1000)),2)',
             'формула'],
            [None, '3.3.1.', 'Удельное водоизмещение подвески НКТ (ступень 1 верхняя)', 'дм3/м',
             '=ROUND(10*3.14*((E21*0.01)^2-(E21*0.01-E20*2*0.01)^2)/4,2)',
             'формула'],
            [None, '3.3.2.', 'Удельное водоизмещение подвески НКТ (ступень 2 нижняя)', 'дм3/м',
             '=ROUND(10*3.14*((E24*0.01)^2-(E24*0.01-E23*2*0.01)^2)/4,2)', 'формула', ],
            [None, '3.3.3.', 'Водоизмещение подвески НКТ (объем жидкости притока при СПО)', 'м3/СПО',
             '=ROUND((E22*E35/1000) +(E25*E36/1000),2)', 'формула'],
            [None, '3.3.4.', 'Водоизмещение подвески штанг (объем жидкости притока при СПО)', 'м3/СПО',
             '=ROUND((10*3.14*((25*0.01)^2/4)*E27/1000)+(10*3.14*((22*0.01)^2/4)*E28/1000)+(10*3.14*((19*0.01)^2/4)*E29/1000),2)',
             'формула'],
            [None, None, None, None, None, None],
            [None, 4, 'Параметры добываемой жидкости и газа', None, None, None, None],
            [None, 4.1, 'Газосодержание нефти',
             'м3/тонну', self.data_well.dict_category[plast]['по газовому фактору'].data, 'ввод'],
            [None, 4.2,
             'Содержание сероводорода в газе (по данным проекта разработки)', '% (об)',
             self.data_well.dict_category[plast]['по сероводороду'].data_percent,
             'ввод'],
            [None, 4.3, 'Обводенность продукции', '% (масс.)', self.data_well.percent_water, 'ввод'],
            [None, 4.4, 'Содержание сероводорода в пластовом флюиде (устьевая проба, вода+нефть)', 'мг/дм3',
             self.data_well.dict_category[plast]['по сероводороду'].data_mg_l, 'ввод'],
            [None, 4.5, 'Плотность воды', 'г/см3', 1.17, 'ввод'],
            [None, 4.6, 'Плотность нефти', 'г/см3', 0.9, 'ввод'],
            [None, None, None, None, None, None],
            [None, 5,
             'Расчет массы сероводорода в жидкости притока (в объеме водоизмещения подвески НКТ и штанг)', None,
             None, None],
            [None, 5.1, 'Масса нефти ', 'т', '=ROUND((E37+E38)*(100-E43)*E46/100,2)', 'формула'],
            [None, 5.2, 'Объем сероводорода, м3', 'м3', '=ROUND(E41*E49*E42/100,3)', 'формула'],
            [None, 5.3, 'Масса сероводорода в нефти (выделяющаяся в газовую фазу при снижении давления)', 'г',
             '=(34*E50*1000/22.14)', 'формула'],
            [None, 5.4, 'Масса сероводорода в жидкости (остаточная растворенная часть)', 'г', '=ROUND((E37+E38)*E44,2)',
             'формула'],
            [None, None, None, None, None, None],
            [None, 6, 'Расчет массы сероводорода в поднасосной жидкости (в объеме скважины под ГНО)', None, None, None],
            [None, 6.1, 'Масса нефти ', 'т', '=ROUND(E33*(100-E43)*E46/100,2)    ', 'формула'],
            [None, 6.2, 'Объем сероводорода, м3', 'м3', '=ROUND(E41*E55*E42/100,3)   ', 'формула'],
            [None, 6.3, 'Масса сероводорода в нефти (доля, выделяющаяся в газовую фазу при снижении давления)', 'г',
             '=ROUND((34*E56*1000/22.14),0)  ', 'формула'],
            [None, 6.4, 'Масса сероводорода в жидкости (остаточная растворенная часть)', 'г', '=ROUND(E33*E44,0)',
             'формула'],
            [None, None, None, None, None, None],
            [None, 7, 'Масса сероводорода общая', 'г', '=ROUND(E51+E52+E57+E58,0)', 'формула'],
            [None, None, None, None, None, None],
            [None, 8, 'Параметр нейтрализатора сероводорода', None, None, None],
            [None, 8.1, 'Емкость реагента по сероводороду (определяется по результатам ЛИ конкретной марки)', 'г/г H2S',
             24, 'ввод'],
            [None, 8.2, 'Плотность товарной формы (марки) реагента (по ТУ)', 'г/см3', 1.065, 'ввод'],
            [None, None, None, None, None, None],
            [None, 9, 'Результат расчета расхода нейтрализатора сероводорода', None, None, None],
            [None, 9.1, 'Расчетная масса реагента', 'кг', '=ROUND(E63*E60/1000,3)', 'формула'],
            [None, 9.2, 'Коэффициент запаса по реагенту (решение ОГ)', 'крат', koeff_zapas, 'ввод'],
            [None, 9.3, 'Масса реагента с запасом', 'кг', '=E67*E68', 'формула', None, None, None],
            [None, None, None, None, None, None],
            [None, 10, 'Удельный расход нейтрализатора сероводорода в растворе глушения', None, None, None],
            [None, 10.1, 'Удельный массовый расход нейтрализатора сероводорода ', 'кг/м3', '=ROUND(E69/E34,3)', 'формула'],
            [None, 10.2, 'Удельный объемный расход нейтрализатора сероводорода ', 'л/м3', '=ROUND(E72/E64,3)', 'формула'],
            ['Примечание: * - расчет приблизительный, поскольку нет информации по содержанию H2S в каждой фазе '
                '(вода, нефть, газ) в одинаковых условиях',
                None, None, None, None, None, ],
            [None, None, None, None, None, None],
            [None, 'Ячейки для заполнения данными параметров скважины', None, None, None, None],
            [None, 'По данным конструкции скважины', None, None, None, None],
            [None, None, None, None, None, None],
            [None, None, None, None, None, None],
            [None, 'По данным лабораторных исследований марки нейтрализатора сероводорода', None, None, None, None,
             None],
        ]

        max_row_h2s = len(SNPKH)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ws3.column_dimensions['B'].width = 15
        ws3.column_dimensions['C'].width = 80
        ws3.column_dimensions['D'].width = 25
        ws3.column_dimensions['f'].width = 30
        ws3.column_dimensions['e'].width = 25

        for row in range(1, max_row_h2s):
            for col in range(1, 7):
                ws3.cell(row=row, column=col).value = SNPKH[row - 1][col - 1]

                if 5 <= row <= 73 and 1 < col < 6:
                    ws3.cell(row=row, column=col).border = thin_border
                    ws3.cell(row=row, column=col).font = Font(name='Arial', size=13, bold=True)
                    ws3.cell(row=row, column=col).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                        vertical='center')

        ws3.hide = True
        ws3.page_setup.fitToPage = True
        ws3.page_setup.fitToHeight = False

        ws3.page_setup.fitToWidth = True
        ws3.print_area = 'A1:F77'
def calv_h2s(self, category_h2s, h2s_mg, h2s_pr):
    if '2' in str(category_h2s) or '1' in str(category_h2s):
        if self.data_well.type_absorbent == 'EVASORB марки 121':
            koeff_zapas = 1.05
        else:
            koeff_zapas = 1

        volume_pod_nkt = volume_well_pod_nkt_calculate(self.data_well)
        udel_vodoiz_nkt = 0
        volume_well = volume_work(self.data_well)
        vodoiz_nkt = 0

        for nkt_key, nkt_values in self.data_well.dict_nkt_before.items():
            if '73' in nkt_key:
                nkt_1 = 73
                nkt_width = 5.5
            elif '60' in nkt_key:
                nkt_1 = 60
                nkt_width = 5
            elif '48' in nkt_key:
                nkt_1 = 48
                nkt_width = 4
            elif '89' in nkt_key:
                nkt_1 = 89
                nkt_width = 6.5

            vodoiz_nkt = round(10 * 3.14 * ((nkt_1 * 0.01) ** 2 - (nkt_1 * 0.01 - nkt_width * 2 * 0.01) ** 2) / 4, 2)
            udel_vodoiz_nkt += vodoiz_nkt*nkt_values/1000

        sucker_rod_l_25 = 0
        sucker_rod_l_22 = 0
        sucker_rod_l_19 = 0

        for sucker_key, sucker_value in self.data_well.dict_sucker_rod.items():
            if '25' in sucker_key:
                sucker_rod_l_25 = sucker_value
            elif '22' in sucker_key:
                sucker_rod_l_22 = sucker_value
            elif '19' in sucker_key:
                sucker_rod_l_19 = sucker_value

        vodoiz_sucker = (10 * 3.14 * ((25 * 0.01) ** 2 / 4) * sucker_rod_l_25 / 1000) + (
                    10 * 3.14 * ((25 * 0.01) ** 2 / 4) * sucker_rod_l_22 / 1000) + (
                                    10 * 3.14 * ((25 * 0.01) ** 2 / 4) * sucker_rod_l_19 / 1000)

        oil_mass = round(float(udel_vodoiz_nkt * (100 - self.data_well.percent_water) * 0.9 / 100), 2)

        try:
            volume_h2s = self.data_well.gaz_factor_percent[0] * oil_mass * (float(h2s_pr)) / 100
        except Exception:
            self.data_well.gaz_factor_percent = [11]
            volume_h2s = self.data_well.gaz_factor_percent[0] * oil_mass * (float(h2s_pr)) / 100

        h2s_mass_in_oil = round(34 * volume_h2s * 1000 / 22.14, 0)

        h2s_mass_in_water = round(float(vodoiz_sucker + vodoiz_nkt) * h2s_mg, 0)

        mass_oil_pog_gno = volume_pod_nkt * (100 - self.data_well.percent_water) * 0.9 / 100
        h2s_volume_pod_gno = mass_oil_pog_gno * self.data_well.gaz_factor_percent[0] * h2s_pr / 100
        mass_h2s_gas = round(34 * h2s_volume_pod_gno * 1000 / 22.14, 0)
        mass_h2s_water = round(volume_pod_nkt * h2s_mg, 0)

        mass_h2s_all = h2s_mass_in_water + h2s_mass_in_oil + mass_h2s_gas + mass_h2s_water

        emk_reag = 24
        plotn_reag = 1.065
        raschet_mass = mass_h2s_all * emk_reag / 1000
        mass_reag_s_zapas = raschet_mass * koeff_zapas
        # print(f'mass{mass_reag_s_zapas}')
        udel_mas_raskhod = mass_reag_s_zapas / volume_well / plotn_reag
        # print(udel_mas_raskhod)

        if udel_mas_raskhod <= 0.01:
            udel_mas_raskhod = 0.01
        return round(udel_mas_raskhod, 3)
    else:
        return 0
