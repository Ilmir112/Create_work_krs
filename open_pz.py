import json
from abc import ABC, abstractmethod
from typing import Optional, cast

from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet

import data_list
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

from cdng import events_gnvp, add_itog, events_gnvp_gnkt
from main import MyMainWindow
from find import FindIndexPZ
from block_name import curator_sel, pop_down, current_datetime
from work_py.dop_plan_py import DopPlanWindow


class WorkWithPZ(ABC):
    @abstractmethod
    def open_excel_file(self, fname):
        pass


class PzInDatabase(WorkWithPZ):
    def __init__(self, parent=None):
        super().__init__()
        self.data_well = parent
        self.ws = getattr(parent, "ws", None) if parent else None

    def open_excel_file(self, ws: Worksheet, work_plan: str, ws2: Worksheet = None):  # type: ignore[override]
        contractor = "РН"
        if data_list.contractor and "Ойл" in data_list.contractor:
            contractor = "ОЙЛ"
        elif data_list.contractor and "РН" in data_list.contractor:
            contractor = "РН"

        if self.data_well is None:
            return ws
        if self.data_well.work_plan == "plan_change":
            DopPlanWindow.extraction_data(
                self,  # type: ignore[arg-type]
                str(self.data_well.well_number.get_value) + " " +
                self.data_well.well_area.get_value + " " + "krs" + " " + contractor, 1)
            if self.ws is not None:
                plan_idx = data_list.plan_correct_index.get_value  # type: ignore[union-attr]
                self.ws.delete_rows(int(plan_idx), self.ws.max_row)
            return self.ws  # type: ignore[return-value]


class CreatePZ(MyMainWindow):
    def __init__(self, data_well: FindIndexPZ, ws: Worksheet, parent=None):
        super(CreatePZ, self).__init__()
        self.wb = parent.wb if parent is not None else None
        self.ws = ws
        self.data_well = data_well

    @staticmethod
    def copy_data_excel_in_excel(source_sheet, target_sheet, start_row, end_row, start_col, end_col, target_start_row):
        col_width = [source_sheet.column_dimensions[get_column_letter(i + 1)].width for i in range(0, 15)]
        # Копируем данные, стили и размеры ячеек
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = source_sheet.cell(row=row, column=col)
                target_cell = target_sheet.cell(row=target_start_row + row - start_row + 1, column=col)
                if cell.value not in ['Ведущий технолог Подрядной организации', 'ф.и.о.', '(Подпись)', '(чч.мм.гг)']:
                    target_cell.value = cell.value

                # Копируем стиль по отдельным свойствам
                if cell.font:
                    target_cell.font = styles.Font(name=cell.font.name,
                                                   size=cell.font.size,
                                                   bold=cell.font.bold,
                                                   italic=cell.font.italic,
                                                   vertAlign=cell.font.vertAlign,
                                                   underline=cell.font.underline,
                                                   strike=cell.font.strike,
                                                   color=cell.font.color)
                if cell.fill:
                    target_cell.fill = styles.PatternFill(fill_type=cell.fill.fill_type,
                                                          fgColor=cell.fill.fgColor,
                                                          bgColor=cell.fill.bgColor)
                if cell.border:
                    target_cell.border = styles.Border(left=cell.border.left,
                                                       right=cell.border.right,
                                                       top=cell.border.top,
                                                       bottom=cell.border.bottom,
                                                       diagonal=cell.border.diagonal,
                                                       diagonal_direction=cell.border.diagonal_direction,
                                                       outline=cell.border.outline,
                                                       vertical=cell.border.vertical,
                                                       horizontal=cell.border.horizontal)
                if cell.number_format:
                    target_cell.number_format = cell.number_format
                if cell.protection:
                    target_cell.protection = styles.Protection(locked=cell.protection.locked,
                                                               hidden=cell.protection.hidden)
                    # Копируем выравнивание
                    if cell.alignment:
                        target_cell.alignment = styles.Alignment(horizontal=cell.alignment.horizontal,
                                                                 vertical=cell.alignment.vertical,
                                                                 text_rotation=cell.alignment.text_rotation,
                                                                 wrap_text=cell.alignment.wrap_text,
                                                                 shrink_to_fit=cell.alignment.shrink_to_fit,
                                                                 indent=cell.alignment.indent)

            # Копирование высоты строки
            target_sheet.row_dimensions[target_start_row + row - start_row + 1].height = \
                source_sheet.row_dimensions[row].height

        for col_ind in range(1, 15):
            target_sheet.column_dimensions[get_column_letter(col_ind)].width = col_width[col_ind - 1]

        # Копируем объединение ячеек

        for merged_range in source_sheet.merged_cells.ranges:
            if merged_range.bounds[0] >= start_col and merged_range.bounds[2] <= end_col and \
                    merged_range.bounds[1] >= start_row and merged_range.bounds[3] <= end_row:
                target_sheet.merge_cells(start_row=target_start_row + merged_range.bounds[1] - start_row + 1,
                                         start_column=merged_range.bounds[0],
                                         end_row=target_start_row + merged_range.bounds[3] - start_row + 1,
                                         end_column=merged_range.bounds[2])



    def append_podpisant_up(self, ws):
        razdel = self.work_podpisant_list(self.data_well.region, data_list.contractor)
        for i in range(1, len(razdel)):  # Добавлением подписантов на вверху
            for j in range(1, 13):
                if razdel[i - 1][j - 1]:
                    ws.cell(row=i, column=j).value = razdel[i - 1][j - 1]
                    ws.cell(row=i, column=j).font = Font(name='Arial Cyr', size=13, bold=True)
                    ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                   vertical='center')

            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
            ws.merge_cells(start_row=i, start_column=8, end_row=i, end_column=12)
            if len(str(razdel[i][1])) > 50 or len(str(razdel[i][7])):
                ws.row_dimensions[i+1].height = 33
            else:
                ws.row_dimensions[i+1].height = 20

    def insert_events_gnvp(self, ws, dict_events_gnvp, merge_count=0):
        # if work_plan != 'dop_plan':
        text_width_dict = {30: (0, 100), 40: (101, 200), 50: (201, 300), 70: (301, 400), 80: (401, 500),
                           100: (501, 600), 120: (601, 700), 130: (701, 800), 140: (801, 900),
                           200: (901, 1500), 270: (1500, 2300)}

        # Устанавливаем параметры границы
        red = 'FF0000'  # Красный цвет в формате HEX
        thin_border = Border(left=Side(style='thin', color=red),
                             right=Side(style='thin', color=red),
                             top=Side(style='thin', color=red),
                             bottom=Side(style='thin', color=red))
        max_row = ws.max_row


        self.data_well.insert_index = max_row + 1
        # Устанавливаем красный цвет для текста
        red_font = Font(name='Arial Cyr', size=13, color='FF0000', bold=True)
        for i in range(self.data_well.insert_index,
                       self.data_well.insert_index + len(dict_events_gnvp[self.data_well.work_plan])):
            for col in range(12):
                data = ws.cell(row=i, column=col + 1)
                data.border = thin_border
                data.value = dict_events_gnvp[self.data_well.work_plan][i - self.data_well.insert_index][col]

                ws.cell(row=i, column=col + 1).font = Font(name='Arial Cyr', size=13, bold=False)
            data_2 = ws.cell(row=i, column=3).value
            data_1 = ws.cell(row=i, column=2).value

            if 'Мероприятия' in str(data_1):
                ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=12)
                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                               vertical='center')

                ws.cell(row=i, column=2).font = Font(name='Arial Cyr', size=13, bold=True)
            elif 'При работе с вертлюгами обеспечить' in str(data_2) \
                    or 'На основании приказа' in str(data_2) \
                    or 'Согласно мероприятий по снижению а' in str(data_2) \
                    or 'Во время нештатных ' in str(data_2) \
                    or 'Для предотвращения падения ' in str(data_2) \
                    or 'После герметизации устья' in str(data_2) \
                    or 'При свинчивании и развинчивании' in str(data_2) \
                    or 'Сборку фрезерующего, ' in str(data_2) \
                    or 'При нулевых и отрицательных' in str(data_2):
                ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=11)
                ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                               vertical='center')
                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='left',
                                                               vertical='center')
                ws.cell(row=i, column=12).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                vertical='center')
                ws.cell(row=i, column=3).font = Font(name='Arial Cyr', size=13, bold=True)
                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                               vertical='center')
            else:
                ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=11)
                ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                               vertical='center')

                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                               vertical='center')
                ws.cell(row=i, column=12).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                vertical='center')
                ws.cell(row=i, column=3).font = Font(name='Arial Cyr', size=13, bold=False)

            if 'ВЫ ДОЛЖНЫ ОТКАЗАТЬСЯ' in str(data_2):
                ws.cell(row=i, column=3).font = red_font

            if data_2:
                text = ws.cell(row=i, column=3).value
                text1 = ws.cell(row=i, column=2).value
                for key, value in text_width_dict.items():
                    text_length = len(text)
                    if value[0] <= text_length <= value[1]:
                        if '\n' in text:
                            row_dimension_value = int(len(text) / 4 + text.count('\n') * 3)
                        else:
                            row_dimension_value = int(len(text) / 4)
                        ws.row_dimensions[i].height = row_dimension_value

        self.data_well.insert_index += len(dict_events_gnvp[self.data_well.work_plan]) - 1

        ws.row_dimensions[2].height = 30

    def open_excel_file(self, ws: Worksheet, work_plan: str, ws2: Optional[Worksheet] = None):

        # print(f' индекс вставки ГНВП{self.data_well.insert_index}')
        dict_events_gnvp = {}
        dict_events_gnvp['krs'] = events_gnvp(self, data_list.contractor)
        dict_events_gnvp['gnkt_opz'] = events_gnvp_gnkt(self)
        dict_events_gnvp['gnkt_bopz'] = events_gnvp_gnkt(self)
        dict_events_gnvp['dop_plan'] = events_gnvp(self, data_list.contractor)
        dict_events_gnvp['prs'] = events_gnvp(self, data_list.contractor)

        if work_plan not in ['application_pvr', 'application_gis', 'gnkt_bopz', 'gnkt_opz', 'gnkt_after_grp',
                             'gnkt_frez']:
            if work_plan != 'plan_change':
                for row_ind, row in enumerate(ws.iter_rows(values_only=True, max_col=13)):
                    ws.row_dimensions[row_ind].hidden = False
                    if 'ПЛАН РАБОТ' in str(row[1]) \
                            and work_plan == 'dop_plan':
                        ws.cell(row=row_ind + 1, column=2).value = (  # type: ignore[assignment]
                            f'ДОПОЛНИТЕЛЬНЫЙ ПЛАН РАБОТ № {self.data_well.number_dp}')
                        aswdw = ws.cell(row=row_ind + 1, column=2).value
                    elif 'План-заказ' in str(row[1]):
                        if work_plan != 'dop_plan':
                            ws.cell(row=row_ind + 1, column=2).value = 'ПЛАН РАБОТ'  # type: ignore[assignment]
                        else:
                            ws.cell(row=row_ind + 1, column=2).value = (  # type: ignore[assignment]
                                f'ДОПОЛНИТЕЛЬНЫЙ ПЛАН РАБОТ № {self.data_well.number_dp}')

            if self.data_well.work_plan not in ['gnkt_frez', 'application_pvr',
                                                'application_gis', 'gnkt_after_grp', 'gnkt_opz', 'gnkt_bopz',
                                                'plan_change', 'prs']:
                if ws2 is None:
                    return ws
                # print(f'план работ {self.data_well.work_plan}')
                self.append_podpisant_up(cast(Worksheet, ws2))

                self.copy_data_excel_in_excel(
                    ws, ws2, self.data_well.cat_well_min.get_value, self.data_well.data_well_max.get_value, 1, 16,
                    ws2.max_row + 1)

                # self.copy_data_excel_in_excel(
                #     ws, ws2, self.data_well.data_fond_min.get_value, self.data_well.data_well_max.get_value, 1, 16,
                #     ws2.max_row + 1)

                self.insert_events_gnvp(ws2, dict_events_gnvp)

                self.copy_data_excel_in_excel(
                    ws, ws2, self.data_well.data_x_min.get_value, self.data_well.data_x_max.get_value, 1, 16,
                    ws2.max_row + 1)

                # Для доп. плана КРС добавляем блок заголовков "Порядок работы" / "п/п" в конец ws2,
                # если его там ещё нет. С границами как у основной таблицы.
                if self.data_well.work_plan in ["dop_plan", 'dop_plan_in_base']:
                    header_exists = False
                    for row in range(1, ws2.max_row + 1):
                        cell = ws2.cell(row=row, column=2)  # колонка B
                        if cell.value and str(cell.value).strip().lower() == "порядок работы":
                            header_exists = True
                            break

                    if not header_exists:
                        krs_begin = [
                            [None, 'Порядок работы', None, None, None, None, None, None, None, None, None, None, None,
                             None, None, None],
                            [None, 'п/п', 'Наименование работ', None, None, None, None, None, None, None,
                             'Ответственный',
                             'Нормы времени \n мин/час.'],
                        ]
                        start_row = ws2.max_row + 1
                        for i, row_vals in enumerate(krs_begin):
                            for j, val in enumerate(row_vals, start=1):
                                cell = ws2.cell(row=start_row + i, column=j)
                                cell.value = val
                                if j != 1:  # как в других местах: рамка для всех, кроме первой колонки
                                    cell.border = data_list.thin_border
                                cell.alignment = Alignment(
                                    wrap_text=True,
                                    horizontal='center' if j in (2, 11, 12) else 'left',
                                    vertical='center',
                                )

                        # Объединения:
                        # 1-я строка ("Порядок работы") — со 2 по 13 колонку
                        ws2.merge_cells(
                            start_row=start_row,
                            start_column=2,
                            end_row=start_row,
                            end_column=13,
                        )
                        # 2-я строка ("Наименование работ") — с 3 по 11 колонку
                        ws2.merge_cells(
                            start_row=start_row + 1,
                            start_column=3,
                            end_row=start_row + 1,
                            end_column=10,
                        )

                # фиксируем индекс вставки для дальнейшей работы в таблице
                self.data_well.insert_index = ws2.max_row

                return ws2

            elif 'prs' in self.data_well.work_plan:
                if ws2 is None:
                    return ws
                self.append_podpisant_up(ws2)
                self.data_well.insert_index = ws2.max_row

                self.copy_data_excel_in_excel(
                    ws, ws2, self.data_well.cat_well_min.get_value, self.data_well.data_well_max.get_value, 1, 17,
                    ws2.max_row + 1)

                self.copy_data_excel_in_excel(
                    ws, ws2, self.data_well.data_fond_min.get_value, self.data_well.condition_of_wells.get_value, 1, 17,
                    ws2.max_row + 1)

                self.insert_events_gnvp(ws2, dict_events_gnvp, 3)

                data_x_min = self.data_well.data_x_min.get_value
                start_row_x = data_x_min if data_x_min is not None else 1
                end_row_x = (data_x_min + 2) if data_x_min is not None else 2
                self.copy_data_excel_in_excel(
                    ws, ws2, start_row_x, end_row_x, 1, 17,
                    ws2.max_row + 1,
                )

                self.data_well.insert_index = ws2.max_row

                return ws2

    def add_itog(self, ws, insert_index, work_plan, ws2=None):
        if ws.merged_cells.ranges:
            merged_cells_copy = list(ws.merged_cells.ranges)  # Создаем копию множества объединенных ячеек
            for merged_cell in merged_cells_copy:
                if merged_cell.min_row > insert_index + 5:
                    try:
                        ws.unmerge_cells(str(merged_cell))
                    except:
                        pass

        if 'prs' not in self.data_well.work_plan:
            merge_column = 11
            size_font = 12
            font_type = 'Arial'
        else:
            merge_column = 14
            size_font = 14
            font_type = 'Times New Roman'

        if work_plan not in ['gnkt_frez', 'application_pvr', 'gnkt_after_grp', 'gnkt_opz', 'gnkt_bopz']:
            itog_list = add_itog(self)
            for i in range(insert_index, len(itog_list) + insert_index):  # Добавлением итогов
                row_list = itog_list[i - insert_index]
                if 'prs' in self.data_well.work_plan:
                    row_list.insert(-8, None)
                    row_list.insert(-8, None)
                    row_list.insert(-8, None)
                j = 1
                if i < insert_index + 6:
                    for j in range(1, len(itog_list[i - insert_index]) + 1):
                        awded = row_list[j - 1]
                        ws.cell(row=i, column=j).value = row_list[j - 1]
                        if j != 1:
                            ws.cell(row=i, column=j).border = data_list.thin_border
                            ws.cell(row=i, column=j).font = Font(name=font_type, size=size_font, bold=False)

                    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=merge_column)
                    ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                   vertical='center')
                else:
                    ws.row_dimensions[insert_index + 6].height = 50
                    ws.row_dimensions[insert_index + 8].height = 50
                    for j in range(1, 13):
                        ws.cell(row=i, column=j).value = row_list[j - 1]
                        ws.cell(row=i, column=j).border = data_list.thin_border
                        ws.cell(row=i, column=j).font = Font(name=font_type, size=size_font, bold=False)
                        ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                       vertical='center')

                    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=merge_column + 1)
                    ws.cell(row=i, column=j).alignment = Alignment(wrap_text=False, horizontal='left',
                                                                   vertical='center')

            insert_index += len(add_itog(self)) + 2

        curator_s = curator_sel(self.data_well.curator, self.data_well.region)
        # print(f'куратор {curator_sel, self.data_well.curator}')
        if curator_s is False:
            return
        if 'prs' not in self.data_well.work_plan:
            podp_down = pop_down(self, self.data_well.region, curator_s)
        else:
            podp_down = pop_down(self, self.data_well.region, curator_s)[:3]

        for i in range(1 + insert_index, 1 + insert_index + len(podp_down)):
            # Добавлением подписантов внизу
            for j in range(1, 11):
                asdd = i - 1 - insert_index
                ws.cell(row=i, column=j).value = podp_down[i - 1 - insert_index][j - 1]
                ws.cell(row=i, column=j).font = Font(name=font_type, size=size_font, bold=False)

            if i in range(insert_index + 7, 1 + insert_index + 15):
                ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=False, vertical='center',
                                                               horizontal='left')
            else:
                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=False, vertical='center',
                                                               horizontal='left')

            ws.row_dimensions[insert_index + 7].height = 30
            ws.row_dimensions[insert_index + 9].height = 25

        insert_index += len(podp_down)
        aaa = ws.max_row

        ws.delete_rows(insert_index, aaa - insert_index)
        if 'prs' in self.data_well.work_plan:
            CreatePZ.copy_data_excel_in_excel(
                self.data_well.ws, ws, self.data_well.prs_copy_index.get_value, self.data_well.data_fond_min.get_value,
                1, 17,
                ws.max_row + 1)

            CreatePZ.copy_data_excel_in_excel(
                self.data_well.ws, ws, self.data_well.condition_of_wells.get_value, self.data_well.ws.max_row, 1, 17,
                ws.max_row + 1)

    @staticmethod
    def is_valid_date(date_string: str):
        try:
            datetime.strptime(date_string, "%Y-%m-%d")
            return True
        except (ValueError, TypeError):
            return False


def work_pz_excel_file(work_plan, parent, ws=None):
    if work_plan in ["krs", "dop_plan"] and parent is not None:
        data_well = getattr(parent, "data_well", parent)
        ws_sheet = ws if ws is not None else getattr(parent, "ws", None)
        if data_well is not None and ws_sheet is not None:
            open_class = CreatePZ(data_well, ws_sheet, parent)
        else:
            open_class = PzInDatabase(parent)
    else:
        open_class = PzInDatabase(parent)

    return open_class

#
